from __future__ import annotations

from pathlib import Path
from datetime import date, datetime
from io import BytesIO, StringIO
import io
import re
import textwrap
import hashlib
import csv
import math

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except Exception:
    plt = None
    MATPLOTLIB_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except Exception:
    colors = None
    TA_CENTER = TA_JUSTIFY = TA_LEFT = None
    A4 = landscape = None
    ParagraphStyle = getSampleStyleSheet = None
    inch = 72
    PageBreak = Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None
    REPORTLAB_AVAILABLE = False


st.set_page_config(
    page_title="Records List Intelligence Dashboard",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
)

CUSTOM_HEADERS = [
    "Distributor name",
    "Instrument type",
    "Installation date",
    "Customer name",
    "In Blood Bank",
    "Address",
    "ZipCode",
    "City",
    "Country",
    "World Region",
    "Commercial Region",
    "Latitude",
    "Longitude",
    "Product Line",
    "Serial number",
    "Machine Configurations",
    "Asset condition",
    "PM plan",
    "Number of tests per day",
    "Operational status",
    "Type of contract",
    "Contract duration",
    "Tag",
    "Notes",
    "PM last date",
    "PM frequency",
    "PM next date",
    "PM performed On",
    "CLIA - Adrenal function",
    "CLIA - Autoimmunity",
    "CLIA - Bone turnover",
    "CLIA - Cardiac Markers",
    "CLIA - Diabetes",
    "CLIA - EBV",
    "CLIA - Fertility",
    "CLIA - Gastroenterology",
    "CLIA - Growth",
    "CLIA - Hematology",
    "CLIA - Hepatitis and Retrovirus",
    "CLIA - Hypertension",
    "CLIA - Infectious diseases",
    "CLIA - PTH",
    "CLIA - Sepsis",
    "CLIA - Thrombosis",
    "CLIA - Thyroid",
    "CLIA - Torch",
    "CLIA - Tumor Markers",
    "CLIA - Vitamin D",
    "ELISA - Autoimmunity",
    "ELISA - Hepatitis",
    "ELISA - Infection Diseases",
    "ELISA - Murex",
    "MOLECULAR ASR",
    "MOLECULAR DAD - Simplexa C Diff Direct kit",
    "MOLECULAR DAD - Simplexa Flu A/B &RSV Direct kit",
    "MOLECULAR DAD - Simplexa Group A Strep Direct kit",
    "MOLECULAR DAD - Simplexa HSV1&2 Direct kit",
    "MOLECULAR UD - Simplexa BKV kit",
    "MOLECULAR UD - Simplexa Bordetella Universal Direct",
    "MOLECULAR UD - Simplexa C Diff Universal Direct",
    "MOLECULAR UD - Simplexa CMV kit",
    "MOLECULAR UD - Simplexa Dengue kit",
    "MOLECULAR UD - Simplexa EBV kit",
    "MOLECULAR UD - Simplexa Flu A/B & RSV kit",
    "MOLECULAR UD - Simplexa Influenza A N1N1 (2009) kit",
    "Other - specify in note field",
    "_blank",
]

ASSAY_COLS = CUSTOM_HEADERS[28:-1]

PLOT_TEMPLATE = "plotly_dark"
PLOT_BG = "rgba(12, 19, 30, 0.02)"
GRID = "rgba(170, 224, 255, 0.10)"
ACCENT = "#56d8ff"
ACCENT_2 = "#8fa8ff"
ACCENT_3 = "#59f0d0"
WARNING = "#ffb454"
DANGER = "#ff5d8f"
TEXT = "#f8fcff"
MUTED = "rgba(238,245,255,0.92)"

APP_CSS = """
<style>
:root {
    --bg-1: #050912;
    --bg-2: #08111b;
    --bg-3: #0a1622;
    --panel: rgba(28, 42, 64, 0.34);
    --panel-strong: rgba(34, 50, 76, 0.46);
    --panel-soft: rgba(255, 255, 255, 0.06);
    --glass-white: rgba(255,255,255,0.12);
    --stroke: rgba(205, 232, 255, 0.22);
    --stroke-2: rgba(255, 255, 255, 0.10);
    --txt: #f8fcff;
    --txt-strong: #ffffff;
    --txt-soft: rgba(241, 248, 255, 0.96);
    --muted: rgba(214, 228, 245, 0.82);
    --cyan: #71e1ff;
    --cyan-2: #35c8ff;
    --cyan-3: #8de8ff;
    --blue: #8ea9ff;
    --mint: #58efd1;
    --amber: #ffbe57;
    --danger: #ff6a8c;
    --shadow: rgba(0, 0, 0, 0.36);
}

html, body, [class*="css"] {
    font-family: "Inter", "Segoe UI", sans-serif;
}

.stApp {
    background:
        radial-gradient(circle at 12% 18%, rgba(133, 214, 255, 0.16), transparent 16%),
        radial-gradient(circle at 78% 10%, rgba(255,255,255,0.11), transparent 16%),
        radial-gradient(circle at 80% 78%, rgba(103, 229, 255, 0.08), transparent 18%),
        radial-gradient(circle at 30% 82%, rgba(122, 148, 255, 0.10), transparent 20%),
        linear-gradient(180deg, #040912 0%, #07101a 18%, #0a1520 44%, #08111a 72%, #060b12 100%);
    background-attachment: fixed;
    color: var(--txt);
}

.stApp::before {
    content: "";
    position: fixed;
    inset: 0;
    pointer-events: none;
    background:
        linear-gradient(180deg, rgba(255,255,255,0.025), transparent 18%, transparent 82%, rgba(255,255,255,0.02)),
        radial-gradient(circle at 50% 0%, rgba(255,255,255,0.08), transparent 18%),
        radial-gradient(circle at 55% 45%, rgba(113,225,255,0.05), transparent 26%);
    z-index: 0;
}

.block-container {
    position: relative;
    z-index: 1;
    padding-top: 0.85rem;
    padding-bottom: 2rem;
    max-width: 98rem;
}

section[data-testid="stSidebar"] {
    background:
        linear-gradient(180deg, rgba(28, 42, 64, 0.28) 0%, rgba(20, 32, 50, 0.36) 100%);
    border-right: 1px solid rgba(255,255,255,0.08);
    backdrop-filter: blur(22px);
    -webkit-backdrop-filter: blur(22px);
    box-shadow:
        inset -1px 0 0 rgba(255,255,255,0.05),
        10px 0 28px rgba(0, 0, 0, 0.18);
}

section[data-testid="stSidebar"] > div {
    background: transparent;
}

section[data-testid="stSidebar"] .stMarkdown,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div {
    color: var(--txt-soft) !important;
    text-shadow: 0 0 6px rgba(74, 203, 255, 0.08);
}

.hero {
    position: relative;
    overflow: hidden;
    padding: 1.2rem 1.45rem;
    border-radius: 28px;
    border: 1px solid rgba(255,255,255,0.16);
    background:
        linear-gradient(180deg, rgba(198, 223, 255, 0.13) 0%, rgba(72, 101, 145, 0.10) 18%, rgba(25, 38, 58, 0.18) 100%);
    backdrop-filter: blur(24px);
    -webkit-backdrop-filter: blur(24px);
    box-shadow:
        0 16px 40px rgba(0, 0, 0, 0.24),
        inset 0 1px 0 rgba(255,255,255,0.22),
        inset 0 -1px 0 rgba(255,255,255,0.03);
    margin-bottom: 1rem;
}

.hero::before {
    content: "";
    position: absolute;
    inset: 0;
    pointer-events: none;
    background:
        linear-gradient(180deg, rgba(255,255,255,0.10), transparent 28%),
        radial-gradient(circle at 18% 12%, rgba(255,255,255,0.18), transparent 16%),
        radial-gradient(circle at 72% 12%, rgba(113,225,255,0.08), transparent 20%);
}

.hero h1 {
    margin: 0;
    font-size: 2.2rem;
    line-height: 1.05;
    letter-spacing: 0.02em;
    color: #f9fdff !important;
    text-shadow:
        0 0 6px rgba(146, 235, 255, 0.34),
        0 0 18px rgba(58, 199, 255, 0.18);
}

.code-stamp {
    font-size: 0.58rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.02em !important;
    color: rgba(235,245,255,0.62) !important;
    margin-left: 0.55rem;
    vertical-align: middle;
    text-shadow: none !important;
    white-space: nowrap;
}

.hero p {
    margin: 0.45rem 0 0 0;
    color: rgba(239,247,255,0.88) !important;
    font-size: 1rem;
    text-shadow: 0 0 8px rgba(83, 205, 255, 0.12);
}

.badge-row {
    display: flex;
    flex-wrap: wrap;
    gap: 0.55rem;
    margin-top: 0.85rem;
}

.badge {
    padding: 0.35rem 0.75rem;
    border-radius: 999px;
    font-size: 0.77rem;
    color: #f8fcff;
    background: rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.14);
    box-shadow:
        inset 0 1px 0 rgba(255,255,255,0.14),
        0 0 12px rgba(83,205,255,0.08);
    backdrop-filter: blur(16px);
    -webkit-backdrop-filter: blur(16px);
}

.metric-shell,
div[data-testid="stMetric"] {
    position: relative;
    overflow: hidden;
    border-radius: 22px !important;
    background:
        linear-gradient(180deg, rgba(181, 214, 255, 0.12) 0%, rgba(25, 38, 58, 0.22) 20%, rgba(18, 28, 44, 0.28) 100%) !important;
    border: 1px solid rgba(255,255,255,0.14) !important;
    backdrop-filter: blur(22px);
    -webkit-backdrop-filter: blur(22px);
    box-shadow:
        0 12px 28px rgba(0,0,0,0.22),
        inset 0 1px 0 rgba(255,255,255,0.18);
}

.metric-shell {
    padding: 0.95rem 1rem;
    min-height: 118px;
}

.metric-shell::before,
div[data-testid="stMetric"]::before {
    content: "";
    position: absolute;
    inset: 0;
    pointer-events: none;
    background:
        linear-gradient(180deg, rgba(255,255,255,0.08), transparent 24%),
        radial-gradient(circle at 18% 0%, rgba(255,255,255,0.12), transparent 18%);
}

.metric-label {
    font-size: 0.78rem;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: rgba(229,243,255,0.78) !important;
    text-shadow: 0 0 8px rgba(77, 204, 255, 0.10);
}

.metric-value {
    font-size: 1.95rem;
    font-weight: 700;
    margin-top: 0.22rem;
    color: #ffffff !important;
    text-shadow:
        0 0 8px rgba(115,228,255,0.24),
        0 0 18px rgba(53,200,255,0.14);
}

.metric-sub {
    margin-top: 0.2rem;
    font-size: 0.86rem;
    color: rgba(235,245,255,0.88) !important;
}

.stTabs [data-baseweb="tab-list"] {
    gap: 0.55rem;
    padding: 0.28rem;
    background: rgba(255,255,255,0.04);
    border-radius: 18px;
    border: 1px solid rgba(255,255,255,0.10);
    backdrop-filter: blur(14px);
    -webkit-backdrop-filter: blur(14px);
}

.stTabs [data-baseweb="tab"] {
    border-radius: 14px;
    padding: 0.52rem 0.92rem;
    background: rgba(255,255,255,0.03);
    color: rgba(238,247,255,0.88);
    border: 1px solid transparent;
    transition: all 0.2s ease;
    text-shadow: 0 0 8px rgba(53,200,255,0.10);
}

.stTabs [data-baseweb="tab"]:hover {
    background: rgba(255,255,255,0.07);
    border-color: rgba(255,255,255,0.10);
}

.stTabs [aria-selected="true"] {
    background:
        linear-gradient(180deg, rgba(129, 212, 255, 0.18), rgba(54, 93, 150, 0.14)) !important;
    border: 1px solid rgba(119, 221, 255, 0.36) !important;
    box-shadow:
        inset 0 1px 0 rgba(255,255,255,0.18),
        0 0 18px rgba(53,200,255,0.16) !important;
    color: #ffffff !important;
    text-shadow:
        0 0 8px rgba(136,234,255,0.34),
        0 0 16px rgba(53,200,255,0.18);
}

div[data-testid="stPlotlyChart"],
div[data-testid="stDataFrame"],
div[data-testid="stTable"],
div[data-testid="stExpander"],
div[data-testid="stForm"] {
    border-radius: 26px;
    overflow: hidden;
    background:
        linear-gradient(180deg, rgba(193, 221, 255, 0.10) 0%, rgba(24, 36, 54, 0.18) 18%, rgba(17, 27, 42, 0.24) 100%);
    border: 1px solid rgba(255,255,255,0.12);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    box-shadow:
        0 12px 30px rgba(0,0,0,0.20),
        inset 0 1px 0 rgba(255,255,255,0.12);
}

div[data-testid="stPlotlyChart"] {
    padding: 0.24rem;
}

div[data-testid="stDataFrame"] table,
div[data-testid="stTable"] table {
    color: #f8fcff !important;
}

.stButton > button,
.stDownloadButton > button {
    border-radius: 16px;
    border: 1px solid rgba(137, 228, 255, 0.34);
    background:
        linear-gradient(180deg, rgba(114, 214, 255, 0.20), rgba(52, 94, 145, 0.18));
    color: #ffffff !important;
    font-weight: 600;
    box-shadow:
        inset 0 1px 0 rgba(255,255,255,0.18),
        0 0 18px rgba(53,200,255,0.12);
    text-shadow: 0 0 8px rgba(110, 227, 255, 0.22);
}

.stButton > button:hover,
.stDownloadButton > button:hover {
    border-color: rgba(160, 237, 255, 0.44);
    box-shadow:
        inset 0 1px 0 rgba(255,255,255,0.22),
        0 0 22px rgba(53,200,255,0.18);
    transform: translateY(-1px);
}

.stTextInput input,
.stNumberInput input,
.stDateInput input,
.stTextArea textarea,
.stSelectbox div[data-baseweb="select"] > div,
.stMultiSelect div[data-baseweb="select"] > div {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 15px !important;
    color: #ffffff !important;
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.08);
}

.stTextInput input::placeholder,
.stTextArea textarea::placeholder {
    color: rgba(235,245,255,0.54) !important;
}

h1, h2, h3, h4, h5, h6,
div[data-testid="stMarkdownContainer"] h1,
div[data-testid="stMarkdownContainer"] h2,
div[data-testid="stMarkdownContainer"] h3,
div[data-testid="stMarkdownContainer"] h4 {
    color: #fbfdff !important;
    letter-spacing: 0.01em;
    text-shadow:
        0 0 8px rgba(140, 235, 255, 0.30),
        0 0 20px rgba(53,200,255,0.14);
}

h2 {
    font-size: 1.9rem !important;
    font-weight: 700 !important;
}

h3 {
    font-size: 1.35rem !important;
    font-weight: 700 !important;
}

p, label, span, div {
    color: var(--txt-soft);
}

div[data-testid="stCaptionContainer"],
.small-note,
.stCaption,
small {
    color: rgba(231,243,255,0.84) !important;
    text-shadow: 0 0 8px rgba(53,200,255,0.08);
}

hr {
    border: none;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.18), transparent);
}

[data-testid="stSidebar"] .stFileUploader,
[data-testid="stSidebar"] .stSelectbox,
[data-testid="stSidebar"] .stMultiSelect {
    background: rgba(255,255,255,0.04);
    border-radius: 18px;
}

[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h1,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 {
    color: #fafdff !important;
    text-shadow:
        0 0 8px rgba(140,235,255,0.30),
        0 0 18px rgba(53,200,255,0.16);
}
</style>
"""
st.markdown(APP_CSS, unsafe_allow_html=True)


def glow_layout(fig: go.Figure, height: int = 420, title_size: int = 18) -> go.Figure:
    fig.update_layout(
        template=PLOT_TEMPLATE,
        height=height,
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        margin=dict(l=18, r=18, t=78, b=18),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            bgcolor="rgba(16, 28, 46, 0.42)",
            bordercolor="rgba(124,221,255,0.26)",
            borderwidth=1,
            font=dict(color="#f8fbff", size=12),
        ),
        font=dict(color=TEXT),
        title_font=dict(size=title_size, color=TEXT),
        title=dict(x=0.02, xanchor="left"),
        hovermode="closest",
        hoverlabel=dict(
            bgcolor="rgba(13, 24, 42, 0.96)",
            bordercolor="rgba(255,255,255,0.22)",
            font=dict(color=TEXT),
        ),
    )
    fig.update_xaxes(
        showgrid=True,
        gridcolor=GRID,
        zeroline=False,
        automargin=True,
        linecolor="rgba(255,255,255,0.12)",
        tickfont=dict(color=TEXT),
        title_font=dict(color=TEXT),
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor=GRID,
        zeroline=False,
        automargin=True,
        linecolor="rgba(255,255,255,0.12)",
        tickfont=dict(color=TEXT),
        title_font=dict(color=TEXT),
    )
    return fig



def compress_value_distribution(series: pd.Series, max_slices: int = 4) -> pd.DataFrame:
    work = (
        series.fillna("No informado")
        .astype(str)
        .str.strip()
        .replace("", "No informado")
        .value_counts()
        .reset_index()
    )
    if work.empty:
        return pd.DataFrame(columns=["Label", "Count"])
    work.columns = ["Label", "Count"]
    if len(work) > max_slices:
        top = work.head(max_slices).copy()
        other_count = int(work.iloc[max_slices:]["Count"].sum())
        if other_count > 0:
            top = pd.concat([top, pd.DataFrame([{"Label": "Otros", "Count": other_count}])], ignore_index=True)
        work = top
    return work


def build_config_donut(field_name: str, series: pd.Series, total_assets: int) -> go.Figure:
    dist = compress_value_distribution(series, max_slices=4)
    fig = go.Figure()

    if dist.empty:
        fig.add_annotation(
            text="Sin datos",
            x=0.5,
            y=0.5,
            xref="paper",
            yref="paper",
            showarrow=False,
            font=dict(size=15, color=TEXT),
        )
        fig.update_layout(title=f"{field_name}")
        return glow_layout(fig, 340, 15)

    palette = [ACCENT, ACCENT_2, ACCENT_3, WARNING, "rgba(255,255,255,0.32)"]
    fig.add_trace(
        go.Pie(
            labels=dist["Label"],
            values=dist["Count"],
            customdata=np.column_stack([dist["Label"], dist["Count"]]),
            hole=0.68,
            sort=False,
            marker=dict(colors=palette[:len(dist)], line=dict(color="rgba(255,255,255,0.18)", width=1.2)),
            textinfo="percent",
            textfont=dict(color="#ffffff", size=13),
            hovertemplate="Valor: %{customdata[0]}<br>Equipos: %{customdata[1]}<br>Participación: %{percent}<extra></extra>",
        )
    )
    fig.add_annotation(
        text=f"<b>{total_assets:,}</b><br><span style='font-size:11px'>equipos</span>",
        x=0.5,
        y=0.5,
        xref="paper",
        yref="paper",
        showarrow=False,
        font=dict(color='#ffffff', size=18),
    )
    fig.update_layout(
        title=field_name,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5, bgcolor="rgba(14,26,42,0.36)", bordercolor="rgba(124,221,255,0.22)", borderwidth=1, font=dict(color="#f8fbff", size=11)),
    )
    return glow_layout(fig, 340, 15)


def compute_mapbox_center_zoom(df: pd.DataFrame, lat_col: str = "Latitude", lon_col: str = "Longitude") -> tuple[dict, float]:
    geo = df.dropna(subset=[lat_col, lon_col]).copy()
    if geo.empty:
        return {"lat": 0.0, "lon": 0.0}, 1.0

    lats = pd.to_numeric(geo[lat_col], errors="coerce").dropna()
    lons = pd.to_numeric(geo[lon_col], errors="coerce").dropna()
    if lats.empty or lons.empty:
        return {"lat": 0.0, "lon": 0.0}, 1.0

    min_lat, max_lat = float(lats.min()), float(lats.max())
    min_lon, max_lon = float(lons.min()), float(lons.max())
    center = {"lat": (min_lat + max_lat) / 2, "lon": (min_lon + max_lon) / 2}

    lat_span = max(max_lat - min_lat, 0.01)
    lon_span = max(max_lon - min_lon, 0.01)
    max_span = max(lat_span, lon_span)

    if len(geo) == 1:
        zoom = 9.5
    elif max_span <= 0.05:
        zoom = 9.0
    elif max_span <= 0.12:
        zoom = 8.2
    elif max_span <= 0.25:
        zoom = 7.2
    elif max_span <= 0.5:
        zoom = 6.3
    elif max_span <= 1.0:
        zoom = 5.5
    elif max_span <= 2.0:
        zoom = 4.7
    elif max_span <= 4.0:
        zoom = 4.0
    elif max_span <= 8.0:
        zoom = 3.2
    elif max_span <= 16.0:
        zoom = 2.6
    elif max_span <= 35.0:
        zoom = 2.0
    elif max_span <= 70.0:
        zoom = 1.45
    else:
        zoom = 1.0

    return center, zoom


EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Paleta ejecutiva para los informes Excel. Se mantiene separada de la paleta
# oscura del dashboard Streamlit para privilegiar legibilidad al imprimir,
# presentar y trabajar directamente en Microsoft Excel.
EXCEL_THEME = {
    "navy": "17324D",          # títulos y encabezados principales
    "royal": "2F6BFF",         # acción / dato principal
    "teal": "00A6A6",          # cobertura y distribución
    "green": "2E9D62",         # condición favorable / rutina
    "amber": "F2B134",         # atención / próximos vencimientos
    "coral": "E85D75",         # riesgo / cero procesamiento
    "purple": "7A5AF8",        # categorías complementarias
    "slate": "5F6B7A",         # información neutral
    "canvas": "F4F7FB",        # fondo general claro
    "panel": "FFFFFF",         # tarjetas y gráficos
    "panel_alt": "EAF1F8",     # panel secundario
    "text": "243447",          # texto principal
    "muted": "61758A",         # texto secundario
    "border": "D7E1EA",        # divisores y bordes
    "white": "FFFFFF",
}

EXCEL_CHART_PALETTE = [
    EXCEL_THEME["royal"],
    EXCEL_THEME["teal"],
    EXCEL_THEME["purple"],
    EXCEL_THEME["green"],
    EXCEL_THEME["amber"],
    EXCEL_THEME["coral"],
    "4B8BBE",
    EXCEL_THEME["slate"],
]


def _excel_safe_sheet_name(sheet_name: str) -> str:
    safe_name = re.sub(r"[\\/*?:\[\]]", "_", str(sheet_name)).strip()[:31]
    return safe_name or "Sheet1"


def _excel_safe_cell_value(value):
    """Return a value that openpyxl can safely write to a worksheet cell.

    Some dashboard DataFrames may contain pandas extension values, dictionaries,
    lists, tuples, sets, intervals or other Python objects created by Plotly /
    Streamlit / pandas processing. openpyxl cannot write those objects directly
    and raises errors such as ``Cannot convert {0: ...} to Excel``. This
    sanitizer keeps numbers/dates as native Excel-compatible values and converts
    complex objects to readable text before writing.
    """
    import json
    from datetime import date as _date, datetime as _datetime

    try:
        if value is None or pd.isna(value):
            return ""
    except Exception:
        # pd.isna(dict/list) can return a non-scalar result. Continue with the
        # explicit object handling below.
        pass

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        try:
            if value.tzinfo is not None:
                value = value.tz_convert(None)
        except Exception:
            pass
        return value.to_pydatetime()

    if isinstance(value, (_datetime, _date)):
        return value

    if isinstance(value, pd.Timedelta):
        return str(value)

    if isinstance(value, np.generic):
        try:
            native = value.item()
            if isinstance(native, float) and (math.isnan(native) or math.isinf(native)):
                return ""
            return native
        except Exception:
            return str(value)

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return value

    if isinstance(value, (int, bool, str)):
        if isinstance(value, str):
            # Remove control characters not accepted by Excel XML.
            return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
        return value

    if isinstance(value, (dict, list, tuple, set)):
        try:
            if isinstance(value, set):
                value = sorted(list(value), key=lambda x: str(x))
            text_value = json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            text_value = str(value)
        return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text_value)

    text_value = str(value)
    return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text_value)


def _excel_clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    clean_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    if clean_df.empty:
        return clean_df

    # Sanitize every column, not only object columns. Pandas categorical /
    # extension columns can still contain values that openpyxl cannot bind.
    for col in clean_df.columns:
        if str(col).startswith("FLAG::"):
            continue
        clean_df[col] = clean_df[col].map(_excel_safe_cell_value)
    return clean_df


def _preferred_export_columns(df: pd.DataFrame) -> list[str]:
    preferred = [
        "Commercial Region", "Country", "Distributor name", "Customer name", "City", "Address",
        "Instrument type", "Serial number", "Installation date", "Age (years)",
        "Operational status grouped", "Operational status", "Asset condition", "Type of contract",
        "Contract duration", "In Blood Bank", "Blood Bank Flag", "Product Line",
        "Number of tests per day", "PM plan", "PM frequency", "PM last date", "PM next date", "PM performed On",
        "Operating System", "Operating System Raw", "Machine config fields populated", "Data completeness %",
        "Latitude", "Longitude", "Machine Configurations", "Tag", "Notes",
    ]
    manufacturing = [
        "Manufacturing Date", "Manufacturing year", "Manufacturing age (years)",
        "Manufacturing age bucket", "Manufacturing Source", "Manufacturing Sheet",
        "Manufacturing Product", "Manufacturing matched", "Manufacturing date conflict",
    ]
    cfg_cols = sorted([c for c in df.columns if str(c).startswith("CFG::")])
    hidden_prefixes = ("FLAG::",)
    ordered = [c for c in preferred + manufacturing if c in df.columns]
    ordered += [c for c in cfg_cols if c not in ordered]
    ordered += [c for c in df.columns if c not in ordered and not any(str(c).startswith(p) for p in hidden_prefixes)]
    return ordered



def _excel_style_workbook(wb) -> None:
    """Normaliza vistas, zoom y colores de pestaña sin tocar los datos."""
    from openpyxl.worksheet.views import Selection
    from openpyxl.utils.cell import coordinate_to_tuple

    tab_colors = {
        "00_Dashboard": EXCEL_THEME["royal"],
        "00_Resumen": EXCEL_THEME["navy"],
        "01_Filtros": EXCEL_THEME["teal"],
        "02_Datos_clave": EXCEL_THEME["green"],
        "03_Datos_completos": EXCEL_THEME["slate"],
        "04_Base_instalada": EXCEL_THEME["royal"],
        "05_Modelo_estado": EXCEL_THEME["green"],
        "06_Machine_config": EXCEL_THEME["purple"],
        "07_OS_PM": EXCEL_THEME["amber"],
        "08_Fabricacion": EXCEL_THEME["slate"],
        "09_Carstock": EXCEL_THEME["coral"],
    }

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90 if ws.title != "00_Dashboard" else 85
        if ws.title in tab_colors:
            ws.sheet_properties.tabColor = tab_colors[ws.title]

        freeze = ws.freeze_panes
        if freeze:
            try:
                row_idx, col_idx = coordinate_to_tuple(str(freeze))
            except Exception:
                row_idx, col_idx = 1, 1

            if row_idx > 1 and col_idx > 1:
                pane = "bottomRight"
            elif row_idx > 1:
                pane = "bottomLeft"
            elif col_idx > 1:
                pane = "topRight"
            else:
                pane = None

            if pane:
                ws.sheet_view.selection = [Selection(pane=pane, activeCell="A1", sqref="A1")]
            else:
                ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        else:
            ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]


def _excel_autofit_and_style(ws, start_row: int, start_col: int, n_rows: int, n_cols: int, title_row: int | None = None) -> None:
    """Aplica una presentación clara, ejecutiva y consistente a cada tabla."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    title_fill = PatternFill("solid", fgColor=EXCEL_THEME["royal"])
    body_fill = PatternFill("solid", fgColor=EXCEL_THEME["panel"])
    stripe_fill = PatternFill("solid", fgColor=EXCEL_THEME["canvas"])
    header_font = Font(color=EXCEL_THEME["white"], bold=True, name="Aptos", size=10)
    title_font = Font(color=EXCEL_THEME["white"], bold=True, size=13, name="Aptos Display")
    body_font = Font(color=EXCEL_THEME["text"], name="Aptos", size=10)
    thin = Side(style="thin", color=EXCEL_THEME["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if n_cols <= 0:
        return

    if title_row is not None:
        for col_idx in range(start_col, start_col + n_cols):
            cell = ws.cell(title_row, col_idx)
            cell.fill = title_fill
            cell.border = border
        title_cell = ws.cell(title_row, start_col)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[title_row].height = 25

    header_row = start_row
    ws.row_dimensions[header_row].height = 30
    for col_idx in range(start_col, start_col + n_cols):
        cell = ws.cell(header_row, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    max_row = start_row + max(n_rows, 1)
    for row_idx in range(start_row + 1, max_row + 1):
        row_fill = stripe_fill if (row_idx - start_row) % 2 == 0 else body_fill
        for col_idx in range(start_col, start_col + n_cols):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = row_fill
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for col_idx in range(start_col, start_col + n_cols):
        letter = get_column_letter(col_idx)
        values = []
        for row_idx in range(max(1, header_row), min(ws.max_row, max_row) + 1):
            value = ws.cell(row_idx, col_idx).value
            values.append(len(str(value)) if value is not None else 0)
        width = min(max(max(values or [12]) + 2, 12), 42)
        ws.column_dimensions[letter].width = width

    if not ws.freeze_panes and start_col == 1:
        ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    ws.auto_filter.ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(max_row, start_col + n_cols - 1).coordinate}"

def _excel_write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, title: str | None = None) -> tuple[int, int, int, int]:
    from openpyxl.utils.dataframe import dataframe_to_rows

    clean_df = _excel_clean_dataframe(df)
    title_row = None
    if title:
        title_row = start_row
        ws.cell(start_row, start_col, title)
        start_row += 1

    for r_offset, row in enumerate(dataframe_to_rows(clean_df, index=False, header=True), start=0):
        for c_offset, value in enumerate(row, start=0):
            ws.cell(start_row + r_offset, start_col + c_offset, _excel_safe_cell_value(value))

    n_rows = len(clean_df)
    n_cols = max(len(clean_df.columns), 1)
    _excel_autofit_and_style(ws, start_row, start_col, n_rows, n_cols, title_row=title_row)
    return start_row, start_col, n_rows, n_cols



def _excel_add_bar_chart(ws, data_start_row: int, data_start_col: int, n_rows: int, n_cols: int, title: str, anchor: str, stacked: bool = False) -> None:
    if n_rows <= 0 or n_cols < 2:
        return
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Categoría"
    chart.x_axis.title = "Cantidad"
    chart.height = 8
    chart.width = 16
    chart.gapWidth = 55
    if stacked:
        chart.grouping = "stacked"
        chart.overlap = 100

    data = Reference(ws, min_col=data_start_col + 1, max_col=data_start_col + n_cols - 1, min_row=data_start_row, max_row=data_start_row + n_rows)
    cats = Reference(ws, min_col=data_start_col, min_row=data_start_row + 1, max_row=data_start_row + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    _excel_apply_series_palette(chart, EXCEL_CHART_PALETTE)
    if not stacked and chart.series:
        points = []
        for idx in range(n_rows):
            point = DataPoint(idx=idx)
            point.graphicalProperties.solidFill = EXCEL_CHART_PALETTE[idx % len(EXCEL_CHART_PALETTE)]
            points.append(point)
        chart.series[0].data_points = points
        chart.legend = None

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    ws.add_chart(chart, anchor)


def _excel_add_pie_chart(ws, data_start_row: int, data_start_col: int, n_rows: int, title: str, anchor: str) -> None:
    if n_rows <= 0:
        return
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint

    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.height = 8
    chart.width = 12
    chart.legend.position = "b"
    labels = Reference(ws, min_col=data_start_col, min_row=data_start_row + 1, max_row=data_start_row + n_rows)
    data = Reference(ws, min_col=data_start_col + 1, min_row=data_start_row, max_row=data_start_row + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    if chart.series:
        points = []
        for idx in range(n_rows):
            point = DataPoint(idx=idx)
            point.graphicalProperties.solidFill = EXCEL_CHART_PALETTE[idx % len(EXCEL_CHART_PALETTE)]
            points.append(point)
        chart.series[0].data_points = points

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showLeaderLines = True
    ws.add_chart(chart, anchor)

def _excel_value_counts_label(value) -> str:
    """Etiqueta segura para conteos en Excel.

    Evita errores con dtypes pandas nullable, por ejemplo Int64/Float64/Categorical,
    donde fillna("No informado") puede fallar porque el dtype no acepta strings.
    """
    try:
        missing = pd.isna(value)
        if isinstance(missing, (bool, np.bool_)) and bool(missing):
            return "No informado"
    except Exception:
        pass

    if isinstance(value, (pd.Timestamp, datetime, date)):
        try:
            if pd.isna(value):
                return "No informado"
        except Exception:
            pass
        try:
            return pd.to_datetime(value).strftime("%Y-%m-%d")
        except Exception:
            return str(value).strip() or "No informado"

    if isinstance(value, (list, tuple, set)):
        text = ", ".join(str(v) for v in value)
    elif isinstance(value, dict):
        text = "; ".join(f"{k}: {v}" for k, v in value.items())
    else:
        text = str(value)

    text = text.strip()
    if not text or text.lower() in {"nan", "none", "nat", "<na>"}:
        return "No informado"
    return text


def _excel_value_counts_df(df: pd.DataFrame, column: str, label_name: str = "Categoría", top_n: int | None = None) -> pd.DataFrame:
    if df is None or df.empty or column not in df.columns:
        return pd.DataFrame(columns=[label_name, "Cantidad"])

    series = df[column]
    if isinstance(series, pd.DataFrame):
        series = series.bfill(axis=1).iloc[:, 0]

    labels = series.astype("object").map(_excel_value_counts_label)
    counts = labels.value_counts(dropna=False).reset_index()
    counts.columns = [label_name, "Cantidad"]
    counts[label_name] = counts[label_name].map(_excel_value_counts_label)
    counts["Cantidad"] = pd.to_numeric(counts["Cantidad"], errors="coerce").fillna(0).astype(int)
    if top_n:
        counts = counts.head(top_n)
    return counts


def _excel_prepare_model_status_matrix(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    if "Instrument type" not in df.columns or "Operational status grouped" not in df.columns:
        return pd.DataFrame()
    work = df.copy()
    work["Instrument type"] = work["Instrument type"].fillna("No informado").astype(str)
    work["Operational status grouped"] = work["Operational status grouped"].fillna("No informado").astype(str)
    pivot = pd.pivot_table(work, index="Instrument type", columns="Operational status grouped", values="Serial number", aggfunc="count", fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    total = pivot.pop("Total")
    pivot.insert(0, "Modelo", pivot.index)
    pivot["Total"] = total.values
    return pivot.reset_index(drop=True)


def _excel_prepare_config_summary(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    cfg_cols = [c for c in df.columns if str(c).startswith("CFG::")]
    coverage_rows = []
    value_rows = []
    for col in cfg_cols:
        field_name = str(col).replace("CFG::", "")
        series = df[col].dropna().astype(str).str.strip()
        series = series[series.ne("")]
        if series.empty:
            continue
        coverage_rows.append({"Campo": field_name, "Equipos con dato": int(series.shape[0]), "% del filtro": round(series.shape[0] * 100 / max(len(df), 1), 1)})
        counts = series.value_counts().head(10).reset_index()
        counts.columns = ["Valor", "Cantidad"]
        for _, row in counts.iterrows():
            value_rows.append({"Campo": field_name, "Valor": row["Valor"], "Cantidad": int(row["Cantidad"])})
    return pd.DataFrame(coverage_rows).sort_values("Equipos con dato", ascending=False) if coverage_rows else pd.DataFrame(columns=["Campo", "Equipos con dato", "% del filtro"]), pd.DataFrame(value_rows)



def _excel_add_readme(ws, filter_summary: dict[str, str], total_rows: int, active_tab: str, source_label_value: str = "") -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = EXCEL_THEME["navy"]
    ws.sheet_view.zoomScale = 95

    navy_fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    royal_fill = PatternFill("solid", fgColor=EXCEL_THEME["royal"])
    canvas_fill = PatternFill("solid", fgColor=EXCEL_THEME["canvas"])
    white_fill = PatternFill("solid", fgColor=EXCEL_THEME["panel"])
    label_fill = PatternFill("solid", fgColor=EXCEL_THEME["panel_alt"])
    thin = Side(style="thin", color=EXCEL_THEME["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 40):
        for col in range(1, 9):
            ws.cell(row, col).fill = canvas_fill
            ws.cell(row, col).font = Font(color=EXCEL_THEME["text"], name="Aptos", size=10)

    ws.merge_cells("A1:H2")
    ws["A1"] = "RECORDS LIST INTELLIGENCE DASHBOARD · RESUMEN DE EXPORTACIÓN"
    ws["A1"].font = Font(bold=True, size=17, color=EXCEL_THEME["white"], name="Aptos Display")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    metadata = [
        ("A4", "Código / versión", f"{CODE_CREATED_AT} · {CODE_VERSION_LABEL}"),
        ("A5", "Build", PARSER_VERSION),
        ("A6", "Fecha de exportación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("A7", "Pestaña activa", active_tab),
        ("A8", "Fuente activa", source_label_value),
        ("A9", "Registros incluidos", f"{total_rows:,}"),
    ]
    for label_cell, label, value in metadata:
        row = ws[label_cell].row
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 1).fill = label_fill
        ws.cell(row, 1).font = Font(color=EXCEL_THEME["navy"], bold=True, name="Aptos")
        ws.cell(row, 2).fill = white_fill
        ws.cell(row, 2).font = Font(color=EXCEL_THEME["text"], name="Aptos")
        ws.cell(row, 1).border = border
        ws.cell(row, 2).border = border
        ws.cell(row, 1).alignment = Alignment(vertical="center")
        ws.cell(row, 2).alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells("A11:B11")
    ws["A11"] = "FILTROS APLICADOS"
    ws["A11"].font = Font(bold=True, color=EXCEL_THEME["white"], name="Aptos Display", size=12)
    ws["A11"].fill = royal_fill
    ws["A11"].alignment = Alignment(horizontal="left", vertical="center")

    row = 12
    for key, value in (filter_summary or {}).items():
        ws.cell(row, 1, key)
        ws.cell(row, 2, value)
        ws.cell(row, 1).fill = label_fill
        ws.cell(row, 1).font = Font(color=EXCEL_THEME["navy"], bold=True, name="Aptos")
        ws.cell(row, 2).fill = white_fill
        ws.cell(row, 2).font = Font(color=EXCEL_THEME["text"], name="Aptos")
        ws.cell(row, 1).border = border
        ws.cell(row, 2).border = border
        ws.cell(row, 2).alignment = Alignment(wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 58
    for column in "CDEFGH":
        ws.column_dimensions[column].width = 3
    ws.freeze_panes = "A4"

def dataframe_to_excel_bytes(sheet_map: dict[str, pd.DataFrame]) -> bytes:
    """Exporta múltiples tablas a un Excel ordenado y con formato básico.

    Se mantiene como función genérica para los exports de carstock existentes,
    pero ahora entrega un archivo más claro que un volcado plano.
    """
    from openpyxl import Workbook

    output = BytesIO()
    wb = Workbook()
    first_sheet = True
    for sheet_name, df in sheet_map.items():
        safe_name = _excel_safe_sheet_name(sheet_name)
        if first_sheet:
            ws = wb.active
            ws.title = safe_name
            first_sheet = False
        else:
            ws = wb.create_sheet(safe_name)
        _excel_write_df(ws, df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(), start_row=1, start_col=1, title=safe_name)
    _excel_style_workbook(wb)
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def _excel_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Devuelve una copia con encabezados válidos y únicos para Excel Tables."""
    work = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    seen: dict[str, int] = {}
    unique_columns: list[str] = []
    for raw_column in work.columns:
        base = str(raw_column).strip() or "Column"
        occurrence = seen.get(base, 0) + 1
        seen[base] = occurrence
        unique_columns.append(base if occurrence == 1 else f"{base}_{occurrence}")
    work.columns = unique_columns
    return work



def _excel_add_native_table(
    ws,
    start_row: int,
    start_col: int,
    n_rows: int,
    n_cols: int,
    table_name: str = "DashboardData",
) -> str:
    """Convierte el rango de datos en una tabla nativa con filtros editables.

    Devuelve el nombre final de la tabla para que el dashboard dinámico pueda
    construir fórmulas estructuradas sin depender de coordenadas frágiles.
    """
    if n_rows <= 0 or n_cols <= 0:
        return ""

    from openpyxl.worksheet.table import Table, TableStyleInfo

    safe_table_name = re.sub(r"[^A-Za-z0-9_]", "_", str(table_name)).strip("_") or "DashboardData"
    if safe_table_name[0].isdigit():
        safe_table_name = f"T_{safe_table_name}"

    end_row = start_row + n_rows
    end_col = start_col + n_cols - 1
    table_ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"
    table = Table(displayName=safe_table_name[:250], ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return table.displayName



EXCEL_INTERACTIVE_HELPER_COLUMNS = (
    "__RowMarker",
    "__Visible",
    "__ProcessingBucket",
    "__BloodBucket",
)


def _excel_table_column_ref(table_name: str, column_name: str) -> str:
    """Crea una referencia estructurada segura a una columna de Excel Table."""
    safe_column = str(column_name).replace("]", "]]" )
    return f"{table_name}[{safe_column}]"


def _excel_add_interactive_helper_columns(export_df: pd.DataFrame) -> pd.DataFrame:
    """Añade columnas ocultas que permiten recalcular el dashboard al filtrar.

    No modifica ninguna columna funcional de la base. Los campos auxiliares se
    escriben al final de la tabla, se ocultan en Excel y solo alimentan fórmulas.
    """
    work = export_df.copy() if isinstance(export_df, pd.DataFrame) else pd.DataFrame()
    work = work.drop(columns=[c for c in EXCEL_INTERACTIVE_HELPER_COLUMNS if c in work.columns], errors="ignore")

    tests = pd.to_numeric(
        work.get("Number of tests per day", pd.Series(index=work.index, dtype=float)),
        errors="coerce",
    ).fillna(0)
    work["__ProcessingBucket"] = np.where(tests.gt(0), "> 0 tests/día", "0 tests/día")

    if "Blood Bank Flag" in work.columns:
        blood_flags = work["Blood Bank Flag"].fillna(False).astype(bool)
    elif "In Blood Bank" in work.columns:
        blood_flags = work["In Blood Bank"].map(is_blood_bank_yes).fillna(False).astype(bool)
    else:
        blood_flags = pd.Series(False, index=work.index, dtype=bool)
    work["__BloodBucket"] = np.where(blood_flags, "Banco de sangre", "Laboratorio")

    work["__RowMarker"] = 1
    work["__Visible"] = 1

    # Deja las columnas técnicas al final para que las columnas de negocio
    # mantengan exactamente su orden actual.
    business_columns = [c for c in work.columns if c not in EXCEL_INTERACTIVE_HELPER_COLUMNS]
    return work[business_columns + list(EXCEL_INTERACTIVE_HELPER_COLUMNS)]


def _excel_apply_visibility_formulas(
    ws,
    header_row: int,
    start_col: int,
    n_rows: int,
    columns: list[str],
) -> None:
    """Convierte __Visible en un indicador 1/0 sensible al AutoFilter.

    SUBTOTAL + OFFSET es una técnica nativa de Excel: cuando una fila queda
    oculta por el filtro de la tabla, la fórmula devuelve 0; cuando permanece
    visible, devuelve 1. No requiere macros ni vínculos externos.
    """
    if n_rows <= 0:
        return

    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    column_positions = {str(name): start_col + idx for idx, name in enumerate(columns)}
    marker_col = column_positions.get("__RowMarker")
    visible_col = column_positions.get("__Visible")
    if marker_col is None or visible_col is None:
        return

    marker_letter = get_column_letter(marker_col)
    first_data_row = header_row + 1
    last_data_row = header_row + n_rows
    for row_idx in range(first_data_row, last_data_row + 1):
        ws.cell(row_idx, marker_col, 1)
        ws.cell(row_idx, visible_col, f"=SUBTOTAL(103,OFFSET(${marker_letter}{row_idx},0,0))")
        ws.cell(row_idx, visible_col).number_format = "0"

    technical_fill = PatternFill("solid", fgColor=EXCEL_THEME["panel_alt"])
    for helper_name in EXCEL_INTERACTIVE_HELPER_COLUMNS:
        col_idx = column_positions.get(helper_name)
        if col_idx is None:
            continue
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].hidden = True
        ws.column_dimensions[letter].width = 3
        ws.cell(header_row, col_idx).fill = technical_fill
        ws.cell(header_row, col_idx).font = Font(color=EXCEL_THEME["muted"], italic=True, size=8)


def _excel_category_pairs(df: pd.DataFrame, column: str) -> list[tuple[str, str]]:
    """Obtiene categorías ordenadas por frecuencia, preservando valores vacíos."""
    if df is None or df.empty or column not in df.columns:
        return []

    series = df[column]
    if isinstance(series, pd.DataFrame):
        series = series.bfill(axis=1).iloc[:, 0]

    def criterion(value) -> str:
        try:
            if value is None or pd.isna(value):
                return ""
        except Exception:
            pass
        text_value = str(value).strip()
        if text_value.lower() in {"nan", "none", "nat", "<na>"}:
            return ""
        return text_value

    criteria = series.astype("object").map(criterion)
    counts = criteria.value_counts(dropna=False)
    pairs: list[tuple[str, str]] = []
    for raw_value, _count in counts.items():
        raw_text = str(raw_value)
        pairs.append((raw_text if raw_text else "No informado", raw_text))
    return pairs


def _excel_write_dynamic_category_section(
    ws,
    export_df: pd.DataFrame,
    range_map: dict[str, str],
    source_column: str,
    section_title: str,
    start_row: int,
    start_col: int,
    top_n: int,
) -> dict:
    """Crea una tabla auxiliar dinámica y un Top-N sensible a filtros."""
    from openpyxl.utils import get_column_letter

    pairs = _excel_category_pairs(export_df, source_column)
    available = bool(pairs and source_column in export_df.columns)
    raw_rows = max(len(pairs), 1)

    headers = [
        f"{section_title} · Etiqueta",
        f"{section_title} · Criterio",
        "Visible",
        "Puntaje",
        f"Top {top_n} · Categoría",
        "Top visible",
        "Top puntaje",
    ]
    for offset, header in enumerate(headers):
        ws.cell(start_row, start_col + offset, header)

    raw_start = start_row + 1
    raw_end = raw_start + raw_rows - 1
    if available:
        visible_ref = _excel_table_column_ref(table_name, "__Visible")
        source_ref = range_map[source_column]
        for idx, (display_label, criteria_value) in enumerate(pairs):
            row_idx = raw_start + idx
            ws.cell(row_idx, start_col, display_label)
            ws.cell(row_idx, start_col + 1, criteria_value)
            criteria_cell = ws.cell(row_idx, start_col + 1).coordinate
            ws.cell(row_idx, start_col + 2, f"=SUMIFS({visible_ref},{source_ref},{criteria_cell})")
            ws.cell(row_idx, start_col + 3, f"=IF({ws.cell(row_idx, start_col + 2).coordinate}>0,{ws.cell(row_idx, start_col + 2).coordinate}+ROW()/1000000,0)")
    else:
        ws.cell(raw_start, start_col, "No disponible")
        ws.cell(raw_start, start_col + 1, "")
        ws.cell(raw_start, start_col + 2, "=0")
        ws.cell(raw_start, start_col + 3, "=0")

    score_range = f"${get_column_letter(start_col + 3)}${raw_start}:${get_column_letter(start_col + 3)}${raw_end}"
    label_range = f"${get_column_letter(start_col)}${raw_start}:${get_column_letter(start_col)}${raw_end}"
    count_range = f"${get_column_letter(start_col + 2)}${raw_start}:${get_column_letter(start_col + 2)}${raw_end}"

    top_start = start_row + 1
    for rank in range(1, top_n + 1):
        row_idx = top_start + rank - 1
        top_score_cell = ws.cell(row_idx, start_col + 6)
        top_label_cell = ws.cell(row_idx, start_col + 4)
        top_value_cell = ws.cell(row_idx, start_col + 5)
        top_score_cell.value = f"=IFERROR(LARGE({score_range},{rank}),0)"
        top_label_cell.value = f'=IF({top_score_cell.coordinate}=0,"",INDEX({label_range},MATCH({top_score_cell.coordinate},{score_range},0)))'
        top_value_cell.value = f"=IF({top_score_cell.coordinate}=0,NA(),INT({top_score_cell.coordinate}))"

    return {
        "available": available,
        "raw_count_range": count_range,
        "top_start_row": top_start,
        "top_n": top_n,
        "top_label_col": start_col + 4,
        "top_value_col": start_col + 5,
    }


def _excel_build_dynamic_helper_sheet(wb, export_df: pd.DataFrame, table_name: str) -> tuple[object, dict[str, dict]]:
    """Construye las fórmulas que alimentan KPIs y gráficos interactivos."""
    helper_ws = wb.create_sheet("_DashboardData")
    helper_ws.sheet_state = "veryHidden"
    helper_ws.sheet_view.showGridLines = False

    specifications = [
        ("models", "Instrument type", "Modelos", 12),
        ("status", "Operational status grouped", "Estado", 12),
        ("countries", "Country", "Países", 10),
        ("os", "Operating System", "Sistema operativo", 10),
        ("processing", "__ProcessingBucket", "Procesamiento", 2),
        ("blood", "__BloodBucket", "Banco de sangre", 2),
        ("distributors", "Distributor name", "Distribuidores", 1),
    ]

    sections: dict[str, dict] = {}
    cursor_col = 1
    for key, source_column, title, top_n in specifications:
        sections[key] = _excel_write_dynamic_category_section(
            helper_ws,
            export_df=export_df,
            table_name=table_name,
            source_column=source_column,
            section_title=title,
            start_row=1,
            start_col=cursor_col,
            top_n=top_n,
        )
        cursor_col += 9

    return helper_ws, sections


def _excel_export_signature(
    df: pd.DataFrame,
    filter_summary: dict[str, str] | None,
    active_tab: str,
    source_label_value: str,
) -> str:
    """Firma rápida para invalidar un Excel cuando cambia la vista filtrada.

    Evita ordenar y serializar la lista completa de seriales en cada rerun de
    Streamlit. El hash vectorizado de pandas mantiene la comprobación del
    universo filtrado con un costo lineal y una huella de memoria mucho menor.
    """
    digest = hashlib.sha256()
    digest.update(str(active_tab).encode("utf-8", errors="ignore"))
    digest.update(str(source_label_value).encode("utf-8", errors="ignore"))
    digest.update(str(len(df) if isinstance(df, pd.DataFrame) else 0).encode("ascii"))
    digest.update(repr(sorted((filter_summary or {}).items())).encode("utf-8", errors="ignore"))

    if isinstance(df, pd.DataFrame) and not df.empty:
        identity_columns = [
            col for col in (
                "Serial number",
                "Country",
                "Distributor name",
                "Instrument type",
                "Operational status grouped",
            )
            if col in df.columns
        ]
        if identity_columns:
            identity_frame = df[identity_columns].astype("string").fillna("")
            row_hashes = pd.util.hash_pandas_object(identity_frame, index=False).to_numpy(dtype="uint64", copy=False)
            digest.update(row_hashes.tobytes())

    return digest.hexdigest()


def resolve_excel_report_dataframe(
    filtered_df: pd.DataFrame,
    active_tab: str,
    source_label_value: str = "",
) -> tuple[pd.DataFrame, str, bool]:
    """Resuelve la misma fuente visual del dashboard para la exportación Excel."""
    report_df = filtered_df.copy() if isinstance(filtered_df, pd.DataFrame) else pd.DataFrame()
    report_source = source_label_value
    using_manufacturing = False

    manufacturing_df = st.session_state.get(MANUFACTURING_EXCEL_EXPORT_SESSION_KEY)
    if isinstance(manufacturing_df, pd.DataFrame) and not manufacturing_df.empty:
        has_manufacturing_age = _has_valid_numeric_column(manufacturing_df, "Manufacturing age (years)")
        if active_tab == "Antigüedad / fabricación" and has_manufacturing_age and _same_serial_universe(report_df, manufacturing_df):
            report_df = manufacturing_df.copy()
            report_source = st.session_state.get(MANUFACTURING_EXCEL_EXPORT_SOURCE_KEY, source_label_value)
            using_manufacturing = True

    return report_df, report_source, using_manufacturing


def _excel_write_helper_table(ws, df: pd.DataFrame, start_row: int, start_col: int) -> tuple[int, int, int, int]:
    """Escribe datos auxiliares para gráficos sin aplicar formato de hoja visible."""
    from openpyxl.utils.dataframe import dataframe_to_rows

    work = _excel_unique_columns(_excel_clean_dataframe(df))
    for r_offset, row in enumerate(dataframe_to_rows(work, index=False, header=True)):
        for c_offset, value in enumerate(row):
            ws.cell(start_row + r_offset, start_col + c_offset, _excel_safe_cell_value(value))
    return start_row, start_col, len(work), max(len(work.columns), 1)


def _excel_apply_series_palette(chart, palette: list[str]) -> None:
    """Aplica la paleta visual del dashboard a las series de gráficos Excel."""
    for idx, series in enumerate(getattr(chart, "series", [])):
        color = palette[idx % len(palette)].replace("#", "")
        try:
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color
        except Exception:
            pass




def _excel_build_visual_dashboard(
    wb,
    export_df: pd.DataFrame,
    filter_summary: dict[str, str] | None,
    active_tab: str,
    source_label_value: str,
    stock_context: dict | None = None,
    table_name: str = "RecordsListFilteredData",
) -> None:
    """Crea un dashboard Excel que se recalcula al modificar AutoFilters.

    Los gráficos y KPIs se alimentan de __Visible, una fórmula oculta que cambia
    entre 1 y 0 según la visibilidad de cada fila de la tabla de datos. Así, el
    usuario puede filtrar directamente en Excel sin macros y el dashboard se
    actualiza cuando Excel recalcula el libro.
    """
    from openpyxl.chart import BarChart, DoughnutChart, Reference
    from openpyxl.chart.data_source import AxDataSource, StrRef
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    helper_ws, sections = _excel_build_dynamic_helper_sheet(wb, export_df, table_name)

    ws = wb.create_sheet("00_Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = EXCEL_THEME["royal"]
    ws.sheet_view.zoomScale = 85
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for column, width in {
        "A": 4, "B": 14, "C": 14, "D": 14,
        "E": 4, "F": 14, "G": 14, "H": 14,
        "I": 4, "J": 14, "K": 14, "L": 14,
        "M": 4, "N": 14, "O": 14, "P": 14,
    }.items():
        ws.column_dimensions[column].width = width

    canvas_fill = PatternFill("solid", fgColor=EXCEL_THEME["canvas"])
    panel_fill = PatternFill("solid", fgColor=EXCEL_THEME["panel"])
    panel_alt_fill = PatternFill("solid", fgColor=EXCEL_THEME["panel_alt"])
    navy_fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    thin = Side(style="thin", color=EXCEL_THEME["border"])
    panel_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 65):
        ws.row_dimensions[row].height = 22
        for col in range(1, 17):
            ws.cell(row, col).fill = canvas_fill
            ws.cell(row, col).font = Font(color=EXCEL_THEME["text"], name="Aptos")

    ws.merge_cells("A1:P2")
    ws["A1"] = "RECORDS LIST INTELLIGENCE DASHBOARD · EXCEL INTERACTIVO"
    ws["A1"].fill = navy_fill
    ws["A1"].font = Font(color=EXCEL_THEME["white"], bold=True, size=22, name="Aptos Display")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A3:P3")
    ws["A3"] = (
        f"Vista inicial: {active_tab}   |   Fuente: {source_label_value or 'No informada'}   |   "
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Sin macros"
    )
    ws["A3"].fill = navy_fill
    ws["A3"].font = Font(color="DCE7F1", size=10, name="Aptos")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    visible_ref = _excel_table_column_ref(table_name, "__Visible")
    total_formula = f"=SUM({visible_ref})"
    countries_formula = f'=COUNTIF(\'_DashboardData\'!{sections["countries"]["raw_count_range"]},">0")'
    distributors_formula = f'=COUNTIF(\'_DashboardData\'!{sections["distributors"]["raw_count_range"]},">0")'
    models_formula = f'=COUNTIF(\'_DashboardData\'!{sections["models"]["raw_count_range"]},">0")'

    cards = [
        ("A5:D5", "A6:D7", "EQUIPOS VISIBLES", total_formula, EXCEL_THEME["royal"]),
        ("E5:H5", "E6:H7", "PAÍSES VISIBLES", countries_formula, EXCEL_THEME["teal"]),
        ("I5:L5", "I6:L7", "DISTRIBUIDORES VISIBLES", distributors_formula, EXCEL_THEME["purple"]),
        ("M5:P5", "M6:P7", "MODELOS VISIBLES", models_formula, EXCEL_THEME["amber"]),
    ]
    for label_range, value_range, label, formula, accent in cards:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=accent)
        label_cell.font = Font(color=EXCEL_THEME["white"], bold=True, size=10, name="Aptos")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = panel_border
        value_cell.value = formula
        value_cell.fill = panel_fill
        value_cell.font = Font(color=accent, bold=True, size=24, name="Aptos Display")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = "#,##0"
        value_cell.border = panel_border

    ws.merge_cells("A9:H9")
    ws["A9"] = "FILTROS APLICADOS EN LA APP AL EXPORTAR"
    ws["A9"].fill = PatternFill("solid", fgColor=EXCEL_THEME["royal"])
    ws["A9"].font = Font(color=EXCEL_THEME["white"], bold=True, size=11, name="Aptos Display")
    ws["A9"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A9"].border = panel_border
    ws.merge_cells("A10:H13")
    filter_lines = [f"• {key}: {value}" for key, value in (filter_summary or {}).items()]
    filter_lines.append(f"• Registros iniciales: {len(export_df):,}")
    ws["A10"] = "\n".join(filter_lines)
    ws["A10"].fill = panel_fill
    ws["A10"].border = panel_border
    ws["A10"].font = Font(color=EXCEL_THEME["text"], size=10, name="Aptos")
    ws["A10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("I9:P9")
    ws["I9"] = "CÓMO USAR LOS FILTROS INTERACTIVOS"
    ws["I9"].fill = PatternFill("solid", fgColor=EXCEL_THEME["teal"])
    ws["I9"].font = Font(color=EXCEL_THEME["white"], bold=True, size=11, name="Aptos Display")
    ws["I9"].alignment = Alignment(horizontal="left", vertical="center")
    ws["I9"].border = panel_border
    ws.merge_cells("I10:P13")
    ws["I10"] = (
        "1. Abre 01_Datos_filtrados y usa las flechas de los encabezados.\n"
        "2. Filtra país, distribuidor, modelo, estado, cliente o cualquier otra columna.\n"
        "3. Regresa a 00_Dashboard: tarjetas y seis gráficos se recalculan automáticamente.\n"
        "4. No elimines las columnas auxiliares ocultas; son el motor del dashboard.\n"
        "5. Si Excel está en cálculo manual, presiona Ctrl + Alt + F9 para recalcular."
    )
    ws["I10"].fill = panel_alt_fill
    ws["I10"].border = panel_border
    ws["I10"].font = Font(color=EXCEL_THEME["text"], size=10, name="Aptos")
    ws["I10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A14:D14")
    ws["A14"] = "ABRIR DATOS Y MODIFICAR FILTROS"
    ws["A14"].hyperlink = "#'01_Datos_filtrados'!A1"
    ws["A14"].fill = PatternFill("solid", fgColor="DDE8FF")
    ws["A14"].font = Font(color=EXCEL_THEME["royal"], bold=True, underline="single", name="Aptos")
    ws["A14"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A14"].border = panel_border

    ws.merge_cells("E14:H14")
    ws["E14"] = "ABRIR RESUMEN DE EXPORTACIÓN"
    ws["E14"].hyperlink = "#'00_Resumen'!A1"
    ws["E14"].fill = PatternFill("solid", fgColor="DDF4F2")
    ws["E14"].font = Font(color=EXCEL_THEME["teal"], bold=True, underline="single", name="Aptos")
    ws["E14"].alignment = Alignment(horizontal="center", vertical="center")
    ws["E14"].border = panel_border

    ws.merge_cells("I14:P14")
    ws["I14"] = "DASHBOARD DINÁMICO · LOS GRÁFICOS RESPONDEN A LOS FILTROS DE 01_Datos_filtrados"
    ws["I14"].fill = PatternFill("solid", fgColor="E7F7EF")
    ws["I14"].font = Font(color=EXCEL_THEME["green"], bold=True, name="Aptos")
    ws["I14"].alignment = Alignment(horizontal="center", vertical="center")
    ws["I14"].border = panel_border

    def add_bar(section_key: str, title: str, anchor: str, point_colors: list[str]) -> None:
        section = sections[section_key]
        if not section.get("available"):
            return
        start_row = section["top_start_row"]
        top_n = section["top_n"]
        label_col = section["top_label_col"]
        value_col = section["top_value_col"]
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        chart.height = 8.5
        chart.width = 15.5
        chart.legend = None
        chart.y_axis.title = "Categoría"
        chart.x_axis.title = "Cantidad visible"
        chart.gapWidth = 55
        chart.display_blanks = "gap"
        data = Reference(helper_ws, min_col=value_col, min_row=start_row - 1, max_row=start_row + top_n - 1)
        chart.add_data(data, titles_from_data=True)
        if chart.series:
            category_formula = (
                f"'{helper_ws.title}'!${get_column_letter(label_col)}${start_row}:"
                f"${get_column_letter(label_col)}${start_row + top_n - 1}"
            )
            chart.series[0].cat = AxDataSource(strRef=StrRef(f=category_formula))
            points = []
            for idx in range(top_n):
                point = DataPoint(idx=idx)
                point.graphicalProperties.solidFill = point_colors[idx % len(point_colors)]
                points.append(point)
            chart.series[0].data_points = points
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        ws.add_chart(chart, anchor)

    def add_doughnut(section_key: str, title: str, anchor: str, colors_for_points: list[str]) -> None:
        section = sections[section_key]
        if not section.get("available"):
            return
        start_row = section["top_start_row"]
        top_n = section["top_n"]
        label_col = section["top_label_col"]
        value_col = section["top_value_col"]
        chart = DoughnutChart()
        chart.title = title
        chart.style = 10
        chart.height = 8.5
        chart.width = 12.5
        chart.holeSize = 62
        chart.firstSliceAng = 270
        chart.legend.position = "b"
        chart.display_blanks = "gap"
        data = Reference(helper_ws, min_col=value_col, min_row=start_row - 1, max_row=start_row + top_n - 1)
        chart.add_data(data, titles_from_data=True)
        if chart.series:
            category_formula = (
                f"'{helper_ws.title}'!${get_column_letter(label_col)}${start_row}:"
                f"${get_column_letter(label_col)}${start_row + top_n - 1}"
            )
            chart.series[0].cat = AxDataSource(strRef=StrRef(f=category_formula))
            points = []
            for idx in range(top_n):
                point = DataPoint(idx=idx)
                point.graphicalProperties.solidFill = colors_for_points[idx % len(colors_for_points)]
                points.append(point)
            chart.series[0].data_points = points
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLeaderLines = True
        ws.add_chart(chart, anchor)

    model_colors = [EXCEL_CHART_PALETTE[i % len(EXCEL_CHART_PALETTE)] for i in range(12)]
    status_colors = [EXCEL_THEME["green"], EXCEL_THEME["amber"], EXCEL_THEME["royal"], EXCEL_THEME["coral"], EXCEL_THEME["purple"], EXCEL_THEME["slate"]]
    country_colors = [EXCEL_CHART_PALETTE[i % len(EXCEL_CHART_PALETTE)] for i in range(10)]
    os_colors = [EXCEL_THEME["green"], EXCEL_THEME["amber"], EXCEL_THEME["coral"], EXCEL_THEME["teal"], EXCEL_THEME["slate"]]

    add_bar("models", "Base instalada visible por modelo", "A16", model_colors)
    add_bar("status", "Estado operativo visible", "I16", status_colors)
    add_bar("countries", "Top países visibles", "A31", country_colors)
    add_bar("os", "Sistema operativo visible", "I31", os_colors)
    add_doughnut("processing", "Procesamiento diario visible", "A46", [EXCEL_THEME["coral"], EXCEL_THEME["green"]])
    add_doughnut("blood", "Banco de sangre visible", "I46", [EXCEL_THEME["royal"], EXCEL_THEME["slate"]])

    stock_context = stock_context or {}
    if stock_context.get("available"):
        ws["M15"] = f"Carstock: {stock_context.get('detected_distributor', 'N/A')}"
        ws["M15"].font = Font(color=EXCEL_THEME["coral"], bold=True)

    ws.sheet_view.selection[0].activeCell = "A1"
    ws.sheet_view.selection[0].sqref = "A1"



# =============================================================================
# EXCEL V50 · PORTADA ESTABLE + EXPLORADOR CON FILTROS
# =============================================================================
EXCEL_V50_FILTER_SPECS = [
    ("Región comercial", "Commercial Region", "$B$5"),
    ("País", "Country", "$B$6"),
    ("Distribuidor", "Distributor name", "$B$7"),
    ("Modelo", "Instrument type", "$B$8"),
    ("Estado operativo", "Operational status grouped", "$B$9"),
    ("Banco de sangre", "Blood Bank group", "$B$10"),
    ("Sistema operativo", "Operating System", "$B$11"),
    ("Procesamiento", "Processing group", "$B$12"),
    ("Rango de antigüedad", "Age group", "$B$13"),
]


EXCEL_V50_KEY_COLUMNS = [
    "Commercial Region",
    "Country",
    "Distributor name",
    "Customer name",
    "City",
    "Instrument type",
    "Serial number",
    "Installation date",
    "Age (years)",
    "Operational status grouped",
    "Operational status",
    "Asset condition",
    "Type of contract",
    "Blood Bank group",
    "Number of tests per day",
    "Processing group",
    "PM plan",
    "PM last date",
    "PM next date",
    "PM status",
    "Operating System",
    "Manufacturing Date",
    "Manufacturing age (years)",
    "Manufacturing age bucket",
    "Manufacturing matched",
]


EXCEL_V50_TECH_COLUMNS = [
    "__RoutineFlag",
    "__ZeroTestsFlag",
    "__PMOverdueFlag",
    "__ManufacturingFlag",
]


def _excel_v50_safe_text_series(df: pd.DataFrame, column: str, default: str = "No informado") -> pd.Series:
    if column not in df.columns:
        return pd.Series(default, index=df.index, dtype="object")
    series = df[column].astype("object")
    series = series.where(~pd.isna(series), default)
    series = series.astype(str).str.strip()
    return series.replace({"": default, "nan": default, "None": default, "<NA>": default, "NaT": default})


def _excel_v50_age_group(values: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce")
    labels = pd.cut(
        numeric,
        bins=[-np.inf, 2.999, 4.999, 7.999, 9.999, 14.999, np.inf],
        labels=["0–2 años", "3–4 años", "5–7 años", "8–9 años", "10–14 años", "15+ años"],
    )
    result = labels.astype("object")
    result = result.where(~pd.isna(result), "No informado")
    return result.astype(str)


def _excel_v50_prepare_key_data(export_df: pd.DataFrame) -> pd.DataFrame:
    """Construye una vista manejable de 25 campos y auxiliares ocultos.

    La hoja de datos completos conserva todas las columnas. Esta vista existe
    para que el usuario pueda filtrar sin navegar horizontalmente por más de
    cien campos técnicos.
    """
    work = export_df.copy() if isinstance(export_df, pd.DataFrame) else pd.DataFrame()

    for column in EXCEL_V50_KEY_COLUMNS:
        if column not in work.columns:
            work[column] = ""

    work["Commercial Region"] = _excel_v50_safe_text_series(work, "Commercial Region")
    work["Country"] = _excel_v50_safe_text_series(work, "Country")
    work["Distributor name"] = _excel_v50_safe_text_series(work, "Distributor name")
    work["Customer name"] = _excel_v50_safe_text_series(work, "Customer name")
    work["City"] = _excel_v50_safe_text_series(work, "City")
    work["Instrument type"] = _excel_v50_safe_text_series(work, "Instrument type")
    work["Operational status grouped"] = _excel_v50_safe_text_series(work, "Operational status grouped")
    work["Operational status"] = _excel_v50_safe_text_series(work, "Operational status")
    work["Asset condition"] = _excel_v50_safe_text_series(work, "Asset condition")
    work["Type of contract"] = _excel_v50_safe_text_series(work, "Type of contract")
    work["Operating System"] = _excel_v50_safe_text_series(work, "Operating System")

    tests = pd.to_numeric(work["Number of tests per day"], errors="coerce").fillna(0)
    work["Number of tests per day"] = tests
    work["Processing group"] = np.where(tests.gt(0), "> 0 tests/día", "0 tests/día")

    if "Blood Bank Flag" in export_df.columns:
        blood = export_df["Blood Bank Flag"].fillna(False).astype(bool)
    elif "In Blood Bank" in export_df.columns:
        blood = export_df["In Blood Bank"].map(is_blood_bank_yes).fillna(False).astype(bool)
    else:
        blood = pd.Series(False, index=work.index, dtype=bool)
    work["Blood Bank group"] = np.where(blood, "Banco de sangre", "Laboratorio")

    pm_next = pd.to_datetime(work["PM next date"], errors="coerce")
    today = pd.Timestamp.today().normalize()
    work["PM status"] = np.select(
        [
            pm_next.notna() & pm_next.lt(today),
            pm_next.notna() & pm_next.le(today + pd.Timedelta(days=90)),
            pm_next.notna(),
        ],
        ["Vencido", "Próximos 90 días", "Planificado"],
        default="No informado",
    )

    manufacturing_age = pd.to_numeric(work["Manufacturing age (years)"], errors="coerce")
    installation_age = pd.to_numeric(work["Age (years)"], errors="coerce")
    age_for_filter = manufacturing_age.where(manufacturing_age.notna(), installation_age)
    work["Age group"] = _excel_v50_age_group(age_for_filter)

    routine_text = work["Operational status grouped"].astype(str).str.lower().str.strip()
    work["__RoutineFlag"] = routine_text.eq("routine").astype(int)
    work["__ZeroTestsFlag"] = tests.le(0).astype(int)
    work["__PMOverdueFlag"] = (pm_next.notna() & pm_next.lt(today)).astype(int)

    matched = work["Manufacturing matched"]
    try:
        matched_bool = matched.fillna(False).astype(bool)
    except Exception:
        matched_bool = matched.astype(str).str.lower().isin({"true", "1", "yes", "si", "sí"})
    work["__ManufacturingFlag"] = matched_bool.astype(int)

    visible_columns = [c for c in EXCEL_V50_KEY_COLUMNS if c in work.columns]
    # Age group is used as a filter and is intentionally visible near age fields.
    insert_at = visible_columns.index("Age (years)") + 1 if "Age (years)" in visible_columns else len(visible_columns)
    visible_columns.insert(insert_at, "Age group")
    return _excel_unique_columns(work[visible_columns + EXCEL_V50_TECH_COLUMNS])


def _excel_v50_navigation(ws, current: str = "") -> None:
    """Crea navegación interna funcional en cada hoja visible."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    links = [
        ("Dashboard", "00_Dashboard", EXCEL_THEME["royal"]),
        ("Filtros", "01_Filtros", EXCEL_THEME["teal"]),
        ("Datos clave", "02_Datos_clave", EXCEL_THEME["green"]),
        ("Datos completos", "03_Datos_completos", EXCEL_THEME["slate"]),
        ("Resumen", "00_Resumen", EXCEL_THEME["navy"]),
    ]
    thin = Side(style="thin", color=EXCEL_THEME["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for idx, (label, sheet_name, color) in enumerate(links, start=1):
        cell = ws.cell(3, idx)
        cell.value = label if sheet_name != current else f"● {label}"
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.fill = PatternFill("solid", fgColor=color if sheet_name == current else EXCEL_THEME["panel_alt"])
        cell.font = Font(
            color=EXCEL_THEME["white"] if sheet_name == current else color,
            bold=True,
            underline=None if sheet_name == current else "single",
            name="Aptos",
            size=9,
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22


def _excel_v50_sheet_header(ws, title: str, subtitle: str, current: str, tab_color: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    ws["A1"].font = Font(color=EXCEL_THEME["white"], bold=True, size=17, name="Aptos Display")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A2:L2")
    ws["A2"] = subtitle
    ws["A2"].fill = PatternFill("solid", fgColor=EXCEL_THEME["canvas"])
    ws["A2"].font = Font(color=EXCEL_THEME["muted"], italic=True, size=10, name="Aptos")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28
    _excel_v50_navigation(ws, current=current)


def _excel_v50_static_banner(ws, row: int = 5) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row, 1)
    cell.value = "FOTOGRAFÍA ESTÁTICA DE LOS FILTROS APLICADOS EN LA APP AL MOMENTO DE EXPORTAR"
    cell.fill = PatternFill("solid", fgColor="FFF3CD")
    cell.font = Font(color="7A4E00", bold=True, name="Aptos", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="E7C96A")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 22


def _excel_v50_summary_values(df: pd.DataFrame) -> dict[str, object]:
    total = int(len(df))
    status = _excel_v50_safe_text_series(df, "Operational status grouped", "No informado")
    routine = int(status.str.lower().str.strip().eq("routine").sum())
    tests = pd.to_numeric(df.get("Number of tests per day", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0)
    zero_tests = int(tests.le(0).sum())
    countries = int(_excel_v50_safe_text_series(df, "Country").nunique()) if total else 0
    distributors = int(_excel_v50_safe_text_series(df, "Distributor name").nunique()) if total else 0
    blood = count_blood_bank_yes(df) if total else 0
    pm_next = pd.to_datetime(df.get("PM next date", pd.Series(index=df.index, dtype="object")), errors="coerce")
    pm_overdue = int((pm_next.notna() & pm_next.lt(pd.Timestamp.today().normalize())).sum())
    matched = df.get("Manufacturing matched", pd.Series(False, index=df.index))
    try:
        manuf = int(matched.fillna(False).astype(bool).sum())
    except Exception:
        manuf = int(matched.astype(str).str.lower().isin({"true", "1", "yes", "si", "sí"}).sum())
    return {
        "Total equipos": total,
        "En rutina": routine,
        "Fuera de rutina": max(total - routine, 0),
        "Países": countries,
        "Distribuidores": distributors,
        "0 tests/día": zero_tests,
        "PM vencidos": pm_overdue,
        "Con fabricación": manuf,
        "Banco de sangre": blood,
    }


def _excel_v50_chart_image(
    df: pd.DataFrame,
    label_col: str,
    value_col: str,
    title: str,
    chart_type: str = "bar",
    base_color: str | None = None,
    semantic: str | None = None,
    max_rows: int = 10,
):
    """Genera una imagen PNG estable para que el dashboard nunca abra vacío."""
    if not MATPLOTLIB_AVAILABLE or df is None or df.empty:
        return None

    work = df[[label_col, value_col]].copy()
    work[value_col] = pd.to_numeric(work[value_col], errors="coerce").fillna(0)
    work = work[work[value_col] > 0].head(max_rows)
    if work.empty:
        return None

    color_map = {
        "royal": f"#{EXCEL_THEME['royal']}",
        "teal": f"#{EXCEL_THEME['teal']}",
        "green": f"#{EXCEL_THEME['green']}",
        "amber": f"#{EXCEL_THEME['amber']}",
        "coral": f"#{EXCEL_THEME['coral']}",
        "purple": f"#{EXCEL_THEME['purple']}",
        "slate": f"#{EXCEL_THEME['slate']}",
    }
    default_color = base_color or color_map["royal"]

    def semantic_color(label: str) -> str:
        low = str(label).lower()
        if semantic == "status":
            if "routine" in low and "not" not in low and "no rutina" not in low:
                return color_map["green"]
            if "scrap" in low or "discard" in low or "baja" in low:
                return color_map["coral"]
            if "warehouse" in low or "almac" in low:
                return color_map["amber"]
            return color_map["slate"]
        if semantic == "os":
            if "windows 10" in low or "win 10" in low:
                return color_map["green"]
            if any(token in low for token in ("windows 7", "windows xp", "legacy", "vista")):
                return color_map["coral"]
            if "no informado" in low or "unknown" in low:
                return color_map["amber"]
            return color_map["teal"]
        if semantic == "processing":
            return color_map["green"] if str(label).startswith(">") else color_map["coral"]
        if semantic == "blood":
            return color_map["royal"] if "banco" in low else color_map["slate"]
        return default_color

    colors = [semantic_color(label) for label in work[label_col]]
    buffer = BytesIO()
    if chart_type == "donut":
        fig, ax = plt.subplots(figsize=(6.8, 3.5))
        wedges, texts, autotexts = ax.pie(
            work[value_col].astype(float),
            labels=work[label_col].astype(str),
            colors=colors,
            autopct=lambda pct: f"{pct:.0f}%" if pct >= 4 else "",
            startangle=90,
            counterclock=False,
            wedgeprops={"width": 0.42, "edgecolor": "white"},
            textprops={"fontsize": 8, "color": f"#{EXCEL_THEME['text']}"},
        )
        for autotext in autotexts:
            autotext.set_fontsize(8)
            autotext.set_color("white")
            autotext.set_fontweight("bold")
        ax.set_title(title, fontsize=12, fontweight="bold", color=f"#{EXCEL_THEME['navy']}", pad=12)
        ax.axis("equal")
    else:
        work = work.sort_values(value_col, ascending=True)
        colors = [semantic_color(label) for label in work[label_col]]
        fig, ax = plt.subplots(figsize=(7.4, 3.8))
        bars = ax.barh(work[label_col].astype(str), work[value_col].astype(float), color=colors)
        ax.set_title(title, fontsize=12, fontweight="bold", color=f"#{EXCEL_THEME['navy']}", pad=10)
        ax.grid(axis="x", alpha=0.18)
        ax.tick_params(axis="both", labelsize=8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_visible(False)
        ax.spines["bottom"].set_color(f"#{EXCEL_THEME['border']}")
        max_value = float(work[value_col].max()) if not work.empty else 0
        for bar in bars:
            value = bar.get_width()
            ax.text(value + max(max_value * 0.015, 0.1), bar.get_y() + bar.get_height() / 2, f"{value:,.0f}", va="center", fontsize=8)
        ax.set_xlim(0, max_value * 1.18 if max_value else 1)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    fig.tight_layout()
    fig.savefig(buffer, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buffer.seek(0)
    return buffer


def _excel_v50_add_image(ws, image_buffer, anchor: str, width: int = 610, height: int = 315) -> None:
    if image_buffer is None:
        return
    from openpyxl.drawing.image import Image as XLImage

    image = XLImage(image_buffer)
    image.width = width
    image.height = height
    ws.add_image(image, anchor)


def _excel_v50_build_snapshot_dashboard(
    wb,
    filtered_df: pd.DataFrame,
    filter_summary: dict[str, str],
    active_tab: str,
    source_label_value: str,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet("00_Dashboard", 0)
    _excel_v50_sheet_header(
        ws,
        "RECORDS LIST INTELLIGENCE · DASHBOARD EJECUTIVO",
        "Portada estable: los indicadores y gráficos representan exactamente los filtros activos al exportar y siempre se muestran al abrir.",
        "00_Dashboard",
        EXCEL_THEME["royal"],
    )
    ws.sheet_view.zoomScale = 82
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for col in range(1, 17):
        ws.column_dimensions[chr(64 + col)].width = 11 if col not in {1, 5, 9, 13} else 5
    for row in range(4, 66):
        ws.row_dimensions[row].height = 22

    values = _excel_v50_summary_values(filtered_df)
    card_specs = [
        ("A5:D5", "A6:D8", "TOTAL EQUIPOS", values["Total equipos"], EXCEL_THEME["royal"]),
        ("E5:H5", "E6:H8", "EN RUTINA", values["En rutina"], EXCEL_THEME["green"]),
        ("I5:L5", "I6:L8", "FUERA DE RUTINA", values["Fuera de rutina"], EXCEL_THEME["amber"]),
        ("M5:P5", "M6:P8", "0 TESTS/DÍA", values["0 tests/día"], EXCEL_THEME["coral"]),
        ("A9:D9", "A10:D12", "PAÍSES", values["Países"], EXCEL_THEME["teal"]),
        ("E9:H9", "E10:H12", "DISTRIBUIDORES", values["Distribuidores"], EXCEL_THEME["purple"]),
        ("I9:L9", "I10:L12", "PM VENCIDOS", values["PM vencidos"], EXCEL_THEME["coral"]),
        ("M9:P9", "M10:P12", "CON FABRICACIÓN", values["Con fabricación"], EXCEL_THEME["slate"]),
    ]
    thin = Side(style="thin", color=EXCEL_THEME["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for label_range, value_range, label, value, color in card_specs:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=color)
        label_cell.font = Font(color=EXCEL_THEME["white"], bold=True, size=10, name="Aptos")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = border
        value_cell.value = value
        value_cell.fill = PatternFill("solid", fgColor=EXCEL_THEME["panel"])
        value_cell.font = Font(color=color, bold=True, size=23, name="Aptos Display")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = "#,##0"
        value_cell.border = border

    ws.merge_cells("A13:H13")
    ws["A13"] = "Explorar con filtros propios de Excel"
    ws["A13"].hyperlink = "#'01_Filtros'!A1"
    ws["A13"].fill = PatternFill("solid", fgColor="DDF4F2")
    ws["A13"].font = Font(color=EXCEL_THEME["teal"], bold=True, underline="single", name="Aptos")
    ws["A13"].alignment = Alignment(horizontal="center")
    ws.merge_cells("I13:P13")
    ws["I13"] = f"Vista exportada: {active_tab} · Fuente: {source_label_value or 'No informada'}"
    ws["I13"].fill = PatternFill("solid", fgColor=EXCEL_THEME["panel_alt"])
    ws["I13"].font = Font(color=EXCEL_THEME["muted"], italic=True, name="Aptos")
    ws["I13"].alignment = Alignment(horizontal="center")

    model_df = _excel_value_counts_df(filtered_df, "Instrument type", "Modelo", top_n=10)
    status_df = _excel_value_counts_df(filtered_df, "Operational status grouped", "Estado", top_n=10)
    country_df = _excel_value_counts_df(filtered_df, "Country", "País", top_n=10)
    os_df = _excel_value_counts_df(filtered_df, "Operating System", "Sistema operativo", top_n=10)
    tests = pd.to_numeric(filtered_df.get("Number of tests per day", pd.Series(index=filtered_df.index, dtype=float)), errors="coerce").fillna(0)
    processing_df = pd.DataFrame({"Procesamiento": ["0 tests/día", "> 0 tests/día"], "Cantidad": [int(tests.le(0).sum()), int(tests.gt(0).sum())]})
    blood_yes = count_blood_bank_yes(filtered_df)
    blood_df = pd.DataFrame({"Ubicación": ["Banco de sangre", "Laboratorio"], "Cantidad": [blood_yes, max(len(filtered_df) - blood_yes, 0)]})

    images = [
        (_excel_v50_chart_image(model_df, "Modelo", "Cantidad", "Base instalada por modelo", base_color=f"#{EXCEL_THEME['royal']}"), "A15"),
        (_excel_v50_chart_image(status_df, "Estado", "Cantidad", "Estado operativo", semantic="status"), "I15"),
        (_excel_v50_chart_image(country_df, "País", "Cantidad", "Top países", base_color=f"#{EXCEL_THEME['teal']}"), "A31"),
        (_excel_v50_chart_image(os_df, "Sistema operativo", "Cantidad", "Sistemas operativos", semantic="os"), "I31"),
        (_excel_v50_chart_image(processing_df, "Procesamiento", "Cantidad", "Procesamiento diario", chart_type="donut", semantic="processing"), "A47"),
        (_excel_v50_chart_image(blood_df, "Ubicación", "Cantidad", "Banco de sangre", chart_type="donut", semantic="blood"), "I47"),
    ]
    for image_buffer, anchor in images:
        _excel_v50_add_image(ws, image_buffer, anchor)

    ws.print_area = "A1:P63"


def _excel_v50_key_range_map(key_df: pd.DataFrame, sheet_name: str = "02_Datos_clave", header_row: int = 5) -> dict[str, str]:
    """Mapea cada campo a un rango A1 normal para máxima compatibilidad.

    El explorador no depende de referencias estructuradas ni del motor de tablas
    de un programa específico. Esto funciona de forma más consistente en Excel
    de escritorio, Excel Online y herramientas compatibles.
    """
    from openpyxl.utils import get_column_letter

    first_row = header_row + 1
    last_row = header_row + max(len(key_df), 1)
    escaped_sheet = str(sheet_name).replace("'", "''")
    result: dict[str, str] = {}
    for idx, column in enumerate(key_df.columns, start=1):
        letter = get_column_letter(idx)
        result[str(column)] = f"'{escaped_sheet}'!${letter}${first_row}:${letter}${last_row}"
    return result


def _excel_v50_filter_formula(range_map: dict[str, str], extra_condition: str | None = None) -> str:
    conditions = []
    for _, column, cell_ref in EXCEL_V50_FILTER_SPECS:
        column_ref = range_map[column]
        conditions.append(f'--((({cell_ref}="Todos")+({column_ref}={cell_ref}))>0)')
    if extra_condition:
        conditions.append(extra_condition)
    return "=SUMPRODUCT(" + ",".join(conditions) + ")"


def _excel_v50_add_data_validation_lists(wb, key_df: pd.DataFrame, filter_ws) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    lists_ws = wb.create_sheet("_Listas")
    lists_ws.sheet_state = "veryHidden"
    for idx, (label, column, cell_ref) in enumerate(EXCEL_V50_FILTER_SPECS, start=1):
        values = ["Todos"]
        if column in key_df.columns:
            raw_values = _excel_v50_safe_text_series(key_df, column).tolist()
            values += sorted({str(v).strip() or "No informado" for v in raw_values}, key=lambda x: x.lower())
        values = list(dict.fromkeys(values))
        lists_ws.cell(1, idx, label)
        for row_idx, value in enumerate(values, start=2):
            lists_ws.cell(row_idx, idx, value)
        target = cell_ref.replace("$", "")
        col_letter = get_column_letter(idx)
        formula = f"'_Listas'!${col_letter}$2:${col_letter}${len(values) + 1}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "Selecciona un valor válido de la lista."
        validation.errorTitle = "Filtro no válido"
        validation.prompt = f"Selecciona {label.lower()} o Todos."
        validation.promptTitle = "Filtro del dashboard"
        filter_ws.add_data_validation(validation)
        validation.add(filter_ws[target])
        filter_ws[target] = "Todos"


def _excel_v50_dynamic_summary_table(
    ws,
    key_df: pd.DataFrame,
    range_map: dict[str, str],
    source_column: str,
    title: str,
    start_row: int,
    start_col: int,
    max_categories: int | None = None,
) -> tuple[int, int]:
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    counts = _excel_value_counts_df(key_df, source_column, "Categoría")
    if max_categories:
        counts = counts.head(max_categories)
    categories = counts["Categoría"].astype(str).tolist() if not counts.empty else ["No informado"]
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
    title_cell = ws.cell(start_row, start_col)
    title_cell.value = title
    title_cell.fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    title_cell.font = Font(color=EXCEL_THEME["white"], bold=True, name="Aptos Display", size=11)
    title_cell.alignment = Alignment(horizontal="left")

    header_row = start_row + 1
    ws.cell(header_row, start_col, "Categoría")
    ws.cell(header_row, start_col + 1, "Cantidad filtrada")
    ws.cell(header_row, start_col + 2, "% filtrado")
    for col in range(start_col, start_col + 3):
        ws.cell(header_row, col).fill = PatternFill("solid", fgColor=EXCEL_THEME["royal"])
        ws.cell(header_row, col).font = Font(color=EXCEL_THEME["white"], bold=True, name="Aptos", size=9)
        ws.cell(header_row, col).alignment = Alignment(horizontal="center", wrap_text=True)

    thin = Side(style="thin", color=EXCEL_THEME["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    source_ref = range_map[source_column]
    for offset, category in enumerate(categories, start=1):
        row = header_row + offset
        ws.cell(row, start_col, category)
        category_cell = f"${get_column_letter(start_col)}${row}"
        extra = f"--({source_ref}={category_cell})"
        ws.cell(row, start_col + 1, _excel_v50_filter_formula(range_map, extra))
        ws.cell(row, start_col + 2, f'=IFERROR({get_column_letter(start_col + 1)}{row}/$D$6,0)')
        ws.cell(row, start_col + 2).number_format = "0.0%"
        for col in range(start_col, start_col + 3):
            ws.cell(row, col).border = border
            ws.cell(row, col).fill = PatternFill("solid", fgColor=EXCEL_THEME["panel"] if offset % 2 else EXCEL_THEME["canvas"])
            ws.cell(row, col).font = Font(color=EXCEL_THEME["text"], name="Aptos", size=9)
    end_row = header_row + len(categories)
    count_range = f"{get_column_letter(start_col + 1)}{header_row + 1}:{get_column_letter(start_col + 1)}{end_row}"
    ws.conditional_formatting.add(
        count_range,
        DataBarRule(start_type="num", start_value=0, end_type="max", color=EXCEL_THEME["teal"], showValue=True),
    )
    ws.column_dimensions[get_column_letter(start_col)].width = 30
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 17
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 13
    return header_row, end_row


def _excel_v50_build_filter_explorer(wb, key_df: pd.DataFrame, range_map: dict[str, str]) -> None:
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet("01_Filtros", 1)
    _excel_v50_sheet_header(
        ws,
        "EXPLORADOR DE EXCEL · FILTROS EDITABLES",
        "Selecciona valores en las celdas azules. Las tarjetas y tablas inferiores se recalculan sin macros y sin usar OFFSET.",
        "01_Filtros",
        EXCEL_THEME["teal"],
    )
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A5"

    thin = Side(style="thin", color=EXCEL_THEME["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A4:B4")
    ws["A4"] = "FILTROS"
    ws["A4"].fill = PatternFill("solid", fgColor=EXCEL_THEME["teal"])
    ws["A4"].font = Font(color=EXCEL_THEME["white"], bold=True, name="Aptos Display", size=12)
    ws["A4"].alignment = Alignment(horizontal="left")

    for idx, (label, _, cell_ref) in enumerate(EXCEL_V50_FILTER_SPECS, start=5):
        ws.cell(idx, 1, label)
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=EXCEL_THEME["panel_alt"])
        ws.cell(idx, 1).font = Font(color=EXCEL_THEME["navy"], bold=True, name="Aptos")
        ws.cell(idx, 1).border = border
        target = ws[cell_ref.replace("$", "")]
        target.fill = PatternFill("solid", fgColor="DDE8FF")
        target.font = Font(color=EXCEL_THEME["royal"], bold=True, name="Aptos")
        target.border = border
        target.alignment = Alignment(horizontal="left")

    _excel_v50_add_data_validation_lists(wb, key_df, ws)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34

    ws.merge_cells("D4:I4")
    ws["D4"] = "INDICADORES SEGÚN LOS FILTROS DE ESTA HOJA"
    ws["D4"].fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    ws["D4"].font = Font(color=EXCEL_THEME["white"], bold=True, name="Aptos Display", size=12)
    ws["D4"].alignment = Alignment(horizontal="left")

    kpis = [
        ("D5:F5", "D6:F8", "EQUIPOS", _excel_v50_filter_formula(range_map), EXCEL_THEME["royal"]),
        ("G5:I5", "G6:I8", "EN RUTINA", _excel_v50_filter_formula(range_map, f"--({range_map['__RoutineFlag']}=1)"), EXCEL_THEME["green"]),
        ("D9:F9", "D10:F12", "0 TESTS/DÍA", _excel_v50_filter_formula(range_map, f"--({range_map['__ZeroTestsFlag']}=1)"), EXCEL_THEME["coral"]),
        ("G9:I9", "G10:I12", "PM VENCIDOS", _excel_v50_filter_formula(range_map, f"--({range_map['__PMOverdueFlag']}=1)"), EXCEL_THEME["amber"]),
        ("J5:L5", "J6:L8", "CON FABRICACIÓN", _excel_v50_filter_formula(range_map, f"--({range_map['__ManufacturingFlag']}=1)"), EXCEL_THEME["slate"]),
        ("J9:L9", "J10:L12", "BANCO DE SANGRE", _excel_v50_filter_formula(range_map, f'--({range_map["Blood Bank group"]}="Banco de sangre")'), EXCEL_THEME["purple"]),
    ]
    for label_range, value_range, label, formula, color in kpis:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=color)
        label_cell.font = Font(color=EXCEL_THEME["white"], bold=True, size=9, name="Aptos")
        label_cell.alignment = Alignment(horizontal="center")
        label_cell.border = border
        value_cell.value = formula
        value_cell.fill = PatternFill("solid", fgColor=EXCEL_THEME["panel"])
        value_cell.font = Font(color=color, bold=True, size=21, name="Aptos Display")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = "#,##0"
        value_cell.border = border

    ws.merge_cells("A15:L15")
    ws["A15"] = "Las tablas siguientes se recalculan con los selectores superiores. Las categorías mantienen un orden estable para facilitar comparación."
    ws["A15"].fill = PatternFill("solid", fgColor="E7F7EF")
    ws["A15"].font = Font(color=EXCEL_THEME["green"], italic=True, name="Aptos", size=9)
    ws["A15"].alignment = Alignment(horizontal="left")

    _, model_end = _excel_v50_dynamic_summary_table(ws, key_df, range_map, "Instrument type", "Equipos por modelo", 17, 1)
    _, status_end = _excel_v50_dynamic_summary_table(ws, key_df, range_map, "Operational status grouped", "Estado operativo", 17, 5)
    _, country_end = _excel_v50_dynamic_summary_table(ws, key_df, range_map, "Country", "Países", max(model_end, status_end) + 3, 1)
    _excel_v50_dynamic_summary_table(ws, key_df, range_map, "Operating System", "Sistemas operativos", max(model_end, status_end) + 3, 5)

    ws["E8"].number_format = "#,##0"
    ws.conditional_formatting.add("D6:L12", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FDE2E7")))


def _excel_v50_add_native_data_sheet(
    wb,
    sheet_name: str,
    title: str,
    subtitle: str,
    df: pd.DataFrame,
    table_name: str,
    tab_color: str,
    freeze_cell: str,
    key_sheet: bool = False,
):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(sheet_name)
    _excel_v50_sheet_header(ws, title, subtitle, sheet_name, tab_color)
    clean_df = _excel_unique_columns(_excel_clean_dataframe(df))
    header_row, start_col, n_rows, n_cols = _excel_write_df(ws, clean_df, start_row=5, start_col=1, title=None)
    final_table_name = _excel_add_native_table(ws, header_row, start_col, n_rows, n_cols, table_name)
    ws.freeze_panes = freeze_cell
    ws.sheet_view.zoomScale = 85 if key_sheet else 70

    # Widths are controlled instead of fully autofitting 100+ technical fields.
    for idx, column in enumerate(clean_df.columns, start=1):
        name = str(column)
        letter = get_column_letter(idx)
        if name in {"Customer name", "Distributor name", "Address", "Notes", "Machine Configurations"}:
            ws.column_dimensions[letter].width = 30
        elif "date" in name.lower():
            ws.column_dimensions[letter].width = 14
        elif name in {"Serial number", "Instrument type", "Operational status", "Operating System"}:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = min(max(len(name) + 2, 11), 18)

    for row in range(header_row + 1, header_row + n_rows + 1):
        ws.row_dimensions[row].height = 18
    for tech_column in EXCEL_V50_TECH_COLUMNS:
        if tech_column in clean_df.columns:
            idx = list(clean_df.columns).index(tech_column) + 1
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    ws.auto_filter.ref = None  # la tabla nativa ya contiene sus propios filtros
    ws["L3"] = "Filtros nativos disponibles en cada encabezado"
    ws["L3"].font = Font(color=EXCEL_THEME["teal"], italic=True, name="Aptos", size=9)
    ws["L3"].alignment = Alignment(horizontal="right")
    return ws, final_table_name


def _excel_v50_apply_data_bars(ws, start_row: int, end_row: int, column: int, color: str) -> None:
    if end_row < start_row:
        return
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    letter = get_column_letter(column)
    ws.conditional_formatting.add(
        f"{letter}{start_row}:{letter}{end_row}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=color, showValue=True),
    )


def build_dashboard_excel_export(
    filtered_df: pd.DataFrame,
    filter_summary: dict[str, str] | None,
    active_tab: str | None = None,
    source_label_value: str = "",
    stock_context: dict | None = None,
    active_dashboard_tab: str | None = None,
    include_visual_dashboard: bool = True,
) -> bytes:
    """Genera el Excel v50 con una arquitectura estable y comprensible.

    1. ``00_Dashboard`` es una fotografía ejecutiva garantizada: usa imágenes PNG
       embebidas, por lo que nunca abre con gráficos vacíos.
    2. ``01_Filtros`` ofrece selectores editables y resúmenes mediante fórmulas
       ``SUMPRODUCT`` no volátiles; no usa OFFSET ni depende del AutoFilter.
    3. ``02_Datos_clave`` limita la vista operativa a los campos principales.
    4. ``03_Datos_completos`` conserva la totalidad de la información técnica.
    5. Las hojas analíticas restantes se identifican como resúmenes estáticos de
       la vista exportada para evitar contradicciones sobre su comportamiento.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule

    active_tab_value = active_dashboard_tab or active_tab or "Dashboard"
    filter_summary = dict(filter_summary or {})
    stock_context = stock_context or {}

    source_df = filtered_df.copy() if isinstance(filtered_df, pd.DataFrame) else pd.DataFrame()
    export_df = source_df.drop(columns=[c for c in source_df.columns if str(c).startswith("FLAG::")], errors="ignore").copy()
    export_df = export_df[_preferred_export_columns(export_df)] if not export_df.empty else export_df
    export_df = _excel_unique_columns(export_df)
    key_df = _excel_v50_prepare_key_data(export_df)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "00_Resumen"
    _excel_add_readme(summary_ws, filter_summary, len(export_df), active_tab_value, source_label_value)
    _excel_v50_navigation(summary_ws, current="00_Resumen")

    if include_visual_dashboard:
        _excel_v50_build_snapshot_dashboard(wb, export_df, filter_summary, active_tab_value, source_label_value)

    key_ws, key_table_name = _excel_v50_add_native_data_sheet(
        wb,
        "02_Datos_clave",
        "DATOS CLAVE · VISTA OPERATIVA",
        "Campos esenciales para consulta y filtros rápidos. Las columnas auxiliares del explorador permanecen ocultas.",
        key_df,
        "RecordsKeyData",
        EXCEL_THEME["green"],
        "G6",
        key_sheet=True,
    )

    key_range_map = _excel_v50_key_range_map(key_df, "02_Datos_clave", header_row=5)
    _excel_v50_build_filter_explorer(wb, key_df, key_range_map)

    _excel_v50_add_native_data_sheet(
        wb,
        "03_Datos_completos",
        "DATOS COMPLETOS · DETALLE TÉCNICO",
        "Incluye todos los campos exportados. Usa esta hoja para auditoría o análisis profundo; la operación diaria debe realizarse en Datos clave.",
        export_df,
        "RecordsCompleteData",
        EXCEL_THEME["slate"],
        "G6",
        key_sheet=False,
    )

    # 04 Base instalada · tablas estáticas, sin gráficos nativos que puedan abrir vacíos.
    base_ws = wb.create_sheet("04_Base_instalada")
    _excel_v50_sheet_header(base_ws, "BASE INSTALADA", "Resumen estático de la vista exportada.", "04_Base_instalada", EXCEL_THEME["royal"])
    _excel_v50_static_banner(base_ws, 5)
    type_df = _excel_value_counts_df(export_df, "Instrument type", "Modelo")
    r1, c1, n1, nc1 = _excel_write_df(base_ws, type_df, 7, 1, "Base instalada por modelo")
    _excel_v50_apply_data_bars(base_ws, r1 + 1, r1 + n1, c1 + 1, EXCEL_THEME["royal"])
    country_df = _excel_value_counts_df(export_df, "Country", "País", top_n=30)
    r2, c2, n2, nc2 = _excel_write_df(base_ws, country_df, 7, 5, "Países")
    _excel_v50_apply_data_bars(base_ws, r2 + 1, r2 + n2, c2 + 1, EXCEL_THEME["teal"])
    city_df = _excel_value_counts_df(export_df.assign(CityLabel=build_city_label_series(export_df)), "CityLabel", "Ciudad | País", top_n=40) if "City" in export_df.columns else pd.DataFrame()
    r3, c3, n3, nc3 = _excel_write_df(base_ws, city_df, max(r1 + n1, r2 + n2) + 4, 1, "Ciudades")
    _excel_v50_apply_data_bars(base_ws, r3 + 1, r3 + n3, c3 + 1, EXCEL_THEME["purple"])

    # 05 Modelo por estado · matriz y mapa de calor.
    matrix_ws = wb.create_sheet("05_Modelo_estado")
    _excel_v50_sheet_header(matrix_ws, "MODELO POR ESTADO OPERATIVO", "Matriz estática de la vista exportada.", "05_Modelo_estado", EXCEL_THEME["green"])
    _excel_v50_static_banner(matrix_ws, 5)
    model_status_df = _excel_prepare_model_status_matrix(export_df)
    rm, cm, nm, ncm = _excel_write_df(matrix_ws, model_status_df, 7, 1, "Base instalada por modelo y estado")
    if nm > 0 and ncm > 2:
        from openpyxl.utils import get_column_letter
        start = f"{get_column_letter(cm + 1)}{rm + 1}"
        end = f"{get_column_letter(cm + ncm - 2)}{rm + nm}"
        matrix_ws.conditional_formatting.add(
            f"{start}:{end}",
            ColorScaleRule(start_type="min", start_color="FFFFFF", mid_type="percentile", mid_value=50, mid_color="FFF3CD", end_type="max", end_color="63BE7B"),
        )

    # 06 Machine configuration · cantidades y porcentajes separados.
    cfg_ws = wb.create_sheet("06_Machine_config")
    _excel_v50_sheet_header(cfg_ws, "MACHINE CONFIGURATION", "Resumen estático de configuración; cantidades y porcentajes se presentan en columnas separadas.", "06_Machine_config", EXCEL_THEME["purple"])
    _excel_v50_static_banner(cfg_ws, 5)
    blood_yes = count_blood_bank_yes(export_df)
    blood_df = pd.DataFrame({"Categoría": ["Banco de sangre", "Laboratorio"], "Cantidad": [blood_yes, max(len(export_df) - blood_yes, 0)]})
    rb, cb, nb, ncb = _excel_write_df(cfg_ws, blood_df, 7, 1, "Banco de sangre")
    _excel_v50_apply_data_bars(cfg_ws, rb + 1, rb + nb, cb + 1, EXCEL_THEME["royal"])
    cfg_coverage, cfg_values = _excel_prepare_config_summary(export_df)
    rc, cc, nc, ncc = _excel_write_df(cfg_ws, cfg_coverage, 7, 5, "Cobertura por campo")
    _excel_v50_apply_data_bars(cfg_ws, rc + 1, rc + nc, cc + 1, EXCEL_THEME["purple"])
    rv, cv, nv, ncv = _excel_write_df(cfg_ws, cfg_values, max(rb + nb, rc + nc) + 4, 1, "Valores principales por campo")
    _excel_v50_apply_data_bars(cfg_ws, rv + 1, rv + nv, cv + 2, EXCEL_THEME["teal"])

    # 07 OS y PM · evita mezclar equipos con tests totales en un mismo gráfico.
    os_ws = wb.create_sheet("07_OS_PM")
    _excel_v50_sheet_header(os_ws, "SISTEMA OPERATIVO · PROCESAMIENTO · PM", "Resumen estático con unidades separadas para mantener lectura correcta.", "07_OS_PM", EXCEL_THEME["amber"])
    _excel_v50_static_banner(os_ws, 5)
    os_df = _excel_value_counts_df(export_df, "Operating System", "Sistema operativo")
    ro, co, no, nco = _excel_write_df(os_ws, os_df, 7, 1, "Sistemas operativos")
    _excel_v50_apply_data_bars(os_ws, ro + 1, ro + no, co + 1, EXCEL_THEME["teal"])
    tests = pd.to_numeric(export_df.get("Number of tests per day", pd.Series(index=export_df.index, dtype=float)), errors="coerce").fillna(0)
    tests_summary = pd.DataFrame({
        "Métrica": ["Equipos", "Equipos con 0 tests/día", "% con 0 tests/día", "Promedio tests/día", "Máximo tests/día"],
        "Valor": [len(export_df), int(tests.le(0).sum()), round(tests.le(0).mean() * 100, 1) if len(tests) else 0, round(float(tests.mean()), 1) if len(tests) else 0, round(float(tests.max()), 1) if len(tests) else 0],
    })
    rt, ct, nt, nct = _excel_write_df(os_ws, tests_summary, 7, 5, "Indicadores de procesamiento")
    tests_by_model = export_df.assign(**{"Number of tests per day": tests}).groupby("Instrument type", dropna=False)["Number of tests per day"].agg(["count", "sum", "mean", "max"]).reset_index()
    tests_by_model.columns = ["Modelo", "Equipos", "Tests/día total", "Promedio tests/día", "Máximo tests/día"]
    rpm, cpm, npm, ncpm = _excel_write_df(os_ws, tests_by_model, max(ro + no, rt + nt) + 4, 1, "Procesamiento por modelo · métricas separadas")
    _excel_v50_apply_data_bars(os_ws, rpm + 1, rpm + npm, cpm + 2, EXCEL_THEME["green"])
    pm_next = pd.to_datetime(export_df.get("PM next date", pd.Series(index=export_df.index, dtype="object")), errors="coerce")
    today = pd.Timestamp.today().normalize()
    pm_status = pd.Series(np.select([pm_next.notna() & pm_next.lt(today), pm_next.notna() & pm_next.le(today + pd.Timedelta(days=90)), pm_next.notna()], ["Vencido", "Próximos 90 días", "Planificado"], default="No informado"), index=export_df.index)
    pm_df = pm_status.value_counts().reset_index()
    pm_df.columns = ["Estado PM", "Cantidad"]
    rps, cps, nps, ncps = _excel_write_df(os_ws, pm_df, rpm + npm + 4, 1, "Estado PM")
    _excel_v50_apply_data_bars(os_ws, rps + 1, rps + nps, cps + 1, EXCEL_THEME["amber"])

    # 08 Fabricación si existe información.
    manuf_cols = ["Manufacturing Date", "Manufacturing year", "Manufacturing age (years)", "Manufacturing age bucket", "Manufacturing matched"]
    if any(col in export_df.columns for col in manuf_cols):
        manuf_ws = wb.create_sheet("08_Fabricacion")
        _excel_v50_sheet_header(manuf_ws, "ANTIGÜEDAD POR FABRICACIÓN", "Resumen estático y detalle del cruce por serial.", "08_Fabricacion", EXCEL_THEME["slate"])
        _excel_v50_static_banner(manuf_ws, 5)
        matched = export_df.get("Manufacturing matched", pd.Series(False, index=export_df.index))
        try:
            matched_bool = matched.fillna(False).astype(bool)
        except Exception:
            matched_bool = matched.astype(str).str.lower().isin({"true", "1", "yes", "si", "sí"})
        age = pd.to_numeric(export_df.get("Manufacturing age (years)", pd.Series(index=export_df.index, dtype=float)), errors="coerce")
        manuf_summary = pd.DataFrame({
            "Métrica": ["Equipos evaluados", "Con fecha de fabricación", "Sin coincidencia", "% con fecha", "Edad promedio", "Más antiguo", "Más nuevo"],
            "Valor": [len(export_df), int(matched_bool.sum()), int(len(export_df) - matched_bool.sum()), round(matched_bool.mean() * 100, 1) if len(export_df) else 0, round(float(age.mean()), 1) if age.notna().any() else "No informado", round(float(age.max()), 1) if age.notna().any() else "No informado", round(float(age.min()), 1) if age.notna().any() else "No informado"],
        })
        rs, cs, ns, ncs = _excel_write_df(manuf_ws, manuf_summary, 7, 1, "Resumen de fabricación")
        bucket_df = _excel_value_counts_df(export_df, "Manufacturing age bucket", "Rango de edad")
        rbu, cbu, nbu, ncbu = _excel_write_df(manuf_ws, bucket_df, 7, 5, "Rangos de antigüedad")
        _excel_v50_apply_data_bars(manuf_ws, rbu + 1, rbu + nbu, cbu + 1, EXCEL_THEME["slate"])
        year_df = _excel_value_counts_df(export_df, "Manufacturing year", "Año fabricación")
        ry, cy, ny, ncy = _excel_write_df(manuf_ws, year_df, max(rs + ns, rbu + nbu) + 4, 1, "Equipos por año de fabricación")
        _excel_v50_apply_data_bars(manuf_ws, ry + 1, ry + ny, cy + 1, EXCEL_THEME["teal"])
        detail_cols = [c for c in ["Commercial Region", "Country", "Distributor name", "Customer name", "Instrument type", "Serial number", "Operational status grouped", "Manufacturing Date", "Manufacturing year", "Manufacturing age (years)", "Manufacturing age bucket", "Manufacturing Source", "Manufacturing matched"] if c in export_df.columns]
        detail_df = export_df[detail_cols].copy() if detail_cols else pd.DataFrame()
        if "Manufacturing age (years)" in detail_df.columns:
            detail_df = detail_df.sort_values("Manufacturing age (years)", ascending=False, na_position="last")
        _excel_write_df(manuf_ws, detail_df, ry + ny + 4, 1, "Detalle completo de fabricación")

    if stock_context.get("available"):
        stock_ws = wb.create_sheet("09_Carstock")
        _excel_v50_sheet_header(stock_ws, "CARSTOCK", "Resumen estático del análisis disponible en la app.", "09_Carstock", EXCEL_THEME["coral"])
        _excel_v50_static_banner(stock_ws, 5)
        stock_pairs = pd.DataFrame({
            "Métrica": ["Distribuidor", "SKUs requeridos", "SKUs OK", "SKUs LOW", "SKUs faltantes", "Costo opción 2"],
            "Valor": [stock_context.get("detected_distributor", "N/A"), stock_context.get("required_skus", 0), stock_context.get("ok_skus", 0), stock_context.get("low_skus", 0), stock_context.get("missing_skus", 0), stock_context.get("option2_cost", 0)],
        })
        _excel_write_df(stock_ws, stock_pairs, 7, 1, "Resumen carstock")
        comparison = stock_context.get("full_comparison_df", pd.DataFrame())
        if isinstance(comparison, pd.DataFrame) and not comparison.empty:
            _excel_write_df(stock_ws, comparison, 17, 1, "Comparación completa")

    _excel_style_workbook(wb)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.active = 0
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()



# =============================================================================
# INFORME A CORPORATIVO MENSUAL - LATAM
# =============================================================================
LATAM_COUNTRY_NAMES = {
    "argentina", "bolivia", "brazil", "brasil", "chile", "colombia", "costa rica",
    "cuba", "dominican republic", "republica dominicana", "república dominicana",
    "ecuador", "el salvador", "guatemala", "haiti", "honduras", "mexico", "méxico",
    "nicaragua", "panama", "panamá", "paraguay", "peru", "perú", "uruguay",
    "venezuela", "puerto rico", "jamaica", "trinidad and tobago", "barbados",
    "belize", "guyana", "suriname", "french guiana", "guayana francesa",
}

CORPORATE_PURCHASED_STATUS_KEYWORDS = (
    "warehouse new system",
    "warehouse ready to be installed",
    "warehouse transit customs",
    "ready to be installed",
    "transit customs",
    "new system",
)


def _corporate_latam_mask(df: pd.DataFrame) -> pd.Series:
    """Detecta LATAM por región comercial y, como respaldo, por país.

    El informe corporativo debe ser solo LATAM y por eso no depende de que el
    usuario haya aplicado o no filtros laterales en el dashboard.
    """
    if df is None or df.empty:
        return pd.Series(False, index=getattr(df, "index", []))

    country_series = df["Country"].fillna("").astype(str) if "Country" in df.columns else pd.Series("", index=df.index, dtype="object")
    country_norm = country_series.map(lambda value: normalize_column_label(value))
    country_mask = country_norm.isin({normalize_column_label(c) for c in LATAM_COUNTRY_NAMES})

    if "Commercial Region" in df.columns:
        commercial = df["Commercial Region"].fillna("").astype(str)
        commercial_norm = commercial.str.lower().str.strip()
        commercial_useful = commercial_norm.ne("") & ~commercial_norm.isin({"_not applicable", "not applicable", "n/a", "na", "none"})
        commercial_latam = commercial_norm.str.contains(r"latam|latin america|latinoamerica|latinoamérica", regex=True, na=False)

        # Cuando el export trae Commercial Region, se toma ese campo como fuente
        # primaria para evitar mezclar subsidiarias fuera del alcance Export LATAM.
        # El país solo se usa como respaldo si Commercial Region viene vacío/no útil.
        if commercial_useful.any():
            fallback_mask = (~commercial_useful) & country_mask
            return commercial_latam | fallback_mask

    region_parts = []
    for col in ["World Region", "Region", "Sales Region"]:
        if col in df.columns:
            region_parts.append(df[col].fillna("").astype(str))
    if region_parts:
        region_text = region_parts[0]
        for part in region_parts[1:]:
            region_text = region_text + " " + part
    else:
        region_text = pd.Series("", index=df.index, dtype="object")

    region_norm = region_text.astype(str).str.lower()
    region_mask = region_norm.str.contains(r"latam|latin america|latinoamerica|latinoamérica|central america|south america|caribbean", regex=True, na=False)
    return region_mask | country_mask


def _corporate_status_type(status_value) -> str:
    text = safe_text(status_value, "").strip()
    if not text:
        return "Otros estados"
    normalized = normalize_operational_status(text)
    low = normalized.lower()
    raw_low = text.lower()
    if normalized == "Routine":
        return "En rutina"
    if any(keyword in low or keyword in raw_low for keyword in CORPORATE_PURCHASED_STATUS_KEYWORDS):
        return "Recién comprados / pipeline"
    return "Otros estados"


def _corporate_prepare_latam_report_df(records_df: pd.DataFrame) -> pd.DataFrame:
    if records_df is None or records_df.empty:
        return pd.DataFrame()

    work = records_df.copy()
    for col in ["Commercial Region", "Country", "Distributor name", "Instrument type", "Serial number", "Installation date", "Operational status grouped", "Operational status", "Number of tests per day"]:
        if col not in work.columns:
            work[col] = pd.NA

    latam_mask = _corporate_latam_mask(work)
    work = work[latam_mask].copy()
    if work.empty:
        return work

    work["Country"] = work["Country"].fillna("No informado").astype(str).str.strip().replace("", "No informado")
    work["Distributor name"] = work["Distributor name"].fillna("No informado").astype(str).str.strip().replace("", "No informado")
    work["Instrument type"] = work["Instrument type"].fillna("No informado").astype(str).str.strip().replace("", "No informado")
    work["Serial number"] = work["Serial number"].fillna("No informado").astype(str).str.strip().replace("", "No informado")

    if "Operational status grouped" not in work.columns or work["Operational status grouped"].isna().all():
        work["Operational status grouped"] = work["Operational status"].map(normalize_operational_status)
    else:
        work["Operational status grouped"] = work["Operational status grouped"].fillna(work["Operational status"].map(normalize_operational_status))

    work["Operational status grouped"] = work["Operational status grouped"].fillna("No informado").astype(str).str.strip().replace("", "No informado")
    work["Corporate status type"] = work["Operational status grouped"].map(_corporate_status_type)

    install_dates = pd.to_datetime(work["Installation date"], errors="coerce")
    work["Corporate month"] = install_dates.dt.to_period("M").astype(str)
    work.loc[install_dates.isna(), "Corporate month"] = "Sin fecha"
    work["Installation year"] = install_dates.dt.year.astype("Int64")

    tests = pd.to_numeric(work["Number of tests per day"], errors="coerce").fillna(0).clip(lower=0)
    work["Tests/day actual"] = tests

    routine_mask = work["Corporate status type"].eq("En rutina")
    purchased_mask = work["Corporate status type"].eq("Recién comprados / pipeline")
    routine_avg_by_model = work.loc[routine_mask].groupby("Instrument type")["Tests/day actual"].mean()
    global_routine_avg = float(work.loc[routine_mask, "Tests/day actual"].mean()) if routine_mask.any() else 0.0
    if math.isnan(global_routine_avg):
        global_routine_avg = 0.0

    work["Model routine avg tests/day"] = work["Instrument type"].map(routine_avg_by_model).fillna(global_routine_avg).fillna(0)
    work["Routine count"] = routine_mask.astype(int)
    work["Recently purchased count"] = purchased_mask.astype(int)
    work["Corporate projection units"] = (routine_mask | purchased_mask).astype(int)
    work["Routine actual tests/day"] = np.where(routine_mask, work["Tests/day actual"], 0.0)
    work["New systems projected tests/day"] = np.where(purchased_mask, work["Model routine avg tests/day"], 0.0)
    work["Projected tests/day"] = work["Routine actual tests/day"] + work["New systems projected tests/day"]
    return work


def _corporate_group_summary(work: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if work is None or work.empty:
        return pd.DataFrame()
    grouped = (
        work.groupby(group_cols, dropna=False)
        .agg(
            **{
                "Base instalada actualizada": ("Serial number", "count"),
                "Equipos en rutina": ("Routine count", "sum"),
                "Recién comprados / pipeline": ("Recently purchased count", "sum"),
                "Base comercial estimada": ("Corporate projection units", "sum"),
                "Tests/día rutina actual": ("Routine actual tests/day", "sum"),
                "Tests/día nuevos estimados": ("New systems projected tests/day", "sum"),
                "Proyección tests/día aproximada": ("Projected tests/day", "sum"),
            }
        )
        .reset_index()
    )
    numeric_cols = [
        "Base instalada actualizada", "Equipos en rutina", "Recién comprados / pipeline",
        "Base comercial estimada", "Tests/día rutina actual", "Tests/día nuevos estimados",
        "Proyección tests/día aproximada",
    ]
    for col in numeric_cols:
        if col in grouped.columns:
            grouped[col] = pd.to_numeric(grouped[col], errors="coerce").fillna(0).round(1)
            if col.startswith("Base") or col.startswith("Equipos") or col.startswith("Recién"):
                grouped[col] = grouped[col].astype(int)
    if "Base comercial estimada" in grouped.columns:
        grouped["% rutina sobre base actual"] = (grouped["Equipos en rutina"] * 100 / grouped["Base instalada actualizada"].replace(0, np.nan)).round(1).fillna(0)
        grouped["% pipeline sobre base actual"] = (grouped["Recién comprados / pipeline"] * 100 / grouped["Base instalada actualizada"].replace(0, np.nan)).round(1).fillna(0)
    sort_cols = [c for c in ["Corporate month", "Base comercial estimada", "Distributor name", "Instrument type"] if c in grouped.columns]
    if sort_cols:
        ascending = [True if c == "Corporate month" else False if c == "Base comercial estimada" else True for c in sort_cols]
        grouped = grouped.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)
    return grouped


def _corporate_sort_monthly_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "Corporate month" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    work = df.copy()
    work["_month_sort"] = pd.to_datetime(work["Corporate month"].astype(str) + "-01", errors="coerce")
    work["_month_sort"] = work["_month_sort"].fillna(pd.Timestamp.max)
    work = work.sort_values(["_month_sort", "Corporate month"]).drop(columns=["_month_sort"]).reset_index(drop=True)
    return work


def _corporate_add_cumulative_columns(monthly_df: pd.DataFrame) -> pd.DataFrame:
    if monthly_df is None or monthly_df.empty:
        return pd.DataFrame()
    work = _corporate_sort_monthly_df(monthly_df)
    numeric_cols = [
        "Base instalada actualizada", "Equipos en rutina", "Recién comprados / pipeline",
        "Base comercial estimada", "Tests/día rutina actual", "Tests/día nuevos estimados",
        "Proyección tests/día aproximada",
    ]
    for col in numeric_cols:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0)
    if "Corporate month" in work.columns:
        sin_fecha_mask = work["Corporate month"].astype(str).eq("Sin fecha")
    else:
        sin_fecha_mask = pd.Series(False, index=work.index)
    regular = work[~sin_fecha_mask].copy()
    no_date = work[sin_fecha_mask].copy()
    for col in numeric_cols:
        if col in regular.columns:
            regular[f"{col} acumulado"] = regular[col].cumsum()
    if not no_date.empty:
        # Sin fecha se mantiene como bloque separado; no se mezcla con la línea temporal.
        for col in numeric_cols:
            if col in no_date.columns:
                no_date[f"{col} acumulado"] = no_date[col]
    return pd.concat([regular, no_date], ignore_index=True)


def _corporate_spanish_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    rename_map = {
        "Corporate month": "Mes de instalación",
        "Instrument type": "Modelo",
        "Distributor name": "Distribuidor",
        "Country": "País",
        "Corporate status type": "Clasificación corporativa",
        "Base instalada actualizada": "Base instalada actualizada",
        "Equipos en rutina": "Equipos en rutina",
        "Recién comprados / pipeline": "Recién comprados / pipeline",
        "Base comercial estimada": "Base comercial estimada",
        "% rutina sobre base actual": "% rutina sobre base actual",
        "% pipeline sobre base actual": "% pipeline sobre base actual",
        "Tests/día rutina actual": "Tests/día rutina actual",
        "Tests/día nuevos estimados": "Tests/día nuevos estimados",
        "Proyección tests/día aproximada": "Proyección venta/consumo aproximada (tests/día)",
        "Base instalada actualizada acumulado": "Base instalada acumulada",
        "Equipos en rutina acumulado": "Equipos en rutina acumulados",
        "Recién comprados / pipeline acumulado": "Recién comprados/pipeline acumulados",
        "Base comercial estimada acumulado": "Base comercial estimada acumulada",
        "Tests/día rutina actual acumulado": "Tests/día rutina actual acumulado",
        "Tests/día nuevos estimados acumulado": "Tests/día nuevos estimados acumulado",
        "Proyección tests/día aproximada acumulado": "Proyección venta/consumo acumulada (tests/día)",
    }
    out = df.copy().rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    return out


def _corporate_global_metrics(work: pd.DataFrame) -> dict[str, float | int]:
    if work is None or work.empty:
        return {
            "base": 0, "routine": 0, "pipeline": 0, "commercial_base": 0,
            "routine_pct": 0.0, "pipeline_pct": 0.0, "routine_tests": 0.0,
            "new_tests": 0.0, "projected_tests": 0.0, "distributors": 0,
            "countries": 0, "models": 0,
        }
    base = int(len(work))
    routine = int(pd.to_numeric(work["Routine count"], errors="coerce").fillna(0).sum())
    pipeline = int(pd.to_numeric(work["Recently purchased count"], errors="coerce").fillna(0).sum())
    commercial_base = int(pd.to_numeric(work["Corporate projection units"], errors="coerce").fillna(0).sum())
    return {
        "base": base,
        "routine": routine,
        "pipeline": pipeline,
        "commercial_base": commercial_base,
        "routine_pct": round(routine * 100 / max(base, 1), 1),
        "pipeline_pct": round(pipeline * 100 / max(base, 1), 1),
        "routine_tests": round(float(pd.to_numeric(work["Routine actual tests/day"], errors="coerce").fillna(0).sum()), 1),
        "new_tests": round(float(pd.to_numeric(work["New systems projected tests/day"], errors="coerce").fillna(0).sum()), 1),
        "projected_tests": round(float(pd.to_numeric(work["Projected tests/day"], errors="coerce").fillna(0).sum()), 1),
        "distributors": int(work["Distributor name"].nunique(dropna=True)),
        "countries": int(work["Country"].nunique(dropna=True)),
        "models": int(work["Instrument type"].nunique(dropna=True)),
    }


def _corporate_write_text_block(ws, rows: list[tuple[str, str]], start_row: int, start_col: int = 1, title: str | None = None) -> int:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    row = start_row
    if title:
        ws.cell(row, start_col, title)
        ws.cell(row, start_col).font = Font(bold=True, size=13, color="FFFFFF")
        ws.cell(row, start_col).fill = PatternFill("solid", fgColor="0B1B2A")
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 5)
        row += 1
    label_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Side(style="thin", color="D9EAF7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for label, value in rows:
        ws.cell(row, start_col, label)
        ws.cell(row, start_col + 1, value)
        ws.cell(row, start_col).font = Font(bold=True, color="0B1B2A")
        ws.cell(row, start_col).fill = label_fill
        ws.cell(row, start_col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, start_col + 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, start_col).border = border
        ws.cell(row, start_col + 1).border = border
        row += 1
    return row


def _corporate_add_column_chart(ws, data_start_row: int, data_start_col: int, n_rows: int, n_cols: int, title: str, anchor: str, stacked: bool = False) -> None:
    if n_rows <= 0 or n_cols < 2:
        return
    from openpyxl.chart import BarChart, Reference
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Cantidad / tests por día"
    chart.x_axis.title = "Mes / categoría"
    chart.height = 8
    chart.width = 17
    if stacked:
        chart.grouping = "stacked"
        chart.overlap = 100
    data = Reference(ws, min_col=data_start_col + 1, max_col=data_start_col + n_cols - 1, min_row=data_start_row, max_row=data_start_row + n_rows)
    cats = Reference(ws, min_col=data_start_col, min_row=data_start_row + 1, max_row=data_start_row + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def build_monthly_corporate_latam_report_excel(records_df: pd.DataFrame, source_label_value: str = "") -> bytes:
    """Genera un informe corporativo LATAM simple y ejecutivo.

    Enfoque exclusivo solicitado:
    - Base instalada actualizada.
    - Cuántos equipos están en rutina por mes.
    - Recién comprados por mes.
    - Base proyectada = equipos en rutina + recién comprados.

    No incluye proyección monetaria ni tests/día para evitar confusión en la lectura corporativa.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output = BytesIO()
    wb = Workbook()

    work = _corporate_prepare_latam_report_df(records_df)

    def _safe_int(value) -> int:
        try:
            return int(round(float(value)))
        except Exception:
            return 0

    def _projection_summary(group_cols: list[str]) -> pd.DataFrame:
        if work is None or work.empty:
            columns = list(group_cols) + [
                "Base instalada actualizada",
                "Equipos en rutina",
                "Recién comprados",
                "Base proyectada",
                "% base proyectada sobre base instalada",
            ]
            return pd.DataFrame(columns=columns)
        summary = (
            work.groupby(group_cols, dropna=False)
            .agg(
                **{
                    "Base instalada actualizada": ("Serial number", "count"),
                    "Equipos en rutina": ("Routine count", "sum"),
                    "Recién comprados": ("Recently purchased count", "sum"),
                }
            )
            .reset_index()
        )
        for col in ["Base instalada actualizada", "Equipos en rutina", "Recién comprados"]:
            summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0).astype(int)
        summary["Base proyectada"] = summary["Equipos en rutina"] + summary["Recién comprados"]
        summary["% base proyectada sobre base instalada"] = (
            summary["Base proyectada"] * 100 / summary["Base instalada actualizada"].replace(0, np.nan)
        ).round(1).fillna(0)
        if "Corporate month" in summary.columns:
            summary = _corporate_sort_monthly_df(summary)
        else:
            sort_cols = [c for c in ["Base proyectada", "Distributor name", "Instrument type"] if c in summary.columns]
            if sort_cols:
                asc = [False if c == "Base proyectada" else True for c in sort_cols]
                summary = summary.sort_values(sort_cols, ascending=asc).reset_index(drop=True)
        return summary

    def _rename_clear(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame()
        rename_map = {
            "Corporate month": "Mes",
            "Instrument type": "Modelo",
            "Distributor name": "Distribuidor",
            "Country": "País",
            "Corporate status type": "Clasificación",
        }
        out = df.copy().rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
        if "Clasificación" in out.columns:
            out["Clasificación"] = out["Clasificación"].replace({"Recién comprados / pipeline": "Recién comprados"})
        return out

    def _add_note(ws, cell: str, text: str) -> None:
        ws[cell] = text
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
        ws[cell].font = Font(italic=True, color="44546A")

    def _add_summary_cards(ws, metrics_rows: list[tuple[str, str]], start_row: int = 5) -> int:
        row = start_row
        fill_label = PatternFill("solid", fgColor="D9EAF7")
        fill_value = PatternFill("solid", fgColor="F8FBFD")
        thin = Side(style="thin", color="C9DDED")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for label, value in metrics_rows:
            ws.cell(row, 1, label)
            ws.cell(row, 2, value)
            ws.cell(row, 1).font = Font(bold=True, color="0B1B2A")
            ws.cell(row, 1).fill = fill_label
            ws.cell(row, 2).fill = fill_value
            ws.cell(row, 1).border = border
            ws.cell(row, 2).border = border
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        return row

    ws = wb.active
    ws.title = "00_Resumen"
    ws["A1"] = "Informe a corporativo mensual - LATAM"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B1B2A")
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Código creado: {CODE_CREATED_AT} · {CODE_VERSION_LABEL}"
    ws["A3"] = f"Fuente activa: {source_label_value}"
    ws["A4"] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    _add_note(
        ws,
        "A12",
        "Lectura ejecutiva: este informe usa únicamente LATAM. La columna 'Mes' se toma de Installation date. "
        "La base proyectada se calcula con una fórmula simple: Equipos en rutina + Recién comprados. "
        "No se incluye proyección monetaria ni tests/día para mantener el informe claro y auditable.",
    )
    ws.merge_cells("A12:H13")

    if work.empty:
        _excel_write_df(ws, pd.DataFrame({"Mensaje": ["No se encontraron registros LATAM para generar el informe."]}), 15, 1, "Resultado")
        _excel_style_workbook(wb)
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    total_base = _safe_int(len(work))
    total_routine = _safe_int(pd.to_numeric(work["Routine count"], errors="coerce").fillna(0).sum())
    total_purchased = _safe_int(pd.to_numeric(work["Recently purchased count"], errors="coerce").fillna(0).sum())
    total_projected = total_routine + total_purchased
    projected_pct = round(total_projected * 100 / max(total_base, 1), 1)
    routine_pct = round(total_routine * 100 / max(total_base, 1), 1)
    purchased_pct = round(total_purchased * 100 / max(total_base, 1), 1)

    metrics_rows = [
        ("Base instalada actualizada LATAM", f"{total_base:,} equipos"),
        ("Equipos en rutina", f"{total_routine:,} equipos ({routine_pct}% de la base instalada)"),
        ("Recién comprados", f"{total_purchased:,} equipos ({purchased_pct}% de la base instalada)"),
        ("Base proyectada", f"{total_projected:,} equipos = rutina + recién comprados"),
        ("% base proyectada sobre base instalada", f"{projected_pct}%"),
        ("Distribuidores LATAM", f"{work['Distributor name'].nunique(dropna=True):,}"),
        ("Países LATAM", f"{work['Country'].nunique(dropna=True):,}"),
        ("Modelos incluidos", f"{work['Instrument type'].nunique(dropna=True):,}"),
    ]
    _add_summary_cards(ws, metrics_rows, 15)

    formula_df = pd.DataFrame({
        "Concepto": [
            "Base instalada actualizada",
            "Equipos en rutina",
            "Recién comprados",
            "Base proyectada",
        ],
        "Qué significa": [
            "Total de equipos LATAM presentes en el Records List actual.",
            "Equipos con estado operativo Routine. Es la base actualmente productiva.",
            "Equipos comprados o próximos a entrar en operación. Incluye estados tipo Warehouse New System, Warehouse Ready To Be Installed y Warehouse Transit Customs.",
            "Suma simple de equipos en rutina + recién comprados. Esta es la base de proyección solicitada.",
        ],
    })
    _excel_write_df(ws, formula_df, 15, 4, "Definiciones del informe")

    overview_df = pd.DataFrame({
        "Indicador": ["Base instalada", "En rutina", "Recién comprados", "Base proyectada"],
        "Cantidad": [total_base, total_routine, total_purchased, total_projected],
    })
    r_over, c_over, n_over, nc_over = _excel_write_df(ws, overview_df, 25, 1, "Vista ejecutiva de la proyección")
    _excel_add_bar_chart(ws, r_over, c_over, n_over, nc_over, "Base instalada vs base proyectada", "D25", stacked=False)

    monthly = _projection_summary(["Corporate month"])
    monthly_out = _rename_clear(monthly)
    monthly_out = monthly_out[[c for c in [
        "Mes",
        "Base instalada actualizada",
        "Equipos en rutina",
        "Recién comprados",
        "Base proyectada",
        "% base proyectada sobre base instalada",
    ] if c in monthly_out.columns]].copy()
    if "Mes" in monthly_out.columns:
        monthly_out["Lectura"] = monthly_out.apply(
            lambda row: f"En {row['Mes']}: {int(row['Equipos en rutina'])} en rutina + {int(row['Recién comprados'])} recién comprados = {int(row['Base proyectada'])} equipos proyectados.",
            axis=1,
        )
    ws_month = wb.create_sheet("01_Mensual_LATAM")
    _add_note(
        ws_month,
        "A1",
        "Esta hoja responde la pregunta principal por mes: cuántos equipos hay en rutina, cuántos son recién comprados y cuál es la base proyectada resultante.",
    )
    ws_month.merge_cells("A1:H2")
    r, c, n, nc = _excel_write_df(ws_month, monthly_out, 4, 1, "Proyección mensual LATAM")
    chart_cols = ["Mes", "Equipos en rutina", "Recién comprados", "Base proyectada"]
    chart_df = monthly_out[[col for col in chart_cols if col in monthly_out.columns]].copy() if not monthly_out.empty else pd.DataFrame()
    r_chart, c_chart, n_chart, nc_chart = _excel_write_df(ws_month, chart_df, r + n + 4, 1, "Rutina + recién comprados por mes")
    _corporate_add_column_chart(ws_month, r_chart, c_chart, n_chart, nc_chart, "Base proyectada por mes", "J4", stacked=False)

    monthly_model = _projection_summary(["Corporate month", "Instrument type"])
    monthly_model_out = _rename_clear(monthly_model)
    monthly_model_out = monthly_model_out[[c for c in [
        "Mes", "Modelo", "Base instalada actualizada", "Equipos en rutina", "Recién comprados", "Base proyectada", "% base proyectada sobre base instalada"
    ] if c in monthly_model_out.columns]].copy()
    ws_model_month = wb.create_sheet("02_Mensual_modelo")
    _add_note(ws_model_month, "A1", "Misma lectura mensual, separada por modelo para entender qué plataforma aporta a la base proyectada.")
    ws_model_month.merge_cells("A1:H2")
    _excel_write_df(ws_model_month, monthly_model_out, 4, 1, "Proyección mensual por modelo")

    dist_model = _projection_summary(["Distributor name", "Country", "Instrument type"])
    dist_model_out = _rename_clear(dist_model)
    dist_model_out = dist_model_out[[c for c in [
        "Distribuidor", "País", "Modelo", "Base instalada actualizada", "Equipos en rutina", "Recién comprados", "Base proyectada", "% base proyectada sobre base instalada"
    ] if c in dist_model_out.columns]].copy()
    if not dist_model_out.empty and "Base proyectada" in dist_model_out.columns:
        dist_model_out = dist_model_out.sort_values(["Base proyectada", "Distribuidor", "Modelo"], ascending=[False, True, True]).reset_index(drop=True)
    ws_dist = wb.create_sheet("03_Distribuidor_modelo")
    _add_note(ws_dist, "A1", "Resumen para seguimiento con distribuidores. Cada fila muestra la base instalada, rutina, recién comprados y base proyectada por modelo.")
    ws_dist.merge_cells("A1:H2")
    r, c, n, nc = _excel_write_df(ws_dist, dist_model_out, 4, 1, "Proyección por distribuidor y modelo")
    top_dist = (
        dist_model_out.groupby("Distribuidor", as_index=False)["Base proyectada"].sum().sort_values("Base proyectada", ascending=False).head(15)
        if not dist_model_out.empty and "Distribuidor" in dist_model_out.columns and "Base proyectada" in dist_model_out.columns
        else pd.DataFrame()
    )
    r_top, c_top, n_top, nc_top = _excel_write_df(ws_dist, top_dist, r + n + 4, 1, "Top distribuidores por base proyectada")
    _excel_add_bar_chart(ws_dist, r_top, c_top, n_top, nc_top, "Top distribuidores por base proyectada", "J4", stacked=False)

    detail_cols = [
        "Country", "Distributor name", "Customer name", "City", "Instrument type", "Serial number",
        "Installation date", "Corporate month", "Operational status grouped", "Corporate status type",
        "Asset condition", "Type of contract", "In Blood Bank", "Operating System",
    ]
    detail_cols = [c for c in detail_cols if c in work.columns]
    detail = work[detail_cols].copy()
    detail = detail.rename(columns={
        "Country": "País",
        "Distributor name": "Distribuidor",
        "Customer name": "Cliente",
        "City": "Ciudad",
        "Instrument type": "Modelo",
        "Serial number": "Serial",
        "Installation date": "Fecha de instalación",
        "Corporate month": "Mes",
        "Operational status grouped": "Estado operativo",
        "Corporate status type": "Clasificación",
        "Asset condition": "Condición del activo",
        "Type of contract": "Tipo de contrato",
        "In Blood Bank": "Banco de sangre",
        "Operating System": "Sistema operativo",
    })
    if "Clasificación" in detail.columns:
        detail["Incluye en base proyectada"] = detail["Clasificación"].isin(["En rutina", "Recién comprados / pipeline"]).map({True: "Sí", False: "No"})
        detail["Clasificación"] = detail["Clasificación"].replace({"Recién comprados / pipeline": "Recién comprados"})
    sort_cols = [c for c in ["País", "Distribuidor", "Modelo", "Clasificación", "Serial"] if c in detail.columns]
    if sort_cols:
        detail = detail.sort_values(sort_cols).reset_index(drop=True)
    ws_detail = wb.create_sheet("04_Detalle")
    _add_note(ws_detail, "A1", "Detalle de los equipos que soportan el informe. Esta hoja sirve para auditoría y validación operativa.")
    ws_detail.merge_cells("A1:H2")
    _excel_write_df(ws_detail, detail, 4, 1, "Detalle LATAM usado en la proyección")

    for sheet in wb.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or "A4"

    _excel_style_workbook(wb)
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def build_equipment_detail_excel(all_rows_df: pd.DataFrame, serial_label: str = "selected") -> bytes:
    from openpyxl import Workbook
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle_equipo"
    _excel_write_df(ws, all_rows_df.copy(), 1, 1, f"Detalle completo del equipo {serial_label}")
    _excel_style_workbook(wb)
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def normalize_operational_status(value) -> str:
    if pd.isna(value):
        return "No informado"
    text = str(value).strip()
    if not text:
        return "No informado"

    upper = re.sub(r"\s+", " ", text.upper()).strip()
    lower = re.sub(r"\s+", " ", text.lower()).strip()

    if upper in {"NOT IN ROUTINE", "NOT ROUTINE", "NO ROUTINE"}:
        return "Not in routine"
    if upper in {"IN ROUTINE", "ROUTINE"}:
        return "Routine"

    # Separar equipos ya descartados de equipos pendientes por descartar.
    # El export real usa variantes como "WAREHOUSE to be scrapped" y
    # "CUSTOMER to be scrapped". Antes se agrupaban dentro de "Scraped"
    # porque la regla genérica buscaba cualquier texto con "scrap".
    if re.search(r"\bto be scrap(?:ped|ed)?\b", lower):
        return "To Be Scrapped"
    if "scrap" in lower:
        return "Scraped"

    if "warehouse" in lower:
        return text.title()
    if "refurb" in lower:
        return text.title()

    return text.title()


def compute_state_filter_counts(df: pd.DataFrame) -> list[tuple[str, int]]:
    if df.empty or "Operational status grouped" not in df.columns:
        return []

    state_series = df["Operational status grouped"].fillna("No informado").astype(str)
    grouped = state_series.value_counts()

    items = []

    # "No rutina" es un acceso rápido: excluye Routine y también excluye No informado.
    non_routine_mask = (~state_series.eq("Routine")) & (~state_series.eq("No informado"))
    non_routine_count = int(non_routine_mask.sum())
    if non_routine_count > 0:
        items.append(("No rutina", non_routine_count))

    preferred_order = [
        "Routine",
        "Not in routine",
        "Scraped",
        "To Be Scrapped",
        "Warehouse New System",
        "Warehouse To Be Refurbished",
        "Warehouse Ready To Be Installed",
        "No informado",
    ]
    seen = set()

    for name in preferred_order:
        count = int(grouped.get(name, 0))
        if count > 0:
            items.append((name, count))
            seen.add(name)

    for name, count in grouped.items():
        if name not in seen and int(count) > 0:
            items.append((name, int(count)))

    return items


def apply_operational_status_filter(df: pd.DataFrame, selected_states: list[str]) -> pd.DataFrame:
    if df.empty or not selected_states or "Operational status grouped" not in df.columns:
        return df

    mask = pd.Series(False, index=df.index)
    state_series = df["Operational status grouped"].fillna("No informado").astype(str)

    for state in selected_states:
        if state == "No rutina":
            mask = mask | ((~state_series.eq("Routine")) & (~state_series.eq("No informado")))
        else:
            mask = mask | state_series.eq(state)

    return df[mask].copy()


def clean_filter_value(values) -> str:
    if values is None:
        return "Todos"
    if isinstance(values, (list, tuple, set)):
        clean = [str(v) for v in values if str(v).strip()]
        return ", ".join(clean) if clean else "All"
    text = str(values).strip()
    return text or "All"


def build_filter_summary(
    selected_regions: list[str],
    selected_countries: list[str],
    selected_distributors: list[str],
    selected_instruments: list[str],
    selected_states: list[str],
) -> dict[str, str]:
    return {
        "Región comercial": clean_filter_value(selected_regions),
        "País": clean_filter_value(selected_countries),
        "Distribuidor": clean_filter_value(selected_distributors),
        "Tipo de instrumento": clean_filter_value(selected_instruments),
        "Estado operativo": clean_filter_value(selected_states),
    }


def translate_status_value(value: str) -> str:
    mapping = {
        'Routine': 'En rutina',
        'NOT IN ROUTINE': 'No en rutina',
        'IN ROUTINE': 'En rutina',
        'Scrapped': 'Descartado',
        'To Be Scrapped': 'Por descartar',
        'Warehouse To Be Refurbished': 'Bodega por reacondicionar',
        'WAREHOUSE to be refurbished': 'Bodega por reacondicionar',
        'Warehouse Ready To Be Installed': 'Bodega lista para instalar',
        'WAREHOUSE ready to be installed': 'Bodega lista para instalar',
        'WAREHOUSE to be scrapped': 'Bodega por descartar',
        'Refurbisched': 'Reacondicionado',
        'Refurbished': 'Reacondicionado',
        'Unknown': 'No informado',
        'Not installed': 'No informado',
        'Missing': 'Faltante',
        'LOW': 'Bajo',
        'OK': 'OK',
    }
    txt = safe_text(value, 'No informado')
    return mapping.get(txt, txt)


def format_pdf_numeric_value(value):
    if pd.isna(value):
        return 'No informado'
    if isinstance(value, (int, np.integer)):
        return f"{int(value):,}".replace(',', '.')
    if isinstance(value, (float, np.floating)):
        if abs(value - round(value)) < 1e-9:
            return f"{int(round(value)):,}".replace(',', '.')
        return f"{float(value):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    text_value = str(value).strip()
    return text_value if text_value else 'No informado'


def prepare_pdf_report_table(df: pd.DataFrame) -> pd.DataFrame:
    preferred_columns = [
        "Commercial Region",
        "Country",
        "Distributor name",
        "Customer name",
        "Instrument type",
        "Serial number",
        "Operational status grouped",
        "Operational status",
        "Operating System",
        "Asset condition",
        "Installation date",
        "Type of contract",
    ]
    available = [col for col in preferred_columns if col in df.columns]
    report_df = df[available].copy()

    if "Installation date" in report_df.columns:
        report_df["Installation date"] = pd.to_datetime(report_df["Installation date"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("No informado")

    for col in report_df.columns:
        if col == "Installation date":
            continue
        report_df[col] = report_df[col].fillna("No informado").astype(str).str.strip().replace("", "No informado")

    for col in ["Operational status grouped", "Operational status", "Asset condition", "Operating System"]:
        if col in report_df.columns:
            report_df[col] = report_df[col].map(translate_status_value)

    report_df = report_df.rename(columns={
        "Commercial Region": "Región",
        "Country": "País",
        "Distributor name": "Distribuidor",
        "Customer name": "Cliente",
        "Instrument type": "Instrumento",
        "Serial number": "Serial",
        "Operational status grouped": "Estado",
        "Operational status": "Estado detallado",
        "Operating System": "Sistema operativo",
        "Asset condition": "Condición",
        "Installation date": "Fecha de instalación",
        "Type of contract": "Tipo de contrato",
    })

    ordered_cols = [c for c in ["Región", "País", "Distribuidor", "Cliente", "Instrumento", "Serial", "Estado", "Estado detallado", "Sistema operativo", "Condición", "Fecha de instalación", "Tipo de contrato"] if c in report_df.columns]
    return report_df[ordered_cols]


def _pdf_header_footer(canvas, doc, short_title: str):
    canvas.saveState()
    width, height = landscape(A4)

    canvas.setStrokeColor(colors.HexColor("#D9D9D9"))
    canvas.setLineWidth(0.6)
    canvas.line(doc.leftMargin, height - 34, width - doc.rightMargin, height - 34)
    canvas.line(doc.leftMargin, 28, width - doc.rightMargin, 28)

    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(colors.HexColor("#222222"))
    canvas.drawString(doc.leftMargin, height - 24, short_title[:90])
    canvas.drawRightString(width - doc.rightMargin, height - 24, f"Page {doc.page}")

    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawString(doc.leftMargin, 16, "Records List Intelligence Dashboard | APA-style filtered report")
    canvas.drawRightString(width - doc.rightMargin, 16, datetime.now().strftime("%Y-%m-%d %H:%M"))
    canvas.restoreState()


def _escape_pdf_text(value) -> str:
    text = safe_text(value, "No informado")
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _paragraph_cell(value, style):
    return Paragraph(_escape_pdf_text(value), style)


def _df_to_wrapped_table(df: pd.DataFrame, styles, col_widths=None, max_rows=None):
    work = df.copy()
    if max_rows is not None:
        work = work.head(max_rows)
    if work.empty:
        return Paragraph("No hay datos disponibles para esta sección.", styles["APA_Body"])

    for col in work.columns:
        work[col] = work[col].map(format_pdf_numeric_value).astype(str).str.slice(0, 140)

    cell_style = styles["APA_Cell_Tiny"] if len(work.columns) >= 8 else styles["APA_Cell"]
    header_style = styles["APA_Cell_Header_Tiny"] if len(work.columns) >= 8 else styles["APA_Cell_Header"]
    header_row = [Paragraph(f"<b>{_escape_pdf_text(c)}</b>", header_style) for c in work.columns]
    body_rows = [[_paragraph_cell(v, cell_style) for v in row] for row in work.values.tolist()]
    if col_widths is None:
        usable_width = 10.9 * inch
        per_col = usable_width / max(len(work.columns), 1)
        col_widths = [per_col] * len(work.columns)
    table = Table([header_row] + body_rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#203864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table



def _summary_table_from_pairs(title: str, pairs: list[tuple[str, str]], styles, col1='Métrica', col2='Valor'):
    data = [[Paragraph(f"<b>{_escape_pdf_text(col1)}</b>", styles["APA_Cell_Header"]), Paragraph(f"<b>{_escape_pdf_text(col2)}</b>", styles["APA_Cell_Header"])]]
    for k, v in pairs:
        data.append([_paragraph_cell(k, styles["APA_Cell"]), _paragraph_cell(v, styles["APA_Cell"])])
    table = Table(data, colWidths=[2.8 * inch, 2.8 * inch], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3b64")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#1f3b64")),
        ("LINEABOVE", (0, 1), (-1, -1), 0.25, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return [Paragraph(title, styles["APA_Heading"]), table]


def _wrap_label(text_value, max_chars: int = 28) -> str:
    text_value = safe_text(text_value, "No informado")
    words = str(text_value).split()
    lines, current = [], ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) <= max_chars:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return "\n".join(lines[:3])


def _clean_spare_qty(series: pd.Series) -> pd.Series:
    vals = pd.to_numeric(series, errors="coerce").fillna(0.0)
    vals = vals.clip(lower=0.0)
    return vals


def _safe_share_pct(part, total) -> float:
    try:
        total = float(total)
        part = float(part)
        if total <= 0:
            return 0.0
        return round(part * 100.0 / total, 1)
    except Exception:
        return 0.0


MANUFACTURING_AGE_BUCKET_ORDER = ["0–3 años", "3–5 años", "5–8 años", "8–10 años", "10–15 años", "15+ años"]
INSTALLATION_AGE_BUCKET_ORDER = ["0-5 años", "5-8 años", "8-10 años", "10+ años"]


def _has_valid_numeric_column(df: pd.DataFrame, column: str) -> bool:
    if df is None or df.empty or column not in df.columns:
        return False
    values = pd.to_numeric(df[column], errors="coerce")
    return bool(values.notna().any())


def _ordered_value_counts(series: pd.Series, order: list[str], label_col: str = "Rango", value_col: str = "Cantidad") -> pd.DataFrame:
    if series is None:
        return pd.DataFrame(columns=[label_col, value_col])
    clean = series.dropna().astype(str).str.strip()
    clean = clean[clean.ne("") & ~clean.str.lower().isin({"nan", "none", "nat", "<na>"})]
    counts = clean.value_counts(dropna=False)
    rows = [{label_col: label, value_col: int(counts.get(label, 0))} for label in order]
    out = pd.DataFrame(rows)
    return out[out[value_col] > 0].reset_index(drop=True)


def _build_pdf_age_profile(filtered_df: pd.DataFrame) -> tuple[pd.DataFrame, str, str]:
    """Devuelve el perfil de antigüedad correcto para el PDF.

    Prioridad:
    1. Fecha de fabricación, cuando `Manufacturing age (years)` está disponible.
    2. Fecha de instalación, solo como respaldo cuando no existe cruce de fabricación.

    Esto corrige la diferencia entre el dashboard de Antigüedad / fabricación y
    el PDF, que antes usaba siempre `Age (years)` calculado por fecha de instalación.
    """
    age_df = filtered_df.copy() if isinstance(filtered_df, pd.DataFrame) else pd.DataFrame()

    if _has_valid_numeric_column(age_df, "Manufacturing age (years)"):
        age_values = pd.to_numeric(age_df["Manufacturing age (years)"], errors="coerce")
        if "Manufacturing age bucket" in age_df.columns:
            bucket_series = age_df["Manufacturing age bucket"]
        else:
            bucket_series = build_age_bucket(age_values)

        valid_bucket_series = bucket_series[age_values.notna()]
        age_counts = _ordered_value_counts(valid_bucket_series, MANUFACTURING_AGE_BUCKET_ORDER)
        return age_counts, "Perfil de antigüedad por fabricación", "Fecha de fabricación"

    age_values = pd.to_numeric(age_df.get("Age (years)", pd.Series(dtype=float)), errors="coerce")
    install_bucket = pd.cut(
        age_values,
        bins=[-np.inf, 5, 8, 10, np.inf],
        labels=INSTALLATION_AGE_BUCKET_ORDER,
        right=False,
    )
    age_counts = _ordered_value_counts(install_bucket, INSTALLATION_AGE_BUCKET_ORDER)
    return age_counts, "Perfil de antigüedad por instalación", "Fecha de instalación"




def _build_pdf_manufacturing_age_analysis(filtered_df: pd.DataFrame) -> tuple[dict | None, dict | None]:
    """Construye una sección PDF dedicada al análisis de edad por fabricación.

    El resumen general del PDF ya contiene un gráfico de perfil de antigüedad,
    pero el usuario necesita que el informe también replique la lectura ejecutiva
    de la pestaña Antigüedad / fabricación: cobertura del cruce, edad promedio,
    equipo más antiguo/nuevo, distribución por rangos, distribución por año y
    listado de los equipos más antiguos.
    """
    if filtered_df is None or filtered_df.empty:
        return None, None
    if not _has_valid_numeric_column(filtered_df, "Manufacturing age (years)"):
        return None, None

    work = filtered_df.copy()
    age_values = pd.to_numeric(work["Manufacturing age (years)"], errors="coerce")
    matched = work[age_values.notna()].copy()
    matched["Manufacturing age (years)"] = age_values.loc[matched.index]
    if matched.empty:
        return None, None

    total_assets = int(len(work))
    matched_count = int(len(matched))
    unmatched_count = max(total_assets - matched_count, 0)
    match_pct = _safe_share_pct(matched_count, total_assets)
    average_age = round(float(matched["Manufacturing age (years)"].mean()), 1)
    oldest_age = round(float(matched["Manufacturing age (years)"].max()), 1)
    newest_age = round(float(matched["Manufacturing age (years)"].min()), 1)

    conflict_count = 0
    if "Manufacturing date conflict" in matched.columns:
        try:
            conflict_count = int(matched["Manufacturing date conflict"].fillna(False).astype(bool).sum())
        except Exception:
            conflict_count = int(matched["Manufacturing date conflict"].astype(str).str.lower().isin({"true", "1", "yes", "si", "sí"}).sum())

    source_count = int(matched["Manufacturing Source"].nunique(dropna=True)) if "Manufacturing Source" in matched.columns else 0
    year_min = "No informado"
    year_max = "No informado"
    if "Manufacturing year" in matched.columns:
        years = pd.to_numeric(matched["Manufacturing year"], errors="coerce").dropna()
        if not years.empty:
            year_min = str(int(years.min()))
            year_max = str(int(years.max()))

    age_counts, _, _ = _build_pdf_age_profile(matched)

    annual_df = pd.DataFrame(columns=["Año fabricación", "Cantidad"])
    if "Manufacturing year" in matched.columns:
        years = pd.to_numeric(matched["Manufacturing year"], errors="coerce").dropna().astype(int)
        if not years.empty:
            annual_df = (
                years.value_counts()
                .sort_index()
                .reset_index()
            )
            annual_df.columns = ["Año fabricación", "Cantidad"]
            annual_df["Año fabricación"] = annual_df["Año fabricación"].astype(str)

    top_oldest = matched.copy()
    if "Manufacturing Date" in top_oldest.columns:
        top_oldest = top_oldest.sort_values(["Manufacturing age (years)", "Manufacturing Date", "Serial number"], ascending=[False, True, True])
    else:
        top_oldest = top_oldest.sort_values(["Manufacturing age (years)", "Serial number"], ascending=[False, True])
    top_oldest = top_oldest.head(15).copy()
    top_oldest["Equipo"] = (
        top_oldest.get("Serial number", pd.Series("N/A", index=top_oldest.index)).fillna("N/A").astype(str)
        + " | "
        + top_oldest.get("Instrument type", pd.Series("N/A", index=top_oldest.index)).fillna("N/A").astype(str)
    )
    top_chart_df = top_oldest[["Equipo", "Manufacturing age (years)"]].rename(columns={"Manufacturing age (years)": "Edad fabricación (años)"})

    detail_columns = [
        "Country",
        "Distributor name",
        "Customer name",
        "Instrument type",
        "Serial number",
        "Manufacturing Date",
        "Manufacturing year",
        "Manufacturing age (years)",
        "Manufacturing age bucket",
        "Operational status grouped",
        "Manufacturing Source",
    ]
    detail_columns = [col for col in detail_columns if col in top_oldest.columns]
    top_table = top_oldest[detail_columns].copy()
    if "Manufacturing Date" in top_table.columns:
        top_table["Manufacturing Date"] = pd.to_datetime(top_table["Manufacturing Date"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("No informado")
    rename_map = {
        "Country": "País",
        "Distributor name": "Distribuidor",
        "Customer name": "Cliente",
        "Instrument type": "Instrumento",
        "Serial number": "Serial",
        "Manufacturing Date": "Fecha fabricación",
        "Manufacturing year": "Año fabricación",
        "Manufacturing age (years)": "Edad fabricación (años)",
        "Manufacturing age bucket": "Rango edad fabricación",
        "Operational status grouped": "Estado operativo",
        "Manufacturing Source": "Fuente fabricación",
    }
    top_table = top_table.rename(columns={k: v for k, v in rename_map.items() if k in top_table.columns})

    full_detail = matched[[c for c in detail_columns if c in matched.columns]].copy()
    if not full_detail.empty:
        full_detail = full_detail.sort_values("Manufacturing age (years)", ascending=False)
        if "Manufacturing Date" in full_detail.columns:
            full_detail["Manufacturing Date"] = pd.to_datetime(full_detail["Manufacturing Date"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("No informado")
        full_detail = full_detail.rename(columns={k: v for k, v in rename_map.items() if k in full_detail.columns})

    summary_pairs = [
        ("Equipos filtrados", f"{total_assets:,}"),
        ("Seriales cruzados con fecha de fabricación", f"{matched_count:,} de {total_assets:,} ({match_pct:.1f}%)"),
        ("Seriales sin coincidencia de fabricación", f"{unmatched_count:,}"),
        ("Edad promedio por fabricación", f"{average_age:.1f} años"),
        ("Equipo más antiguo", f"{oldest_age:.1f} años"),
        ("Equipo más nuevo", f"{newest_age:.1f} años"),
        ("Rango de años de fabricación", f"{year_min} – {year_max}"),
        ("Fuentes de fabricación usadas", f"{source_count:,}"),
        ("Seriales con conflicto de fecha", f"{conflict_count:,}"),
    ]

    charts = [
        _make_pdf_barh(age_counts, "Rango", "Cantidad", "Perfil de antigüedad por fabricación", max_rows=len(age_counts), preserve_order=True),
        _make_pdf_barh(annual_df, "Año fabricación", "Cantidad", "Equipos por año de fabricación", max_rows=min(max(len(annual_df), 1), 24), preserve_order=True, label_wrap=12) if not annual_df.empty else None,
        _make_pdf_barh(top_chart_df, "Equipo", "Edad fabricación (años)", "Top 15 equipos más antiguos por fabricación", xlabel="Años desde fabricación", max_rows=15, color="#ffb454", label_wrap=38),
    ]

    section = {
        "title": "Análisis de antigüedad por fabricación",
        "intro": (
            "Esta sección usa el cruce por serial contra los archivos de fabricación cargados manualmente. "
            "La edad se calcula desde Manufacturing Date y no desde Installation date, por lo que debe coincidir "
            "con la pestaña Antigüedad / fabricación del dashboard."
        ),
        "summary_pairs": summary_pairs,
        "charts": charts,
        "table_title": "Top 15 equipos más antiguos por fabricación",
        "table_df": top_table,
        "table_max_rows": min(len(top_table), 15),
    }

    annex = {
        "title": "Anexo. Detalle completo de antigüedad por fabricación",
        "intro": "Detalle de todos los equipos filtrados que tuvieron coincidencia válida con fecha de fabricación.",
        "summary_pairs": [("Filas incluidas", f"{len(full_detail):,}"), ("Alcance", "Equipos con Manufacturing Date válida")],
        "charts": [],
        "table_title": "Detalle completo de fabricación",
        "table_df": full_detail,
        "table_max_rows": max(len(full_detail), 1),
    }
    return section, annex


def _make_pdf_barh(
    df: pd.DataFrame,
    label_col: str,
    value_col: str,
    title: str,
    xlabel: str = "Cantidad",
    max_rows: int = 10,
    color: str = "#2F80ED",
    preserve_order: bool = False,
    label_wrap: int = 30,
):
    """Genera una barra horizontal para PDF.

    Mejora v44:
    - `preserve_order=True` evita que gráficos ordinales, como rangos de edad,
      se reordenen por cantidad. Esto permite que el PDF conserve la misma lógica
      de lectura del dashboard.
    - `label_wrap` permite controlar el corte de etiquetas sin modificar el resto
      de gráficas ejecutivas.
    """
    if not MATPLOTLIB_AVAILABLE or df is None or df.empty or label_col not in df.columns or value_col not in df.columns:
        return None

    work = df[[label_col, value_col]].copy().dropna(subset=[label_col])
    if work.empty:
        return None

    work[value_col] = pd.to_numeric(work[value_col], errors="coerce")
    work = work.dropna(subset=[value_col])
    work = work[work[value_col] > 0]
    if work.empty:
        return None

    if preserve_order:
        # Matplotlib barh dibuja la primera fila abajo; se invierte para que el
        # orden del DataFrame se lea de arriba hacia abajo en el PDF.
        work = work.head(max_rows).iloc[::-1].copy()
    else:
        work = work.sort_values(value_col, ascending=False).head(max_rows).sort_values(value_col, ascending=True)

    work[label_col] = work[label_col].map(lambda x: _wrap_label(x, label_wrap))
    height = max(2.8, 0.48 * len(work) + 1.25)
    fig, ax = plt.subplots(figsize=(8.6, height))
    bars = ax.barh(work[label_col].astype(str), work[value_col].astype(float), color=color)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.set_xlabel(xlabel, fontsize=9)
    ax.tick_params(axis="y", labelsize=8)
    ax.tick_params(axis="x", labelsize=8)
    ax.grid(axis="x", alpha=0.22)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    max_value = float(work[value_col].max()) if not work.empty else 0.0
    for bar in bars:
        value = bar.get_width()
        ax.text(
            value + max(max_value * 0.01, 0.1),
            bar.get_y() + bar.get_height() / 2,
            safe_number_text(value, "0"),
            va="center",
            fontsize=8,
        )

    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf



def _make_pdf_stacked_barh(df: pd.DataFrame, category_col: str, segment_col: str, value_col: str, title: str, max_categories: int = 8, max_segments: int = 6):
    if df is None or df.empty or category_col not in df.columns or segment_col not in df.columns or value_col not in df.columns:
        return None

    work = df.copy()
    work[category_col] = work[category_col].fillna('No informado').astype(str)
    work[segment_col] = work[segment_col].fillna('No informado').astype(str)
    work[value_col] = pd.to_numeric(work[value_col], errors='coerce').fillna(0)
    work = work[work[value_col] > 0]
    if work.empty:
        return None

    cat_order = (
        work.groupby(category_col, as_index=False)[value_col]
        .sum()
        .sort_values(value_col, ascending=False)[category_col]
        .tolist()
    )[:max_categories]
    work = work[work[category_col].isin(cat_order)].copy()

    seg_order = (
        work.groupby(segment_col, as_index=False)[value_col]
        .sum()
        .sort_values(value_col, ascending=False)[segment_col]
        .tolist()
    )
    if len(seg_order) > max_segments:
        keep = seg_order[: max_segments - 1]
        work[segment_col] = np.where(work[segment_col].isin(keep), work[segment_col], 'Otros')
        seg_order = keep + ['Otros']
        work = work.groupby([category_col, segment_col], as_index=False)[value_col].sum()

    pivot = work.pivot_table(index=category_col, columns=segment_col, values=value_col, aggfunc='sum', fill_value=0)
    pivot = pivot.reindex(index=cat_order, fill_value=0)
    ordered_cols = [c for c in seg_order if c in pivot.columns] + [c for c in pivot.columns if c not in seg_order]
    pivot = pivot[ordered_cols]

    fig, ax = plt.subplots(figsize=(8.6, 4.6))
    left = np.zeros(len(pivot))
    palette = ['#5BC0EB', '#9BB1FF', '#63E0C9', '#FDBA5A', '#6B7280', '#C084FC']

    for i, col in enumerate(pivot.columns):
        values = pivot[col].to_numpy(dtype=float)
        bars = ax.barh(pivot.index.astype(str), values, left=left, color=palette[i % len(palette)], label=str(col))
        for b, v, l in zip(bars, values, left):
            if v >= 1:
                ax.text(l + v + 0.3, b.get_y() + b.get_height()/2, f'{int(v)}', va='center', ha='left', fontsize=8, color='#0F172A')
        left = left + values

    ax.set_title(title, fontsize=13, fontweight='bold')
    ax.set_xlabel('Cantidad')
    ax.set_ylabel('Modelo')
    ax.grid(axis='x', alpha=0.22)
    ax.invert_yaxis()
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.14), ncol=min(3, max(1, len(pivot.columns))), frameon=False, fontsize=8)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=220, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf
def _make_pdf_donut(df: pd.DataFrame, label_col: str, value_col: str, title: str, max_rows: int = 5):
    if not MATPLOTLIB_AVAILABLE or df is None or df.empty or label_col not in df.columns or value_col not in df.columns:
        return None
    work = df[[label_col, value_col]].copy().dropna()
    if work.empty:
        return None
    work[value_col] = pd.to_numeric(work[value_col], errors='coerce')
    work = work.dropna()
    work = work[work[value_col] > 0]
    if work.empty:
        return None
    work = work.sort_values(value_col, ascending=False)
    if len(work) > max_rows:
        top = work.head(max_rows - 1).copy()
        others = work.iloc[max_rows - 1:][value_col].sum()
        top = pd.concat([top, pd.DataFrame({label_col: ['Otros'], value_col: [others]})], ignore_index=True)
        work = top
    labels = work[label_col].astype(str).map(lambda x: _wrap_label(x, 22)).tolist()
    values = work[value_col].astype(float).tolist()
    total = sum(values)
    colors_list = ['#2F80ED', '#56CCF2', '#27AE60', '#F2C94C', '#EB5757', '#9B51E0']
    fig, ax = plt.subplots(figsize=(4.8, 3.6))
    wedges, texts, autotexts = ax.pie(
        values,
        labels=None,
        colors=colors_list[:len(values)],
        startangle=90,
        counterclock=False,
        wedgeprops=dict(width=0.35, edgecolor='white'),
        autopct=lambda pct: f'{pct:.1f}%' if pct >= 8 else ''
    )
    ax.text(0, 0.05, safe_number_text(total, '0'), ha='center', va='center', fontsize=16, fontweight='bold')
    ax.text(0, -0.15, 'equipos', ha='center', va='center', fontsize=9)
    ax.set_title(_wrap_label(title, 34), fontsize=10.5, fontweight='bold', pad=10)
    ax.legend(wedges, labels, loc='lower center', bbox_to_anchor=(0.5, -0.22), ncol=2, fontsize=7, frameon=False, columnspacing=1.2, handletextpad=0.6)
    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=180, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_pdf_hist_categories(df: pd.DataFrame, label_col: str, order: list[str], title: str, xlabel: str = 'Cantidad'):
    counts = df[label_col].fillna('No informado').astype(str).value_counts()
    chart_df = pd.DataFrame({label_col: order, 'Count': [int(counts.get(v, 0)) for v in order]})
    chart_df = chart_df[chart_df['Count'] > 0]
    return _make_pdf_barh(chart_df, label_col, 'Count', title, xlabel=xlabel, max_rows=len(chart_df))


def _pdf_image_flowables(image_buffers: list, max_per_row: int = 2, image_width: float = 4.7 * inch, image_height: float = 2.9 * inch):
    from reportlab.platypus import Image
    flowables = []
    valid = [img for img in image_buffers if img is not None]
    if not valid:
        return flowables
    for idx in range(0, len(valid), max_per_row):
        row = valid[idx:idx + max_per_row]
        imgs = [Image(img, width=image_width, height=image_height) for img in row]
        if len(imgs) == 1:
            flowables.append(imgs[0])
        else:
            tbl = Table([[imgs[0], imgs[1]]], colWidths=[image_width, image_width])
            tbl.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 8)]))
            flowables.append(tbl)
        flowables.append(Spacer(1, 0.10 * inch))
    return flowables


def _build_machine_config_summary(filtered_df: pd.DataFrame):
    cfg_cols = [c for c in filtered_df.columns if c.startswith('CFG::')]
    cfg_cov_rows = []
    value_summary_rows = []
    chart_buffers = []
    for col in cfg_cols:
        non_null = filtered_df[col].dropna()
        count_non_null = int(non_null.shape[0])
        if count_non_null <= 0:
            continue
        field_name = col.replace('CFG::', '')
        if field_name.strip().lower() in {'operative system', 'operating system'}:
            continue
        vc = non_null.astype(str).str.strip()
        vc = vc[vc != '']
        if vc.empty:
            continue
        counts = vc.value_counts().reset_index()
        counts.columns = ['Value', 'Count']
        counts['Share %'] = counts['Count'].map(lambda x: _safe_share_pct(x, counts['Count'].sum()))
        top_value = safe_text(counts.iloc[0]['Value'])
        top_count = int(counts.iloc[0]['Count'])
        cfg_cov_rows.append({'Campo de configuración': field_name, 'Equipos con dato': count_non_null})
        value_summary_rows.append({
            'Campo de configuración': field_name,
            'Equipos con dato': count_non_null,
            'Valores únicos': int(counts.shape[0]),
            'Valor principal': top_value,
            'Conteo principal': top_count,
        })
        for _, row in counts.head(5).iterrows():
            value_summary_rows.append({
                'Campo de configuración': f"{field_name} — valor",
                'Equipos con dato': '',
                'Valores únicos': '',
                'Valor principal': safe_text(row['Value']),
                'Conteo principal': f"{int(row['Count'])} ({row['Share %']:.1f}%)",
            })
        if counts.shape[0] <= 5 and counts.iloc[0]['Count'] / counts['Count'].sum() < 0.86:
            chart = _make_pdf_donut(counts.rename(columns={'Value': 'Categoría'}), 'Categoría', 'Count', field_name, max_rows=5)
        else:
            chart = _make_pdf_barh(counts.rename(columns={'Value': 'Categoría'}), 'Categoría', 'Count', field_name, xlabel='Equipos', max_rows=6, color='#1f77b4')
        chart_buffers.append((count_non_null, chart))
    cov_df = pd.DataFrame(cfg_cov_rows).sort_values('Equipos con dato', ascending=False) if cfg_cov_rows else pd.DataFrame(columns=['Campo de configuración', 'Equipos con dato'])
    value_df = pd.DataFrame(value_summary_rows)
    charts = [c for _, c in sorted(chart_buffers, key=lambda x: x[0], reverse=True)[:6]]
    return cov_df, value_df, charts


def _build_executive_insights(filtered_df: pd.DataFrame, stock_context: dict | None = None) -> tuple[list[str], list[str]]:
    insights = []
    recommendations = []
    total_records = len(filtered_df)
    countries = filtered_df['Country'].fillna('No informado').value_counts()
    if not countries.empty:
        top_country = countries.index[0]
        top_country_pct = _safe_share_pct(countries.iloc[0], total_records)
        insights.append(f"La base instalada filtrada contiene {total_records} equipos y se concentra principalmente en {top_country} ({top_country_pct}%).")
    routine_assets = int(filtered_df.get('Is in routine', pd.Series(dtype=bool)).sum())
    insights.append(f"Se identificaron {routine_assets} equipos en rutina dentro del universo filtrado.")
    os_series = filtered_df.get('Operating System', pd.Series(dtype=object)).fillna('No informado').astype(str)
    legacy_count = int(os_series.isin(['Windows XP', 'Windows Vista', 'Windows 7', 'Windows 2000']).sum())
    unknown_os = int(os_series.isin(['Unknown', 'No informado', 'Not installed']).sum())
    if legacy_count > 0:
        insights.append(f"Existen {legacy_count} equipos con sistemas operativos legados que deben priorizarse en el plan de actualización.")
    if unknown_os > 0:
        insights.append(f"Hay {unknown_os} equipos sin visibilidad clara del sistema operativo, lo que limita la planeación técnica.")
    if 'PM next date' in filtered_df.columns:
        pm_next = pd.to_datetime(filtered_df['PM next date'], errors='coerce')
        overdue_pm = int((pm_next < pd.Timestamp.today().normalize()).fillna(False).sum())
        if overdue_pm > 0:
            insights.append(f"Se detectaron {overdue_pm} mantenimientos preventivos vencidos en la vista actual.")
    if stock_context and stock_context.get('available'):
        missing = int(stock_context.get('missing_skus', 0))
        low = int(stock_context.get('low_skus', 0))
        cost = float(stock_context.get('option2_cost', 0) or 0)
        insights.append(f"La revisión de carstock identificó {missing} SKUs faltantes y {low} SKUs en nivel bajo, con una exposición estimada de EUR {cost:,.2f}.")

        if missing > 0:
            recommendations.append('Priorizar la compra de repuestos faltantes y de bajo stock con base en el costo estimado y la criticidad operativa.')
    if legacy_count > 0:
        recommendations.append('Ejecutar un plan de migración para equipos con Windows Vista/XP/7 y validar de inmediato los activos sin dato de sistema operativo.')
    if unknown_os > 0:
        recommendations.append('Completar los campos vacíos de sistema operativo y configuración de equipo para mejorar la trazabilidad del parque instalado.')
    if 'PM next date' in filtered_df.columns:
        pm_next = pd.to_datetime(filtered_df['PM next date'], errors='coerce')
        overdue_pm = int((pm_next < pd.Timestamp.today().normalize()).fillna(False).sum())
        if overdue_pm > 0:
            recommendations.append('Reprogramar los mantenimientos preventivos vencidos y ordenar la ejecución por volumen de pruebas y criticidad del cliente.')
    recommendations.append('Usar este informe como base para una revisión ejecutiva del distribuidor, combinando base instalada, OS, PM y cobertura de repuestos.')
    return insights[:6], recommendations[:5]


def _build_pdf_sections(filtered_df: pd.DataFrame, stock_context: dict | None = None):
    sections = []
    annexes = []

    base_pairs = [
        ('Registros filtrados', f"{len(filtered_df):,}"),
        ('Países', f"{filtered_df['Country'].nunique(dropna=True):,}"),
        ('Distribuidores', f"{filtered_df['Distributor name'].nunique(dropna=True):,}"),
        ('Tipos de instrumento', f"{filtered_df['Instrument type'].nunique(dropna=True):,}"),
        ('Equipos en rutina', f"{int(filtered_df.get('Is in routine', pd.Series(dtype=bool)).sum()):,}"),
    ]
    top_country = filtered_df['Country'].fillna('No informado').value_counts().reset_index()
    top_country.columns = ['País', 'Cantidad']
    top_inst = filtered_df['Instrument type'].fillna('No informado').value_counts().reset_index()
    top_inst.columns = ['Instrumento', 'Cantidad']
    state_counts = filtered_df['Operational status grouped'].fillna('No informado').value_counts().reset_index()
    state_counts.columns = ['Estado', 'Cantidad']
    age_counts, age_chart_title, age_source_label = _build_pdf_age_profile(filtered_df)
    base_pairs.append(('Base usada para antigüedad', age_source_label))

    corporate_model_charts = []
    corporate_model_df = filtered_df.copy()
    corporate_model_df['Instrument type'] = corporate_model_df['Instrument type'].fillna('No informado').astype(str)
    corporate_model_df['Distributor name'] = corporate_model_df['Distributor name'].fillna('No informado').astype(str)
    model_rank = corporate_model_df['Instrument type'].value_counts().index.tolist()
    global_dist = (
        corporate_model_df.groupby(['Instrument type', 'Distributor name'], dropna=False)
        .size()
        .reset_index(name='Cantidad')
        .rename(columns={'Instrument type': 'Modelo', 'Distributor name': 'Distribuidor'})
    )
    if not global_dist.empty:
        top_global = global_dist.groupby('Distribuidor', as_index=False)['Cantidad'].sum().sort_values(['Cantidad', 'Distribuidor'], ascending=[False, True]).head(5)['Distribuidor'].tolist()
        global_dist_main = global_dist[global_dist['Distribuidor'].isin(top_global)].copy()
        global_dist_main['Distribuidor'] = global_dist_main['Distribuidor'].astype(str).map(lambda x: distributor_display_name(x, 18))
        corporate_model_charts.append(_make_pdf_stacked_barh(global_dist_main, 'Modelo', 'Distribuidor', 'Cantidad', 'Vista global por distribuidor | resumen (Top 5)', max_categories=8, max_segments=6))
    detail_corporate_rows = []
    for model_name in model_rank[:6]:
        model_slice = corporate_model_df[corporate_model_df['Instrument type'] == model_name].copy()
        counts = model_slice['Distributor name'].value_counts().reset_index()
        counts.columns = ['Distribuidor', 'Cantidad']
        counts = counts.sort_values(['Cantidad', 'Distribuidor'], ascending=[False, True]).reset_index(drop=True)
        if counts.empty:
            continue
        top_counts = counts.head(5).copy()
        top_counts['Distribuidor'] = top_counts['Distribuidor'].astype(str).map(lambda x: distributor_display_name(x, 20))
        corporate_model_charts.append(_make_pdf_donut(top_counts, 'Distribuidor', 'Cantidad', f'Distribución por distribuidor | {model_name} | Top 5', max_rows=5))
        full_counts = counts.copy()
        full_counts['Modelo'] = model_name
        detail_corporate_rows.append(full_counts)

    sections.append({
        'title': 'Resumen de base instalada',
        'intro': 'Esta sección resume la base instalada filtrada y destaca la concentración geográfica, el mix de instrumentos, el estado operativo, el perfil de antigüedad y la distribución por distribuidor para cada modelo visible.',
        'summary_pairs': base_pairs,
        'charts': [
            _make_pdf_barh(top_country, 'País', 'Cantidad', 'Países con mayor concentración', max_rows=10),
            _make_pdf_barh(top_inst, 'Instrumento', 'Cantidad', 'Mix de instrumentos', max_rows=10),
            _make_pdf_barh(state_counts, 'Estado', 'Cantidad', 'Distribución por estado operativo', max_rows=10),
            _make_pdf_barh(age_counts, 'Rango', 'Cantidad', age_chart_title, max_rows=len(age_counts), preserve_order=True),
        ] + corporate_model_charts,
        'table_title': 'Muestra resumida de equipos filtrados',
        'table_df': prepare_pdf_report_table(filtered_df),
        'table_max_rows': 10,
    })
    manufacturing_age_section, manufacturing_age_annex = _build_pdf_manufacturing_age_analysis(filtered_df)
    if manufacturing_age_section is not None:
        sections.append(manufacturing_age_section)
    if manufacturing_age_annex is not None:
        annexes.append(manufacturing_age_annex)

    annexes.append({
        'title': 'Anexo A. Base instalada detallada',
        'intro': 'Detalle tabular de la base instalada filtrada.',
        'summary_pairs': [('Filas incluidas', f"{len(filtered_df):,}"), ('Alcance', 'Detalle completo de la base instalada filtrada')],
        'charts': [],
        'table_title': 'Detalle completo de equipos filtrados',
        'table_df': prepare_pdf_report_table(filtered_df),
        'table_max_rows': max(len(filtered_df), 1),
    })
    if detail_corporate_rows:
        detail_corporate_df = pd.concat(detail_corporate_rows, ignore_index=True)
        detail_charts = []
        for model_name in model_rank[:6]:
            model_detail = detail_corporate_df[detail_corporate_df['Modelo'].eq(model_name)].copy()
            if model_detail.empty:
                continue
            model_detail['Distribuidor'] = model_detail['Distribuidor'].astype(str).map(lambda x: distributor_display_name(x, 28))
            detail_charts.append(_make_pdf_barh(model_detail, 'Distribuidor', 'Cantidad', f'Detalle completo | {model_name}', xlabel='Cantidad de equipos', max_rows=max(12, len(model_detail)), color='#2F80ED'))
        annexes.append({
            'title': 'Anexo B. Distribución completa por distribuidor y modelo',
            'intro': 'Detalle completo de distribuidores por modelo. En el cuerpo principal solo se muestra el resumen Top 5 para mantener la lectura ejecutiva.',
            'summary_pairs': [('Filas incluidas', f"{len(detail_corporate_df):,}"), ('Alcance', 'Detalle completo de distribuidores por modelo')],
            'charts': detail_charts,
            'table_title': 'Detalle completo por modelo',
            'table_df': detail_corporate_df[['Modelo', 'Distribuidor', 'Cantidad']].sort_values(['Modelo', 'Cantidad', 'Distribuidor'], ascending=[True, False, True]),
            'table_max_rows': max(len(detail_corporate_df), 1),
        })

    blood_bank_yes = count_blood_bank_yes(filtered_df)
    cfg_pairs = [
        ('Equipos con configuración', f"{int(filtered_df['Machine Configurations'].notna().sum()):,}"),
        ('Equipos de banco de sangre', f"{blood_bank_yes:,} de {len(filtered_df):,} ({_safe_share_pct(blood_bank_yes, len(filtered_df)):.1f}% del total)"),
        ('Campos activos de configuración', f"{sum(int(filtered_df[c].notna().sum()) > 0 for c in filtered_df.columns if c.startswith('CFG::')):,}"),
        ('Promedio de campos poblados', f"{filtered_df.get('Machine config fields populated', pd.Series([0])).fillna(0).mean():.1f}"),
    ]
    cfg_cov, cfg_value_df, cfg_charts = _build_machine_config_summary(filtered_df)
    cfg_charts = cfg_charts[:4]
    sections.append({
        'title': 'Configuración de equipo',
        'intro': 'Se consolidan los campos detectados en configuración de equipo y se muestran las distribuciones de los ítems con mayor visibilidad en el filtro activo. Banco de sangre se presenta primero como indicador ejecutivo principal.',
        'summary_pairs': cfg_pairs,
        'charts': [_make_pdf_donut(pd.DataFrame({'Categoría':['Banco de sangre','Equipos en laboratorio'],'Count':[blood_bank_yes,max(len(filtered_df)-blood_bank_yes,0)]}), 'Categoría', 'Count', 'Banco de sangre', max_rows=2), _make_pdf_barh(cfg_cov, 'Campo de configuración', 'Equipos con dato', 'Cobertura de campos de configuración', max_rows=10)] + cfg_charts,
        'table_title': 'Resumen de configuración de equipo',
        'table_df': cfg_value_df,
        'table_max_rows': 12,
    })
    if not cfg_value_df.empty:
        annexes.append({
            'title': 'Anexo C. Valores de configuración',
            'intro': 'Valores principales por campo de configuración.',
            'summary_pairs': [('Filas incluidas', f"{len(cfg_value_df):,}"), ('Alcance', 'Resumen ampliado de campos y valores de configuración')],
            'charts': [],
            'table_title': 'Valores principales por campo',
            'table_df': cfg_value_df,
            'table_max_rows': max(len(cfg_value_df), 1),
        })

    os_df = filtered_df.copy()
    os_df['Operating System'] = os_df['Operating System'].fillna('No informado').replace({'Unknown':'No informado','Not installed':'No informado'})
    os_df['Bucket de actualización'] = os_df['Operating System'].map(os_upgrade_bucket).replace({
        'Windows 10 / OK': 'Windows 10 / OK',
        'Legacy / urgente migrar': 'Legado / migración urgente',
        'Revisar campo OS': 'Revisar campo OS',
        'Otro OS / validar': 'Otro OS / validar',
    })
    urgent_table = os_df[os_df['Operating System'].isin(['Windows XP', 'Windows Vista', 'Windows 7', 'Windows 2000'])][['Country','Distributor name','Customer name','Instrument type','Serial number','Operating System']].copy()
    urgent_table.columns = ['País', 'Distribuidor', 'Cliente', 'Instrumento', 'Serial', 'Sistema operativo']
    os_pairs = [
        ('Equipos con OS identificado', f"{int(filtered_df['Operating System'].notna().sum()):,}"),
        ('Valores únicos de OS', f"{filtered_df['Operating System'].nunique(dropna=True):,}"),
        ('OS legado / migración urgente', f"{int(os_df['Operating System'].isin(['Windows XP','Windows Vista','Windows 7','Windows 2000']).sum()):,}"),
        ('OS no informado', f"{int(os_df['Operating System'].isin(['Unknown','No informado','Not installed']).sum()):,}"),
    ]
    os_counts = os_df['Operating System'].value_counts().reset_index()
    os_counts.columns = ['Sistema operativo', 'Cantidad']
    os_bucket = os_df['Bucket de actualización'].value_counts().reset_index()
    os_bucket.columns = ['Prioridad', 'Cantidad']
    sections.append({
        'title': 'Sistema operativo',
        'intro': 'Esta sección identifica equipos con sistemas operativos legados, visibilidad incompleta y prioridades de actualización.',
        'summary_pairs': os_pairs,
        'charts': [
            _make_pdf_barh(os_counts, 'Sistema operativo', 'Cantidad', 'Distribución de sistema operativo', max_rows=10),
            _make_pdf_barh(os_bucket, 'Prioridad', 'Cantidad', 'Priorización de actualización', max_rows=10, color='#1f77b4'),
        ],
        'table_title': 'Equipos que requieren actualización de Windows',
        'table_df': urgent_table,
        'table_max_rows': 10,
    })
    if not urgent_table.empty:
        annexes.append({
            'title': 'Anexo D. Equipos con OS legado',
            'intro': 'Detalle de equipos con sistema operativo legado.',
            'summary_pairs': [('Filas incluidas', f"{len(urgent_table):,}"), ('Alcance', 'Equipos con Windows XP/Vista/7/2000')],
            'charts': [],
            'table_title': 'Detalle de equipos con OS legado',
            'table_df': urgent_table,
            'table_max_rows': max(len(urgent_table), 1),
        })

    proc_df = filtered_df.copy()
    proc_df['Pruebas por día'] = pd.to_numeric(proc_df['Number of tests per day'], errors='coerce').fillna(0)
    today = pd.Timestamp.today().normalize()
    if 'PM next date' in proc_df.columns:
        pm_next = pd.to_datetime(proc_df['PM next date'], errors='coerce')
        proc_df['Estado PM'] = np.where(pm_next < today, 'Vencido', np.where(pm_next <= today + pd.Timedelta(days=90), 'Próximos 90 días', 'Planificado más adelante'))
    else:
        proc_df['Estado PM'] = 'No informado'
    pm_status = proc_df['Estado PM'].value_counts().reset_index()
    pm_status.columns = ['Estado PM', 'Cantidad']
    top_tests = proc_df[['Serial number', 'Pruebas por día', 'Instrument type']].copy()
    top_tests = top_tests.sort_values('Pruebas por día', ascending=False).head(10)
    top_tests['Equipo'] = top_tests['Serial number'].astype(str) + ' | ' + top_tests['Instrument type'].astype(str)
    proc_pairs = [
        ('Promedio de pruebas por día', safe_number_text(proc_df['Pruebas por día'].mean(), '0')),
        ('Máximo de pruebas por día', safe_number_text(proc_df['Pruebas por día'].max(), '0')),
        ('PM próximos 90 días', f"{int((proc_df['Estado PM'] == 'Próximos 90 días').sum()):,}"),
        ('PM vencidos', f"{int((proc_df['Estado PM'] == 'Vencido').sum()):,}"),
    ]
    proc_table = proc_df[['Country', 'Distributor name', 'Instrument type', 'Serial number', 'Pruebas por día', 'Estado PM']].copy()
    proc_table.columns = ['País', 'Distribuidor', 'Instrumento', 'Serial', 'Pruebas por día', 'Estado PM']
    sections.append({
        'title': 'Procesamiento y planificación de PM',
        'intro': 'Se prioriza la carga operativa y el estado del mantenimiento preventivo mediante visuales ejecutivas más legibles.',
        'summary_pairs': proc_pairs,
        'charts': [
            _make_pdf_barh(top_tests.rename(columns={'Equipo': 'Equipo'}), 'Equipo', 'Pruebas por día', 'Top 10 equipos por pruebas por día', xlabel='Pruebas/día', max_rows=10),
            _make_pdf_barh(pm_status, 'Estado PM', 'Cantidad', 'Estado del plan de mantenimiento preventivo', max_rows=10, color='#2D9CDB'),
        ],
        'table_title': 'Resumen de equipos con mayor volumen y estado PM',
        'table_df': proc_table.sort_values('Pruebas por día', ascending=False),
        'table_max_rows': 10,
    })
    annexes.append({
        'title': 'Anexo E. Detalle de procesamiento y PM',
        'intro': 'Detalle ampliado de pruebas por día y estado de PM.',
        'summary_pairs': [('Filas incluidas', f"{len(proc_table):,}"), ('Alcance', 'Detalle ampliado de procesamiento y mantenimiento preventivo')],
        'charts': [],
        'table_title': 'Detalle ampliado de procesamiento y PM',
        'table_df': proc_table.sort_values('Pruebas por día', ascending=False),
        'table_max_rows': max(len(proc_table), 1),
    })

    stock_context = stock_context or {}
    if stock_context.get('available'):
        full_comparison_df = stock_context.get('full_comparison_df', pd.DataFrame()).copy()
        purchase_df = stock_context.get('purchase_df', pd.DataFrame()).copy()
        extra_df = stock_context.get('extra_df', pd.DataFrame()).copy()
        stock_top_gap = stock_context.get('top_gap_df', pd.DataFrame()).copy()
        for df_ in [full_comparison_df, purchase_df, extra_df, stock_top_gap]:
            if df_ is not None and not df_.empty:
                if 'Uploaded Qty' in df_.columns:
                    df_['Uploaded Qty'] = _clean_spare_qty(df_['Uploaded Qty'])
                if 'Coverage %' in df_.columns:
                    df_['Coverage %'] = pd.to_numeric(df_['Coverage %'], errors='coerce').fillna(0.0).clip(lower=0.0, upper=999.0)
        main_status = pd.DataFrame({'Estado': ['OK', 'Bajo', 'Faltante'], 'Cantidad': [int(stock_context.get('ok_skus', 0)), int(stock_context.get('low_skus', 0)), int(stock_context.get('missing_skus', 0))]})
        extras_status = pd.DataFrame({'Estado': ['Extras'], 'Cantidad': [int(stock_context.get('extra_skus', 0))]})
        stock_pairs = [
            ('Distribuidor detectado', stock_context.get('detected_distributor', 'N/A')),
            ('Familias comparadas', ', '.join(stock_context.get('families', [])) or 'N/A'),
            ('SKUs requeridos', f"{stock_context.get('required_skus', 0):,}"),
            ('SKUs OK', f"{stock_context.get('ok_skus', 0):,}"),
            ('SKUs LOW', f"{stock_context.get('low_skus', 0):,}"),
            ('SKUs faltantes', f"{stock_context.get('missing_skus', 0):,}"),
            ('Costo estimado opción 2', f"EUR {float(stock_context.get('option2_cost', 0) or 0):,.2f}"),
        ]
        if not stock_top_gap.empty:
            stock_top_gap = stock_top_gap.copy()
            stock_top_gap['Parte'] = stock_top_gap['Required Part Number'].astype(str) + ' | ' + stock_top_gap['Required Description'].fillna('').astype(str).str.slice(0, 28)
        gap_table = pd.DataFrame()
        if not stock_top_gap.empty:
            gap_table = stock_top_gap[[c for c in ['Required Part Number','Required Description','Required Qty','Uploaded Qty','Qty Gap','Status','Option 2 Estimated Cost','Currency'] if c in stock_top_gap.columns]].copy()
            gap_table.columns = ['Parte requerida', 'Descripción', 'Cant. requerida', 'Cant. cargada', 'Brecha', 'Estado', 'Costo estimado opción 2', 'Moneda']
            if 'Estado' in gap_table.columns:
                gap_table['Estado'] = gap_table['Estado'].map(translate_status_value)
        sections.append({
            'title': 'Repuestos y brecha de carstock',
            'intro': 'Se resume la cobertura del stock requerido y la brecha estimada de compra. Los ítems extra se muestran por separado para no distorsionar la lectura principal del gap.',
            'summary_pairs': stock_pairs,
            'charts': [
                _make_pdf_barh(main_status, 'Estado', 'Cantidad', 'Cobertura del carstock requerido', max_rows=3, color='#2F80ED'),
                _make_pdf_barh(extras_status, 'Estado', 'Cantidad', 'Ítems extra no requeridos por el maestro', max_rows=1, color='#56CCF2') if int(stock_context.get('extra_skus', 0)) > 0 else None,
                _make_pdf_barh(stock_top_gap.rename(columns={'Parte': 'Parte', 'Qty Gap': 'Brecha'}), 'Parte', 'Brecha', 'Principales repuestos faltantes', xlabel='Brecha de cantidad', max_rows=10, color='#EB5757') if not stock_top_gap.empty else None,
            ],
            'table_title': 'Principales brechas de repuestos',
            'table_df': gap_table,
            'table_max_rows': 10,
        })
        if not full_comparison_df.empty:
            annex_table = full_comparison_df[[c for c in ['Required Part Number','Required Description','Required Qty','Uploaded Qty','Qty Gap','Coverage %','Status','Option 2 Unit Price','Option 2 Estimated Cost','Currency'] if c in full_comparison_df.columns]].copy()
            annex_table.columns = ['Parte requerida', 'Descripción', 'Cant. requerida', 'Cant. cargada', 'Brecha', 'Cobertura %', 'Estado', 'Precio unitario opción 2', 'Costo estimado opción 2', 'Moneda']
            if 'Estado' in annex_table.columns:
                annex_table['Estado'] = annex_table['Estado'].map(translate_status_value)
            annexes.append({
                'title': 'Anexo F. Comparación completa de repuestos',
                'intro': 'Comparación completa entre el maestro de carstock y el stock cargado por el distribuidor.',
                'summary_pairs': [('Filas incluidas', f"{len(annex_table):,}"), ('Alcance', 'Comparación completa de repuestos')],
                'charts': [],
                'table_title': 'Comparación completa de repuestos',
                'table_df': annex_table,
                'table_max_rows': max(len(annex_table), 1),
            })
        if not purchase_df.empty:
            pur_cols = [c for c in ['Required Part Number','Required Description','Qty Gap','Option 2 Unit Price','Option 2 Estimated Cost','Currency','Status'] if c in purchase_df.columns]
            pur_table = purchase_df[pur_cols].copy()
            pur_table.columns = ['Parte requerida', 'Descripción', 'Brecha', 'Precio unitario opción 2', 'Costo estimado opción 2', 'Moneda', 'Estado']
            if 'Estado' in pur_table.columns:
                pur_table['Estado'] = pur_table['Estado'].map(translate_status_value)
            annexes.append({
                'title': 'Anexo G. Lista sugerida de compra',
                'intro': 'Compra sugerida para cerrar la brecha actual de carstock.',
                'summary_pairs': [('Filas incluidas', f"{len(pur_table):,}"), ('Alcance', 'Lista sugerida de compra basada en opción 2')],
                'charts': [],
                'table_title': 'Lista sugerida de compra',
                'table_df': pur_table,
                'table_max_rows': max(len(pur_table), 1),
            })
        if not extra_df.empty:
            ex_cols = [c for c in ['Uploaded Part Number','Uploaded Description','Uploaded Qty','Status'] if c in extra_df.columns]
            ex_table = extra_df[ex_cols].copy()
            ex_table.columns = ['Parte cargada', 'Descripción cargada', 'Cantidad cargada', 'Estado']
            annexes.append({
                'title': 'Anexo H. Ítems extra no requeridos',
                'intro': 'Repuestos reportados por el distribuidor que no pertenecen al maestro de carstock seleccionado.',
                'summary_pairs': [('Filas incluidas', f"{len(ex_table):,}"), ('Alcance', 'Ítems extra no requeridos por el maestro')],
                'charts': [],
                'table_title': 'Ítems extra no requeridos',
                'table_df': ex_table,
                'table_max_rows': max(len(ex_table), 1),
            })
    return sections, annexes


def build_pdf_report(
    filtered_df: pd.DataFrame,
    filter_summary: dict[str, str],
    report_title: str,
    author_name: str,
    author_role: str,
    signature_date: str,
    references_text: str = "",
    stock_context: dict | None = None,
) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab no está instalado en el entorno.")
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.65 * inch,
        title=report_title,
        author=author_name,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="APA_Title", parent=styles["Title"], fontName="Times-Bold", fontSize=18, leading=24, alignment=TA_CENTER, spaceAfter=12, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="APA_Subtitle", parent=styles["Normal"], fontName="Times-Roman", fontSize=12, leading=16, alignment=TA_CENTER, spaceAfter=6, textColor=colors.HexColor("#222222")))
    styles.add(ParagraphStyle(name="APA_Heading", parent=styles["Heading2"], fontName="Times-Bold", fontSize=13, leading=16, alignment=TA_LEFT, spaceBefore=4, spaceAfter=6, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="APA_Body", parent=styles["BodyText"], fontName="Times-Roman", fontSize=11, leading=16, alignment=TA_JUSTIFY, spaceAfter=6))
    styles.add(ParagraphStyle(name="APA_Cell", parent=styles["BodyText"], fontName="Times-Roman", fontSize=8, leading=10, alignment=TA_LEFT, wordWrap='CJK'))
    styles.add(ParagraphStyle(name="APA_Cell_Header", parent=styles["BodyText"], fontName="Times-Bold", fontSize=8, leading=10, alignment=TA_LEFT, textColor=colors.white))
    styles.add(ParagraphStyle(name="APA_Cell_Tiny", parent=styles["BodyText"], fontName="Times-Roman", fontSize=7, leading=8.5, alignment=TA_LEFT, wordWrap='CJK'))
    styles.add(ParagraphStyle(name="APA_Cell_Header_Tiny", parent=styles["BodyText"], fontName="Times-Bold", fontSize=7, leading=8.5, alignment=TA_LEFT, textColor=colors.white))
    styles.add(ParagraphStyle(name="APA_Signature", parent=styles["BodyText"], fontName="Times-Roman", fontSize=11, leading=16, alignment=TA_LEFT, spaceAfter=3))

    elements = []
    short_title = re.sub(r"\s+", " ", (report_title.strip() or "Informe de base instalada"))[:80]
    generated_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    org_name = "DiaSorin S.p.A."
    title_for_cover = report_title or "Informe de base instalada"

    def page_header_footer(canvas, doc):
        canvas.saveState()
        width, height = landscape(A4)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#555555"))
        canvas.drawString(doc.leftMargin, height - 24, short_title)
        canvas.drawRightString(width - doc.rightMargin, height - 24, f"Página {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#D7D7D7"))
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, height - 30, width - doc.rightMargin, height - 30)
        canvas.line(doc.leftMargin, 24, width - doc.rightMargin, 24)
        canvas.drawString(doc.leftMargin, 12, generated_date)
        canvas.drawRightString(width - doc.rightMargin, 12, "Formato APA")
        canvas.restoreState()

    elements.append(Spacer(1, 1.3 * inch))
    elements.append(Paragraph(title_for_cover, styles["APA_Title"]))
    elements.append(Spacer(1, 0.18 * inch))
    elements.append(Paragraph(author_name, styles["APA_Subtitle"]))
    elements.append(Paragraph(author_role, styles["APA_Subtitle"]))
    elements.append(Paragraph(org_name, styles["APA_Subtitle"]))
    elements.append(Paragraph(signature_date, styles["APA_Subtitle"]))
    elements.append(PageBreak())

    insights, recommendations = _build_executive_insights(filtered_df, stock_context=stock_context)
    elements.append(Paragraph("Resumen ejecutivo", styles["APA_Heading"]))
    elements.append(Paragraph(
        "Este informe consolida la información visible en el dashboard filtrado y resume la base instalada, la configuración de equipo, el sistema operativo, la planificación de mantenimiento preventivo y la cobertura de repuestos del distribuidor seleccionado.",
        styles["APA_Body"],
    ))
    for text_line in insights:
        elements.append(Paragraph(f"• {_escape_pdf_text(text_line)}", styles["APA_Body"]))
    elements.append(Spacer(1, 0.06 * inch))
    elements.append(Paragraph("Acciones recomendadas", styles["APA_Heading"]))
    for text_line in recommendations:
        elements.append(Paragraph(f"• {_escape_pdf_text(text_line)}", styles["APA_Body"]))

    elements.append(Spacer(1, 0.08 * inch))
    filters_pairs = [(k, v) for k, v in filter_summary.items()]
    for block in _summary_table_from_pairs("Filtros aplicados", filters_pairs, styles):
        elements.append(block)

    sections, annexes = _build_pdf_sections(filtered_df, stock_context=stock_context)
    from reportlab.platypus import Image
    for section in sections:
        elements.append(PageBreak())
        elements.append(Paragraph(section['title'], styles['APA_Heading']))
        if section.get('intro'):
            elements.append(Paragraph(section['intro'], styles['APA_Body']))
        for block in _summary_table_from_pairs("Resumen de la sección", section['summary_pairs'], styles):
            elements.append(block)
        charts = [c for c in section.get('charts', []) if c is not None]
        if charts:
            elements.append(Spacer(1, 0.05 * inch))
            for fl in _pdf_image_flowables(charts, max_per_row=2):
                elements.append(fl)
        table_df = section.get('table_df', pd.DataFrame())
        if table_df is not None and not table_df.empty:
            elements.append(Paragraph(section.get('table_title', 'Tabla de apoyo'), styles['APA_Heading']))
            col_widths = None
            if section['title'] == 'Resumen de base instalada':
                width_map = {'Región': 0.75 * inch, 'País': 0.75 * inch, 'Distribuidor': 1.1 * inch, 'Cliente': 1.25 * inch, 'Instrumento': 1.0 * inch, 'Serial': 0.8 * inch, 'Estado': 0.8 * inch, 'Estado detallado': 0.95 * inch, 'Sistema operativo': 0.85 * inch, 'Condición': 0.75 * inch, 'Fecha de instalación': 0.9 * inch, 'Tipo de contrato': 1.25 * inch}
                col_widths = [width_map.get(c, 0.9 * inch) for c in table_df.columns]
            elif section['title'] == 'Repuestos y brecha de carstock':
                width_map = {'Parte requerida': 1.0 * inch, 'Descripción': 1.7 * inch, 'Cant. requerida': 0.8 * inch, 'Cant. cargada': 0.8 * inch, 'Brecha': 0.7 * inch, 'Estado': 0.8 * inch, 'Costo estimado opción 2': 1.0 * inch, 'Moneda': 0.55 * inch}
                col_widths = [width_map.get(c, 0.9 * inch) for c in table_df.columns]
            max_rows = section.get('table_max_rows', len(table_df))
            elements.append(_df_to_wrapped_table(table_df, styles, col_widths=col_widths, max_rows=max_rows))
            if isinstance(max_rows, int) and len(table_df) > max_rows:
                elements.append(Paragraph(f"Nota. En el cuerpo principal solo se muestran las primeras {max_rows} filas. El detalle completo se encuentra en los anexos.", styles['APA_Body']))

    elements.append(PageBreak())
    elements.append(Paragraph("Conclusiones", styles["APA_Heading"]))
    for text_line in insights[:4]:
        elements.append(Paragraph(f"• {_escape_pdf_text(text_line)}", styles["APA_Body"]))
    elements.append(Spacer(1, 0.06 * inch))
    elements.append(Paragraph("Fuente de datos", styles["APA_Heading"]))
    elements.append(Paragraph("Fuente de datos: registros filtrados del dashboard y, cuando aplica, archivo de stock cargado en la sesión actual.", styles["APA_Body"]))
    elements.append(Spacer(1, 0.08 * inch))
    elements.append(Paragraph("Firma", styles["APA_Heading"]))
    elements.append(Paragraph(_escape_pdf_text(author_name), styles["APA_Signature"]))
    elements.append(Paragraph(_escape_pdf_text(author_role), styles["APA_Signature"]))
    elements.append(Paragraph(_escape_pdf_text(org_name), styles["APA_Signature"]))
    elements.append(Paragraph(f"Fecha: {_escape_pdf_text(signature_date)}", styles["APA_Signature"]))

    for annex in annexes:
        elements.append(PageBreak())
        elements.append(Paragraph(annex['title'], styles['APA_Heading']))
        if annex.get('intro'):
            elements.append(Paragraph(annex['intro'], styles['APA_Body']))
        for block in _summary_table_from_pairs("Resumen del anexo", annex['summary_pairs'], styles):
            elements.append(block)
        if annex.get('charts'):
            for fl in _pdf_image_flowables(annex['charts'], max_per_row=2):
                elements.append(fl)
        table_df = annex.get('table_df', pd.DataFrame())
        if table_df is not None and not table_df.empty:
            elements.append(Paragraph(annex.get('table_title', 'Tabla del anexo'), styles['APA_Heading']))
            elements.append(_df_to_wrapped_table(table_df, styles, max_rows=annex.get('table_max_rows', len(table_df))))

    def cover_page(canvas, doc):
        canvas.saveState()
        canvas.restoreState()

    doc.build(elements, onFirstPage=cover_page, onLaterPages=page_header_footer)
    return buffer.getvalue()

def metric_card(label: str, value: str, subtitle: str = "") -> None:
    st.markdown(
        f"""
        <div class="metric-shell">
            <div class="metric-orb"></div>
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-sub">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def safe_text(value, fallback: str = "N/A") -> str:
    if pd.isna(value):
        return fallback
    text = str(value).strip()
    return text if text else fallback


def safe_number_text(value, fallback: str = "0") -> str:
    if pd.isna(value):
        return fallback
    try:
        val = float(value)
    except Exception:
        return fallback
    return f"{int(val):,}" if float(val).is_integer() else f"{val:,.1f}"


BLOOD_BANK_HEADER_ALIASES = {
    "in blood bank",
    "in blook bank",
    "in bloock bank",
    "in blod bank",
    "blood bank",
    "bloodbank",
    "blook bank",
    "banco de sangre",
    "banco sangre",
}


def is_blood_bank_yes(value) -> bool:
    if pd.isna(value):
        return False
    txt = str(value).strip().lower()
    txt = txt.replace('="', '').replace('"', '').strip()
    txt = re.sub(r"\s+", " ", txt)
    if not txt or txt in {"no", "n", "false", "0", "0.0", "nan", "none", "unknown", "data not available", "not applicable", "n.a.", "na"}:
        return False
    return txt in {"yes", "y", "true", "1", "1.0", "si", "sí", "s", "x"}


def _first_valid_series_from_columns(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    result = pd.Series(pd.NA, index=df.index, dtype="object")
    for col in columns:
        data = df[col]
        if isinstance(data, pd.DataFrame):
            data = data.bfill(axis=1).iloc[:, 0]
        data = data.replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        result = result.fillna(data)
    return result


def standardize_blood_bank_column(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    candidate_cols = []
    for col in df.columns:
        normalized = normalize_column_label(col)
        if normalized in BLOOD_BANK_HEADER_ALIASES:
            candidate_cols.append(col)
    # El export real de Records List trae el typo "In Blook Bank"; aquí se consolida
    # siempre en la columna canónica usada por el dashboard.
    if candidate_cols:
        df["In Blood Bank"] = _first_valid_series_from_columns(df, candidate_cols)
    elif "In Blood Bank" not in df.columns:
        df["In Blood Bank"] = pd.NA
    df["Blood Bank Raw"] = df["In Blood Bank"]
    df["Blood Bank Flag"] = df["In Blood Bank"].map(is_blood_bank_yes).fillna(False).astype(bool)
    return df


def count_blood_bank_yes(df: pd.DataFrame) -> int:
    if df is None or df.empty:
        return 0
    if "Blood Bank Flag" in df.columns:
        return int(df["Blood Bank Flag"].fillna(False).astype(bool).sum())
    if "In Blood Bank" in df.columns:
        return int(df["In Blood Bank"].map(is_blood_bank_yes).fillna(False).sum())
    return 0


def build_blood_bank_donut(df: pd.DataFrame) -> go.Figure:
    total_assets = int(len(df))
    yes_count = count_blood_bank_yes(df) if total_assets else 0
    no_count = max(total_assets - yes_count, 0)
    summary = pd.DataFrame({"Label": ["Banco de sangre", "Equipos en laboratorio"], "Count": [yes_count, no_count]})

    fig = go.Figure()
    fig.add_trace(
        go.Pie(
            labels=summary["Label"],
            values=summary["Count"],
            customdata=np.column_stack([summary["Label"], summary["Count"]]),
            hole=0.68,
            sort=False,
            marker=dict(colors=[ACCENT_3, "rgba(255,255,255,0.28)"], line=dict(color="rgba(255,255,255,0.20)", width=1.2)),
            textinfo="percent",
            textfont=dict(color="#ffffff", size=12),
            hovertemplate="%{customdata[0]}<br>Equipos: %{customdata[1]}<br>Participación: %{percent}<extra></extra>",
        )
    )
    fig.add_annotation(
        text=f"<b>{yes_count:,}</b><br><span style='font-size:11px'>de {total_assets:,}</span>",
        x=0.5,
        y=0.5,
        xref="paper",
        yref="paper",
        showarrow=False,
        font=dict(color="#ffffff", size=18),
    )
    fig.update_layout(
        title="Banco de sangre",
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5, bgcolor="rgba(14,26,42,0.36)", bordercolor="rgba(124,221,255,0.22)", borderwidth=1, font=dict(color="#f8fbff", size=11)),
    )
    return glow_layout(fig, 340, 15)


def normalize_part_number(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.startswith('="') and text.endswith('"'):
        text = text[2:-1]
    text = text.replace(".0", "").replace(" ", "").replace("\n", "").replace("\t", "")
    return text.upper()


def normalize_key_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def normalize_search_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()



def shorten_distributor_name(name: str, max_len: int = 22) -> str:
    text_name = safe_text(name, "No informado")
    exact_map = {
        "annar diagnostica import sas": "Annar",
        "laboratorios cienvar s a": "Cienvar",
        "wm argentina s a": "WM Argentina",
        "grupo bios": "Grupo Bios",
        "bio nuclear": "Bio-Nuclear",
        "diagnostico ual": "Diag. UAL",
        "biotec del paraguay s r l": "Biotec Paraguay",
        "biotec del paraguay": "Biotec Paraguay",
        "islalab products llc": "IslaLab",
        "capris médica": "Capris",
        "capris medica": "Capris",
        "dimex medica": "Dimex",
        "caribbean medical supplies inc": "Caribbean Medical",
        "simed ecuador": "Simed Ecuador",
        "simed ecuador": "Simed Ecuador",
    }
    if text_name in exact_map:
        short = exact_map[text_name]
        return short if len(short) <= max_len else short[: max_len - 1] + "…"

    cleaned = re.sub(r"\b(s\.a\.?|s\.a\.s\.?|s\.r\.l\.?|ltd\.?|llc|inc\.?|corp\.?|corporation|company|import|imports|diagnostica|diagnostics|medical|medica|laboratorios|laboratorio)\b", "", text_name, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,.-")
    if not cleaned:
        cleaned = text_name

    words = cleaned.split()
    if len(cleaned) <= max_len:
        return cleaned
    if len(words) >= 2:
        candidate = " ".join(words[:2]).strip()
        if len(candidate) <= max_len:
            return candidate
    return cleaned[: max_len - 1].rstrip() + "…"


def wrap_chart_title(text_value: str, width: int = 28) -> str:
    return "<br>".join(textwrap.wrap(safe_text(text_value, ""), width=width)) if safe_text(text_value, "") else ""

def build_long_palette(n: int) -> list[str]:
    base = [ACCENT, ACCENT_2, ACCENT_3, WARNING, "#9BB1FF", "#C084FC", "#F472B6", "#60A5FA", "#34D399", "#F59E0B", "#A78BFA", "#F87171", "#22D3EE", "#4ADE80"]
    if n <= len(base):
        return base[:n]
    repeats = (n // len(base)) + 1
    return (base * repeats)[:n]


def distributor_display_name(name: str, max_len: int = 22) -> str:
    text_name = safe_text(name, "No informado")
    return shorten_distributor_name(text_name, max_len=max_len)


def summarize_distributor_counts(summary_df: pd.DataFrame, top_n: int = 5) -> pd.DataFrame:
    """
    Mantiene solo los distribuidores más relevantes para la vista ejecutiva.
    No agrupa en "Otros"; simplemente corta el dataset al top_n para que la
    visual principal permanezca legible y ordenada.
    """
    if summary_df is None or summary_df.empty:
        return pd.DataFrame(columns=["Distributor name", "Count"])

    work = summary_df.copy()
    if "Distributor name" not in work.columns or "Count" not in work.columns:
        return work

    work["Distributor name"] = work["Distributor name"].fillna("No informado").astype(str)
    work["Count"] = pd.to_numeric(work["Count"], errors="coerce").fillna(0)

    work = (
        work.sort_values(["Count", "Distributor name"], ascending=[False, True])
        .reset_index(drop=True)
    )

    if isinstance(top_n, int) and top_n > 0:
        work = work.head(top_n).copy()

    return work.reset_index(drop=True)

def build_distributor_detail_table(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Modelo",
                "Distribuidor",
                "Cantidad",
                "% del modelo",
                "% del total filtrado",
            ]
        )

    work = df.copy()
    work["Instrument type"] = work["Instrument type"].fillna("No informado").astype(str).str.strip()
    work["Distributor name"] = work["Distributor name"].fillna("No informado").astype(str).str.strip()

    summary = (
        work.groupby(["Instrument type", "Distributor name"], dropna=False)
        .size()
        .reset_index(name="Cantidad")
    )

    if summary.empty:
        return pd.DataFrame(
            columns=[
                "Modelo",
                "Distribuidor",
                "Cantidad",
                "% del modelo",
                "% del total filtrado",
            ]
        )

    total_filtered = int(summary["Cantidad"].sum())

    model_totals = (
        summary.groupby("Instrument type", as_index=False)["Cantidad"]
        .sum()
        .rename(columns={"Cantidad": "Total modelo"})
    )

    summary = summary.merge(model_totals, on="Instrument type", how="left")
    summary["% del modelo"] = (summary["Cantidad"] / summary["Total modelo"] * 100).round(1)
    summary["% del total filtrado"] = (summary["Cantidad"] / total_filtered * 100).round(1)

    summary = summary.rename(columns={"Instrument type": "Modelo", "Distributor name": "Distribuidor"})
    summary = summary.sort_values(by=["Modelo", "Cantidad", "Distribuidor"], ascending=[True, False, True]).reset_index(drop=True)

    return summary[["Modelo", "Distribuidor", "Cantidad", "% del modelo", "% del total filtrado"]]


def build_distributor_detail_bar(df: pd.DataFrame, selected_model: str) -> go.Figure:
    fig = go.Figure()

    if df is None or df.empty or not selected_model:
        fig.update_layout(title="Detalle completo por distribuidor")
        return glow_layout(fig, 520, 15)

    work = df.copy()
    work["Instrument type"] = work["Instrument type"].fillna("No informado").astype(str).str.strip()
    work["Distributor name"] = work["Distributor name"].fillna("No informado").astype(str).str.strip()

    model_df = work[work["Instrument type"].eq(selected_model)].copy()
    if model_df.empty:
        fig.update_layout(title=f"Detalle completo | {selected_model}")
        return glow_layout(fig, 520, 15)

    summary = (
        model_df.groupby("Distributor name", dropna=False)
        .size()
        .reset_index(name="Count")
    )
    if summary.empty:
        fig.update_layout(title=f"Detalle completo | {selected_model}")
        return glow_layout(fig, 520, 15)

    summary["Distributor name"] = summary["Distributor name"].fillna("No informado").astype(str)
    summary["Count"] = pd.to_numeric(summary["Count"], errors="coerce").fillna(0)
    summary = summary.sort_values(["Count", "Distributor name"], ascending=[False, True]).reset_index(drop=True)
    summary["Display name"] = summary["Distributor name"].map(lambda x: distributor_display_name(x, 28))

    palette = build_long_palette(len(summary))
    color_map = {row["Display name"]: palette[i % len(palette)] for i, (_, row) in enumerate(summary.iterrows())}
    order = summary["Display name"].tolist()[::-1]

    fig = px.bar(
        summary,
        x="Count",
        y="Display name",
        orientation="h",
        text="Count",
        title=f"Detalle completo | {selected_model}",
        custom_data=["Distributor name", "Count"],
        color="Display name",
        color_discrete_map=color_map,
        category_orders={"Display name": order},
    )
    fig.update_traces(
        textposition="outside",
        hovertemplate="<b>Modelo:</b> " + selected_model + "<br><b>Distribuidor:</b> %{customdata[0]}<br><b>Cantidad:</b> %{customdata[1]}<extra></extra>",
        showlegend=False,
    )
    fig.update_layout(
        xaxis_title="Cantidad de equipos",
        yaxis_title="Distribuidor",
        margin=dict(t=72, b=28, l=8, r=8),
        height=max(420, 80 + 32 * len(summary)),
    )
    return glow_layout(fig, max(420, 80 + 32 * len(summary)), 15)

DISTRIBUTOR_ALIASES = {
    "annar": "Annar Diagnostica Import sas",
    "bio nuclear": "Bio-Nuclear",
    "bionuclear": "Bio-Nuclear",
    "biotec": "Biotec del Paraguay",
    "biotec paraguay": "Biotec del Paraguay",
    "biotec del paraguay": "Biotec del Paraguay",
    "grupo bios": "Grupo Bios",
    "qls": "QLS",
    "simed ecuador": "Simed Ecuador",
    "simed peru": "Simed Perú",
    "wiener": "Wiener Lab",
    "wm argentina": "WM Argentina",
}


def infer_distributor_from_filename_strict(filename: str, distributor_options: list[str]) -> tuple[str | None, list[str]]:
    base = normalize_search_text(Path(filename).stem)
    if not base:
        return None, []

    alias_hits = []
    for alias, official_name in DISTRIBUTOR_ALIASES.items():
        if re.search(rf"\b{re.escape(alias)}\b", base):
            if official_name in distributor_options:
                alias_hits.append(official_name)

    alias_hits = list(dict.fromkeys(alias_hits))
    if len(alias_hits) == 1:
        return alias_hits[0], alias_hits
    if len(alias_hits) > 1:
        return None, alias_hits

    strong_hits = []
    for distributor in distributor_options:
        norm = normalize_search_text(distributor)
        if not norm:
            continue
        if re.search(rf"\b{re.escape(norm)}\b", base):
            strong_hits.append(distributor)

    strong_hits = list(dict.fromkeys(strong_hits))
    if len(strong_hits) == 1:
        return strong_hits[0], strong_hits
    if len(strong_hits) > 1:
        return None, strong_hits

    candidates = []
    weak_tokens = {"bio", "lab", "labs", "import", "sas", "ltd", "pte", "sa", "srl", "corp", "group"}

    for distributor in distributor_options:
        tokens = [t for t in normalize_search_text(distributor).split() if len(t) >= 4 and t not in weak_tokens]
        if not tokens:
            continue
        hits = sum(1 for t in tokens if re.search(rf"\b{re.escape(t)}\b", base))
        if hits > 0:
            candidates.append((hits, len(tokens), distributor))

    candidates.sort(key=lambda x: (-x[0], -x[1], x[2]))
    if not candidates:
        return None, []

    top_score = candidates[0][0]
    tied = [c[2] for c in candidates if c[0] == top_score]
    if len(tied) == 1:
        return tied[0], tied
    return None, tied


def os_upgrade_bucket(value) -> str:
    text = safe_text(value, "No informado")
    if text == "Windows 10":
        return "Windows 10 / OK"
    if text in {"Windows XP", "Windows Vista", "Windows 7", "Windows 2000"}:
        return "Legacy / urgente migrar"
    if text in {"Unknown", "No informado", "Not installed"}:
        return "Revisar campo OS"
    return "Otro OS / validar"


def format_date_for_hover(value) -> str:
    if pd.isna(value):
        return "N/A"
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return safe_text(value)


def to_numeric_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.fillna("")
        .astype(str)
        .str.strip()
        .str.replace(r"[^0-9,\.\-]", "", regex=True)
        .str.replace(",", ".", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def normalize_instrument_type(value) -> str:
    text = safe_text(value, "")
    if ":" in text:
        text = text.split(":", 1)[1]
    return text.strip() or safe_text(value)


# =============================================================================
# FILTRO INTERACTIVO DESDE GRÁFICAS
# =============================================================================
CHART_DRILL_SESSION_KEY = "active_chart_drill_filter_stack_v30"
CHART_DRILL_RESET_COUNTER_KEY = "chart_drill_reset_counter_v30"
SIDEBAR_CLEAR_PENDING_KEY = "sidebar_filter_clear_pending_v30"
SIDEBAR_CLEAR_COUNTER_KEY = "sidebar_filter_clear_counter_v30"

SIDEBAR_REGION_KEY = "sidebar_regions_v30"
SIDEBAR_COUNTRY_KEY = "sidebar_countries_v30"
SIDEBAR_DISTRIBUTOR_KEY = "sidebar_distributors_v30"
SIDEBAR_INSTRUMENT_KEY = "sidebar_instruments_v30"
SIDEBAR_STATE_KEY = "sidebar_states_v30"
ACTIVE_DASHBOARD_TAB_KEY = "active_dashboard_tab_v30"
MANUFACTURING_EXCEL_EXPORT_SESSION_KEY = "manufacturing_excel_export_df_v38"
MANUFACTURING_EXCEL_EXPORT_SOURCE_KEY = "manufacturing_excel_export_source_v38"
DASHBOARD_TABS = [
    "Base instalada",
    "Machine configuration",
    "Sistema operativo",
    "Procesamiento / PM",
    "Stock / Carstock gap",
    "Antigüedad / fabricación",
    "Detalle por equipo",
]
SIDEBAR_FILTER_KEYS = [
    SIDEBAR_REGION_KEY,
    SIDEBAR_COUNTRY_KEY,
    SIDEBAR_DISTRIBUTOR_KEY,
    SIDEBAR_INSTRUMENT_KEY,
    SIDEBAR_STATE_KEY,
]


def ensure_chart_drill_state() -> None:
    if CHART_DRILL_SESSION_KEY not in st.session_state:
        st.session_state[CHART_DRILL_SESSION_KEY] = []
    active = st.session_state.get(CHART_DRILL_SESSION_KEY)
    if isinstance(active, dict):
        st.session_state[CHART_DRILL_SESSION_KEY] = [active] if active.get("filters") else []
    elif not isinstance(active, list):
        st.session_state[CHART_DRILL_SESSION_KEY] = []
    if CHART_DRILL_RESET_COUNTER_KEY not in st.session_state:
        st.session_state[CHART_DRILL_RESET_COUNTER_KEY] = 0


def compute_installation_stage(df: pd.DataFrame) -> pd.Series:
    status_series = df.get("Operational status", pd.Series("No informado", index=df.index)).fillna("No informado").astype(str)
    stage = pd.Series("Installed / Active", index=df.index, dtype="object")
    stage = stage.mask(status_series.str.contains("ready", case=False, na=False), "Ready to install")
    stage = stage.mask(status_series.str.contains("transit|customs|shipping", case=False, na=False), "Transit / Customs")
    stage = stage.mask(status_series.str.contains("warehouse|stock", case=False, na=False), "Warehouse")
    return stage


def build_city_label_series(df: pd.DataFrame) -> pd.Series:
    city = df.get("City", pd.Series("No informado", index=df.index)).fillna("No informado").astype(str).str.strip()
    country = df.get("Country", pd.Series("No country", index=df.index)).fillna("No country").astype(str).str.strip()
    city = city.replace("", "No informado")
    country = country.replace("", "No country")
    return city + " | " + country


def get_active_chart_drill_filters() -> list[dict]:
    ensure_chart_drill_state()
    active = st.session_state.get(CHART_DRILL_SESSION_KEY, [])
    if isinstance(active, dict):
        active = [active] if active.get("filters") else []
    if not isinstance(active, list):
        return []
    return [item for item in active if isinstance(item, dict) and item.get("filters")]


def get_active_chart_drill_filter() -> dict | None:
    filters = get_active_chart_drill_filters()
    if not filters:
        return None
    return filters[-1]


def clear_chart_drill_filter() -> None:
    ensure_chart_drill_state()
    st.session_state[CHART_DRILL_SESSION_KEY] = []
    st.session_state[CHART_DRILL_RESET_COUNTER_KEY] = int(st.session_state.get(CHART_DRILL_RESET_COUNTER_KEY, 0)) + 1
    st.rerun()


def pop_last_chart_drill_filter() -> None:
    ensure_chart_drill_state()
    active = get_active_chart_drill_filters()
    if active:
        st.session_state[CHART_DRILL_SESSION_KEY] = active[:-1]
        st.session_state[CHART_DRILL_RESET_COUNTER_KEY] = int(st.session_state.get(CHART_DRILL_RESET_COUNTER_KEY, 0)) + 1
        st.rerun()


def remove_chart_drill_filter_at(index: int) -> None:
    ensure_chart_drill_state()
    active = get_active_chart_drill_filters()
    if 0 <= index < len(active):
        del active[index]
        st.session_state[CHART_DRILL_SESSION_KEY] = active
        st.session_state[CHART_DRILL_RESET_COUNTER_KEY] = int(st.session_state.get(CHART_DRILL_RESET_COUNTER_KEY, 0)) + 1
        st.rerun()


def request_sidebar_filter_clear() -> None:
    """Marca los filtros laterales para limpieza segura en el siguiente rerun.

    Streamlit no permite modificar st.session_state de widgets ya renderizados
    durante el mismo ciclo. Por eso no se limpian directamente desde botones
    ubicados después de los multiselect; se deja una bandera y se aplica antes
    de crear los widgets en el siguiente rerun.
    """
    st.session_state[SIDEBAR_CLEAR_PENDING_KEY] = True
    st.session_state[SIDEBAR_CLEAR_COUNTER_KEY] = int(st.session_state.get(SIDEBAR_CLEAR_COUNTER_KEY, 0)) + 1


def consume_pending_sidebar_filter_clear() -> None:
    """Limpia los multiselect del sidebar antes de que sean instanciados."""
    if bool(st.session_state.get(SIDEBAR_CLEAR_PENDING_KEY, False)):
        for key in SIDEBAR_FILTER_KEYS:
            st.session_state[key] = []
        st.session_state[SIDEBAR_CLEAR_PENDING_KEY] = False


def clear_sidebar_filter_widgets() -> None:
    request_sidebar_filter_clear()
    st.rerun()


def clear_all_dashboard_filters() -> None:
    # No modificar directamente las keys de widgets ya renderizados.
    request_sidebar_filter_clear()
    st.session_state[CHART_DRILL_SESSION_KEY] = []
    st.session_state[CHART_DRILL_RESET_COUNTER_KEY] = int(st.session_state.get(CHART_DRILL_RESET_COUNTER_KEY, 0)) + 1
    st.rerun()


def _drill_payload_signature(payload: dict | None) -> tuple:
    if not isinstance(payload, dict):
        return tuple()
    normalized_filters = []
    for item in payload.get("filters", []):
        if isinstance(item, dict):
            normalized_filters.append((str(item.get("column", "")), str(item.get("value", ""))))
    return (str(payload.get("source", "")), tuple(normalized_filters))


def set_chart_drill_filter(payload: dict | None) -> None:
    if not payload or not payload.get("filters"):
        return
    ensure_chart_drill_state()
    active = get_active_chart_drill_filters()
    signature = _drill_payload_signature(payload)
    existing_signatures = [_drill_payload_signature(item) for item in active]
    if signature not in existing_signatures:
        active.append(payload)
        st.session_state[CHART_DRILL_SESSION_KEY] = active
        st.rerun()


def make_chart_drill_payload(source: str, filters: list[dict], label: str) -> dict:
    clean_filters = []
    for item in filters:
        if not isinstance(item, dict):
            continue
        column = safe_text(item.get("column"), "").strip()
        value = safe_text(item.get("value"), "").strip()
        if column and value:
            clean_filters.append({"column": column, "value": value})
    return {"source": source, "filters": clean_filters, "label": label}


def _series_as_clean_text(df: pd.DataFrame, column: str, default: str = "No informado") -> pd.Series:
    if column in df.columns:
        return df[column].fillna(default).astype(str).str.strip().replace("", default)
    return pd.Series(default, index=df.index, dtype="object")


def apply_single_chart_filter(df: pd.DataFrame, item: dict) -> pd.Series:
    column = safe_text(item.get("column"), "")
    value = safe_text(item.get("value"), "")
    mask = pd.Series(True, index=df.index)

    if column == "Installation year":
        years = pd.to_datetime(df.get("Installation date", pd.Series(pd.NaT, index=df.index)), errors="coerce").dt.year
        try:
            return mask & years.eq(int(float(value)))
        except Exception:
            return pd.Series(False, index=df.index)
    if column == "Installation stage":
        return mask & compute_installation_stage(df).astype(str).eq(value)
    if column == "CityLabel":
        return mask & build_city_label_series(df).astype(str).eq(value)
    if column == "Config field populated":
        cfg_col = f"CFG::{value}"
        if cfg_col not in df.columns:
            return pd.Series(False, index=df.index)
        return mask & df[cfg_col].notna() & df[cfg_col].astype(str).str.strip().ne("")
    if column == "Blood Bank Flag":
        desired = value.lower() in {"yes", "true", "1", "banco de sangre", "blood bank"}
        if "Blood Bank Flag" in df.columns:
            return mask & df["Blood Bank Flag"].fillna(False).astype(bool).eq(desired)
        detected = df.apply(lambda row: detect_blood_bank_from_row(row)[0], axis=1)
        return mask & detected.eq(desired)
    if column == "OS Upgrade Bucket":
        os_values = _series_as_clean_text(df, "Operating System")
        return mask & os_values.map(os_upgrade_bucket).eq(value)
    if column == "Product Line contains":
        product_values = _series_as_clean_text(df, "Product Line", "")
        escaped = re.escape(value)
        return mask & product_values.str.contains(escaped, case=False, na=False)
    if column == "Manufacturing year":
        if "Manufacturing year" in df.columns:
            years = pd.to_numeric(df["Manufacturing year"], errors="coerce")
            try:
                return mask & years.eq(int(float(value)))
            except Exception:
                return pd.Series(False, index=df.index)
        return pd.Series(False, index=df.index)
    if column == "Manufacturing age bucket":
        if "Manufacturing age bucket" in df.columns:
            return mask & _series_as_clean_text(df, "Manufacturing age bucket").eq(value)
        return pd.Series(False, index=df.index)
    if column.startswith("CFG::") and column in df.columns:
        return mask & _series_as_clean_text(df, column).eq(value)
    if column in df.columns:
        return mask & _series_as_clean_text(df, column).eq(value)
    return pd.Series(False, index=df.index)


def apply_chart_drill_filter(df: pd.DataFrame) -> tuple[pd.DataFrame, bool]:
    active = get_active_chart_drill_filters()
    if not active:
        return df.copy(), False

    mask = pd.Series(True, index=df.index)
    for payload in active:
        for item in payload.get("filters", []):
            mask = mask & apply_single_chart_filter(df, item)
    return df[mask].copy(), True


def _render_filter_chip(text_value: str) -> str:
    safe = safe_text(text_value, "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return f'<span class="badge" style="margin-right:0.35rem;margin-bottom:0.35rem;display:inline-block;">{safe}</span>'


def render_chart_drill_filter_banner(base_df: pd.DataFrame, current_df: pd.DataFrame, sidebar_filter_summary: dict[str, str] | None = None) -> None:
    active = get_active_chart_drill_filters()
    sidebar_filter_summary = sidebar_filter_summary or {}
    sidebar_active = {k: v for k, v in sidebar_filter_summary.items() if safe_text(v, "Todos") not in {"Todos", "All", ""}}

    if not active and not sidebar_active:
        return

    total_before = len(base_df)
    total_after = len(current_df)
    pct = _safe_share_pct(total_after, total_before)

    st.markdown(
        f'''
        <div style="border:1px solid rgba(113,225,255,0.30);border-radius:18px;padding:0.85rem 1rem;margin:0.7rem 0 1rem 0;background:rgba(113,225,255,0.08);box-shadow:0 0 18px rgba(53,200,255,0.08);">
            <div style="font-size:0.78rem;text-transform:uppercase;letter-spacing:0.12em;color:rgba(235,245,255,0.70);">Filtros aplicados</div>
            <div style="font-size:0.90rem;color:rgba(235,245,255,0.86);margin-top:0.18rem;">Mostrando {total_after:,} de {total_before:,} registros después de los filtros interactivos ({pct:.1f}%).</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )

    if sidebar_active:
        st.markdown("**Filtros laterales activos**")
        chips = "".join(_render_filter_chip(f"{k}: {v}") for k, v in sidebar_active.items())
        st.markdown(chips, unsafe_allow_html=True)

    if active:
        st.markdown("**Filtros aplicados desde gráficas**")
        for idx, payload in enumerate(active):
            label = safe_text(payload.get("label"), "Filtro gráfico")
            source = safe_text(payload.get("source"), "Gráfica")
            cols = st.columns([0.84, 0.16])
            with cols[0]:
                st.markdown(f"{idx + 1}. **{label}**  \nOrigen: `{source}`")
            with cols[1]:
                if st.button("Quitar", key=f"remove_chart_filter_{idx}_v30"):
                    remove_chart_drill_filter_at(idx)

    action_cols = st.columns(3)
    with action_cols[0]:
        if active and st.button("← Deshacer último filtro gráfico", key="undo_last_chart_filter_button_v30"):
            pop_last_chart_drill_filter()
    with action_cols[1]:
        if active and st.button("Limpiar filtros gráficos", key="clear_chart_drill_filter_button_v30"):
            clear_chart_drill_filter()
    with action_cols[2]:
        if st.button("Limpiar todos los filtros", key="clear_all_filters_button_v30"):
            clear_all_dashboard_filters()


def extract_plotly_selected_point(event) -> dict | None:
    if event is None:
        return None
    points = None
    try:
        points = event.selection.points
    except Exception:
        pass
    if points is None and isinstance(event, dict):
        points = event.get("selection", {}).get("points", [])
    if points is None:
        return None
    try:
        if len(points) == 0:
            return None
        point = points[0]
    except Exception:
        return None
    if isinstance(point, dict):
        return point
    try:
        return dict(point)
    except Exception:
        return None


def _get_customdata(point: dict, index: int, default=""):
    custom = point.get("customdata")
    try:
        if custom is not None and len(custom) > index:
            return custom[index]
    except Exception:
        pass
    return default


def render_drilldown_plotly_chart(fig: go.Figure, key: str, source_label: str, payload_builder, help_text: str | None = None) -> None:
    ensure_chart_drill_state()
    counter = int(st.session_state.get(CHART_DRILL_RESET_COUNTER_KEY, 0))
    widget_key = f"{key}_r{counter}"
    if help_text is None:
        help_text = "Filtro disponible: selecciona/clic en una barra, punto o segmento para aplicar un filtro al dashboard. Usa el panel de filtros aplicados para volver atrás o quitarlo."
    st.caption(help_text)
    try:
        event = st.plotly_chart(
            fig,
            use_container_width=True,
            key=widget_key,
            on_select="rerun",
            selection_mode="points",
        )
    except TypeError:
        st.plotly_chart(fig, use_container_width=True, key=widget_key)
        st.caption("La versión actual de Streamlit no expone selección directa para esta gráfica; la visualización queda normal, sin filtro por clic.")
        return

    point = extract_plotly_selected_point(event)
    if point:
        payload = payload_builder(point)
        if payload:
            set_chart_drill_filter(payload)


def render_pie_filter_fallback_buttons(
    options,
    key_prefix: str,
    payload_from_value,
    max_options: int = 8,
    excluded_labels: set[str] | None = None,
) -> None:
    """Fallback seguro para gráficas circulares de Plotly/Pie.

    En Streamlit, la selección por clic es mucho más estable en barras que en
    gráficas tipo Pie. Para evitar que Machine Configuration quede sin filtro,
    se muestran botones por segmento debajo de cada gráfica circular.
    """
    excluded_labels = excluded_labels or {"Otros"}
    clean_options = []
    seen = set()
    for value in list(options or []):
        label = safe_text(value, "").strip()
        if not label or label in excluded_labels or label.lower() in {"nan", "none", "<na>"}:
            continue
        if label not in seen:
            clean_options.append(label)
            seen.add(label)
    if not clean_options:
        return

    st.caption("Filtro alternativo de esta gráfica circular: selecciona un valor y aplica el filtro. Este control no depende del clic directo sobre el donut.")
    visible_options = clean_options[:max_options]
    selector_digest = hashlib.md5(str(key_prefix).encode("utf-8", errors="ignore")).hexdigest()[:10]
    selected_option = st.selectbox(
        "Valor para filtrar",
        options=["— Selecciona un valor —"] + visible_options,
        key=f"pie_filter_select_{key_prefix}_{selector_digest}_v30",
        label_visibility="collapsed",
    )
    if selected_option != "— Selecciona un valor —":
        if st.button(f"Aplicar filtro: {selected_option}", key=f"pie_filter_apply_{key_prefix}_{selector_digest}_v30", use_container_width=True):
            payload = payload_from_value(selected_option)
            if payload:
                set_chart_drill_filter(payload)

    n_cols = min(3, max(1, len(visible_options)))
    cols = st.columns(n_cols)
    for idx, label in enumerate(visible_options):
        short_label = label if len(label) <= 26 else label[:23] + "..."
        digest = hashlib.md5(f"{key_prefix}|{label}".encode("utf-8", errors="ignore")).hexdigest()[:10]
        with cols[idx % n_cols]:
            if st.button(f"Filtrar: {short_label}", key=f"pie_filter_{key_prefix}_{digest}_v30", use_container_width=True):
                payload = payload_from_value(label)
                if payload:
                    set_chart_drill_filter(payload)

    if len(clean_options) > max_options:
        st.caption(f"Se muestran {max_options} filtros principales de {len(clean_options)} valores. Los valores agrupados como 'Otros' no se filtran como segmento único para evitar mezclar categorías heterogéneas.")


def payload_from_axis_value(point: dict, column: str, source_label: str, label_prefix: str, axis: str = "x") -> dict | None:
    value = safe_text(point.get(axis), "").strip()
    if not value:
        return None
    return make_chart_drill_payload(source_label, [{"column": column, "value": value}], f"{label_prefix}: {value}")


def payload_from_serial_axis(point: dict, source_label: str, axis: str = "x") -> dict | None:
    serial = safe_text(point.get(axis), "").strip()
    if not serial:
        return None
    return make_chart_drill_payload(source_label, [{"column": "Serial number", "value": serial}], f"Serial: {serial}")


def payload_from_geo_point(point: dict) -> dict | None:
    serial = safe_text(_get_customdata(point, 0, ""), "").strip()
    if not serial:
        return None
    return make_chart_drill_payload(
        "Mapa geográfico de base instalada",
        [{"column": "Serial number", "value": serial}],
        f"Serial en mapa: {serial}",
    )



# =============================================================================
# MAPA DE BASE INSTALADA — SECCIÓN CRÍTICA
# =============================================================================
# Esta función usa Plotly ScatterGeo. NO requiere token de Mapbox, Google Maps,
# Leaflet, tile server ni clave API. Las capas geográficas (tierra, océano,
# países, costas, lagos y retícula) son parte del layout geo de Plotly.
#
# IMPORTANTE: no envuelva st.plotly_chart entre etiquetas HTML <div> abiertas y
# cerradas en llamadas st.markdown separadas. Streamlit renderiza cada llamada
# como un bloque React independiente y no garantiza que esas etiquetas envuelvan
# widgets posteriores. Ese patrón puede dejar el contenedor del mapa en blanco.
# =============================================================================

def _prepare_installed_base_geo_data(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Normaliza coordenadas y separa registros inválidos sin mutar el DataFrame original."""
    required = ["Latitude", "Longitude"]
    missing = [column for column in required if column not in df.columns]
    if missing:
        raise KeyError(f"Faltan columnas requeridas para el mapa: {', '.join(missing)}")

    geo = df.copy()
    geo["Latitude"] = pd.to_numeric(geo["Latitude"], errors="coerce")
    geo["Longitude"] = pd.to_numeric(geo["Longitude"], errors="coerce")

    coordinate_mask = (
        geo["Latitude"].notna()
        & geo["Longitude"].notna()
        & geo["Latitude"].between(-90, 90, inclusive="both")
        & geo["Longitude"].between(-180, 180, inclusive="both")
    )
    valid_geo = geo.loc[coordinate_mask].copy()
    invalid_geo = geo.loc[~coordinate_mask].copy()
    return valid_geo, invalid_geo


def _build_installed_base_geo_figure(geo_df: pd.DataFrame) -> go.Figure:
    """Construye el mapa conservando proyección, capas, marcadores y tooltips estables."""
    hover_columns = [
        "Serial number",
        "Instrument type",
        "Country",
        "Distributor name",
        "Operational status",
        "Commercial Region",
    ]
    geo_df = geo_df.copy()
    for column in hover_columns + ["Customer name"]:
        if column not in geo_df.columns:
            geo_df[column] = "No informado"

    # custom_data se declara explícitamente para garantizar que el serial siempre
    # permanezca en customdata[0], requerido por payload_from_geo_point().
    fig_geo = px.scatter_geo(
        geo_df,
        lat="Latitude",
        lon="Longitude",
        hover_name="Customer name",
        custom_data=hover_columns,
        height=560,
        projection="mollweide",
    )
    fig_geo.update_traces(
        marker=dict(
            size=7.0,
            color=ACCENT,
            opacity=0.98,
            line=dict(color="rgba(255,255,255,0.96)", width=1.25),
        ),
        hovertemplate=(
            "<b>%{hovertext}</b><br>"
            "Serie: %{customdata[0]}<br>"
            "Instrumento: %{customdata[1]}<br>"
            "País: %{customdata[2]}<br>"
            "Distribuidor: %{customdata[3]}<br>"
            "Estado: %{customdata[4]}<br>"
            "Región comercial: %{customdata[5]}<extra></extra>"
        ),
        selected=dict(marker=dict(size=10, opacity=1.0)),
        unselected=dict(marker=dict(opacity=0.55)),
    )
    fig_geo.update_geos(
        projection_type="mollweide",
        projection_scale=0.92,
        center=dict(lat=8, lon=0),
        showframe=False,
        bgcolor="rgba(255,255,255,0)",
        showocean=True,
        oceancolor="rgba(14,28,46,0.18)",
        showland=True,
        landcolor="rgba(255,255,255,0.14)",
        showcountries=True,
        countrycolor="rgba(190,235,255,0.28)",
        countrywidth=0.7,
        showcoastlines=True,
        coastlinecolor="rgba(190,235,255,0.22)",
        coastlinewidth=0.7,
        showlakes=True,
        lakecolor="rgba(30,52,80,0.12)",
        lataxis_showgrid=True,
        lonaxis_showgrid=True,
        lataxis_gridcolor="rgba(190,235,255,0.06)",
        lonaxis_gridcolor="rgba(190,235,255,0.06)",
        lataxis_dtick=15,
        lonaxis_dtick=30,
        domain=dict(x=[0.10, 0.90], y=[0.14, 0.86]),
    )
    fig_geo.update_layout(
        height=560,
        autosize=True,
        paper_bgcolor=PLOT_BG,
        plot_bgcolor=PLOT_BG,
        margin=dict(l=0, r=0, t=0, b=0),
        font=dict(color=TEXT),
        dragmode="pan",
        # Conserva el zoom/paneo elegido por el usuario durante reruns de filtros.
        uirevision="installed-base-geo-map-stable-v53",
    )
    return fig_geo


def render_installed_base_geo_map(filtered_df: pd.DataFrame) -> None:
    """Renderiza el mapa con selección por serial y fallback visible ante errores."""
    try:
        geo_df, invalid_geo_df = _prepare_installed_base_geo_data(filtered_df)
    except Exception as exc:
        st.error("No fue posible preparar las coordenadas del mapa.")
        st.caption(f"Detalle técnico: {exc}")
        return

    if geo_df.empty:
        st.info("No hay coordenadas válidas para mostrar en el mapa con los filtros actuales.")
        if not invalid_geo_df.empty:
            st.caption(f"Registros descartados por coordenadas vacías o fuera de rango: {len(invalid_geo_df):,}.")
        return

    # Cabecera visual balanceada en UNA sola llamada HTML. No encierra el widget.
    st.markdown(
        '''
        <div style="border:1px solid rgba(255,255,255,0.12); border-radius:20px;
                    padding:0.8rem 1rem; margin:0.2rem 0 0.55rem 0;
                    background:linear-gradient(180deg, rgba(193,221,255,0.10), rgba(17,27,42,0.22));">
            <div style="font-size:1.05rem; font-weight:700; color:#ffffff;">Vista global de la base instalada</div>
            <div style="font-size:0.82rem; color:rgba(231,243,255,0.84); margin-top:0.22rem;">
                Proyección global Mollweide · utiliza zoom, paneo y selección de puntos para filtrar por serial.
            </div>
        </div>
        ''',
        unsafe_allow_html=True,
    )

    try:
        fig_geo = _build_installed_base_geo_figure(geo_df)
    except Exception as exc:
        st.error("No fue posible construir la figura geográfica.")
        st.caption(f"Detalle técnico: {exc}")
        return

    ensure_chart_drill_state()
    counter = int(st.session_state.get(CHART_DRILL_RESET_COUNTER_KEY, 0))
    widget_key = f"geo_map_serial_chart_v53_r{counter}"
    plotly_config = {
        "responsive": True,
        "scrollZoom": True,
        "displayModeBar": True,
        "displaylogo": False,
        "showTips": True,
    }

    st.caption(
        "Filtro disponible: haz clic o selecciona un punto para filtrar ese serial. "
        "Los controles de Plotly permiten zoom, paneo, selección y restablecimiento."
    )

    try:
        event = st.plotly_chart(
            fig_geo,
            use_container_width=True,
            key=widget_key,
            on_select="rerun",
            selection_mode="points",
            config=plotly_config,
        )
        point = extract_plotly_selected_point(event)
        if point:
            payload = payload_from_geo_point(point)
            if payload:
                set_chart_drill_filter(payload)
    except TypeError:
        # Compatibilidad con versiones de Streamlit anteriores a on_select.
        try:
            st.plotly_chart(
                fig_geo,
                use_container_width=True,
                key=f"{widget_key}_compat",
                config=plotly_config,
            )
            st.caption(
                "El mapa está operativo. Esta versión de Streamlit no admite el filtro por clic; "
                "zoom, paneo y controles permanecen disponibles."
            )
        except Exception as fallback_exc:
            st.error("El mapa no pudo renderizarse en modo de compatibilidad.")
            st.caption(f"Detalle técnico: {fallback_exc}")
    except Exception as exc:
        # Nunca dejar un contenedor silenciosamente vacío.
        st.error("El mapa encontró un error durante el renderizado.")
        st.caption(f"Detalle técnico: {exc}")
        with st.expander("Diagnóstico de coordenadas", expanded=False):
            st.write(f"Registros válidos para el mapa: {len(geo_df):,}")
            st.write(f"Registros descartados: {len(invalid_geo_df):,}")
            diagnostic_columns = [
                column for column in [
                    "Customer name", "Serial number", "Country", "Latitude", "Longitude"
                ] if column in invalid_geo_df.columns
            ]
            if diagnostic_columns and not invalid_geo_df.empty:
                st.dataframe(
                    invalid_geo_df[diagnostic_columns].head(100),
                    use_container_width=True,
                    hide_index=True,
                )



def payload_from_instrument_point(point: dict, source_label: str) -> dict | None:
    instrument = safe_text(_get_customdata(point, 0, point.get("y")), "").strip()
    if not instrument:
        return None
    return make_chart_drill_payload(
        source_label,
        [{"column": "Instrument type", "value": instrument}],
        f"Instrumento: {instrument}",
    )


def payload_from_installation_year_point(point: dict) -> dict | None:
    year_value = _get_customdata(point, 0, point.get("x"))
    try:
        year_text = str(int(float(year_value)))
    except Exception:
        return None
    return make_chart_drill_payload(
        "Instalaciones por año",
        [{"column": "Installation year", "value": year_text}],
        f"Año de instalación: {year_text}",
    )


def payload_from_pipeline_point(point: dict) -> dict | None:
    instrument = safe_text(_get_customdata(point, 0, point.get("x")), "").strip()
    stage = safe_text(_get_customdata(point, 1, ""), "").strip()
    if not instrument or not stage:
        return None
    return make_chart_drill_payload(
        "Sistemas instalados vs listos / pipeline",
        [
            {"column": "Instrument type", "value": instrument},
            {"column": "Installation stage", "value": stage},
        ],
        f"Instrumento: {instrument} · Etapa: {stage}",
    )



def payload_from_model_status_point(point: dict) -> dict | None:
    instrument = safe_text(_get_customdata(point, 0, point.get("y")), "").strip()
    status_value = safe_text(_get_customdata(point, 1, ""), "").strip()
    if not instrument or not status_value:
        return None
    return make_chart_drill_payload(
        "Base instalada por modelo y estado operativo",
        [
            {"column": "Instrument type", "value": instrument},
            {"column": "Operational status grouped", "value": status_value},
        ],
        f"Modelo: {instrument} · Estado operativo: {status_value}",
    )


def payload_from_city_point(point: dict) -> dict | None:
    city_label = safe_text(_get_customdata(point, 0, point.get("y")), "").strip()
    if not city_label:
        return None
    return make_chart_drill_payload(
        "Análisis por ciudad",
        [{"column": "CityLabel", "value": city_label}],
        f"Ciudad / país: {city_label}",
    )


def payload_from_global_distributor_point(point: dict) -> dict | None:
    instrument = safe_text(_get_customdata(point, 0, ""), "").strip()
    distributor = safe_text(_get_customdata(point, 1, ""), "").strip()
    if not instrument or not distributor:
        return None
    return make_chart_drill_payload(
        "Vista global por distribuidor",
        [
            {"column": "Instrument type", "value": instrument},
            {"column": "Distributor name", "value": distributor},
        ],
        f"Modelo: {instrument} · Distribuidor: {distributor}",
    )


def payload_from_detail_distributor_point(selected_model: str, point: dict) -> dict | None:
    distributor = safe_text(_get_customdata(point, 0, ""), "").strip()
    if not selected_model or not distributor:
        return None
    return make_chart_drill_payload(
        f"Detalle completo | {selected_model}",
        [
            {"column": "Instrument type", "value": selected_model},
            {"column": "Distributor name", "value": distributor},
        ],
        f"Modelo: {selected_model} · Distribuidor: {distributor}",
    )


def payload_from_distributor_model_donut(selected_model: str, point: dict) -> dict | None:
    distributor = safe_text(_get_customdata(point, 0, ""), "").strip()
    if not selected_model or not distributor or distributor.lower().startswith("otros"):
        return None
    return make_chart_drill_payload(
        f"Distribución por distribuidor | {selected_model}",
        [
            {"column": "Instrument type", "value": selected_model},
            {"column": "Distributor name", "value": distributor},
        ],
        f"Modelo: {selected_model} · Distribuidor: {distributor}",
    )


def payload_from_blood_bank_point(point: dict) -> dict | None:
    label = safe_text(_get_customdata(point, 0, point.get("label", "")), "").strip()
    if not label:
        return None
    desired = "Yes" if label.lower().strip() == "banco de sangre" else "No"
    return make_chart_drill_payload(
        "Banco de sangre",
        [{"column": "Blood Bank Flag", "value": desired}],
        f"Banco de sangre: {label}",
    )


def payload_from_config_coverage_point(point: dict) -> dict | None:
    field_name = safe_text(point.get("y", _get_customdata(point, 0, "")), "").strip()
    if not field_name:
        return None
    return make_chart_drill_payload(
        "Cobertura por campo aplicable",
        [{"column": "Config field populated", "value": field_name}],
        f"Campo con dato: {field_name}",
    )


def payload_from_config_value_point(field_name: str, point: dict) -> dict | None:
    value = safe_text(_get_customdata(point, 0, point.get("label", "")), "").strip()
    if not field_name or not value or value.lower() == "otros":
        return None
    return make_chart_drill_payload(
        f"Configuración | {field_name}",
        [{"column": f"CFG::{field_name}", "value": value}],
        f"{field_name}: {value}",
    )


def payload_from_product_line_point(point: dict) -> dict | None:
    product_line = safe_text(point.get("y", _get_customdata(point, 0, "")), "").strip()
    if not product_line:
        return None
    return make_chart_drill_payload(
        "Product line performed on the analyzer",
        [{"column": "Product Line contains", "value": product_line}],
        f"Product line: {product_line}",
    )


def payload_from_manufacturing_age_bucket(point: dict) -> dict | None:
    bucket = safe_text(point.get("x", _get_customdata(point, 0, "")), "").strip()
    if not bucket:
        return None
    return make_chart_drill_payload(
        "Estado de la base instalada por rango de edad",
        [{"column": "Manufacturing age bucket", "value": bucket}],
        f"Rango de edad: {bucket}",
    )


def payload_from_manufacturing_year(point: dict) -> dict | None:
    year_value = safe_text(point.get("x", _get_customdata(point, 0, "")), "").strip()
    try:
        year_text = str(int(float(year_value)))
    except Exception:
        return None
    return make_chart_drill_payload(
        "Equipos por año de fabricación",
        [{"column": "Manufacturing year", "value": year_text}],
        f"Año de fabricación: {year_text}",
    )

CODE_CREATED_AT = "2026-07-15 10:54:00 COT"
CODE_VERSION_LABEL = "v53-final-map-excel"
PARSER_VERSION = "records-list-stable-v45-20260625-1715COT-pdf-manufacturing-age-section"


def get_uploaded_file_signature(uploaded_file) -> str:
    if uploaded_file is None:
        return ""
    content = uploaded_file.getvalue()
    raw = f"{PARSER_VERSION}|{uploaded_file.name}|{len(content)}|".encode("utf-8") + content
    return hashlib.md5(raw).hexdigest()


def read_table_any(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    raw = uploaded_file.getvalue()

    if name.endswith(".csv"):
        attempts = [
            {"sep": ";", "encoding": "utf-8-sig"},
            {"sep": ";", "encoding": "latin1"},
            {"sep": ",", "encoding": "utf-8-sig"},
            {"sep": ",", "encoding": "latin1"},
            {"sep": None, "encoding": "utf-8-sig"},
            {"sep": None, "encoding": "latin1"},
        ]
        for att in attempts:
            try:
                text = raw.decode(att["encoding"], errors="replace")
                df = pd.read_csv(StringIO(text), sep=att["sep"], engine="python", on_bad_lines="skip")
                if df.shape[1] >= 3:
                    return df
            except Exception:
                continue
        raise ValueError(f"No fue posible leer el CSV: {uploaded_file.name}")

    if name.endswith(".xlsx") or name.endswith(".xls"):
        book = pd.ExcelFile(BytesIO(raw))
        preferred = None
        best_score = -1
        for sheet in book.sheet_names:
            s = str(sheet).lower()
            score = 0
            if "datos" in s:
                score += 25
            if "combined" in s:
                score += 20
            if "consolidated" in s:
                score += 20
            if "records" in s:
                score += 18
            if "data" in s:
                score += 15
            if score > best_score:
                best_score = score
                preferred = sheet
        if preferred is None:
            preferred = book.sheet_names[0]
        return pd.read_excel(book, sheet_name=preferred)

    raise ValueError(f"Formato no soportado: {uploaded_file.name}")


def _align_records_row(row: list[str], expected_len: int) -> list[str]:
    """Alinea una fila del Records List contra CUSTOM_HEADERS sin mover columnas críticas.
    El export puede traer una columna final vacía no nombrada y algunas notas con ';'.
    Si sobra información, se compacta en Notes para mantener In Blood Bank y Status por posición.
    """
    if len(row) == expected_len:
        return row
    if len(row) < expected_len:
        return row + [""] * (expected_len - len(row))

    # En este export los ';' adicionales aparecen en Notes. Mantener el final alineado.
    try:
        notes_idx = CUSTOM_HEADERS.index("Notes")
    except ValueError:
        notes_idx = 23

    extra = len(row) - expected_len
    merged_note = ";".join(row[notes_idx : notes_idx + extra + 1])
    return row[:notes_idx] + [merged_note] + row[notes_idx + extra + 1 :]


def _operational_status_like_score(series: pd.Series) -> int:
    """Cuenta valores que parecen realmente estados operativos del Records List."""
    if series is None:
        return 0
    values = series.dropna().astype(str).str.strip().str.lower()
    if values.empty:
        return 0
    pattern = r"routine|scrap|warehouse|customer\s+to\s+be|transit|customs|ready\s+to\s+be\s+installed|new\s+system|refurb"
    return int(values.str.contains(pattern, regex=True, na=False).sum())


def _best_operational_status_series(work: pd.DataFrame) -> pd.Series | None:
    """Selecciona la columna que realmente contiene el estado operativo.

    En algunos exports de Records List el encabezado visible viene desplazado:
    la columna llamada "Number of tests per day" contiene valores como
    IN ROUTINE, NOT IN ROUTINE, WAREHOUSE to be scrapped, etc., mientras
    la columna "Status" contiene el tipo de contrato. Esta función evita
    que el filtro de estado operativo use la columna contractual equivocada.
    """
    candidate_names = [
        "Operational status",
        "Status",
        " Status",
        "Number of tests per day",
        " Number of tests per day",
        "Estado operativo",
    ]
    candidates = []
    for name in candidate_names:
        if name in work.columns:
            data = work[name]
            if isinstance(data, pd.DataFrame):
                data = data.bfill(axis=1).iloc[:, 0]
            score = _operational_status_like_score(data)
            non_empty = int(data.replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA}).notna().sum())
            candidates.append((score, non_empty, name, data))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    best_score, _, _, best_data = candidates[0]
    if best_score <= 0:
        return None
    return best_data



def recover_records_core_fields(df: pd.DataFrame) -> pd.DataFrame:
    """Recover critical Records List fields using aliases and fixed export positions.

    This is intentionally redundant with canonicalize_records_columns because Streamlit
    cache/session reuse can preserve an older parsed DataFrame. The function repairs the
    active DataFrame after parsing so the dashboard can still build the correct status matrix.
    """
    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]

    def first_existing_by_alias(alias_names: list[str]) -> pd.Series | None:
        normalized_targets = {normalize_column_label(x) for x in alias_names}
        matched = [c for c in work.columns if normalize_column_label(c) in normalized_targets]
        if not matched:
            return None
        return _first_valid_series_from_columns(work, matched)

    def has_useful_values(column: str) -> bool:
        if column not in work.columns:
            return False
        vals = work[column].replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        return bool(vals.notna().any())

    # Export typo: In Blook Bank -> In Blood Bank.
    bb_series = first_existing_by_alias([
        "In Blood Bank", "In Blook Bank", "In Bloock Bank", "In Blod Bank",
        "Blood Bank", "Bloodbank", "Banco de sangre", "Banco sangre",
    ])
    if bb_series is not None and (not has_useful_values("In Blood Bank")):
        work["In Blood Bank"] = bb_series

    # Real Records List export: choose the column that contains true operational
    # states. Some exports have the values under "Number of tests per day" and
    # the column named "Status" contains contract information.
    status_series = _best_operational_status_series(work)
    if status_series is None:
        status_series = first_existing_by_alias([
            "Operational status", "Status", "Operation status", "Estado operativo",
            "Estado de operacion", "Estado de operación",
        ])
    if status_series is not None:
        work["Operational status"] = status_series

    # Real Records List export: Instrument Status = asset condition.
    asset_series = first_existing_by_alias([
        "Asset condition", "Instrument Status", "Instrument asset status",
        "Asset status", "Condicion del activo", "Condición del activo",
    ])
    if asset_series is not None and (not has_useful_values("Asset condition")):
        work["Asset condition"] = asset_series

    # Real Records List export: Volume is the column used by the dashboard as PM plan.
    pm_series = first_existing_by_alias(["PM plan", "Volume", "Plan PM", "Plan de PM"])
    if pm_series is not None and (not has_useful_values("PM plan")):
        work["PM plan"] = pm_series

    # Positional safety net for the standard semicolon Records List export.
    # Header positions: 4=In Blook Bank, 16=Instrument Status, 17=Volume, 19=Status.
    if (not has_useful_values("Operational status")) and work.shape[1] > 19:
        candidate = work.iloc[:, 19].replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        if candidate.notna().any():
            work["Operational status"] = candidate
    if (not has_useful_values("In Blood Bank")) and work.shape[1] > 4:
        candidate = work.iloc[:, 4].replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        if candidate.notna().any():
            work["In Blood Bank"] = candidate
    if (not has_useful_values("Asset condition")) and work.shape[1] > 16:
        candidate = work.iloc[:, 16].replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        if candidate.notna().any():
            work["Asset condition"] = candidate
    if (not has_useful_values("PM plan")) and work.shape[1] > 17:
        candidate = work.iloc[:, 17].replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        if candidate.notna().any():
            work["PM plan"] = candidate

    return work

def _finalize_records_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Critical standardization for Records List exports:
    # The real CSV uses headers like "Instrument Status", "Volume" and "Status",
    # while the dashboard model expects "Asset condition", "PM plan" and
    # "Operational status". Without this step, all operational status values
    # become blank/No informado and the model-status matrix is wrong.
    df = canonicalize_records_columns(df)
    df = recover_records_core_fields(df)

    if "_blank" in df.columns:
        df = df.drop(columns=["_blank"])

    df = standardize_blood_bank_column(df)

    for missing_col in [c for c in CUSTOM_HEADERS if c != "_blank" and c not in df.columns]:
        df[missing_col] = pd.NA

    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "<NA>": pd.NA})

    def unexcel(value):
        if pd.isna(value):
            return pd.NA
        value = str(value).strip()
        if value.startswith('="') and value.endswith('"'):
            return value[2:-1]
        return value

    for col in ["Latitude", "Longitude", "Serial number"]:
        if col in df.columns:
            df[col] = df[col].map(unexcel)

    for col in ["Latitude", "Longitude", "Number of tests per day", "PM frequency", "Contract duration"]:
        if col in df.columns:
            df[col] = to_numeric_series(df[col])

    for col in ["Installation date", "PM last date", "PM next date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")

    df["Instrument family"] = df["Instrument type"].map(normalize_instrument_type)
    df["Operational status grouped"] = df["Operational status"].map(normalize_operational_status)

    today = pd.Timestamp(date.today())
    df["Age (years)"] = ((today - df["Installation date"]).dt.days / 365.25).round(1)
    df["Is in routine"] = df["Operational status grouped"].eq("Routine")
    df["Has geolocation"] = df["Latitude"].notna() & df["Longitude"].notna()

    yes_map = {"yes", "y", "true", "1", "1.0", "si", "sí", "s", "x"}
    assay_flags = {}
    for col in ASSAY_COLS:
        if col not in df.columns:
            df[col] = pd.NA
        normalized = df[col].fillna("No").astype(str).str.strip()
        df[col] = normalized
        assay_flags[f"FLAG::{col}"] = normalized.str.lower().isin(yes_map)
    if assay_flags:
        assay_flags_df = pd.DataFrame(assay_flags, index=df.index)
        df = pd.concat([df.reset_index(drop=True), assay_flags_df.reset_index(drop=True)], axis=1)
        df["Enabled assay count"] = assay_flags_df.sum(axis=1).to_numpy()
    else:
        df["Enabled assay count"] = 0

    base_cols = [c for c in CUSTOM_HEADERS if c != "_blank" and c in df.columns]
    df["Data completeness %"] = (df[base_cols].notna().sum(axis=1) / max(len(base_cols), 1) * 100).round(1)
    return df


def load_records(file_bytes: bytes, parser_version: str = PARSER_VERSION) -> pd.DataFrame:
    # Parser tolerante para Records List.
    # Nota: este texto se deja como comentario para que Streamlit no lo renderice
    # accidentalmente como bloque visible en el dashboard.
    # parser_version actúa como breaker de firma/sesión entre builds.
    _ = parser_version
    raw_text = file_bytes.decode("utf-8-sig", errors="replace")
    lines = [line for line in raw_text.splitlines() if line.strip()]
    expected_len = len(CUSTOM_HEADERS)

    if not lines:
        return _finalize_records_df(pd.DataFrame(columns=[c for c in CUSTOM_HEADERS if c != "_blank"]))

    sample = "\n".join(lines[:20])
    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except Exception:
        delimiter_counts = {sep: lines[0].count(sep) for sep in [";", ",", "\t", "|"]}
        delimiter = max(delimiter_counts, key=delimiter_counts.get)

    reader = csv.reader(lines, delimiter=delimiter)
    parsed_rows = list(reader)
    if not parsed_rows:
        return _finalize_records_df(pd.DataFrame(columns=[c for c in CUSTOM_HEADERS if c != "_blank"]))

    header = [str(c).strip() for c in parsed_rows[0]]
    normalized_header = [normalize_column_label(c) for c in header]
    normalized_custom = {normalize_column_label(c): c for c in CUSTOM_HEADERS if c != "_blank"}
    header_matches = sum(1 for c in normalized_header if c in normalized_custom)

    if header_matches >= 10:
        body = parsed_rows[1:]
        width = len(header)
        aligned_body = [row + [""] * max(width - len(row), 0) for row in body]
        aligned_body = [row[:width] for row in aligned_body]
        df = pd.DataFrame(aligned_body, columns=header)
        rename_map = {}
        for col in df.columns:
            normalized = normalize_column_label(col)
            if normalized in normalized_custom:
                rename_map[col] = normalized_custom[normalized]
        df = df.rename(columns=rename_map)
        df = adapt_uploaded_records_to_standard(df)
        for missing in [c for c in CUSTOM_HEADERS if c != "_blank" and c not in df.columns]:
            df[missing] = pd.NA
        return _finalize_records_df(df)

    rows = []
    for row in parsed_rows[1:]:
        if not any(str(cell).strip() for cell in row):
            continue
        rows.append(_align_records_row(row, expected_len))

    df = pd.DataFrame(rows, columns=CUSTOM_HEADERS)
    return _finalize_records_df(df)


def adapt_uploaded_records_to_standard(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "_blank" in out.columns:
        out = out.drop(columns=["_blank"])

    exact_matches = sum(1 for c in CUSTOM_HEADERS if c in out.columns)
    if exact_matches >= 20:
        return out

    if out.shape[1] == len(CUSTOM_HEADERS):
        out.columns = CUSTOM_HEADERS
        if "_blank" in out.columns:
            out = out.drop(columns=["_blank"])
        return out

    if out.shape[1] == len(CUSTOM_HEADERS) - 1:
        out.columns = [c for c in CUSTOM_HEADERS if c != "_blank"]
        return out

    return out


def parse_uploaded_records(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    raw = uploaded_file.getvalue()

    if name.endswith(".csv"):
        return load_records(raw, PARSER_VERSION)

    table = read_table_any(uploaded_file)
    table = adapt_uploaded_records_to_standard(table)

    df = table.copy()
    for missing in [c for c in CUSTOM_HEADERS if c != "_blank" and c not in df.columns]:
        df[missing] = pd.NA

    return _finalize_records_df(df)

def get_active_records_dataset(uploaded_file, sample_candidates: list[Path]) -> tuple[pd.DataFrame, str]:
    if uploaded_file is not None:
        current_sig = get_uploaded_file_signature(uploaded_file)
        saved_sig = st.session_state.get("records_active_signature", "")

        if "records_active_df" not in st.session_state or current_sig != saved_sig:
            active_df = parse_uploaded_records(uploaded_file)
            st.session_state["records_active_df"] = active_df.copy()
            st.session_state["records_active_signature"] = current_sig
            st.session_state["records_active_name"] = uploaded_file.name

        return st.session_state["records_active_df"].copy(), st.session_state["records_active_name"]

    if "records_active_df" in st.session_state and st.session_state.get("records_active_name"):
        return st.session_state["records_active_df"].copy(), st.session_state["records_active_name"]

    if sample_candidates:
        sample_path = sample_candidates[0]
        active_df = load_records(sample_path.read_bytes())
        st.session_state["records_active_df"] = active_df.copy()
        st.session_state["records_active_signature"] = f"sample::{sample_path.name}"
        st.session_state["records_active_name"] = sample_path.name
        return active_df, sample_path.name

    return pd.DataFrame(), ""


@st.cache_data(show_spinner=False)
def parse_machine_configuration(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    parsed_rows = []
    key_set = set()
    for raw in df["Machine Configurations"].fillna(""):
        row_dict = {}
        text = str(raw).strip()
        if text:
            for part in [p.strip() for p in re.split(r"\s*\|\s*|\r?\n|;(?=\s*[A-Za-z][A-Za-z0-9 /_-]{1,60}\s*:)", text) if p.strip()]:
                if ":" in part:
                    key, value = part.split(":", 1)
                    key = key.strip()
                    value = value.strip()
                    if value:
                        row_dict[key] = value
                        key_set.add(key)
        parsed_rows.append(row_dict)
    config_cols = sorted(key_set)
    if config_cols:
        cfg_df = pd.DataFrame([{key: row.get(key, pd.NA) for key in config_cols} for row in parsed_rows])
    else:
        cfg_df = pd.DataFrame(index=df.index)
    cfg_df = cfg_df.add_prefix("CFG::")
    return pd.concat([df.reset_index(drop=True), cfg_df.reset_index(drop=True)], axis=1), config_cols


@st.cache_data(show_spinner=False)
def add_operating_system_columns(df: pd.DataFrame, config_cols: list[str]) -> pd.DataFrame:
    os_candidates = ["CFG::Operative System", "CFG::Operating System", "CFG::ETI-Max 3000 Operative System", "CFG::LQS PC OS", "CFG::PC OS", "CFG::OS"]
    existing = [col for col in os_candidates if col in df.columns]

    def normalize_os(value):
        if pd.isna(value):
            return pd.NA
        text = str(value).strip()
        if not text:
            return pd.NA
        low = text.lower()
        if "don't know" in low or "dont know" in low or low == "unknown":
            return "Unknown"
        if "not installed" in low:
            return "Not installed"
        if "win10" in low or "windows 10" in low:
            return "Windows 10"
        if "vista" in low:
            return "Windows Vista"
        if "windows 7" in low or low == "win7":
            return "Windows 7"
        if "windows xp" in low or low == "xp":
            return "Windows XP"
        if "windows 2000" in low:
            return "Windows 2000"
        return text

    if existing:
        os_raw = _first_valid_series_from_columns(df, existing)
        df["Operating System Raw"] = os_raw
        df["Operating System"] = os_raw.map(normalize_os)
    else:
        df["Operating System Raw"] = pd.NA
        df["Operating System"] = pd.NA

    cfg_prefix_cols = [f"CFG::{c}" for c in config_cols if f"CFG::{c}" in df.columns]
    df["Machine config fields populated"] = df[cfg_prefix_cols].notna().sum(axis=1) if cfg_prefix_cols else 0
    return df


@st.cache_data(show_spinner=False)
def to_csv_download(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def load_table_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Carga archivos tabulares usados en Stock/Carstock de forma tolerante.

    Algunos reportes de stock enviados por distribuidores vienen como CSV
    separados por punto y coma, coma, tabulador o con líneas irregulares.
    El lector anterior usaba pd.read_csv con parámetros por defecto y podía
    tumbar la app con ParserError. Esta versión prueba combinaciones seguras
    y omite líneas corruptas antes de fallar.
    """
    name = filename.lower()
    if name.endswith(".csv"):
        attempts = []
        for encoding in ["utf-8-sig", "latin1"]:
            for sep in [None, ";", ",", "\t", "|"]:
                attempts.append({"encoding": encoding, "sep": sep, "quoting": csv.QUOTE_MINIMAL})
            for sep in [";", ",", "\t", "|"]:
                attempts.append({"encoding": encoding, "sep": sep, "quoting": csv.QUOTE_NONE})

        best_df = None
        best_score = -1
        last_error = None

        for attempt in attempts:
            try:
                text = file_bytes.decode(attempt["encoding"], errors="replace")
                read_kwargs = dict(
                    sep=attempt["sep"],
                    engine="python",
                    on_bad_lines="skip",
                    dtype=str,
                )
                if attempt["quoting"] == csv.QUOTE_NONE:
                    read_kwargs["quoting"] = csv.QUOTE_NONE
                    read_kwargs["escapechar"] = "\\"
                df = pd.read_csv(StringIO(text), **read_kwargs)
                df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
                if df.empty or df.shape[1] < 1:
                    continue
                score = df.shape[1] * 100000 + df.shape[0]
                if score > best_score:
                    best_df = df
                    best_score = score
            except Exception as exc:
                last_error = exc
                continue

        if best_df is not None:
            best_df.columns = [str(c).strip() for c in best_df.columns]
            return best_df

        raise ValueError(f"No fue posible leer el CSV de stock: {filename}. Último error: {last_error}")

    return pd.read_excel(BytesIO(file_bytes))


def detect_stock_columns(df: pd.DataFrame) -> tuple[str | None, str | None, str | None]:
    normalized = {col: re.sub(r"[^a-z0-9]+", " ", str(col).lower()).strip() for col in df.columns}
    part_col = qty_col = desc_col = None
    for col, norm in normalized.items():
        if part_col is None and ("part number" in norm or norm == "part no" or "product code" in norm or "material" in norm):
            part_col = col
        if qty_col is None and ("quantity" in norm or norm == "qty" or "stock" in norm or "cantidad" in norm):
            qty_col = col
        if desc_col is None and ("description" in norm or "descripcion" in norm or "product description" in norm):
            desc_col = col
    if part_col is None and len(df.columns) >= 1:
        part_col = df.columns[0]
    if qty_col is None and len(df.columns) >= 2:
        qty_col = df.columns[1]
    if desc_col is None and len(df.columns) >= 3:
        desc_col = df.columns[2]
    return part_col, qty_col, desc_col


@st.cache_data(show_spinner=False)
def load_spare_master_legacy(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    book = pd.ExcelFile(BytesIO(file_bytes))
    mapping = {"LXL Carstock": "LXL", "LXS Carstock": "LXS", "MDX Carstock": "MDX", "EMX Carstock": "EMX"}
    output = {}
    for sheet, family in mapping.items():
        if sheet not in book.sheet_names:
            continue
        df = pd.read_excel(book, sheet_name=sheet)

        part_col = next((c for c in df.columns if "PART NUMBER" in str(c).upper()), None)
        desc_col = next((c for c in df.columns if "DESCRIPTION" in str(c).upper()), None)
        qty_col = next((c for c in df.columns if "QUANTITY" in str(c).upper()), None)
        if not part_col or not qty_col:
            continue

        slim = pd.DataFrame(
            {
                "Required Distributor": "",
                "Required Family": family,
                "Required Part Number": df[part_col],
                "Required Description": df[desc_col] if desc_col else pd.NA,
                "Required Qty": to_numeric_series(df[qty_col]),
            }
        )
        slim["Part Key"] = slim["Required Part Number"].map(normalize_part_number)
        slim = slim[slim["Part Key"] != ""].copy()
        slim["Required Description"] = slim["Required Description"].fillna("").astype(str).str.strip()
        slim["Required Qty"] = pd.to_numeric(slim["Required Qty"], errors="coerce").fillna(0.0)
        slim = slim.groupby(["Part Key", "Required Family"], as_index=False).agg(
            {
                "Required Distributor": "first",
                "Required Part Number": "first",
                "Required Description": "first",
                "Required Qty": "sum",
            }
        )
        output[family] = slim.sort_values(["Required Qty", "Required Part Number"], ascending=[False, True]).reset_index(drop=True)
    return output


def detect_carstock_master_columns(df: pd.DataFrame) -> tuple[str | None, str | None, str | None, str | None, str | None]:
    normalized = {col: re.sub(r"[^a-z0-9]+", " ", str(col).lower()).strip() for col in df.columns}
    distributor_col = family_col = part_col = qty_col = desc_col = None

    for col, norm in normalized.items():
        if distributor_col is None and (
            "distributor" in norm
            or norm in {"dealer", "dealer name", "dist", "distributor name"}
        ):
            distributor_col = col

        if family_col is None and (
            "family" in norm
            or "platform" in norm
            or "carstock family" in norm
            or "instrument family" in norm
            or "family code" in norm
            or norm in {"instrument type", "instrument", "system"}
        ):
            family_col = col

        if part_col is None and (
            "part number" in norm
            or "latest part number" in norm
            or "part no" in norm
            or "pn" == norm
            or "material" in norm
            or "code" == norm
            or "product code" in norm
        ):
            part_col = col

        if qty_col is None and (
            "required qty" in norm
            or "required quantity" in norm
            or "carstock qty" in norm
            or "car stock qty" in norm
            or "quantity" in norm
            or norm in {"qty", "cantidad"}
        ):
            qty_col = col

        if desc_col is None and (
            "description" in norm
            or "descripcion" in norm
            or "product description" in norm
            or "spare part description" in norm
        ):
            desc_col = col

    return distributor_col, family_col, part_col, qty_col, desc_col


def normalize_family_code(value) -> str:
    text = normalize_key_text(value)
    if not text:
        return ""
    if "mdx" in text:
        return "MDX"
    if "emx" in text or "etimax" in text:
        return "EMX"
    if "xs" in text or "liaisonxs" in text:
        return "LXS"
    if "xl" in text or "las" in text or "liaisonxl" in text:
        return "LXL"
    return ""


def normalize_master_instrument_family(value) -> str:
    text = normalize_search_text(value)
    if not text:
        return ""
    if "eti max" in text or "etimax" in text:
        return "EMX"
    if "liaison xs" in text or text == "xs" or " xs " in f" {text} ":
        return "LXS"
    if "liaison xl" in text or "xl las" in text or "las" in text:
        return "LXL"
    if "mdx" in text or "molecular" in text or "murex" in text:
        return "MDX"
    return normalize_family_code(value)


def safe_ceil_qty(value) -> int:
    try:
        if pd.isna(value):
            return 0
        return int(math.ceil(max(float(value), 0.0)))
    except Exception:
        return 0


def detect_advanced_carstock_columns(df: pd.DataFrame) -> dict[str, str | None]:
    normalized = {col: re.sub(r"[^a-z0-9]+", " ", str(col).lower()).strip() for col in df.columns}
    mapping = {
        "part_col": None,
        "desc_col": None,
        "instrument_col": None,
        "carstock_qty_col": None,
        "pps12_col": None,
        "min_stock_col": None,
        "price2_col": None,
        "currency_col": None,
    }
    for col, norm in normalized.items():
        if mapping["part_col"] is None and ("part number" in norm or norm == "pn" or "latest part number" in norm):
            mapping["part_col"] = col
        if mapping["desc_col"] is None and ("description" in norm or "part number description" in norm):
            mapping["desc_col"] = col
        if mapping["instrument_col"] is None and ("instrument type" in norm or norm == "instrument"):
            mapping["instrument_col"] = col
        if mapping["carstock_qty_col"] is None and ("carstock qty" in norm or "car stock qty" in norm):
            mapping["carstock_qty_col"] = col
        if mapping["pps12_col"] is None and ("parts per system 12 months" in norm or "parts per system" in norm):
            mapping["pps12_col"] = col
        if mapping["min_stock_col"] is None and ("minimum stock level required" in norm or "minimum stock level" in norm):
            mapping["min_stock_col"] = col
        if mapping["price2_col"] is None and ("sp price option 2" in norm or "option 2" in norm):
            mapping["price2_col"] = col
        if mapping["currency_col"] is None and "currency" in norm:
            mapping["currency_col"] = col
    return mapping


def build_advanced_master(df: pd.DataFrame) -> pd.DataFrame:
    cols = detect_advanced_carstock_columns(df)
    part_col = cols.get("part_col")
    if not part_col:
        return pd.DataFrame()
    work = pd.DataFrame({
        "Required Part Number": df[part_col],
        "Required Description": df[cols["desc_col"]] if cols.get("desc_col") in df.columns else "",
        "Instrument Type Raw": df[cols["instrument_col"]] if cols.get("instrument_col") in df.columns else "",
        "Carstock Base Qty": to_numeric_series(df[cols["carstock_qty_col"]]) if cols.get("carstock_qty_col") in df.columns else 0,
        "Parts per system (12 months)": to_numeric_series(df[cols["pps12_col"]]) if cols.get("pps12_col") in df.columns else 0,
        "Minimum Stock Level Required": to_numeric_series(df[cols["min_stock_col"]]) if cols.get("min_stock_col") in df.columns else 0,
        "Option 2 Unit Price": to_numeric_series(df[cols["price2_col"]]) if cols.get("price2_col") in df.columns else np.nan,
        "Currency": df[cols["currency_col"]] if cols.get("currency_col") in df.columns else "EUR",
    })
    work["Required Description"] = work["Required Description"].fillna("").astype(str).str.strip()
    work["Instrument Type Raw"] = work["Instrument Type Raw"].fillna("").astype(str).str.strip()
    work["Required Family"] = work["Instrument Type Raw"].map(normalize_master_instrument_family)
    work["Part Key"] = work["Required Part Number"].map(normalize_part_number)
    for c in ["Carstock Base Qty", "Parts per system (12 months)", "Minimum Stock Level Required", "Option 2 Unit Price"]:
        work[c] = pd.to_numeric(work[c], errors="coerce").fillna(0.0)
    work["Currency"] = work["Currency"].fillna("EUR").astype(str).str.strip().replace("", "EUR")
    work = work[work["Part Key"] != ""].copy()
    work = work[
        (work["Carstock Base Qty"] > 0)
        | (work["Parts per system (12 months)"] > 0)
        | (work["Minimum Stock Level Required"] > 0)
    ].copy()
    if work.empty:
        return work
    work = work.groupby(["Required Family", "Part Key"], as_index=False).agg({
        "Required Part Number": "first",
        "Required Description": "first",
        "Instrument Type Raw": "first",
        "Carstock Base Qty": "max",
        "Parts per system (12 months)": "max",
        "Minimum Stock Level Required": "max",
        "Option 2 Unit Price": "max",
        "Currency": "first",
    })
    work["Required Qty"] = work["Minimum Stock Level Required"].map(safe_ceil_qty)
    return work


def compute_installed_base_by_family(df: pd.DataFrame, distributor_name: str) -> dict[str, int]:
    if df is None or df.empty or not distributor_name:
        return {}
    scoped = df[df["Distributor name"].fillna("").astype(str).eq(distributor_name)].copy()
    if scoped.empty:
        return {}
    scoped["Detected Family"] = scoped["Instrument type"].map(normalize_master_instrument_family)
    counts = scoped["Detected Family"].value_counts().to_dict()
    return {str(k): int(v) for k, v in counts.items() if str(k).strip()}


def apply_dynamic_required_qty(master_df: pd.DataFrame, installed_base_by_family: dict[str, int], po_frequency_weeks: float) -> pd.DataFrame:
    if master_df is None or master_df.empty:
        return master_df
    work = master_df.copy()
    if "Required Family" not in work.columns:
        return work
    weeks = 0.0 if pd.isna(po_frequency_weeks) else max(float(po_frequency_weeks), 0.0)
    work["Installed Base Family"] = work["Required Family"].map(lambda fam: int(installed_base_by_family.get(str(fam), 0)))
    if "Carstock Base Qty" not in work.columns:
        work["Carstock Base Qty"] = pd.to_numeric(work.get("Required Qty", 0), errors="coerce").fillna(0.0)
    if "Parts per system (12 months)" not in work.columns:
        work["Parts per system (12 months)"] = 0.0
    if "Minimum Stock Level Required" not in work.columns:
        work["Minimum Stock Level Required"] = pd.to_numeric(work.get("Required Qty", 0), errors="coerce").fillna(0.0)

    work["Dynamic Demand Window"] = (
        pd.to_numeric(work["Parts per system (12 months)"], errors="coerce").fillna(0.0)
        * pd.to_numeric(work["Installed Base Family"], errors="coerce").fillna(0.0)
        * weeks / 52.0
    )
    work["Required Qty Raw"] = work[["Carstock Base Qty", "Minimum Stock Level Required", "Dynamic Demand Window"]].max(axis=1)
    work["Required Qty"] = work["Required Qty Raw"].map(safe_ceil_qty)
    return work
    if "mdx" in text:
        return "MDX"
    if "emx" in text:
        return "EMX"
    if "xs" in text or "liaisonxs" in text:
        return "LXS"
    if "xl" in text or "las" in text or "liaisonxl" in text:
        return "LXL"
    return ""


def infer_families_from_instruments(instruments: list[str]) -> list[str]:
    families = []
    for inst in instruments:
        fam = normalize_family_code(inst)
        if fam:
            families.append(fam)
    return sorted(set(families))


def build_master_slim(
    df: pd.DataFrame,
    distributor_col: str | None,
    family_col: str | None,
    part_col: str,
    qty_col: str,
    desc_col: str | None,
    fallback_distributor: str = "",
    fallback_family: str = "",
) -> pd.DataFrame:
    slim = pd.DataFrame(
        {
            "Required Distributor": df[distributor_col] if distributor_col and distributor_col in df.columns else fallback_distributor,
            "Required Family": df[family_col] if family_col and family_col in df.columns else fallback_family,
            "Required Part Number": df[part_col],
            "Required Description": df[desc_col] if desc_col and desc_col in df.columns else "",
            "Required Qty": to_numeric_series(df[qty_col]),
        }
    )
    slim["Required Distributor"] = slim["Required Distributor"].fillna("").astype(str).str.strip()
    slim["Required Family"] = slim["Required Family"].map(normalize_family_code)
    if fallback_family:
        slim["Required Family"] = slim["Required Family"].replace("", fallback_family)
    slim["Required Description"] = slim["Required Description"].fillna("").astype(str).str.strip()
    slim["Required Qty"] = pd.to_numeric(slim["Required Qty"], errors="coerce").fillna(0.0)
    slim["Part Key"] = slim["Required Part Number"].map(normalize_part_number)
    slim["Distributor Key"] = slim["Required Distributor"].map(normalize_key_text)
    slim = slim[(slim["Part Key"] != "") & (slim["Required Qty"] > 0)].copy()
    slim = slim.groupby(["Distributor Key", "Required Distributor", "Required Family", "Part Key"], as_index=False).agg(
        {
            "Required Part Number": "first",
            "Required Description": "first",
            "Required Qty": "sum",
        }
    )
    return slim


def detect_price_reference_columns(df: pd.DataFrame) -> tuple[str | None, str | None, str | None, str | None]:
    normalized = {col: re.sub(r"[^a-z0-9]+", " ", str(col).lower()).strip() for col in df.columns}
    part_col = desc_col = option2_col = currency_col = None
    for col, norm in normalized.items():
        if part_col is None and (
            "part number" in norm or "latest part number" in norm or norm in {"part no", "pn"} or "material" in norm
        ):
            part_col = col
        if desc_col is None and ("description" in norm or "descripcion" in norm or "product description" in norm):
            desc_col = col
        if option2_col is None and ("option 2" in norm or "opt2" in norm):
            option2_col = col
        if currency_col is None and "currency" in norm:
            currency_col = col
    return part_col, desc_col, option2_col, currency_col


def build_price_reference(df: pd.DataFrame, part_col: str, option2_col: str, desc_col: str | None, currency_col: str | None) -> pd.DataFrame:
    price_df = pd.DataFrame(
        {
            "Price Part Number": df[part_col],
            "Price Description": df[desc_col] if desc_col and desc_col in df.columns else "",
            "Option 2 Unit Price": to_numeric_series(df[option2_col]),
            "Currency": df[currency_col] if currency_col and currency_col in df.columns else "EUR",
        }
    )
    price_df["Part Key"] = price_df["Price Part Number"].map(normalize_part_number)
    price_df["Price Description"] = price_df["Price Description"].fillna("").astype(str).str.strip()
    price_df["Currency"] = price_df["Currency"].fillna("EUR").astype(str).str.strip().replace("", "EUR")
    price_df["Option 2 Unit Price"] = pd.to_numeric(price_df["Option 2 Unit Price"], errors="coerce")
    price_df = price_df[(price_df["Part Key"] != "") & price_df["Option 2 Unit Price"].notna() & (price_df["Option 2 Unit Price"] > 0)].copy()
    if price_df.empty:
        return pd.DataFrame(columns=["Part Key", "Option 2 Unit Price", "Currency", "Price Description"])
    price_df = price_df.groupby("Part Key", as_index=False).agg(
        {
            "Option 2 Unit Price": "first",
            "Currency": "first",
            "Price Description": "first",
        }
    )
    return price_df


@st.cache_data(show_spinner=False)
def load_carstock_master_bundle(file_bytes: bytes, filename: str) -> dict[str, object]:
    path = Path(filename)
    ext = path.suffix.lower()

    legacy_families = {}
    consolidated_frames = []
    price_frames = []
    advanced_master = pd.DataFrame()

    if ext in {".xlsx", ".xls"}:
        try:
            book = pd.ExcelFile(BytesIO(file_bytes))
        except Exception:
            book = None

        if book is not None:
            sheet_names = set(book.sheet_names)
            if "Carstock" in sheet_names:
                try:
                    advanced_candidate = pd.read_excel(book, sheet_name="Carstock")
                    advanced_master = build_advanced_master(advanced_candidate)
                    if not advanced_master.empty:
                        consolidated_frames.append(
                            advanced_master[[
                                "Required Family",
                                "Part Key",
                                "Required Part Number",
                                "Required Description",
                                "Required Qty",
                            ]].assign(**{
                                "Distributor Key": "",
                                "Required Distributor": "",
                            })[
                                [
                                    "Distributor Key",
                                    "Required Distributor",
                                    "Required Family",
                                    "Part Key",
                                    "Required Part Number",
                                    "Required Description",
                                    "Required Qty",
                                ]
                            ]
                        )
                        price_frames.append(
                            advanced_master[["Part Key", "Option 2 Unit Price", "Currency", "Required Description"]]
                            .rename(columns={"Required Description": "Price Description"})
                        )
                except Exception:
                    advanced_master = pd.DataFrame()

            if any(sheet in sheet_names for sheet in {"LXL Carstock", "LXS Carstock", "MDX Carstock", "EMX Carstock"}):
                legacy_families = load_spare_master_legacy(file_bytes)

            for sheet in book.sheet_names:
                try:
                    df = pd.read_excel(book, sheet_name=sheet)
                except Exception:
                    continue
                if df is None or df.empty:
                    continue

                price_part_col, price_desc_col, option2_col, currency_col = detect_price_reference_columns(df)
                if price_part_col and option2_col:
                    price_ref = build_price_reference(df, price_part_col, option2_col, price_desc_col, currency_col)
                    if not price_ref.empty:
                        price_frames.append(price_ref)

                distributor_col, family_col, part_col, qty_col, desc_col = detect_carstock_master_columns(df)
                if not part_col or not qty_col:
                    continue

                fallback_family = normalize_family_code(sheet)
                slim = build_master_slim(
                    df,
                    distributor_col=distributor_col,
                    family_col=family_col,
                    part_col=part_col,
                    qty_col=qty_col,
                    desc_col=desc_col,
                    fallback_family=fallback_family,
                )
                if not slim.empty:
                    consolidated_frames.append(slim)

    else:
        df = load_table_file(file_bytes, filename)
        if df is not None and not df.empty:
            price_part_col, price_desc_col, option2_col, currency_col = detect_price_reference_columns(df)
            if price_part_col and option2_col:
                price_ref = build_price_reference(df, price_part_col, option2_col, price_desc_col, currency_col)
                if not price_ref.empty:
                    price_frames.append(price_ref)
            distributor_col, family_col, part_col, qty_col, desc_col = detect_carstock_master_columns(df)
            if part_col and qty_col:
                consolidated_frames.append(
                    build_master_slim(
                        df,
                        distributor_col=distributor_col,
                        family_col=family_col,
                        part_col=part_col,
                        qty_col=qty_col,
                        desc_col=desc_col,
                    )
                )

    if consolidated_frames:
        consolidated = pd.concat(consolidated_frames, ignore_index=True)
        consolidated["Required Distributor"] = consolidated["Required Distributor"].fillna("").astype(str).str.strip()
        consolidated["Required Family"] = consolidated["Required Family"].fillna("").astype(str).str.strip()
        consolidated["Distributor Key"] = consolidated["Required Distributor"].map(normalize_key_text)
        consolidated = consolidated.groupby(
            ["Distributor Key", "Required Distributor", "Required Family", "Part Key"], as_index=False
        ).agg(
            {
                "Required Part Number": "first",
                "Required Description": "first",
                "Required Qty": "max",
            }
        )
    else:
        consolidated = pd.DataFrame(
            columns=[
                "Distributor Key",
                "Required Distributor",
                "Required Family",
                "Part Key",
                "Required Part Number",
                "Required Description",
                "Required Qty",
            ]
        )

    if price_frames:
        price_reference = pd.concat(price_frames, ignore_index=True)
        price_reference = price_reference.groupby("Part Key", as_index=False).agg(
            {
                "Option 2 Unit Price": "max",
                "Currency": "first",
                "Price Description": "first",
            }
        )
    else:
        price_reference = pd.DataFrame(columns=["Part Key", "Option 2 Unit Price", "Currency", "Price Description"])

    distributor_options = sorted([d for d in consolidated["Required Distributor"].dropna().astype(str).unique().tolist() if d.strip()])
    family_options = sorted([f for f in consolidated["Required Family"].dropna().astype(str).unique().tolist() if f.strip()])

    return {
        "legacy_families": legacy_families,
        "consolidated": consolidated,
        "advanced_master": advanced_master,
        "price_reference": price_reference,
        "master_distributors": distributor_options,
        "master_families": family_options,
    }


def build_required_master_from_scope(
    master_bundle: dict[str, object],
    assigned_distributor: str,
    selected_families: list[str],
) -> tuple[pd.DataFrame, str]:
    advanced_master = master_bundle.get("advanced_master", pd.DataFrame())
    consolidated = master_bundle.get("consolidated", pd.DataFrame())
    legacy_families = master_bundle.get("legacy_families", {})

    if advanced_master is not None and not advanced_master.empty:
        scoped = advanced_master.copy()
        if selected_families:
            scoped = scoped[scoped["Required Family"].isin(selected_families)]
        scoped = scoped.sort_values(["Required Family", "Required Part Number"]).reset_index(drop=True)
        return scoped, "advanced"

    if consolidated is not None and not consolidated.empty:
        scoped = consolidated.copy()
        if assigned_distributor and assigned_distributor != "<sin asignar>" and scoped["Distributor Key"].astype(str).str.len().gt(0).any():
            scoped = scoped[scoped["Distributor Key"].eq(normalize_key_text(assigned_distributor))]
        if selected_families:
            scoped = scoped[scoped["Required Family"].isin(selected_families)]

        scoped = scoped.groupby("Part Key", as_index=False).agg(
            {
                "Required Part Number": "first",
                "Required Description": "first",
                "Required Qty": "sum",
            }
        )
        return scoped.sort_values(["Required Qty", "Required Part Number"], ascending=[False, True]).reset_index(drop=True), "consolidated"

    selected_families = [f for f in selected_families if f in legacy_families]
    if not selected_families:
        return pd.DataFrame(columns=["Part Key", "Required Part Number", "Required Description", "Required Qty"]), "legacy"

    scoped = pd.concat([legacy_families[f] for f in selected_families], ignore_index=True)
    scoped = scoped.groupby("Part Key", as_index=False).agg(
        {
            "Required Part Number": "first",
            "Required Description": "first",
            "Required Qty": "sum",
        }
    )
    return scoped.sort_values(["Required Qty", "Required Part Number"], ascending=[False, True]).reset_index(drop=True), "legacy"


def prepare_uploaded_stock(stock_df: pd.DataFrame, part_col: str, qty_col: str, desc_col: str | None) -> pd.DataFrame:
    work = stock_df.copy()
    work["Uploaded Part Number"] = work[part_col]
    work["Uploaded Qty"] = to_numeric_series(work[qty_col]).fillna(0.0)
    if desc_col is not None and desc_col in work.columns:
        work["Uploaded Description"] = work[desc_col]
    else:
        work["Uploaded Description"] = ""
    work["Uploaded Description"] = work["Uploaded Description"].fillna("").astype(str).str.strip()
    work["Part Key"] = work["Uploaded Part Number"].map(normalize_part_number)
    work = work[work["Part Key"] != ""].copy()
    stock_slim = work.groupby("Part Key", as_index=False).agg(
        {"Uploaded Part Number": "first", "Uploaded Description": "first", "Uploaded Qty": "sum"}
    )
    stock_slim["Uploaded Qty"] = pd.to_numeric(stock_slim["Uploaded Qty"], errors="coerce").fillna(0.0)
    return stock_slim.sort_values(["Uploaded Qty", "Uploaded Part Number"], ascending=[False, True]).reset_index(drop=True)


def compare_stock(
    master_df: pd.DataFrame,
    stock_df: pd.DataFrame,
    part_col: str,
    qty_col: str,
    desc_col: str | None,
    price_reference: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    stock_slim = prepare_uploaded_stock(stock_df, part_col, qty_col, desc_col)

    merged = master_df.copy().merge(stock_slim, on="Part Key", how="left")
    merged["Required Description"] = merged.get("Required Description", "").fillna("").astype(str).str.strip().replace("", "No disponible")
    merged["Required Qty"] = pd.to_numeric(merged.get("Required Qty", 0), errors="coerce").fillna(0.0)
    merged["Uploaded Qty"] = pd.to_numeric(merged.get("Uploaded Qty", 0), errors="coerce").fillna(0.0)
    merged["Uploaded Part Number"] = merged.get("Uploaded Part Number", "").fillna("")
    merged["Uploaded Description"] = merged.get("Uploaded Description", "").fillna("")
    merged["Qty Gap"] = (merged["Required Qty"] - merged["Uploaded Qty"]).clip(lower=0.0)
    denominator = merged["Required Qty"].replace(0, np.nan).astype(float)
    merged["Coverage %"] = ((merged["Uploaded Qty"].astype(float) / denominator) * 100).round(1).fillna(0.0)

    low_mask = (merged["Uploaded Qty"] > 0) & (merged["Uploaded Qty"] < merged["Required Qty"])
    merged["Status"] = np.where(
        merged["Qty Gap"] <= 0,
        "OK",
        np.where(low_mask, "LOW", "Missing"),
    )

    if price_reference is not None and not price_reference.empty:
        merged = merged.merge(price_reference, on="Part Key", how="left", suffixes=("", "_ref"))
    else:
        merged["Option 2 Unit Price"] = np.nan
        merged["Currency"] = "EUR"
        merged["Price Description"] = ""

    if "Option 2 Unit Price_ref" in merged.columns:
        merged["Option 2 Unit Price"] = pd.to_numeric(merged["Option 2 Unit Price"], errors="coerce").fillna(
            pd.to_numeric(merged["Option 2 Unit Price_ref"], errors="coerce")
        )
    merged["Option 2 Unit Price"] = pd.to_numeric(merged.get("Option 2 Unit Price", 0), errors="coerce").fillna(0.0)

    if "Currency_ref" in merged.columns:
        merged["Currency"] = merged.get("Currency", "").fillna("").astype(str).replace("", pd.NA).fillna(merged["Currency_ref"])
    merged["Currency"] = merged.get("Currency", "EUR").fillna("EUR").astype(str).str.strip().replace("", "EUR")

    merged["Purchase Qty Option 2"] = merged["Qty Gap"]
    merged["Option 2 Estimated Cost"] = (merged["Purchase Qty Option 2"] * merged["Option 2 Unit Price"]).round(2)
    merged["Gap Label"] = merged["Required Part Number"].astype(str).str.strip() + " | " + merged["Required Description"].fillna("").astype(str).str.slice(0, 42)

    extra_df = stock_slim[~stock_slim["Part Key"].isin(master_df["Part Key"])].copy()
    if not extra_df.empty:
        extra_df["Status"] = "Extra / no requerido"

    merged = merged.sort_values(["Status", "Qty Gap", "Required Qty"], ascending=[True, False, False]).reset_index(drop=True)
    extra_df = extra_df.sort_values(["Uploaded Qty", "Uploaded Part Number"], ascending=[False, True]).reset_index(drop=True)
    return merged, extra_df, stock_slim


def active_config_fields(df: pd.DataFrame, config_keys: list[str]) -> list[str]:
    active = []
    for key in config_keys:
        col = f"CFG::{key}"
        if col in df.columns and df[col].notna().any():
            active.append(key)
    return active


def build_distributor_status_chart(df: pd.DataFrame, selected_model: str) -> go.Figure:
    fig = go.Figure()

    if df.empty or not selected_model:
        fig.update_layout(title="Estado por distribuidor")
        return glow_layout(fig, 620, 17)

    work = df.copy()
    work["Instrument type"] = work["Instrument type"].fillna("No informado").astype(str)
    work["Distributor name"] = work["Distributor name"].fillna("No informado").astype(str)
    work["Operational status"] = work["Operational status"].fillna("No informado").astype(str).str.strip()
    work["Status for chart"] = np.where(work["Operational status"].eq(""), "No informado", work["Operational status"])

    model_df = work[work["Instrument type"] == selected_model].copy()
    if model_df.empty:
        fig.update_layout(title=f"Estado por distribuidor | {selected_model}")
        return glow_layout(fig, 620, 17)

    summary = (
        model_df.groupby(["Distributor name", "Status for chart"], dropna=False)
        .size()
        .reset_index(name="Count")
    )

    distributor_order = (
        summary.groupby("Distributor name", as_index=False)["Count"]
        .sum()
        .sort_values("Count", ascending=False)["Distributor name"]
        .tolist()
    )
    status_order = (
        summary.groupby("Status for chart", as_index=False)["Count"]
        .sum()
        .sort_values("Count", ascending=False)["Status for chart"]
        .tolist()
    )

    color_sequence = px.colors.qualitative.Set2 + px.colors.qualitative.Bold + px.colors.qualitative.Safe
    color_map = {status: color_sequence[i % len(color_sequence)] for i, status in enumerate(status_order)}

    fig = px.bar(
        summary,
        y="Distributor name",
        x="Count",
        color="Status for chart",
        orientation="h",
        barmode="stack",
        title=f"Estado por distribuidor | {selected_model}",
        custom_data=["Status for chart", "Count"],
        color_discrete_map=color_map,
        category_orders={"Distributor name": distributor_order, "Status for chart": status_order},
    )

    fig.update_traces(
        hovertemplate=(
            "<b>Distribuidor:</b> %{y}<br>"
            "<b>Modelo:</b> " + selected_model + "<br>"
            "<b>Estado:</b> %{customdata[0]}<br>"
            "<b>Cantidad:</b> %{customdata[1]}<extra></extra>"
        )
    )

    fig.update_layout(
        xaxis_title="Cantidad de instrumentos",
        yaxis_title="Distribuidor",
        legend_title="Estado operativo",
    )
    return glow_layout(fig, 620, 17)


def build_distributor_global_overview(df: pd.DataFrame, top_n: int = 5) -> go.Figure:
    fig = go.Figure()

    if df.empty:
        fig.update_layout(title='Vista global por distribuidor')
        return glow_layout(fig, 520, 16)

    work = df.copy()
    work['Instrument type'] = work['Instrument type'].fillna('No informado').astype(str)
    work['Distributor name'] = work['Distributor name'].fillna('No informado').astype(str)

    model_order = work['Instrument type'].value_counts().index.tolist()
    summary = work.groupby(['Instrument type', 'Distributor name'], dropna=False).size().reset_index(name='Count')
    if summary.empty:
        fig.update_layout(title='Vista global por distribuidor')
        return glow_layout(fig, 520, 16)

    dist_order = summary.groupby('Distributor name', as_index=False)['Count'].sum().sort_values(['Count', 'Distributor name'], ascending=[False, True])
    top_distributors = dist_order['Distributor name'].tolist()[:top_n]
    summary = summary[summary['Distributor name'].isin(top_distributors)].copy()
    label_map = {name: distributor_display_name(name, 18) for name in top_distributors}
    legend_order = [label_map[name] for name in top_distributors]
    summary['Legend label'] = summary['Distributor name'].astype(str).map(label_map)
    summary['Legend label'] = pd.Categorical(summary['Legend label'], categories=legend_order, ordered=True)
    summary['Instrument type'] = pd.Categorical(summary['Instrument type'], categories=model_order, ordered=True)
    summary = summary.sort_values(['Instrument type', 'Legend label', 'Count'], ascending=[True, True, False])
    palette = build_long_palette(len(top_distributors))

    fig = px.bar(
        summary,
        y='Instrument type',
        x='Count',
        color='Legend label',
        orientation='h',
        barmode='stack',
        text='Count',
        title='Vista global por distribuidor | resumen ejecutivo (Top 5)',
        category_orders={'Instrument type': model_order, 'Legend label': legend_order},
        color_discrete_sequence=palette,
        custom_data=['Instrument type', 'Distributor name', 'Count'],
    )
    fig.update_traces(
        textposition='inside',
        insidetextanchor='middle',
        hovertemplate='<b>Modelo:</b> %{customdata[0]}<br><b>Distribuidor:</b> %{customdata[1]}<br><b>Cantidad:</b> %{customdata[2]}<extra></extra>'
    )
    fig.update_layout(
        legend_title='Distribuidor',
        xaxis_title='Cantidad de equipos',
        yaxis_title='Modelo',
        margin=dict(t=72, b=48, l=8, r=8),
        height=520,
    )
    fig.update_yaxes(categoryorder='array', categoryarray=model_order[::-1])
    return glow_layout(fig, 520, 16)


def build_distributor_model_donut(df: pd.DataFrame, selected_model: str, top_n: int = 5) -> go.Figure:
    fig = go.Figure()

    if df.empty or not selected_model:
        fig.update_layout(title="Distribución por distribuidor")
        return glow_layout(fig, 430, 15)

    work = df.copy()
    work["Instrument type"] = work["Instrument type"].fillna("No informado").astype(str)
    work["Distributor name"] = work["Distributor name"].fillna("No informado").astype(str)
    model_df = work[work["Instrument type"] == selected_model].copy()

    if model_df.empty:
        fig.update_layout(title=selected_model)
        return glow_layout(fig, 430, 15)

    summary = (
        model_df.groupby("Distributor name", dropna=False)
        .size()
        .reset_index(name="Count")
        .sort_values(["Count", "Distributor name"], ascending=[False, True])
        .reset_index(drop=True)
    )

    if summary.empty:
        fig.update_layout(title=selected_model)
        return glow_layout(fig, 430, 15)

    summary = summarize_distributor_counts(summary, top_n=top_n)
    summary["Legend label"] = summary["Distributor name"].astype(str).map(lambda x: distributor_display_name(x, 20))
    summary = summary.sort_values(["Count", "Legend label"], ascending=[False, True]).reset_index(drop=True)
    palette = build_long_palette(len(summary))

    fig.add_trace(
        go.Pie(
            labels=summary["Legend label"],
            values=summary["Count"],
            hole=0.68,
            sort=False,
            marker=dict(colors=palette[:len(summary)], line=dict(color="rgba(255,255,255,0.20)", width=1.2)),
            textinfo="percent",
            textfont=dict(color="#ffffff", size=12),
            customdata=np.column_stack([summary["Distributor name"], summary["Count"]]),
            hovertemplate="<b>Modelo:</b> "
            + selected_model
            + "<br><b>Distribuidor:</b> %{customdata[0]}<br><b>Cantidad:</b> %{customdata[1]}<br><b>Participación:</b> %{percent}<extra></extra>",
        )
    )
    total_assets = int(model_df.shape[0])
    fig.add_annotation(
        text=f"<b>{total_assets:,}</b><br><span style='font-size:11px'>equipos</span>",
        x=0.5,
        y=0.52,
        xref="paper",
        yref="paper",
        showarrow=False,
        font=dict(color="#ffffff", size=17),
    )
    fig.update_layout(
        title=dict(text=wrap_chart_title(f"{selected_model} | Top 5", 26), x=0.03, y=0.96, xanchor="left", yanchor="top", font=dict(size=14, color="#f9fdff")),
        showlegend=True,
        height=430,
        margin=dict(t=72, b=96, l=8, r=8),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.10,
            xanchor="center",
            x=0.5,
            bgcolor="rgba(14,26,42,0.18)",
            bordercolor="rgba(124,221,255,0.16)",
            borderwidth=1,
            font=dict(color="#f8fbff", size=10),
            itemwidth=90,
            itemsizing="constant",
        ),
    )
    return glow_layout(fig, 430, 15)




# =============================================================================
# MÓDULO ADICIONAL: ANTIGÜEDAD SEGÚN FECHA DE FABRICACIÓN
# Este bloque es independiente y no altera la lógica existente del dashboard.
# =============================================================================

def normalize_serial_match(value) -> str:
    """Normaliza seriales para cruzar Records List con archivos de fabricación."""
    if pd.isna(value):
        return ""
    text_value = str(value).strip()
    if text_value.startswith('="') and text_value.endswith('"'):
        text_value = text_value[2:-1]
    text_value = re.sub(r"\.0$", "", text_value)
    return re.sub(r"[^A-Za-z0-9]+", "", text_value).upper()


def normalize_column_label(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


RECORDS_COLUMN_ALIASES = {
    # Real Records List export headers and common variants mapped to dashboard canonical fields.
    "in blook bank": "In Blood Bank",
    "in bloock bank": "In Blood Bank",
    "in blod bank": "In Blood Bank",
    "blood bank": "In Blood Bank",
    "bloodbank": "In Blood Bank",
    "blook bank": "In Blood Bank",
    "banco de sangre": "In Blood Bank",
    "banco sangre": "In Blood Bank",

    # This is the root cause of the incorrect chart: the CSV header is "Status",
    # not "Operational status". It must be mapped before the dashboard groups states.
    "status": "Operational status",
    "operational status": "Operational status",
    "operation status": "Operational status",
    "estado operativo": "Operational status",
    "estado de operacion": "Operational status",
    "estado de operación": "Operational status",

    # The export uses "Instrument Status" for the asset/new-used condition.
    "instrument status": "Asset condition",
    "instrument asset status": "Asset condition",
    "asset status": "Asset condition",
    "asset condition": "Asset condition",
    "condicion del activo": "Asset condition",
    "condición del activo": "Asset condition",

    # The export uses "Volume" in the position historically used by PM plan.
    "volume": "PM plan",
    "pm plan": "PM plan",
    "plan pm": "PM plan",
    "plan de pm": "PM plan",
}


def canonicalize_records_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize real Records List headers into the canonical names used by the dashboard.

    This function also coalesces duplicate/alias columns by taking the first non-empty
    value per row, so existing canonical columns are preserved and alias columns fill gaps.
    """
    if df is None or df.empty:
        return df.copy()

    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]

    grouped_columns: dict[str, list[str]] = {}
    for col in list(work.columns):
        normalized = normalize_column_label(col)
        target = RECORDS_COLUMN_ALIASES.get(normalized, col)
        grouped_columns.setdefault(target, []).append(col)

    canonical = pd.DataFrame(index=work.index)
    for target, source_cols in grouped_columns.items():
        if len(source_cols) == 1 and source_cols[0] == target:
            data = work[source_cols[0]]
            if isinstance(data, pd.DataFrame):
                data = data.bfill(axis=1).iloc[:, 0]
            canonical[target] = data
        else:
            canonical[target] = _first_valid_series_from_columns(work, source_cols)

    return canonical


def guess_column_index(columns, keywords: list[str]) -> int:
    normalized = [normalize_column_label(col) for col in columns]
    for keyword in keywords:
        key = normalize_column_label(keyword)
        for idx, column_name in enumerate(normalized):
            if column_name == key:
                return idx
    for keyword in keywords:
        key = normalize_column_label(keyword)
        for idx, column_name in enumerate(normalized):
            if key and key in column_name:
                return idx
    return 0


def manufacturing_date_candidate_columns(df: pd.DataFrame) -> list[str]:
    candidates = []
    preferred_terms = [
        "manufacturing date",
        "manufacture date",
        "production date",
        "mfg date",
        "build date",
        "fecha de fabricacion",
        "fecha fabricación",
        "po date",
        "invoice date",
    ]
    for term in preferred_terms:
        term_norm = normalize_column_label(term)
        for col in df.columns:
            col_norm = normalize_column_label(col)
            if (col_norm == term_norm or term_norm in col_norm) and col not in candidates:
                candidates.append(col)
    for col in df.columns:
        if "date" in normalize_column_label(col) or "fecha" in normalize_column_label(col):
            if col not in candidates:
                candidates.append(col)
    return candidates or df.columns.tolist()


def parse_date_flexible(series: pd.Series) -> pd.Series:
    direct = pd.to_datetime(series, errors="coerce", dayfirst=False)
    alternate = pd.to_datetime(series, errors="coerce", dayfirst=True)
    return direct.fillna(alternate)


@st.cache_data(show_spinner=False)
def load_manufacturing_source(file_bytes: bytes, filename: str) -> dict[str, pd.DataFrame]:
    """Lee uno o varios archivos de fabricación subidos manualmente en cada sesión."""
    name = filename.lower()
    try:
        if name.endswith(".csv"):
            raw_text = file_bytes.decode("utf-8-sig", errors="replace")
            for sep in [";", ",", None]:
                try:
                    frame = pd.read_csv(StringIO(raw_text), sep=sep, engine="python")
                    if frame.shape[1] >= 2:
                        return {filename: frame}
                except Exception:
                    continue
            raise ValueError("No se pudo reconocer la estructura del CSV.")
        if name.endswith(".xlsx") or name.endswith(".xls"):
            workbook = pd.read_excel(BytesIO(file_bytes), sheet_name=None)
            return {str(sheet): frame for sheet, frame in workbook.items() if frame is not None and not frame.empty}
    except ImportError as exc:
        if name.endswith(".xls"):
            raise RuntimeError(
                "Para leer archivos .XLS en Streamlit se requiere incluir `xlrd>=2.0.1` en requirements.txt."
            ) from exc
        raise
    raise ValueError(f"Formato no soportado para fecha de fabricación: {filename}")


def build_age_bucket(series: pd.Series) -> pd.Series:
    age_values = pd.to_numeric(series, errors="coerce")
    return pd.cut(
        age_values,
        bins=[-np.inf, 3, 5, 8, 10, 15, np.inf],
        labels=MANUFACTURING_AGE_BUCKET_ORDER,
        right=False,
    )

def build_manufacturing_match(
    filtered_assets: pd.DataFrame,
    manufacturing_frames: list[pd.DataFrame],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Cruza la vista filtrada actual con las referencias de fabricación por serial."""
    assets = filtered_assets.copy()
    assets["Serial match key"] = assets["Serial number"].map(normalize_serial_match)

    if not manufacturing_frames:
        assets["Manufacturing Date"] = pd.NaT
        assets["Manufacturing Source"] = pd.NA
        assets["Manufacturing Product"] = pd.NA
        assets["Manufacturing age (years)"] = np.nan
        assets["Manufacturing age bucket"] = pd.NA
        return assets, pd.DataFrame()

    reference = pd.concat(manufacturing_frames, ignore_index=True)
    reference["Serial match key"] = reference["Manufacturing Serial"].map(normalize_serial_match)
    reference["Manufacturing Date"] = pd.to_datetime(reference["Manufacturing Date"], errors="coerce")
    reference = reference[(reference["Serial match key"] != "") & reference["Manufacturing Date"].notna()].copy()

    duplicate_check = (
        reference.groupby("Serial match key", dropna=False)["Manufacturing Date"]
        .nunique()
        .reset_index(name="Different manufacturing dates")
    )
    conflicting_keys = set(
        duplicate_check.loc[duplicate_check["Different manufacturing dates"] > 1, "Serial match key"].tolist()
    )

    reference = reference.sort_values(["Serial match key", "Manufacturing Date"], ascending=[True, True])
    reference = reference.drop_duplicates(subset=["Serial match key"], keep="first").copy()
    reference["Manufacturing date conflict"] = reference["Serial match key"].isin(conflicting_keys)

    keep_cols = [
        "Serial match key",
        "Manufacturing Serial",
        "Manufacturing Date",
        "Manufacturing Source",
        "Manufacturing Sheet",
        "Manufacturing Product",
        "Manufacturing date conflict",
    ]
    merged = assets.merge(reference[keep_cols], on="Serial match key", how="left")
    today = pd.Timestamp(date.today())
    merged["Manufacturing age (years)"] = ((today - merged["Manufacturing Date"]).dt.days / 365.25).round(1)
    merged["Manufacturing year"] = merged["Manufacturing Date"].dt.year.astype("Int64")
    merged["Manufacturing age bucket"] = build_age_bucket(merged["Manufacturing age (years)"])
    merged["Manufacturing matched"] = merged["Manufacturing Date"].notna()
    return merged, reference


def _same_serial_universe(left_df: pd.DataFrame, right_df: pd.DataFrame) -> bool:
    """Valida que dos dataframes representen el mismo filtro de equipos."""
    if left_df is None or right_df is None or left_df.empty or right_df.empty:
        return False
    if "Serial number" not in left_df.columns or "Serial number" not in right_df.columns:
        return False

    left_serials = set(left_df["Serial number"].map(normalize_serial_match).dropna())
    right_serials = set(right_df["Serial number"].map(normalize_serial_match).dropna())
    left_serials.discard("")
    right_serials.discard("")
    return bool(left_serials) and left_serials == right_serials


def resolve_pdf_report_dataframe(
    filtered_df: pd.DataFrame,
    active_tab: str,
    source_label_value: str = "",
) -> tuple[pd.DataFrame, str, bool]:
    """Selecciona el dataframe correcto para generar PDF.

    v45: el PDF usa el dataframe enriquecido con fabricación siempre que exista
    un cruce válido para el mismo universo de seriales filtrados, no solo cuando
    la pestaña activa sea Antigüedad / fabricación. Esto evita que el informe
    pierda el análisis de edad por fabricación si el usuario prepara el PDF desde
    otra pestaña o desde el sidebar después de haber calculado el cruce.
    """
    report_df = filtered_df.copy() if isinstance(filtered_df, pd.DataFrame) else pd.DataFrame()
    report_source = source_label_value
    using_manufacturing = False

    manufacturing_df = st.session_state.get(MANUFACTURING_EXCEL_EXPORT_SESSION_KEY)
    if isinstance(manufacturing_df, pd.DataFrame) and not manufacturing_df.empty:
        has_manufacturing_age = _has_valid_numeric_column(manufacturing_df, "Manufacturing age (years)")
        if has_manufacturing_age and _same_serial_universe(report_df, manufacturing_df):
            report_df = manufacturing_df.copy()
            report_source = st.session_state.get(MANUFACTURING_EXCEL_EXPORT_SOURCE_KEY, source_label_value)
            using_manufacturing = True
        elif active_tab == "Antigüedad / fabricación":
            st.warning(
                "El PDF usará la vista estándar del Records List porque el cruce de fabricación guardado "
                "no coincide con los filtros activos actuales o no contiene edades de fabricación válidas. "
                "Vuelve a abrir la pestaña Antigüedad / fabricación para recalcular el cruce antes de preparar el PDF."
            )

    return report_df, report_source, using_manufacturing


st.markdown(
    f"""
    <div class="hero">
        <div class="hero-top">
            <div class="hero-brand">
                <div class="brand-chip">DASHBOARD</div>
                <div class="workspace-chip">Hi, Javier · Workspace de base instalada</div>
            </div>
            <div class="workspace-chip">Control visual · Devoryn dark mode</div>
        </div>
        <h1>Records List Intelligence Dashboard <span class="code-stamp">Código creado: {CODE_CREATED_AT} · {CODE_VERSION_LABEL}</span></h1>
        <p>Panel ejecutivo para explorar la base instalada, configuration insights, sistema operativo, procesamiento y gap de repuestos con una apariencia oscura, limpia y premium.</p>
        <div class="badge-row">
            <span class="badge">Base instalada</span>
            <span class="badge">Machine configuration</span>
            <span class="badge">Operating system</span>
            <span class="badge">PM & processing</span>
            <span class="badge">Stock gap analysis</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div style="border:1px solid rgba(113,225,255,0.38); border-radius:14px; padding:0.65rem 0.9rem; margin:0.45rem 0 0.9rem 0; background:rgba(113,225,255,0.08); font-size:0.78rem; color:#eafaff;">
        ✅ BUILD ACTIVO CONFIRMADO: <b>{CODE_VERSION_LABEL}</b> · {CODE_CREATED_AT} · <code>{PARSER_VERSION}</code>
    </div>
    """,
    unsafe_allow_html=True,
)

st.sidebar.markdown(
    """
    <div class="sidebar-top-card">
        <h3>✦ Control center</h3>
        <p>Explora la base instalada, filtra la operación y navega el dashboard con una experiencia visual alineada al estilo oscuro premium que definiste.</p>
        <div class="sidebar-pill">Devoryn dark · active</div>
    </div>
    """,
    unsafe_allow_html=True,
)
uploaded_file = st.sidebar.file_uploader("Sube el archivo Records List", type=["csv", "xlsx", "xls"])

base_dir = Path(__file__).resolve().parent
sample_candidates = sorted(base_dir.glob("Records_List_Report*.csv"))
default_master_candidates = sorted(base_dir.glob("New TP Spare*.xlsx"))

raw_df, source_label = get_active_records_dataset(uploaded_file, sample_candidates)

if raw_df.empty:
    st.info("Sube el archivo Records List para activar el dashboard.")
    st.stop()

ensure_chart_drill_state()

raw_df, CONFIG_KEYS = parse_machine_configuration(raw_df)
raw_df = add_operating_system_columns(raw_df, CONFIG_KEYS)
raw_df = standardize_blood_bank_column(raw_df)
st.sidebar.caption(f"Fuente activa: {source_label}")
st.sidebar.caption(f"Build activo: {PARSER_VERSION}")
st.sidebar.caption(f"Código creado: {CODE_CREATED_AT}")
st.sidebar.markdown('<div class="small-note">Usa los filtros como un panel de control para refinar región, país, distribuidor, instrumento y estado operativo.</div>', unsafe_allow_html=True)

# Aplica limpiezas pendientes antes de crear cualquier multiselect del sidebar.
consume_pending_sidebar_filter_clear()

if st.sidebar.button("Limpiar filtros laterales", key="clear_sidebar_filters_button_v30"):
    clear_sidebar_filter_widgets()

region_options = sorted(raw_df["Commercial Region"].dropna().unique().tolist())
selected_regions = st.sidebar.multiselect("Región comercial", options=region_options, default=[], placeholder="Selecciona una o varias regiones", key=SIDEBAR_REGION_KEY)

country_base = raw_df.copy()
if selected_regions:
    country_base = country_base[country_base["Commercial Region"].isin(selected_regions)]
country_options = sorted(country_base["Country"].dropna().unique().tolist())
selected_countries = st.sidebar.multiselect("País", options=country_options, default=[], placeholder="Selecciona uno o varios países", key=SIDEBAR_COUNTRY_KEY)

dist_base = raw_df.copy()
if selected_regions:
    dist_base = dist_base[dist_base["Commercial Region"].isin(selected_regions)]
if selected_countries:
    dist_base = dist_base[dist_base["Country"].isin(selected_countries)]
distributor_options = sorted(dist_base["Distributor name"].dropna().unique().tolist())
selected_distributors = st.sidebar.multiselect("Nombre de distribuidor", options=distributor_options, default=[], placeholder="Selecciona uno o varios distribuidores", key=SIDEBAR_DISTRIBUTOR_KEY)

instrument_base = raw_df.copy()
if selected_regions:
    instrument_base = instrument_base[instrument_base["Commercial Region"].isin(selected_regions)]
if selected_countries:
    instrument_base = instrument_base[instrument_base["Country"].isin(selected_countries)]
if selected_distributors:
    instrument_base = instrument_base[instrument_base["Distributor name"].isin(selected_distributors)]
instrument_options = sorted(instrument_base["Instrument type"].dropna().unique().tolist())
selected_instruments = st.sidebar.multiselect("Tipo de instrumento", options=instrument_options, default=[], placeholder="Selecciona uno o varios instrumentos", key=SIDEBAR_INSTRUMENT_KEY)

status_base = raw_df.copy()
if selected_regions:
    status_base = status_base[status_base["Commercial Region"].isin(selected_regions)]
if selected_countries:
    status_base = status_base[status_base["Country"].isin(selected_countries)]
if selected_distributors:
    status_base = status_base[status_base["Distributor name"].isin(selected_distributors)]
if selected_instruments:
    status_base = status_base[status_base["Instrument type"].isin(selected_instruments)]

state_count_items = compute_state_filter_counts(status_base)
state_option_map = {f"{state} ({count})": state for state, count in state_count_items}
selected_state_labels = st.sidebar.multiselect(
    "Estado operativo",
    options=list(state_option_map.keys()),
    default=[],
    placeholder="Selecciona uno o varios estados",
    help="Incluye el estado especial 'No rutina' y cualquier otro estado disponible en la vista actual.",
    key=SIDEBAR_STATE_KEY,
)
selected_states = [state_option_map[label] for label in selected_state_labels]

filtered = raw_df.copy()
if selected_regions:
    filtered = filtered[filtered["Commercial Region"].isin(selected_regions)]
if selected_countries:
    filtered = filtered[filtered["Country"].isin(selected_countries)]
if selected_distributors:
    filtered = filtered[filtered["Distributor name"].isin(selected_distributors)]
if selected_instruments:
    filtered = filtered[filtered["Instrument type"].isin(selected_instruments)]
filtered = apply_operational_status_filter(filtered, selected_states)

sidebar_filtered = filtered.copy()
filter_summary_for_panel = build_filter_summary(
    selected_regions,
    selected_countries,
    selected_distributors,
    selected_instruments,
    selected_states,
)
filtered, chart_drill_active = apply_chart_drill_filter(filtered)

if filtered.empty:
    render_chart_drill_filter_banner(sidebar_filtered, filtered, filter_summary_for_panel)
    st.warning("No hay datos para la combinación de filtros actual.")
    st.stop()

render_chart_drill_filter_banner(sidebar_filtered, filtered, filter_summary_for_panel)

st.sidebar.markdown("---")
st.markdown("### Navegación del dashboard")
active_dashboard_tab = st.radio(
    "Selecciona la pestaña activa",
    options=DASHBOARD_TABS,
    horizontal=True,
    key=ACTIVE_DASHBOARD_TAB_KEY,
    label_visibility="collapsed",
)
st.caption("La pestaña activa se conserva al aplicar filtros desde gráficas o botones; ya no debe regresar automáticamente a Base instalada.")

if active_dashboard_tab == "Base instalada":
    st.subheader("Base instalada")
    st.caption("Mapa y analítica de base instalada con enfoque en cobertura geográfica, antigüedad de instalación y estado de despliegue.")

    st.download_button(
        "Informe a corporativo mensual",
        data=build_monthly_corporate_latam_report_excel(raw_df, source_label_value=source_label),
        file_name=f"informe_corporativo_mensual_LATAM_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime=EXCEL_MIME,
        use_container_width=False,
        key="download_monthly_corporate_latam_report_v42",
        help="Genera un Excel solo LATAM con base instalada actualizada, equipos en rutina, recién comprados y base proyectada = rutina + recién comprados.",
    )
    st.caption("Informe corporativo mensual solo LATAM: base instalada actualizada, equipos en rutina, recién comprados y base proyectada = rutina + recién comprados.")
    render_installed_base_geo_map(filtered)

    c1, c2 = st.columns(2)
    with c1:
        type_df = filtered["Instrument type"].fillna("Unknown").value_counts().reset_index()
        type_df.columns = ["Instrument type", "Count"]
        fig_type = px.bar(type_df, x="Count", y="Instrument type", orientation="h", title="Base instalada por tipo de instrumento", text="Count", custom_data=["Instrument type", "Count"])
        fig_type.update_traces(marker_color=ACCENT, textposition="outside", hovertemplate="Instrumento: %{customdata[0]}<br>Activos: %{customdata[1]}<extra></extra>")
        fig_type.update_layout(yaxis=dict(categoryorder="total ascending"))
        render_drilldown_plotly_chart(
            glow_layout(fig_type, 470),
            key="drill_base_type_chart_v30",
            source_label="Base instalada por tipo de instrumento",
            payload_builder=lambda point: payload_from_instrument_point(point, "Base instalada por tipo de instrumento"),
        )

    with c2:
        install_df = filtered.dropna(subset=["Installation date"]).copy()
        if install_df.empty:
            st.info("No hay fechas de instalación válidas para el filtro actual.")
        else:
            install_df["Installation year"] = install_df["Installation date"].dt.year.astype(int)
            yearly = install_df.groupby("Installation year", dropna=False).size().reset_index(name="Count").sort_values("Installation year")
            fig_year = px.bar(yearly, x="Installation year", y="Count", title="Instalaciones por año", text="Count", custom_data=["Installation year", "Count"])
            fig_year.update_traces(marker_color=ACCENT_2, textposition="outside", hovertemplate="Año: %{customdata[0]}<br>Instalaciones: %{customdata[1]}<extra></extra>")
            render_drilldown_plotly_chart(
                glow_layout(fig_year, 470),
                key="drill_installation_year_chart_v30",
                source_label="Instalaciones por año",
                payload_builder=payload_from_installation_year_point,
            )

    c3, c4 = st.columns(2)
    with c3:
        pipeline_df = filtered.copy()
        pipeline_df["Installation stage"] = compute_installation_stage(filtered)
        pipeline_summary = pipeline_df.groupby(["Instrument type", "Installation stage"], dropna=False).size().reset_index(name="Count")
        fig_ready = px.bar(
            pipeline_summary,
            x="Instrument type",
            y="Count",
            color="Installation stage",
            title="Sistemas instalados vs listos / pipeline",
            custom_data=["Instrument type", "Installation stage", "Count"],
        )
        fig_ready.update_traces(hovertemplate="Instrumento: %{customdata[0]}<br>Etapa: %{customdata[1]}<br>Cantidad: %{customdata[2]}<extra></extra>")
        fig_ready.update_xaxes(tickangle=-28)
        render_drilldown_plotly_chart(
            glow_layout(fig_ready, 470),
            key="drill_pipeline_chart_v30",
            source_label="Sistemas instalados vs listos / pipeline",
            payload_builder=payload_from_pipeline_point,
        )

    with c4:
        city_df = (
            filtered.assign(CityLabel=build_city_label_series(filtered))
            .groupby("CityLabel", dropna=False)
            .size()
            .reset_index(name="Count")
            .sort_values("Count", ascending=False)
            .head(15)
        )
        fig_city = px.bar(city_df, x="Count", y="CityLabel", orientation="h", title="Análisis por ciudad", text="Count", custom_data=["CityLabel", "Count"])
        fig_city.update_traces(marker_color=ACCENT_3, textposition="outside", hovertemplate="Ciudad / País: %{customdata[0]}<br>Activos: %{customdata[1]}<extra></extra>")
        fig_city.update_layout(yaxis=dict(categoryorder="total ascending"))
        render_drilldown_plotly_chart(
            glow_layout(fig_city, 470),
            key="drill_city_chart_v30",
            source_label="Análisis por ciudad",
            payload_builder=payload_from_city_point,
        )

    st.markdown("### Base instalada por modelo y estado operativo")
    st.caption("Vista matricial de la base instalada: cada fila representa un modelo y cada segmento muestra el estado operativo. Selecciona un segmento para filtrar ese modelo en ese estado.")
    model_status_df = filtered.copy()
    model_status_df["Instrument type"] = model_status_df["Instrument type"].fillna("No informado").astype(str).str.strip().replace("", "No informado")

    # Defensive rebuild of grouped status. This prevents the matrix from collapsing into
    # a single "No informado" segment if the source file used the export header "Status".
    if "Operational status" in model_status_df.columns:
        model_status_df["Operational status grouped"] = model_status_df["Operational status"].map(normalize_operational_status)
    elif "Status" in model_status_df.columns:
        model_status_df["Operational status grouped"] = model_status_df["Status"].map(normalize_operational_status)

    model_status_df["Operational status grouped"] = model_status_df["Operational status grouped"].fillna("No informado").astype(str).str.strip().replace("", "No informado")
    model_status_summary = (
        model_status_df
        .groupby(["Instrument type", "Operational status grouped"], dropna=False)
        .size()
        .reset_index(name="Count")
    )
    if model_status_summary.empty:
        st.info("No hay datos de modelo y estado operativo para la vista actual.")
    else:
        if model_status_summary["Operational status grouped"].nunique(dropna=True) == 1 and str(model_status_summary["Operational status grouped"].iloc[0]) == "No informado":
            st.warning("La matriz solo encontró 'No informado'. Revisa que estés ejecutando la versión v27 y que el archivo fuente tenga la columna Status/Operational status correctamente cargada.")
        model_order = (
            model_status_summary.groupby("Instrument type", as_index=False)["Count"]
            .sum()
            .sort_values(["Count", "Instrument type"], ascending=[True, True])["Instrument type"]
            .tolist()
        )
        fig_model_status = px.bar(
            model_status_summary,
            x="Count",
            y="Instrument type",
            color="Operational status grouped",
            orientation="h",
            text="Count",
            title="Base instalada por modelo y estado operativo",
            custom_data=["Instrument type", "Operational status grouped", "Count"],
        )
        fig_model_status.update_traces(
            texttemplate="%{x}",
            textposition="inside",
            insidetextanchor="middle",
            hovertemplate="Modelo: %{customdata[0]}<br>Estado operativo: %{customdata[1]}<br>Equipos: %{customdata[2]}<extra></extra>",
        )
        fig_model_status.update_layout(
            barmode="stack",
            yaxis=dict(categoryorder="array", categoryarray=model_order),
            legend_title_text="Estado operativo",
        )
        render_drilldown_plotly_chart(
            glow_layout(fig_model_status, max(430, 82 + 46 * len(model_order))),
            key="drill_model_operational_status_chart_v30",
            source_label="Base instalada por modelo y estado operativo",
            payload_builder=payload_from_model_status_point,
            help_text="Filtro disponible: selecciona/clic en un segmento para ver únicamente ese modelo con ese estado operativo. Usa el panel superior para quitar o deshacer filtros.",
        )

        with st.expander("Ver tabla matriz modelo vs estado operativo"):
            status_matrix = model_status_summary.pivot_table(
                index="Instrument type",
                columns="Operational status grouped",
                values="Count",
                aggfunc="sum",
                fill_value=0,
            )
            status_matrix["Total"] = status_matrix.sum(axis=1)
            status_matrix = status_matrix.sort_values("Total", ascending=False)
            st.dataframe(status_matrix, use_container_width=True)


    st.markdown("### Vista corporativa por distribuidor")
    model_options = (
        filtered["Instrument type"]
        .fillna("No informado")
        .astype(str)
        .value_counts()
        .index
        .tolist()
    )
    if model_options:
        st.markdown(
            '<div class="small-note">Primero se muestra un resumen ejecutivo con los distribuidores más relevantes. Debajo aparecen gráficos circulares Top 5 por modelo. El detalle completo de todos los distribuidores se puede abrir más abajo sin saturar la vista principal.</div>',
            unsafe_allow_html=True,
        )
        render_drilldown_plotly_chart(
            build_distributor_global_overview(filtered, top_n=5),
            key="drill_global_distributor_overview_bar_v30",
            source_label="Vista global por distribuidor",
            payload_builder=payload_from_global_distributor_point,
        )

        cards_per_row = 3
        for start in range(0, len(model_options), cards_per_row):
            row_models = model_options[start:start + cards_per_row]
            cols = st.columns(len(row_models))
            for local_idx, (col, model_name) in enumerate(zip(cols, row_models)):
                model_key = f"{start}_{local_idx}_{hashlib.md5(str(model_name).encode('utf-8', errors='ignore')).hexdigest()[:8]}"
                with col:
                    render_drilldown_plotly_chart(
                        build_distributor_model_donut(filtered, model_name, top_n=5),
                        key=f"donut_model_distributor_{model_key}_v30",
                        source_label=f"Distribución por distribuidor | {model_name}",
                        payload_builder=lambda point, selected_model=model_name: payload_from_distributor_model_donut(selected_model, point),
                    )

        with st.expander("Ver detalle completo de todos los distribuidores por modelo", expanded=False):
            st.markdown(
                '<div class="small-note">Esta vista despliega el detalle completo sin resumir. Se usa una barra horizontal por modelo porque comunica mejor que un donut cuando hay muchos distribuidores.</div>',
                unsafe_allow_html=True,
            )
            st.dataframe(build_distributor_detail_table(filtered), use_container_width=True, hide_index=True)
            for model_idx, model_name in enumerate(model_options):
                model_key = f"{model_idx}_{hashlib.md5(str(model_name).encode('utf-8', errors='ignore')).hexdigest()[:8]}"
                render_drilldown_plotly_chart(
                    build_distributor_detail_bar(filtered, model_name),
                    key=f"drill_detail_bar_model_{model_key}_v30",
                    source_label=f"Detalle completo | {model_name}",
                    payload_builder=lambda point, selected_model=model_name: payload_from_detail_distributor_point(selected_model, point),
                )

    st.markdown("### Tabla general filtrada")
    visible_columns = [
        "Commercial Region",
        "Country",
        "Distributor name",
        "Customer name",
        "Instrument type",
        "Serial number",
        "Operational status grouped",
        "Operational status",
        "Asset condition",
        "Installation date",
        "Number of tests per day",
        "Operating System",
        "Machine Configurations",
    ]
    st.dataframe(filtered[visible_columns].copy(), use_container_width=True, hide_index=True)

if active_dashboard_tab == "Machine configuration":
    st.subheader("Machine configuration")
    st.caption("Vista ejecutiva por ítem de configuración, con Banco de sangre como indicador principal y gráficas separadas para cada campo aplicable.")
    applicable_fields = active_config_fields(filtered, CONFIG_KEYS)
    cfg_cols_prefixed = [f"CFG::{col}" for col in applicable_fields]

    if not cfg_cols_prefixed:
        st.info("No se detectaron campos aplicables dentro de Machine Configurations para el filtro actual.")
    else:
        assets_with_cfg = int(filtered["Machine Configurations"].notna().sum())
        avg_cfg_fields = filtered["Machine config fields populated"].mean()
        unique_cfg_fields = len(applicable_fields)

        blood_bank_yes = count_blood_bank_yes(filtered)

        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1:
            metric_card("Equipos con config", f"{assets_with_cfg:,}", "Con información en Machine Configurations")
        with mc2:
            metric_card("Banco de sangre", f"{blood_bank_yes:,}", f"{_safe_share_pct(blood_bank_yes, len(filtered)):.1f}% del total filtrado")
        with mc3:
            metric_card("Campos aplicables", f"{unique_cfg_fields}", "Solo ítems presentes en el filtro actual")
        with mc4:
            metric_card("Promedio de campos", f"{avg_cfg_fields:.1f}", "Campos poblados por equipo")

        st.markdown("### Banco de sangre")
        st.markdown('<div class="small-note">Conteo validado desde la columna <b>In Blood Bank</b> y aliases del export como <b>In Blook Bank</b>.</div>', unsafe_allow_html=True)
        render_drilldown_plotly_chart(
            build_blood_bank_donut(filtered),
            key="blood_bank_donut_main_v30",
            source_label="Banco de sangre",
            payload_builder=payload_from_blood_bank_point,
        )
        render_pie_filter_fallback_buttons(
            ["Banco de sangre", "Equipos en laboratorio"],
            key_prefix="blood_bank_main",
            payload_from_value=lambda selected_label: payload_from_blood_bank_point({"label": selected_label, "customdata": [selected_label]}),
            max_options=2,
            excluded_labels=set(),
        )

        coverage_df = pd.DataFrame(
            [{"Config field": col.replace("CFG::", ""), "Populated assets": int(filtered[col].notna().sum())} for col in cfg_cols_prefixed]
        )
        coverage_df = coverage_df[coverage_df["Populated assets"] > 0].sort_values("Populated assets", ascending=False)

        fig_cfg_fill = px.bar(
            coverage_df,
            x="Populated assets",
            y="Config field",
            orientation="h",
            title="Cobertura por campo aplicable",
            text="Populated assets",
        )
        fig_cfg_fill.update_traces(
            marker_color=ACCENT,
            textposition="outside",
            hovertemplate="Campo: %{y}<br>Equipos con dato: %{x}<extra></extra>",
        )
        fig_cfg_fill.update_layout(yaxis=dict(categoryorder="total ascending"))
        render_drilldown_plotly_chart(
            glow_layout(fig_cfg_fill, 520),
            key="config_coverage_chart_v30",
            source_label="Cobertura por campo aplicable",
            payload_builder=payload_from_config_coverage_point,
        )

        st.markdown("### Distribución visual por ítem")
        st.markdown(
            '<div class="small-note">Cada gráfico resume la distribución de valores del ítem correspondiente. Se muestran los valores principales y, si aplica, una categoría <b>Otros</b> para simplificar la lectura.</div>',
            unsafe_allow_html=True,
        )

        donut_fields = coverage_df["Config field"].tolist()
        if donut_fields:
            for idx in range(0, len(donut_fields), 3):
                cols = st.columns(3)
                for col_ui, field_name in zip(cols, donut_fields[idx:idx + 3]):
                    selected_cfg_col = f"CFG::{field_name}"
                    item_series = filtered[selected_cfg_col].dropna()
                    item_series = item_series.astype(str).str.strip()
                    item_series = item_series[item_series.ne("")]
                    total_assets = int(item_series.shape[0])
                    with col_ui:
                        render_drilldown_plotly_chart(
                            build_config_donut(field_name, item_series, total_assets),
                            key=f"config_donut_{hashlib.md5(str(field_name).encode('utf-8', errors='ignore')).hexdigest()[:10]}_v30",
                            source_label=f"Configuración | {field_name}",
                            payload_builder=lambda point, selected_field=field_name: payload_from_config_value_point(selected_field, point),
                        )
                        donut_dist = compress_value_distribution(item_series, max_slices=4)
                        render_pie_filter_fallback_buttons(
                            donut_dist["Label"].tolist(),
                            key_prefix=f"config_{hashlib.md5(str(field_name).encode('utf-8', errors='ignore')).hexdigest()[:10]}",
                            payload_from_value=lambda selected_value, selected_field=field_name: payload_from_config_value_point(
                                selected_field,
                                {"label": selected_value, "customdata": [selected_value]},
                            ),
                            max_options=5,
                            excluded_labels={"Otros"},
                        )

        st.markdown("### Top valores por ítem")
        detail_rows = []
        for field_name in donut_fields:
            selected_cfg_col = f"CFG::{field_name}"
            item_series = filtered[selected_cfg_col].dropna().astype(str).str.strip()
            item_series = item_series[item_series.ne("")]
            if item_series.empty:
                continue
            dist = item_series.value_counts().reset_index()
            dist.columns = ["Value", "Count"]
            top_row = dist.iloc[0]
            detail_rows.append(
                {
                    "Config field": field_name,
                    "Top value": str(top_row["Value"]),
                    "Top count": int(top_row["Count"]),
                    "Unique values": int(dist.shape[0]),
                    "Assets with value": int(item_series.shape[0]),
                }
            )

        if detail_rows:
            st.dataframe(pd.DataFrame(detail_rows), use_container_width=True, hide_index=True)

        with st.expander("Ver tabla ampliada de machine configuration"):
            detail_columns = [
                "Commercial Region",
                "Country",
                "Distributor name",
                "Customer name",
                "Instrument type",
                "Serial number",
                "Operating System",
                "Operational status",
            ] + [f"CFG::{field}" for field in donut_fields]
            machine_table = filtered[detail_columns].copy().rename(columns={f"CFG::{field}": field for field in donut_fields})
            st.dataframe(machine_table, use_container_width=True, hide_index=True)

if active_dashboard_tab == "Sistema operativo":
    st.subheader("Sistema operativo")
    st.caption("Vista diseñada para identificar instrumentos con sistemas operativos legacy y priorizar migraciones urgentes a Windows 10.")
    os_df = filtered.copy()
    os_df["Operating System"] = os_df["Operating System"].fillna("No informado")
    os_df["OS Upgrade Bucket"] = os_df["Operating System"].map(os_upgrade_bucket)
    os_df["Installation date display"] = os_df["Installation date"].map(format_date_for_hover)

    systems_with_os = int(filtered["Operating System"].notna().sum())
    unique_os = int(filtered["Operating System"].nunique(dropna=True))
    unknown_os = int(os_df["Operating System"].isin(["Unknown", "No informado"]).sum())
    urgent_mask = os_df["Operating System"].isin(["Windows XP", "Windows Vista", "Windows 7", "Windows 2000"])
    urgent_count = int(urgent_mask.sum())

    o1, o2, o3, o4 = st.columns(4)
    with o1:
        metric_card("Equipos con OS", f"{systems_with_os:,}", "OS identificado desde machine configuration")
    with o2:
        metric_card("OS únicos", f"{unique_os}", "Variedad de sistemas operativos")
    with o3:
        metric_card("Legacy / urgentes", f"{urgent_count:,}", "XP, Vista, Win7 o Win2000")
    with o4:
        metric_card("Sin definir / unknown", f"{unknown_os:,}", "Requiere depuración del dato")

    s1, s2 = st.columns(2)
    with s1:
        os_summary = os_df.groupby("Operating System", dropna=False).size().reset_index(name="Count").sort_values("Count", ascending=False)
        fig_os = px.bar(os_summary, x="Operating System", y="Count", title="Distribución detallada de sistemas operativos", text="Count")
        fig_os.update_traces(
            marker_color=ACCENT,
            textposition="outside",
            hovertemplate="Sistema operativo: %{x}<br>Equipos: %{y}<extra></extra>",
        )
        fig_os.update_xaxes(tickangle=-28)
        render_drilldown_plotly_chart(
            glow_layout(fig_os, 500, title_size=16),
            key="os_distribution_chart_v30",
            source_label="Distribución detallada de sistemas operativos",
            payload_builder=lambda point: payload_from_axis_value(point, "Operating System", "Distribución detallada de sistemas operativos", "Sistema operativo", axis="x"),
        )

    with s2:
        os_points = os_df[[
            "Serial number",
            "Instrument type",
            "Operating System",
            "OS Upgrade Bucket",
            "Distributor name",
            "Customer name",
            "Country",
            "Commercial Region",
            "Operational status",
            "Installation date display",
        ]].copy()
        fig_os_type = px.scatter(
            os_points,
            x="Operating System",
            y="Serial number",
            color="OS Upgrade Bucket",
            title="Qué seriales tienen cada sistema operativo",
            custom_data=[
                "Instrument type",
                "Distributor name",
                "Customer name",
                "Country",
                "Commercial Region",
                "Operational status",
                "Installation date display",
                "OS Upgrade Bucket",
            ],
            category_orders={"Operating System": os_summary["Operating System"].tolist()},
            color_discrete_map={
                "Windows 10 / OK": ACCENT_3,
                "Legacy / urgente migrar": DANGER,
                "Otro OS / validar": WARNING,
                "Revisar campo OS": ACCENT_2,
            },
        )
        fig_os_type.update_traces(
            marker=dict(size=11, opacity=0.88),
            hovertemplate=(
                "Serial: %{y}<br>"
                "OS: %{x}<br>"
                "Instrumento: %{customdata[0]}<br>"
                "Distribuidor: %{customdata[1]}<br>"
                "Cliente: %{customdata[2]}<br>"
                "País: %{customdata[3]}<br>"
                "Región: %{customdata[4]}<br>"
                "Estado operativo: %{customdata[5]}<br>"
                "Instalación: %{customdata[6]}<br>"
                "Prioridad: %{customdata[7]}<extra></extra>"
            ),
        )
        fig_os_type.update_layout(legend_title_text="Prioridad upgrade")
        fig_os_type.update_xaxes(tickangle=-28)
        render_drilldown_plotly_chart(
            glow_layout(fig_os_type, 620, title_size=16),
            key="os_serial_scatter_v30",
            source_label="Qué seriales tienen cada sistema operativo",
            payload_builder=lambda point: payload_from_serial_axis(point, "Qué seriales tienen cada sistema operativo", axis="y"),
        )

    s3, s4 = st.columns(2)
    with s3:
        urgent_points = os_df[urgent_mask][[
            "Serial number",
            "Instrument type",
            "Operating System",
            "Distributor name",
            "Customer name",
            "Country",
            "Commercial Region",
            "Operational status",
            "Installation date display",
        ]].copy()
        if urgent_points.empty:
            st.success("No se detectan instrumentos con Windows legacy dentro del filtro actual.")
        else:
            urgent_points = urgent_points.sort_values(["Country", "Instrument type", "Serial number"])
            fig_urgent = px.scatter(
                urgent_points,
                x="Country",
                y="Serial number",
                color="Instrument type",
                title="Seriales que requieren actualización urgente a Windows 10",
                custom_data=[
                    "Operating System",
                    "Instrument type",
                    "Distributor name",
                    "Customer name",
                    "Commercial Region",
                    "Operational status",
                    "Installation date display",
                ],
            )
            fig_urgent.update_traces(
                marker=dict(size=12, opacity=0.92),
                hovertemplate=(
                    "Serial: %{y}<br>"
                    "País: %{x}<br>"
                    "OS actual: %{customdata[0]}<br>"
                    "Instrumento: %{customdata[1]}<br>"
                    "Distribuidor: %{customdata[2]}<br>"
                    "Cliente: %{customdata[3]}<br>"
                    "Región: %{customdata[4]}<br>"
                    "Estado operativo: %{customdata[5]}<br>"
                    "Instalación: %{customdata[6]}<extra></extra>"
                ),
            )
            fig_urgent.update_layout(legend_title_text="Instrumento")
            fig_urgent.update_xaxes(tickangle=-18)
            render_drilldown_plotly_chart(
                glow_layout(fig_urgent, 620, title_size=16),
                key="os_urgent_serials_chart_v30",
                source_label="Seriales que requieren actualización urgente a Windows 10",
                payload_builder=lambda point: payload_from_serial_axis(point, "Seriales que requieren actualización urgente a Windows 10", axis="y"),
            )

    with s4:
        bucket_df = os_df.groupby("OS Upgrade Bucket", dropna=False).size().reset_index(name="Count")
        order = ["Windows 10 / OK", "Legacy / urgente migrar", "Otro OS / validar", "Revisar campo OS"]
        bucket_df["order"] = bucket_df["OS Upgrade Bucket"].map({k: i for i, k in enumerate(order)}).fillna(999)
        bucket_df = bucket_df.sort_values(["order", "Count"], ascending=[True, False])
        fig_bucket = px.bar(bucket_df, x="OS Upgrade Bucket", y="Count", title="Priorización de acción para upgrade", text="Count")
        fig_bucket.update_traces(marker_color=ACCENT_2, textposition="outside", hovertemplate="Acción: %{x}<br>Equipos: %{y}<extra></extra>")
        fig_bucket.update_xaxes(tickangle=-18)
        render_drilldown_plotly_chart(
            glow_layout(fig_bucket, 520, title_size=16),
            key="os_upgrade_bucket_chart_v30",
            source_label="Priorización de acción para upgrade",
            payload_builder=lambda point: payload_from_axis_value(point, "OS Upgrade Bucket", "Priorización de acción para upgrade", "Prioridad OS", axis="x"),
        )

    st.markdown("### Tabla priorizada para migración a Windows 10")
    urgent_table = os_df[urgent_mask][[
        "Commercial Region", "Country", "Distributor name", "Customer name", "Instrument type", "Serial number", "Operating System", "Operational status", "Installation date display"
    ]].copy()
    if urgent_table.empty:
        st.info("No hay equipos en categoría urgente para el filtro actual.")
    else:
        urgent_table = urgent_table.rename(columns={"Installation date display": "Installation date"})
        st.dataframe(urgent_table, use_container_width=True, hide_index=True)

    st.markdown("### Tabla de soporte OS")
    os_table = filtered[[
        "Commercial Region",
        "Country",
        "Distributor name",
        "Customer name",
        "Instrument type",
        "Serial number",
        "Operating System",
        "Machine Configurations",
    ]].copy()
    st.dataframe(os_table, use_container_width=True, hide_index=True)

if active_dashboard_tab == "Procesamiento / PM":
    st.subheader("Procesamiento, product line y PM planner")
    st.caption("Nueva pestaña para revisar volumen de procesamiento por serie, líneas de producto activas y planeación de mantenimiento preventivo.")

    proc_df = filtered.copy()
    proc_df["Number of tests per day"] = pd.to_numeric(proc_df["Number of tests per day"], errors="coerce")
    proc_df["PM next date display"] = proc_df["PM next date"].map(format_date_for_hover)
    proc_df["PM last date display"] = proc_df["PM last date"].map(format_date_for_hover)
    proc_df["Tests/day display"] = proc_df["Number of tests per day"].map(lambda x: safe_number_text(x, "0"))

    p1, p2, p3, p4 = st.columns(4)
    tests_valid = proc_df["Number of tests per day"].dropna()
    pm_ready = int(proc_df["PM plan"].notna().sum())
    upcoming_pm = int(proc_df["PM next date"].between(pd.Timestamp.today().normalize(), pd.Timestamp.today().normalize() + pd.Timedelta(days=90), inclusive="both").sum()) if proc_df["PM next date"].notna().any() else 0
    product_lines_count = int(proc_df["Product Line"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    with p1:
        metric_card("Tests/día promedio", safe_number_text(tests_valid.mean() if not tests_valid.empty else pd.NA, "0"), "Promedio del filtro")
    with p2:
        metric_card("Tests/día máximos", safe_number_text(tests_valid.max() if not tests_valid.empty else pd.NA, "0"), "Serie con mayor procesamiento")
    with p3:
        metric_card("Product lines", f"{product_lines_count}", "Líneas de producto detectadas")
    with p4:
        metric_card("PM próximos 90 días", f"{upcoming_pm:,}", f"{pm_ready:,} equipos con PM plan")

    g1, g2 = st.columns(2)
    with g1:
        tests_df = proc_df.dropna(subset=["Number of tests per day", "Serial number"]).copy()
        if tests_df.empty:
            st.info("No hay datos válidos en Number of tests per day para el filtro actual.")
        else:
            tests_df = tests_df.sort_values(["Number of tests per day", "Serial number"], ascending=[False, True]).reset_index(drop=True)
            fig_tests = px.scatter(
                tests_df,
                x="Serial number",
                y="Number of tests per day",
                color="Instrument type",
                title="Number of tests/day por cada serie",
                custom_data=["Customer name", "Distributor name", "Country", "Commercial Region", "Product Line", "Operational status"],
            )
            fig_tests.update_traces(
                marker=dict(size=10, opacity=0.9),
                hovertemplate=(
                    "Serie: %{x}<br>"
                    "Tests/día: %{y}<br>"
                    "Instrumento: %{fullData.name}<br>"
                    "Cliente: %{customdata[0]}<br>"
                    "Distribuidor: %{customdata[1]}<br>"
                    "País: %{customdata[2]}<br>"
                    "Región: %{customdata[3]}<br>"
                    "Product line: %{customdata[4]}<br>"
                    "Estado: %{customdata[5]}<extra></extra>"
                )
            )
            zero_tests_total = int(tests_df.shape[0])
            zero_tests_count = int(pd.to_numeric(tests_df["Number of tests per day"], errors="coerce").fillna(np.nan).eq(0).sum())
            zero_tests_pct = _safe_share_pct(zero_tests_count, zero_tests_total)
            zero_tests_text = (
                "<span style='font-size:11px;color:rgba(235,245,255,0.86)'>Tests/día = 0</span><br>"
                f"<b>{zero_tests_count:,}</b> de {zero_tests_total:,}<br>"
                f"<span style='font-size:12px'>{zero_tests_pct:.1f}% del filtro</span>"
            )
            fig_tests.add_annotation(
                text=zero_tests_text,
                x=0.985,
                y=0.875,
                xref="paper",
                yref="paper",
                xanchor="right",
                yanchor="top",
                align="right",
                showarrow=False,
                bordercolor="rgba(113,225,255,0.46)",
                borderwidth=1,
                borderpad=8,
                bgcolor="rgba(8,17,27,0.82)",
                font=dict(color="#f8fcff", size=13),
            )
            fig_tests.update_xaxes(tickangle=-60)
            render_drilldown_plotly_chart(
                glow_layout(fig_tests, 520),
                key="tests_per_day_serial_chart_v30",
                source_label="Number of tests/day por cada serie",
                payload_builder=lambda point: payload_from_serial_axis(point, "Number of tests/day por cada serie", axis="x"),
            )

    with g2:
        product_series = proc_df["Product Line"].fillna("").astype(str).str.strip()
        product_rows = []
        for value in product_series:
            if not value:
                continue
            parts = [p.strip() for p in re.split(r"[\|;,/]", value) if p.strip()]
            if not parts:
                parts = [value]
            product_rows.extend(parts)
        if not product_rows:
            st.info("No hay datos válidos en Product Line para el filtro actual.")
        else:
            product_line_df = pd.Series(product_rows, name="Product line").value_counts().reset_index()
            product_line_df.columns = ["Product line", "Count"]
            product_line_df = product_line_df.head(20)
            fig_product = px.bar(
                product_line_df,
                x="Count",
                y="Product line",
                orientation="h",
                title="Product line performed on the analyzer",
                text="Count",
            )
            fig_product.update_traces(
                marker_color=ACCENT_2,
                textposition="outside",
                hovertemplate="Product line: %{y}<br>Equipos / apariciones: %{x}<extra></extra>",
            )
            fig_product.update_layout(yaxis=dict(categoryorder="total ascending"))
            render_drilldown_plotly_chart(
                glow_layout(fig_product, 520),
                key="product_line_chart_v30",
                source_label="Product line performed on the analyzer",
                payload_builder=payload_from_product_line_point,
            )

    g3, g4 = st.columns(2)
    with g3:
        pm_plan_df = proc_df.copy()
        pm_plan_df["PM Plan label"] = pm_plan_df["PM plan"].fillna("No informado").astype(str)
        pm_summary = pm_plan_df.groupby("PM Plan label", dropna=False).size().reset_index(name="Count").sort_values("Count", ascending=False)
        fig_pm_plan = px.bar(
            pm_summary,
            x="PM Plan label",
            y="Count",
            title="PM planner | distribución de PM plan",
            text="Count",
        )
        fig_pm_plan.update_traces(
            marker_color=ACCENT_3,
            textposition="outside",
            hovertemplate="PM plan: %{x}<br>Equipos: %{y}<extra></extra>",
        )
        fig_pm_plan.update_xaxes(tickangle=-28)
        render_drilldown_plotly_chart(
            glow_layout(fig_pm_plan, 500),
            key="pm_plan_distribution_chart_v30",
            source_label="PM planner | distribución de PM plan",
            payload_builder=lambda point: payload_from_axis_value(point, "PM plan", "PM planner | distribución de PM plan", "PM plan", axis="x"),
        )

    with g4:
        pm_timeline = proc_df.dropna(subset=["PM next date", "Serial number"]).copy()
        if pm_timeline.empty:
            st.info("No hay fechas válidas en PM next date para el filtro actual.")
        else:
            today = pd.Timestamp.today().normalize()
            pm_timeline["PM planner status"] = np.where(
                pm_timeline["PM next date"] < today,
                "Overdue",
                np.where(pm_timeline["PM next date"] <= today + pd.Timedelta(days=90), "Next 90 days", "Planned later"),
            )
            fig_pm_timeline = px.scatter(
                pm_timeline.sort_values("PM next date"),
                x="PM next date",
                y="Serial number",
                color="PM planner status",
                title="PM planner | calendario por serie",
                custom_data=["Instrument type", "Customer name", "Distributor name", "Country", "PM plan", "PM frequency", "PM performed On", "PM last date display", "PM next date display"],
                color_discrete_map={"Overdue": DANGER, "Next 90 days": WARNING, "Planned later": ACCENT},
            )
            fig_pm_timeline.update_traces(
                marker=dict(size=11, opacity=0.9),
                hovertemplate=(
                    "Serie: %{y}<br>"
                    "PM next date: %{customdata[8]}<br>"
                    "Instrumento: %{customdata[0]}<br>"
                    "Cliente: %{customdata[1]}<br>"
                    "Distribuidor: %{customdata[2]}<br>"
                    "País: %{customdata[3]}<br>"
                    "PM plan: %{customdata[4]}<br>"
                    "PM frequency: %{customdata[5]}<br>"
                    "PM performed on: %{customdata[6]}<br>"
                    "PM last date: %{customdata[7]}<br>"
                    "Estado planner: %{fullData.name}<extra></extra>"
                )
            )
            render_drilldown_plotly_chart(
                glow_layout(fig_pm_timeline, 500),
                key="pm_timeline_serial_chart_v30",
                source_label="PM planner | calendario por serie",
                payload_builder=lambda point: payload_from_serial_axis(point, "PM planner | calendario por serie", axis="y"),
            )

    st.markdown("### Tabla de soporte para procesamiento / PM")
    process_table_cols = [
        "Commercial Region",
        "Country",
        "Distributor name",
        "Customer name",
        "Instrument type",
        "Serial number",
        "Number of tests per day",
        "Product Line",
        "PM plan",
        "PM frequency",
        "PM performed On",
        "PM last date",
        "PM next date",
    ]
    st.dataframe(proc_df[process_table_cols].copy(), use_container_width=True, hide_index=True)

if active_dashboard_tab == "Stock / Carstock gap":
    st.subheader("Gap analysis de stock vs carstock requerido")
    st.session_state.setdefault("pdf_stock_context", {"available": False})
    st.caption(
        "Sube el maestro de referencia y luego el archivo trimestral del distribuidor. El dashboard intentará identificar automáticamente el distribuidor a partir del nombre del archivo y hará el análisis de brecha sin guardar histórico."
    )

    default_master_bytes = default_master_candidates[0].read_bytes() if default_master_candidates else None
    default_master_name = default_master_candidates[0].name if default_master_candidates else "No cargado"

    master_upload = st.file_uploader(
        "Archivo maestro de carstock (consolidado o New TP Spare)",
        type=["xlsx", "xls", "csv"],
        key="master_spare_upload",
    )
    stock_upload = st.file_uploader(
        "Archivo de stock reportado por el distribuidor",
        type=["xlsx", "xls", "csv"],
        key="distributor_stock_upload",
    )

    if master_upload is not None:
        master_bytes = master_upload.getvalue()
        master_name = master_upload.name
    elif default_master_bytes is not None:
        master_bytes = default_master_bytes
        master_name = default_master_name
    else:
        master_bytes = None
        master_name = None

    if master_bytes is None:
        st.session_state["pdf_stock_context"] = {"available": False}
        st.info("Sube el archivo maestro de carstock para activar esta pestaña.")
    else:
        master_bundle = load_carstock_master_bundle(master_bytes, master_name)
        has_consolidated = not master_bundle["consolidated"].empty
        has_legacy = bool(master_bundle["legacy_families"])

        if not has_consolidated and not has_legacy:
            st.session_state["pdf_stock_context"] = {"available": False}
            st.warning("No se pudo interpretar el archivo maestro. Debe contener al menos part number y quantity.")
        else:
            st.markdown(f"**Archivo maestro activo:** {master_name}")
            st.markdown(
                "<div class='small-note'>Lógica activa: detección automática del distribuidor desde el nombre del archivo, inferencia de familias según su base instalada y comparación inmediata contra el carstock requerido.</div>",
                unsafe_allow_html=True,
            )

            if stock_upload is None:
                st.session_state["pdf_stock_context"] = {"available": False}
                st.info("Sube ahora el archivo trimestral del distribuidor. Ejemplo recomendado: `ANNAR_stock_Q1_2026.xlsx`.")
            else:
                stock_df_raw = load_table_file(stock_upload.getvalue(), stock_upload.name)
                candidate_distributors = []
                if stock_df_raw is None or stock_df_raw.empty:
                    st.session_state["pdf_stock_context"] = {"available": False}
                    st.warning("El archivo subido no contiene datos legibles.")
                else:
                    part_col_guess, qty_col_guess, desc_col_guess = detect_stock_columns(stock_df_raw)
                    master_distributors = sorted(set(raw_df["Distributor name"].dropna().tolist()) | set(master_bundle["master_distributors"]))
                    detected_distributor, candidate_distributors = infer_distributor_from_filename_strict(stock_upload.name, master_distributors)

                    if detected_distributor is None:
                        st.warning("No fue posible identificar de forma única el distribuidor desde el nombre del archivo.")
                        if candidate_distributors:
                            detected_distributor = st.selectbox(
                                "Selecciona manualmente el distribuidor",
                                options=sorted(candidate_distributors),
                                key="manual_distributor_selection_candidates",
                            )
                        else:
                            detected_distributor = st.selectbox(
                                "Selecciona manualmente el distribuidor",
                                options=sorted(master_distributors),
                                key="manual_distributor_selection_all",
                            )

                    if not detected_distributor:
                        st.session_state["pdf_stock_context"] = {"available": False}
                        st.error("No pude identificar el distribuidor desde el nombre del archivo.")
                        st.caption("Renombra el archivo con un formato claro, por ejemplo: `ANNAR_stock_Q1_2026.xlsx`, `Biotec_del_Paraguay_stock.xlsx` o `Simed_Ecuador_carstock.xlsx`.")
                    else:
                        distributor_scope = raw_df[raw_df["Distributor name"].eq(detected_distributor)].copy()
                        distributor_inst = distributor_scope["Instrument type"].dropna().unique().tolist()
                        family_from_base = infer_families_from_instruments(distributor_inst)

                        if has_consolidated:
                            consolidated = master_bundle["consolidated"].copy()
                            dist_key = normalize_key_text(detected_distributor)
                            dist_specific = consolidated[consolidated["Distributor Key"].eq(dist_key)].copy()
                            if not dist_specific.empty:
                                dist_families = sorted([f for f in dist_specific["Required Family"].dropna().unique().tolist() if f])
                            else:
                                dist_families = []
                            available_families = master_bundle["master_families"] or sorted(consolidated["Required Family"].dropna().unique().tolist())
                        else:
                            dist_specific = pd.DataFrame()
                            dist_families = []
                            available_families = list(master_bundle["legacy_families"].keys())

                        auto_families = [fam for fam in family_from_base if fam in available_families]
                        if dist_families:
                            auto_families = [fam for fam in dist_families if fam in available_families] or auto_families
                        if not auto_families:
                            auto_families = available_families[:1]

                        s1, s2, s3 = st.columns(3)
                        with s1:
                            metric_card("Distribuidor detectado", detected_distributor, "Detectado automáticamente o seleccionado manualmente")
                        with s2:
                            metric_card("Familias inferidas", ", ".join(auto_families) if auto_families else "N/A", "Según base instalada")
                        with s3:
                            metric_card("Instrumentos del distribuidor", f"{len(distributor_inst):,}", "Tipos detectados en la base")

                        info_lines = []
                        if distributor_inst:
                            info_lines.append("Instrumentos en base: " + ", ".join(sorted(distributor_inst)))
                        if candidate_distributors and len(candidate_distributors) > 1:
                            info_lines.append("Coincidencias secundarias detectadas en el nombre: " + ", ".join(candidate_distributors[1:4]))
                        if info_lines:
                            st.caption(" | ".join(info_lines))

                        with st.expander("Ajustes avanzados de lectura del archivo", expanded=False):
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                part_col = st.selectbox(
                                    "Columna de part number",
                                    options=stock_df_raw.columns.tolist(),
                                    index=stock_df_raw.columns.tolist().index(part_col_guess) if part_col_guess in stock_df_raw.columns else 0,
                                )
                            with col2:
                                qty_col = st.selectbox(
                                    "Columna de cantidad",
                                    options=stock_df_raw.columns.tolist(),
                                    index=stock_df_raw.columns.tolist().index(qty_col_guess) if qty_col_guess in stock_df_raw.columns else min(1, len(stock_df_raw.columns) - 1),
                                )
                            with col3:
                                desc_options = ["<sin descripción>"] + stock_df_raw.columns.tolist()
                                default_desc_index = desc_options.index(desc_col_guess) if desc_col_guess in stock_df_raw.columns else 0
                                desc_selection = st.selectbox("Columna de descripción", options=desc_options, index=default_desc_index)
                                desc_col = None if desc_selection == "<sin descripción>" else desc_selection

                            all_master_families = sorted(set([f for f in available_families if str(f).strip()]))
                            safe_auto_families = [fam for fam in auto_families if fam in all_master_families]
                            if not safe_auto_families and all_master_families:
                                safe_auto_families = all_master_families[:1]

                            family_mode = st.radio(
                                "Modo de selección de familias",
                                options=["Automático", "Manual"],
                                horizontal=True,
                                key="stock_family_mode_selector",
                                help="Automático usa las familias inferidas para el distribuidor. Manual te deja escoger cualquier familia disponible en el maestro.",
                            )

                            manual_family_key = "stock_family_selector_manual"
                            previous_manual = st.session_state.get(manual_family_key, safe_auto_families.copy())
                            if not isinstance(previous_manual, list):
                                previous_manual = safe_auto_families.copy()
                            previous_manual = [fam for fam in previous_manual if fam in all_master_families]
                            if not previous_manual and safe_auto_families:
                                previous_manual = safe_auto_families.copy()
                            st.session_state[manual_family_key] = previous_manual

                            selected_families_manual = st.multiselect(
                                "Familias a comparar",
                                options=all_master_families,
                                default=previous_manual,
                                key=manual_family_key,
                                placeholder="Selecciona una o varias familias",
                                help="En modo manual puedes elegir libremente las familias del maestro, aunque no hayan sido inferidas automáticamente.",
                            )

                            if family_mode == "Automático":
                                selected_families_stock = safe_auto_families.copy()
                                st.caption("Modo automático activo: se usarán las familias inferidas para el distribuidor. Cambia a Manual si quieres forzarlas tú mismo.")
                            else:
                                selected_families_stock = [fam for fam in selected_families_manual if fam in all_master_families]
                                st.caption("Modo manual activo: las familias seleccionadas aquí se respetarán exactamente en la comparación.")

                        if 'part_col' not in locals():
                            part_col = part_col_guess
                            qty_col = qty_col_guess
                            desc_col = desc_col_guess
                            selected_families_stock = auto_families

                        po_frequency_weeks = st.number_input(
                            "PO Frequency (weeks)",
                            min_value=1,
                            max_value=52,
                            value=12,
                            step=1,
                            key="po_frequency_weeks_input",
                            help="Cantidad de semanas que el distribuidor normalmente espera entre pedidos de repuestos.",
                        )


                        if not selected_families_stock:
                            st.session_state["pdf_stock_context"] = {"available": False}
                            st.warning("No hay familias seleccionadas para comparar. Ajusta el maestro o la selección avanzada.")
                        else:
                            master_df, master_mode = build_required_master_from_scope(
                                master_bundle=master_bundle,
                                assigned_distributor=detected_distributor,
                                selected_families=selected_families_stock,
                            )

                            if master_df.empty and detected_distributor:
                                master_df, master_mode = build_required_master_from_scope(
                                    master_bundle=master_bundle,
                                    assigned_distributor="<sin asignar>",
                                    selected_families=selected_families_stock,
                                )

                            if master_df.empty:
                                st.session_state["pdf_stock_context"] = {"available": False}
                                st.warning("No encontré carstock requerido para este distribuidor con las familias inferidas. Revisa el maestro o el nombre del archivo.")
                            else:
                                installed_base_by_family = compute_installed_base_by_family(raw_df, detected_distributor)
                                if master_mode == "advanced":
                                    master_df = apply_dynamic_required_qty(
                                        master_df,
                                        installed_base_by_family=installed_base_by_family,
                                        po_frequency_weeks=po_frequency_weeks,
                                    )

                                comparison, extra_df, stock_slim = compare_stock(
                                    master_df,
                                    stock_df_raw,
                                    part_col,
                                    qty_col,
                                    desc_col,
                                    price_reference=master_bundle.get("price_reference", pd.DataFrame()),
                                )
                                missing_skus = int((comparison["Status"] == "Missing").sum())
                                low_skus = int((comparison["Status"] == "LOW").sum())
                                covered_skus = int((comparison["Status"] == "OK").sum())
                                total_gap = float(comparison["Qty Gap"].sum())
                                coverage = (covered_skus / len(comparison) * 100) if len(comparison) else 0
                                extra_skus = int(len(extra_df))
                                option2_cost = float(pd.to_numeric(comparison["Option 2 Estimated Cost"], errors="coerce").fillna(0).sum())
                                option2_currency = next((c for c in comparison["Currency"].dropna().astype(str).tolist() if c.strip()), "EUR")
                                purchase_df = comparison[comparison["Qty Gap"] > 0][[
                                    "Required Part Number",
                                    "Required Description",
                                    "Qty Gap",
                                    "Option 2 Unit Price",
                                    "Option 2 Estimated Cost",
                                    "Currency",
                                    "Status",
                                ]].sort_values(["Option 2 Estimated Cost", "Qty Gap"], ascending=[False, False])

                                st.session_state["pdf_stock_context"] = {
                                    "available": True,
                                    "detected_distributor": detected_distributor,
                                    "families": selected_families_stock,
                                    "required_skus": len(comparison),
                                    "ok_skus": covered_skus,
                                    "low_skus": low_skus,
                                    "missing_skus": missing_skus,
                                    "extra_skus": extra_skus,
                                    "gap_total": total_gap,
                                    "option2_cost": option2_cost,
                                    "currency": option2_currency,
                                    "po_frequency_weeks": int(po_frequency_weeks),
                                    "installed_base_by_family": installed_base_by_family,
                                    "top_gap_df": comparison[comparison["Qty Gap"] > 0].sort_values(["Qty Gap", "Required Part Number"], ascending=[False, True]).head(15).copy(),
                                    "full_comparison_df": comparison.copy(),
                                    "purchase_df": purchase_df.copy() if not purchase_df.empty else pd.DataFrame(columns=["Required Part Number", "Required Description", "Qty Gap", "Option 2 Unit Price", "Option 2 Estimated Cost", "Currency", "Status"]),
                                    "extra_df": extra_df.copy() if not extra_df.empty else pd.DataFrame(columns=["Uploaded Part Number", "Uploaded Description", "Uploaded Qty", "Status"]),
                                }

                                total_installed_scope = int(sum(installed_base_by_family.get(fam, 0) for fam in selected_families_stock))
                                sm1, sm2, sm3, sm4, sm5, sm6, sm7, sm8 = st.columns(8)
                                with sm1:
                                    metric_card("SKUs requeridos", f"{len(comparison):,}", "Carstock esperado")
                                with sm2:
                                    metric_card("SKUs OK", f"{covered_skus:,}", f"{coverage:.1f}% del carstock")
                                with sm3:
                                    metric_card("SKUs LOW", f"{low_skus:,}", "Tienen stock pero insuficiente")
                                with sm4:
                                    metric_card("SKUs Missing", f"{missing_skus:,}", "Sin stock reportado")
                                with sm5:
                                    metric_card("Gap total qty", safe_number_text(total_gap, "0"), "Cantidad faltante acumulada")
                                with sm6:
                                    metric_card("Compra opción 2", f"{option2_currency} {option2_cost:,.2f}", "Costo estimado para cubrir gap")
                                with sm7:
                                    metric_card("PO frequency", f"{int(po_frequency_weeks)} sem", f"Base instalada: {total_installed_scope:,}")
                                with sm8:
                                    metric_card("Extras", f"{extra_skus:,}", "No incluidos en el maestro")

                                master_label = {
                                    "advanced": "Carstock dinámico",
                                    "consolidated": "Consolidado",
                                    "legacy": "Estándar por familia",
                                }.get(master_mode, master_mode)
                                st.caption(
                                    f"Alcance automático: {detected_distributor} | Familias: {', '.join(selected_families_stock)} | Maestro: {master_label} | PO frequency: {int(po_frequency_weeks)} semanas"
                                )

                                g1, g2 = st.columns(2)
                                with g1:
                                    status_df = comparison["Status"].value_counts().reset_index()
                                    status_df.columns = ["Status", "Count"]
                                    color_map = {"OK": ACCENT_3, "LOW": WARNING, "Missing": DANGER}
                                    fig_status = px.pie(status_df, names="Status", values="Count", title="Cobertura del carstock", hole=0.52)
                                    fig_status.update_traces(
                                        marker=dict(colors=[color_map.get(s, ACCENT) for s in status_df["Status"]]),
                                        hovertemplate="Estado: %{label}<br>SKUs: %{value}<br>%{percent}<extra></extra>",
                                    )
                                    fig_status.update_layout(template=PLOT_TEMPLATE, paper_bgcolor=PLOT_BG, plot_bgcolor=PLOT_BG, font=dict(color=TEXT), height=430)
                                    st.caption("Aplicabilidad del filtro: esta gráfica pertenece al análisis de carstock/SKUs, no a filas de equipos. No filtra la base instalada; usa la tabla de brechas para revisar los SKUs.")
                                    st.plotly_chart(fig_status, use_container_width=True)

                                with g2:
                                    gap_df = comparison[comparison["Qty Gap"] > 0].sort_values(["Qty Gap", "Required Part Number"], ascending=[False, True]).head(15)
                                    if gap_df.empty:
                                        st.success("No se detectan faltantes para el distribuidor identificado.")
                                    else:
                                        fig_gap = px.bar(
                                            gap_df,
                                            x="Qty Gap",
                                            y="Gap Label",
                                            orientation="h",
                                            title="Top faltantes por cantidad",
                                            custom_data=["Required Description", "Uploaded Qty", "Required Qty", "Coverage %", "Status", "Option 2 Unit Price", "Option 2 Estimated Cost", "Currency"],
                                            text="Qty Gap",
                                        )
                                        fig_gap.update_traces(
                                            marker_color=DANGER,
                                            textposition="outside",
                                            hovertemplate=(
                                                "Parte: %{y}<br>"
                                                "Descripción: %{customdata[0]}<br>"
                                                "Qty reportada: %{customdata[1]}<br>"
                                                "Qty requerida: %{customdata[2]}<br>"
                                                "Cobertura: %{customdata[3]}%<br>"
                                                "Estado: %{customdata[4]}<br>"
                                                "Precio opción 2: %{customdata[7]} %{customdata[5]:,.2f}<br>"
                                                "Costo compra opción 2: %{customdata[7]} %{customdata[6]:,.2f}<br>"
                                                "Gap: %{x}<extra></extra>"
                                            ),
                                        )
                                        fig_gap.update_layout(yaxis=dict(categoryorder="total ascending"))
                                        st.caption("Aplicabilidad del filtro: esta gráfica muestra brechas de repuestos por SKU. No filtra equipos; se mantiene como análisis de compra/carstock.")
                                        st.plotly_chart(glow_layout(fig_gap, 430), use_container_width=True)

                                st.markdown("### Tabla de brechas")
                                show_cols = ["Required Part Number", "Required Description", "Required Qty", "Uploaded Qty", "Qty Gap", "Coverage %", "Option 2 Unit Price", "Option 2 Estimated Cost", "Currency", "Status"]
                                comparison_display = comparison[show_cols].copy()
                                comparison_display["Required Description"] = comparison_display["Required Description"].fillna("No disponible").replace("", "No disponible")
                                comparison_display["Currency"] = comparison_display["Currency"].fillna("EUR").replace("", "EUR")
                                for num_col in ["Required Qty", "Uploaded Qty", "Qty Gap", "Coverage %", "Option 2 Unit Price", "Option 2 Estimated Cost"]:
                                    comparison_display[num_col] = pd.to_numeric(comparison_display[num_col], errors="coerce").fillna(0)
                                st.dataframe(comparison_display, use_container_width=True, hide_index=True)

                                if not purchase_df.empty:
                                    st.markdown("### Compra sugerida para cerrar el gap (opción 2)")
                                    st.dataframe(purchase_df, use_container_width=True, hide_index=True)
                                    st.markdown(f"**Total estimado de compra sugerida:** {option2_currency} {option2_cost:,.2f}")

                                if not extra_df.empty:
                                    st.markdown("### Partes reportadas que no están en el carstock requerido")
                                    st.dataframe(
                                        extra_df[["Uploaded Part Number", "Uploaded Description", "Uploaded Qty", "Status"]],
                                        use_container_width=True,
                                        hide_index=True,
                                    )

                                export_df = comparison[show_cols].copy()
                                if not extra_df.empty:
                                    extras_export = extra_df.rename(
                                        columns={
                                            "Uploaded Part Number": "Required Part Number",
                                            "Uploaded Description": "Required Description",
                                            "Uploaded Qty": "Uploaded Qty",
                                        }
                                    )
                                    extras_export["Required Qty"] = 0
                                    extras_export["Qty Gap"] = 0
                                    extras_export["Coverage %"] = 0
                                    extras_export["Option 2 Unit Price"] = pd.NA
                                    extras_export["Option 2 Estimated Cost"] = pd.NA
                                    extras_export["Currency"] = "EUR"
                                    export_df = pd.concat(
                                        [
                                            export_df,
                                            extras_export[["Required Part Number", "Required Description", "Required Qty", "Uploaded Qty", "Qty Gap", "Coverage %", "Option 2 Unit Price", "Option 2 Estimated Cost", "Currency", "Status"]],
                                        ],
                                        ignore_index=True,
                                    )

                                export_df = export_df.sort_values(["Status", "Qty Gap", "Required Part Number"], ascending=[True, False, True], na_position="last").reset_index(drop=True)
                                purchase_export = purchase_df.reset_index(drop=True) if not purchase_df.empty else pd.DataFrame(columns=["Required Part Number", "Required Description", "Qty Gap", "Option 2 Unit Price", "Option 2 Estimated Cost", "Currency", "Status"])
                                if not purchase_export.empty:
                                    purchase_total_row = {col: "" for col in purchase_export.columns}
                                    if "Required Part Number" in purchase_total_row:
                                        purchase_total_row["Required Part Number"] = "TOTAL"
                                    if "Option 2 Estimated Cost" in purchase_total_row:
                                        purchase_total_row["Option 2 Estimated Cost"] = round(option2_cost, 2)
                                    if "Currency" in purchase_total_row:
                                        purchase_total_row["Currency"] = option2_currency
                                    purchase_export = pd.concat([purchase_export, pd.DataFrame([purchase_total_row])], ignore_index=True)
                                if not extra_df.empty:
                                    extras_export_final = extra_df[["Uploaded Part Number", "Uploaded Description", "Uploaded Qty", "Status"]].copy()
                                    extras_export_final["__sort_uploaded_part_number"] = extras_export_final["Uploaded Part Number"].astype("string").fillna("")
                                    extras_export_final = extras_export_final.sort_values(["__sort_uploaded_part_number"], ascending=[True], na_position="last").drop(columns=["__sort_uploaded_part_number"]).reset_index(drop=True)
                                else:
                                    extras_export_final = pd.DataFrame(columns=["Uploaded Part Number", "Uploaded Description", "Uploaded Qty", "Status"])

                                excel_bytes = dataframe_to_excel_bytes({
                                    "Gap analysis": export_df,
                                    "Purchase option 2": purchase_export,
                                    "Extras not required": extras_export_final,
                                })

                                st.download_button(
                                    "Descargar análisis de faltantes",
                                    data=excel_bytes,
                                    file_name=f"carstock_gap_{normalize_key_text(detected_distributor) or 'distribuidor'}_{'_'.join(selected_families_stock) or 'familia'}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )

                                with st.expander("Guía para el archivo maestro consolidado"):
                                    st.markdown(
                                        """
                                        Formato recomendado para el archivo maestro único:
                                        - `Distributor name`
                                        - `Instrument family` o `Platform` (`LXL`, `LXS`, `MDX`, `EMX`)
                                        - `Part Number`
                                        - `Description`
                                        - `Required Qty`

                                        El archivo puede estar en una sola hoja o varias hojas. El dashboard intentará reconocer sinónimos de estas columnas automáticamente.
                                        """
                                    )


if active_dashboard_tab == "Antigüedad / fabricación":
    st.subheader("Antigüedad de la base instalada por fecha de fabricación")
    st.caption(
        "Esta pestaña cruza el serial de la vista filtrada actual con los archivos de fabricación que cargues manualmente. "
        "Los filtros globales de región, país, distribuidor, tipo de instrumento y estado operativo se aplican automáticamente."
    )

    manufacturing_uploads = st.file_uploader(
        "Carga manualmente los archivos actualizados de fabricación",
        type=["xls", "xlsx", "csv"],
        accept_multiple_files=True,
        key="manufacturing_date_uploads",
        help="Puedes cargar simultáneamente los archivos LXL/LXL LAS, LXS y ETI-MAX. No quedan almacenados al cerrar la sesión.",
    )

    if not manufacturing_uploads:
        st.info(
            "Carga los archivos de fechas de fabricación para activar el cruce por serial y visualizar la edad real de los equipos."
        )
    else:
        prepared_sources = []
        reading_errors = []

        st.markdown("### Configuración de las fuentes cargadas")
        st.markdown(
            '<div class="small-note">Selecciona la hoja, el serial y la columna que corporativamente corresponde a la fecha de fabricación. '
            'Esta confirmación evita interpretar automáticamente <b>PO Date</b> o <b>Invoice Date</b> como fabricación cuando no corresponda.</div>',
            unsafe_allow_html=True,
        )

        for source_idx, uploaded_manufacturing in enumerate(manufacturing_uploads):
            try:
                source_sheets = load_manufacturing_source(uploaded_manufacturing.getvalue(), uploaded_manufacturing.name)
                sheet_names = list(source_sheets.keys())
                if not sheet_names:
                    reading_errors.append(f"{uploaded_manufacturing.name}: no contiene hojas con datos.")
                    continue

                sheet_scores = []
                for sheet_name in sheet_names:
                    frame = source_sheets[sheet_name]
                    score = 0
                    normalized_cols = [normalize_column_label(c) for c in frame.columns]
                    if any("serial" in c for c in normalized_cols):
                        score += 10
                    if any("date" in c or "fecha" in c for c in normalized_cols):
                        score += 5
                    score += min(len(frame), 1000) / 10000
                    sheet_scores.append((score, sheet_name))
                suggested_sheet = sorted(sheet_scores, reverse=True)[0][1]

                with st.expander(f"Fuente: {uploaded_manufacturing.name}", expanded=True):
                    selected_sheet = st.selectbox(
                        "Hoja con información de seriales",
                        options=sheet_names,
                        index=sheet_names.index(suggested_sheet),
                        key=f"manufacturing_sheet_{source_idx}",
                    )
                    source_frame = source_sheets[selected_sheet].copy()
                    columns_available = source_frame.columns.tolist()

                    if source_frame.empty or not columns_available:
                        st.warning("La hoja seleccionada no contiene registros.")
                        continue

                    serial_default = guess_column_index(
                        columns_available,
                        ["Serial No", "Serial number", "Serial", "Número de serie", "Serie"],
                    )
                    date_options = manufacturing_date_candidate_columns(source_frame)
                    date_default = guess_column_index(
                        date_options,
                        ["Manufacturing Date", "Manufacture Date", "Production Date", "MFG Date", "Fecha de fabricación", "PO Date", "Invoice Date"],
                    )
                    product_default = 0
                    product_terms = ["Product", "Product description", "Material", "Instrument", "Description", "Model"]
                    normalized_product_columns = [normalize_column_label(col) for col in columns_available]
                    for term in product_terms:
                        term_normalized = normalize_column_label(term)
                        matched_product_columns = [
                            idx for idx, column_name in enumerate(normalized_product_columns)
                            if column_name == term_normalized or term_normalized in column_name
                        ]
                        if matched_product_columns:
                            product_default = matched_product_columns[0] + 1
                            break

                    c_source_1, c_source_2, c_source_3 = st.columns(3)
                    with c_source_1:
                        serial_column = st.selectbox(
                            "Columna de serial",
                            options=columns_available,
                            index=serial_default,
                            key=f"manufacturing_serial_col_{source_idx}",
                        )
                    with c_source_2:
                        date_column = st.selectbox(
                            "Columna de fecha de fabricación",
                            options=date_options,
                            index=date_default if date_default < len(date_options) else 0,
                            key=f"manufacturing_date_col_{source_idx}",
                        )
                    with c_source_3:
                        product_choices = ["<sin columna de producto>"] + columns_available
                        selected_product = st.selectbox(
                            "Columna de producto / modelo (opcional)",
                            options=product_choices,
                            index=product_default,
                            key=f"manufacturing_product_col_{source_idx}",
                        )

                    prepared = pd.DataFrame(
                        {
                            "Manufacturing Serial": source_frame[serial_column],
                            "Manufacturing Date": parse_date_flexible(source_frame[date_column]),
                            "Manufacturing Source": uploaded_manufacturing.name,
                            "Manufacturing Sheet": selected_sheet,
                            "Manufacturing Product": (
                                source_frame[selected_product]
                                if selected_product != "<sin columna de producto>"
                                else pd.Series(pd.NA, index=source_frame.index)
                            ),
                        }
                    )
                    prepared["Serial match key"] = prepared["Manufacturing Serial"].map(normalize_serial_match)
                    valid_rows = int(((prepared["Serial match key"] != "") & prepared["Manufacturing Date"].notna()).sum())
                    st.caption(
                        f"Registros válidos para cruce: {valid_rows:,} de {len(prepared):,} | "
                        f"Fecha utilizada: {date_column}"
                    )
                    if valid_rows:
                        prepared_sources.append(prepared)
            except Exception as exc:
                reading_errors.append(f"{uploaded_manufacturing.name}: {exc}")

        if reading_errors:
            for message in reading_errors:
                st.warning(message)

        if not prepared_sources:
            st.info("Aún no hay una fuente válida con serial y fecha para realizar la comparación.")
        else:
            manufacturing_df, manufacturing_reference = build_manufacturing_match(filtered, prepared_sources)
            # Mantener disponible la vista enriquecida de fabricación para el botón
            # global de descarga Excel al final del dashboard. Así, cuando la pestaña
            # activa es Antigüedad / fabricación, el descargable incluye exactamente
            # los campos calculados en esta pestaña y no solo el Records List base.
            st.session_state[MANUFACTURING_EXCEL_EXPORT_SESSION_KEY] = manufacturing_df.drop(columns=["Serial match key"], errors="ignore").copy()
            st.session_state[MANUFACTURING_EXCEL_EXPORT_SOURCE_KEY] = source_label
            matched_df = manufacturing_df[manufacturing_df["Manufacturing matched"]].copy()
            unmatched_df = manufacturing_df[~manufacturing_df["Manufacturing matched"]].copy()
            future_date_count = int((matched_df["Manufacturing Date"] > pd.Timestamp(date.today())).sum()) if not matched_df.empty else 0
            conflicting_date_count = int(matched_df.get("Manufacturing date conflict", pd.Series(dtype=bool)).fillna(False).sum()) if not matched_df.empty else 0
            match_rate = _safe_share_pct(len(matched_df), len(manufacturing_df))

            st.markdown("### Cobertura y edad del parque filtrado")
            a1, a2, a3, a4, a5 = st.columns(5)
            with a1:
                metric_card("Equipos filtrados", f"{len(manufacturing_df):,}", "Vista actual del dashboard")
            with a2:
                metric_card("Seriales cruzados", f"{len(matched_df):,}", f"{match_rate:.1f}% con fecha de fabricación")
            with a3:
                average_age = matched_df["Manufacturing age (years)"].mean() if not matched_df.empty else pd.NA
                metric_card("Edad promedio", f"{average_age:.1f} años" if pd.notna(average_age) else "N/A", "Según fecha seleccionada")
            with a4:
                oldest_age = matched_df["Manufacturing age (years)"].max() if not matched_df.empty else pd.NA
                metric_card("Equipo más antiguo", f"{oldest_age:.1f} años" if pd.notna(oldest_age) else "N/A", "Mayor edad identificada")
            with a5:
                newest_age = matched_df["Manufacturing age (years)"].min() if not matched_df.empty else pd.NA
                metric_card("Equipo más nuevo", f"{newest_age:.1f} años" if pd.notna(newest_age) else "N/A", "Menor edad identificada")

            if future_date_count > 0:
                st.warning(f"Se detectaron {future_date_count} equipos con fecha de fabricación futura; revisa la columna de fecha seleccionada.")
            if conflicting_date_count > 0:
                st.warning(f"Se detectaron {conflicting_date_count} seriales con más de una fecha distinta en los archivos cargados.")

            if matched_df.empty:
                st.warning("No se encontraron coincidencias por serial entre la base filtrada y las fuentes de fabricación.")
            else:
                matched_df["Age label"] = (
                    matched_df["Serial number"].fillna("SIN SERIAL").astype(str)
                    + " | "
                    + matched_df["Instrument type"].fillna("Sin modelo").astype(str)
                )
                matched_df["Manufacturing date display"] = matched_df["Manufacturing Date"].map(format_date_for_hover)
                oldest = matched_df.sort_values(["Manufacturing age (years)", "Serial number"], ascending=[False, True]).head(15)
                annual = (
                    matched_df.groupby("Manufacturing year", dropna=False)
                    .size()
                    .reset_index(name="Count")
                    .dropna(subset=["Manufacturing year"])
                    .sort_values("Manufacturing year")
                )
                age_distribution = (
                    matched_df["Manufacturing age bucket"]
                    .value_counts(sort=False)
                    .reset_index()
                )
                age_distribution.columns = ["Rango de edad", "Count"]
                age_distribution = age_distribution[age_distribution["Count"] > 0]

                chart_left, chart_right = st.columns(2)
                with chart_left:
                    fig_age_distribution = px.bar(
                        age_distribution,
                        x="Rango de edad",
                        y="Count",
                        text="Count",
                        title="Estado de la base instalada por rango de edad",
                        category_orders={"Rango de edad": ["0–3 años", "3–5 años", "5–8 años", "8–10 años", "10–15 años", "15+ años"]},
                    )
                    fig_age_distribution.update_traces(
                        marker_color=ACCENT,
                        textposition="outside",
                        hovertemplate="Rango: %{x}<br>Equipos: %{y}<extra></extra>",
                    )
                    render_drilldown_plotly_chart(
                        glow_layout(fig_age_distribution, 475, title_size=16),
                        key="manufacturing_age_bucket_chart_v30",
                        source_label="Estado de la base instalada por rango de edad",
                        payload_builder=payload_from_manufacturing_age_bucket,
                        help_text="Filtro disponible si la vista ya contiene datos de fabricación cargados; selecciona un rango para aplicar filtro por rango de edad."
                    )

                with chart_right:
                    fig_oldest = px.bar(
                        oldest.sort_values("Manufacturing age (years)", ascending=True),
                        x="Manufacturing age (years)",
                        y="Age label",
                        orientation="h",
                        text="Manufacturing age (years)",
                        title="Top 15 equipos más antiguos",
                        custom_data=["Manufacturing date display", "Customer name", "Distributor name", "Country", "Operational status"],
                    )
                    fig_oldest.update_traces(
                        marker_color=WARNING,
                        texttemplate="%{text:.1f} años",
                        textposition="outside",
                        hovertemplate=(
                            "Equipo: %{y}<br>"
                            "Edad: %{x:.1f} años<br>"
                            "Fabricación: %{customdata[0]}<br>"
                            "Cliente: %{customdata[1]}<br>"
                            "Distribuidor: %{customdata[2]}<br>"
                            "País: %{customdata[3]}<br>"
                            "Estado: %{customdata[4]}<extra></extra>"
                        ),
                    )
                    render_drilldown_plotly_chart(
                        glow_layout(fig_oldest, 475, title_size=16),
                        key="manufacturing_oldest_serial_chart_v30",
                        source_label="Top 15 equipos más antiguos",
                        payload_builder=lambda point: payload_from_serial_axis(point, "Top 15 equipos más antiguos", axis="y"),
                    )

                timeline_left, timeline_right = st.columns(2)
                with timeline_left:
                    fig_timeline = px.scatter(
                        matched_df.sort_values("Manufacturing Date"),
                        x="Manufacturing Date",
                        y="Serial number",
                        color="Manufacturing age bucket",
                        title="Línea de tiempo de fabricación por serial",
                        custom_data=["Instrument type", "Manufacturing age (years)", "Customer name", "Distributor name", "Country", "Operational status", "Manufacturing Source"],
                        color_discrete_map={
                            "0–3 años": ACCENT_3,
                            "3–5 años": ACCENT,
                            "5–8 años": ACCENT_2,
                            "8–10 años": WARNING,
                            "10–15 años": "#ff8b55",
                            "15+ años": DANGER,
                        },
                    )
                    fig_timeline.update_traces(
                        marker=dict(size=10, opacity=0.90),
                        hovertemplate=(
                            "Serial: %{y}<br>"
                            "Fabricación: %{x|%Y-%m-%d}<br>"
                            "Modelo: %{customdata[0]}<br>"
                            "Edad: %{customdata[1]:.1f} años<br>"
                            "Cliente: %{customdata[2]}<br>"
                            "Distribuidor: %{customdata[3]}<br>"
                            "País: %{customdata[4]}<br>"
                            "Estado: %{customdata[5]}<br>"
                            "Fuente: %{customdata[6]}<extra></extra>"
                        ),
                    )
                    fig_timeline.update_layout(legend_title="Rango de edad")
                    render_drilldown_plotly_chart(
                        glow_layout(fig_timeline, 600, title_size=16),
                        key="manufacturing_timeline_serial_chart_v30",
                        source_label="Línea de tiempo de fabricación por serial",
                        payload_builder=lambda point: payload_from_serial_axis(point, "Línea de tiempo de fabricación por serial", axis="y"),
                    )

                with timeline_right:
                    fig_annual = px.bar(
                        annual,
                        x="Manufacturing year",
                        y="Count",
                        text="Count",
                        title="Equipos por año de fabricación",
                    )
                    fig_annual.update_traces(
                        marker_color=ACCENT_2,
                        textposition="outside",
                        hovertemplate="Año de fabricación: %{x}<br>Equipos: %{y}<extra></extra>",
                    )
                    render_drilldown_plotly_chart(
                        glow_layout(fig_annual, 600, title_size=16),
                        key="manufacturing_year_chart_v30",
                        source_label="Equipos por año de fabricación",
                        payload_builder=payload_from_manufacturing_year,
                        help_text="Filtro disponible si la vista ya contiene datos de fabricación cargados; selecciona un año para aplicar filtro por año de fabricación."
                    )

                st.markdown("### Equipos ordenados del más antiguo al más nuevo")
                age_table_columns = [
                    "Commercial Region",
                    "Country",
                    "Distributor name",
                    "Customer name",
                    "Instrument type",
                    "Serial number",
                    "Manufacturing Date",
                    "Manufacturing age (years)",
                    "Manufacturing age bucket",
                    "Installation date",
                    "Operational status",
                    "Manufacturing Source",
                ]
                age_table = matched_df[age_table_columns].sort_values(
                    ["Manufacturing age (years)", "Manufacturing Date"],
                    ascending=[False, True],
                )
                st.dataframe(age_table, use_container_width=True, hide_index=True)

                manufacturing_export_df = manufacturing_df.drop(columns=["Serial match key"], errors="ignore")
                manufacturing_export_signature = _excel_export_signature(
                    manufacturing_export_df,
                    filter_summary_for_panel,
                    "Antigüedad / fabricación",
                    source_label,
                )

                if st.button(
                    "📗 Preparar cruce de fabricación en Excel",
                    use_container_width=False,
                    key="prepare_manufacturing_age_analysis_excel_v49",
                ):
                    try:
                        manufacturing_excel_bytes = build_dashboard_excel_export(
                            manufacturing_export_df,
                            filter_summary_for_panel,
                            active_dashboard_tab="Antigüedad / fabricación",
                            source_label_value=source_label,
                            stock_context=st.session_state.get("pdf_stock_context", {"available": False}),
                            include_visual_dashboard=False,
                        )
                        st.session_state["prepared_manufacturing_excel_bytes"] = manufacturing_excel_bytes
                        st.session_state["prepared_manufacturing_excel_signature"] = manufacturing_export_signature
                        st.session_state["prepared_manufacturing_excel_name"] = (
                            f"installed_base_manufacturing_age_filtered_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                        )
                        st.success("Cruce de fabricación preparado correctamente.")
                    except Exception as e:
                        st.session_state.pop("prepared_manufacturing_excel_bytes", None)
                        st.session_state.pop("prepared_manufacturing_excel_signature", None)
                        st.error(f"No fue posible generar el Excel de fabricación: {e}")

                manufacturing_excel_is_current = (
                    st.session_state.get("prepared_manufacturing_excel_bytes") is not None
                    and st.session_state.get("prepared_manufacturing_excel_signature") == manufacturing_export_signature
                )
                if manufacturing_excel_is_current:
                    st.download_button(
                        "⬇️ Descargar cruce de fechas de fabricación en Excel",
                        data=st.session_state["prepared_manufacturing_excel_bytes"],
                        file_name=st.session_state.get(
                            "prepared_manufacturing_excel_name",
                            f"installed_base_manufacturing_age_filtered_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        ),
                        mime=EXCEL_MIME,
                        use_container_width=False,
                        key="download_manufacturing_age_analysis_excel_v49",
                    )
                elif st.session_state.get("prepared_manufacturing_excel_bytes") is not None:
                    st.info("La vista de fabricación cambió. Prepara nuevamente el archivo antes de descargarlo.")

            if not unmatched_df.empty:
                with st.expander(f"Seriales sin coincidencia en los archivos de fabricación ({len(unmatched_df):,})", expanded=False):
                    missing_columns = [
                        "Country",
                        "Distributor name",
                        "Customer name",
                        "Instrument type",
                        "Serial number",
                        "Operational status",
                    ]
                    st.dataframe(unmatched_df[missing_columns], use_container_width=True, hide_index=True)

if active_dashboard_tab == "Detalle por equipo":
    st.subheader("Detalle por equipo")
    st.caption(
        "Búsqueda ampliada por serial, hospital/cliente, ciudad, país, distribuidor o modelo. "
        "La vista muestra toda la información disponible del registro seleccionado, incluyendo mantenimiento preventivo, antigüedad, configuración y campos técnicos."
    )

    detail_df = filtered.copy().reset_index(drop=False).rename(columns={"index": "_source_index"})

    searchable_columns = [
        "Serial number",
        "Customer name",
        "City",
        "Country",
        "Distributor name",
        "Instrument type",
        "Address",
        "Operational status",
    ]
    for col in searchable_columns:
        if col not in detail_df.columns:
            detail_df[col] = pd.NA

    detail_df["selector"] = (
        detail_df["Serial number"].fillna("SIN SERIAL").astype(str)
        + " | "
        + detail_df["Customer name"].fillna("SIN CLIENTE / HOSPITAL").astype(str)
        + " | "
        + detail_df["City"].fillna("SIN CIUDAD").astype(str)
        + " | "
        + detail_df["Country"].fillna("SIN PAÍS").astype(str)
    )

    search_text = st.text_input(
        "Buscar equipo",
        value="",
        placeholder="Escribe serial, hospital/cliente, ciudad, país, distribuidor o modelo",
        key="detail_equipment_search",
    ).strip()

    if search_text:
        search_norm = normalize_search_text(search_text)
        search_blob = (
            detail_df[searchable_columns]
            .fillna("")
            .astype(str)
            .agg(" | ".join, axis=1)
            .map(normalize_search_text)
        )
        matched_mask = search_blob.str.contains(re.escape(search_norm), na=False)
        detail_options = detail_df.loc[matched_mask, "selector"].tolist()

        if not detail_options:
            st.warning("No encontré equipos con ese criterio dentro del filtro actual. Revisa si los filtros globales están limitando la búsqueda.")
            detail_options = detail_df["selector"].tolist()
        else:
            st.success(f"Coincidencias encontradas: {len(detail_options):,}")
    else:
        detail_options = detail_df["selector"].tolist()

    if not detail_options:
        st.info("No hay equipos disponibles para mostrar en la vista actual.")
        st.stop()

    selected = st.selectbox(
        "Selecciona un equipo",
        options=detail_options,
        key="detail_equipment_selector",
    )
    row = detail_df.loc[detail_df["selector"] == selected].iloc[0]

    install_age = row.get("Age (years)")
    install_age_text = f"{float(install_age):.1f} años" if pd.notna(install_age) else "N/A"

    d1, d2, d3, d4, d5 = st.columns(5)
    with d1:
        metric_card("Serial", safe_text(row.get("Serial number")), safe_text(row.get("Instrument type"), ""))
    with d2:
        metric_card("Cliente / hospital", safe_text(row.get("Customer name")), safe_text(row.get("City"), ""))
    with d3:
        metric_card("Estado operativo", safe_text(row.get("Operational status")), safe_text(row.get("Asset condition"), ""))
    with d4:
        metric_card("Operating System", safe_text(row.get("Operating System")), safe_text(row.get("Country"), ""))
    with d5:
        metric_card("Antigüedad", install_age_text, "Desde fecha de instalación")

    st.markdown("### Resumen ejecutivo del equipo")
    executive_columns = [
        "Commercial Region",
        "Country",
        "Distributor name",
        "Customer name",
        "City",
        "Address",
        "Instrument type",
        "Product Line",
        "Serial number",
        "Installation date",
        "Age (years)",
        "Operational status grouped",
        "Operational status",
        "Asset condition",
        "Type of contract",
        "Contract duration",
        "Operating System",
        "Operating System Raw",
        "Number of tests per day",
        "In Blood Bank",
        "Data completeness %",
    ]

    executive_rows = []
    for c in executive_columns:
        if c not in detail_df.columns:
            continue
        value = row.get(c)
        if "date" in c.lower():
            value = format_date_for_hover(value)
        elif c in {"Age (years)", "Data completeness %", "Number of tests per day", "Contract duration"}:
            value = safe_number_text(value, "N/A")
        else:
            value = safe_text(value, "N/A")
        executive_rows.append({"Campo": c, "Valor": value})
    st.dataframe(pd.DataFrame(executive_rows), use_container_width=True, hide_index=True)

    st.markdown("### Mantenimiento preventivo / PM")
    pm_columns = [
        "PM plan",
        "PM last date",
        "PM frequency",
        "PM next date",
        "PM performed On",
    ]
    pm_rows = []
    for c in pm_columns:
        if c not in detail_df.columns:
            continue
        value = row.get(c)
        if "date" in c.lower() or c == "PM performed On":
            value = format_date_for_hover(value)
        elif c == "PM frequency":
            value = safe_number_text(value, "N/A")
        else:
            value = safe_text(value, "N/A")
        pm_rows.append({"Campo PM": c, "Valor": value})

    pm_next = pd.to_datetime(row.get("PM next date"), errors="coerce")
    if pd.notna(pm_next):
        days_to_pm = int((pm_next.normalize() - pd.Timestamp.today().normalize()).days)
        pm_status = "Vencido" if days_to_pm < 0 else ("Próximos 90 días" if days_to_pm <= 90 else "Planificado")
        pm_rows.append({"Campo PM": "PM status calculated", "Valor": pm_status})
        pm_rows.append({"Campo PM": "Days to next PM", "Valor": str(days_to_pm)})

    if pm_rows:
        st.dataframe(pd.DataFrame(pm_rows), use_container_width=True, hide_index=True)
    else:
        st.info("Este registro no contiene campos de mantenimiento preventivo.")

    st.markdown("### Machine configuration del equipo")
    applicable_row_fields = []
    for key in active_config_fields(detail_df.loc[[row.name]], CONFIG_KEYS):
        col = f"CFG::{key}"
        applicable_row_fields.append({"Campo": key, "Valor": safe_text(row.get(col), "N/A")})

    if applicable_row_fields:
        st.dataframe(pd.DataFrame(applicable_row_fields), use_container_width=True, hide_index=True)
    else:
        st.info("No hay campos estructurados de machine configuration para este equipo.")

    with st.expander("Machine configurations completas", expanded=False):
        st.code(safe_text(row.get("Machine Configurations"), "No disponible"))

    st.markdown("### Información completa del registro")
    hidden_prefixes = ("FLAG::",)
    all_rows = []
    for c in detail_df.columns:
        if c in {"selector", "_source_index"} or any(str(c).startswith(prefix) for prefix in hidden_prefixes):
            continue
        value = row.get(c)
        if "date" in str(c).lower():
            value = format_date_for_hover(value)
        else:
            value = safe_text(value, "N/A")
        all_rows.append({"Campo": c, "Valor": value})
    st.dataframe(pd.DataFrame(all_rows), use_container_width=True, hide_index=True)

    st.download_button(
        "Descargar detalle completo del equipo en Excel",
        data=build_equipment_detail_excel(pd.DataFrame(all_rows), normalize_serial_match(row.get('Serial number')) or 'selected'),
        file_name=f"equipment_detail_{normalize_serial_match(row.get('Serial number')) or 'selected'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime=EXCEL_MIME,
        use_container_width=False,
        key="download_selected_equipment_detail_excel",
    )


with st.sidebar:
    st.subheader("📄 Informe PDF")
    st.caption("Formato ajustado con márgenes APA de 1 pulgada y estructura de informe técnico ejecutiva.")
    pdf_title = st.text_input("Título del informe", value="Installed Base Dashboard Report")
    pdf_author = st.text_input("Nombre para firma", value="Javier Avellaneda")
    pdf_role = st.text_input("Cargo / título", value="Service Leader | Export LATAM")
    pdf_signature_date = st.text_input("Fecha de firma", value=datetime.now().strftime("%Y-%m-%d"))
    pdf_references = st.text_area(
        "Referencias APA (una por línea, opcional)",
        value="",
        height=110,
        placeholder="Ejemplo:\nDiaSorin. (2026). Installed base export dashboard. Internal operational dataset.",
    )

    if not REPORTLAB_AVAILABLE:
        st.warning("La exportación PDF requiere reportlab en el entorno. Agrega `reportlab` a requirements.txt.")
    else:
        base_summary = build_filter_summary(
            selected_regions=selected_regions,
            selected_countries=selected_countries,
            selected_distributors=selected_distributors,
            selected_instruments=selected_instruments,
            selected_states=selected_states,
        )
        pdf_filter_summary = dict(base_summary) if isinstance(base_summary, dict) else {"Filters": str(base_summary)}

        pdf_report_df, pdf_report_source, pdf_using_manufacturing = resolve_pdf_report_dataframe(
            filtered,
            active_dashboard_tab,
            source_label_value=source_label,
        )
        pdf_filter_summary["Total records"] = f"{len(pdf_report_df):,}"
        pdf_filter_summary["Fuente activa PDF"] = pdf_report_source
        if pdf_using_manufacturing:
            pdf_filter_summary["Cruce de fabricación"] = "Incluido: el PDF usa Manufacturing age (years) y Manufacturing age bucket"
            st.caption("PDF listo para usar antigüedad por fabricación en lugar de antigüedad por instalación.")
        elif active_dashboard_tab == "Antigüedad / fabricación":
            st.caption("PDF en modo Records List estándar: no hay cruce de fabricación válido para los filtros actuales.")

        if st.button("🧾 Preparar informe PDF", use_container_width=True, key="prepare_pdf_report"):
            try:
                prepared_bytes = build_pdf_report(
                    filtered_df=pdf_report_df,
                    filter_summary=pdf_filter_summary,
                    report_title=pdf_title,
                    author_name=pdf_author,
                    author_role=pdf_role,
                    signature_date=pdf_signature_date,
                    references_text=pdf_references,
                    stock_context=st.session_state.get("pdf_stock_context", {"available": False}),
                )
                st.session_state["prepared_pdf_bytes"] = prepared_bytes
                st.session_state["prepared_pdf_name"] = f"dashboard_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                st.success("Informe PDF preparado correctamente.")
            except Exception as e:
                st.session_state.pop("prepared_pdf_bytes", None)
                st.error(f"No fue posible generar el PDF: {e}")

        if st.session_state.get("prepared_pdf_bytes") is not None:
            st.download_button(
                "⬇️ Descargar informe PDF (APA)",
                data=st.session_state["prepared_pdf_bytes"],
                file_name=st.session_state.get("prepared_pdf_name", f"dashboard_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"),
                mime="application/pdf",
                use_container_width=True,
                key="download_prepared_pdf",
            )

    st.markdown("---")
    # EXCEL BAJO DEMANDA — NO mover build_dashboard_excel_export fuera del botón.
    # OpenPyXL debe ejecutarse únicamente cuando el usuario pulse Preparar.
    st.subheader("📊 Informe Excel")
    st.caption(
        "Genera un libro con una portada ejecutiva estable, un explorador con listas desplegables, "
        "datos clave y datos completos. Los gráficos de portada siempre se muestran al abrir."
    )

    excel_sidebar_df, excel_sidebar_source, excel_using_manufacturing = resolve_excel_report_dataframe(
        filtered,
        active_dashboard_tab,
        source_label_value=source_label,
    )
    excel_sidebar_filter_summary = dict(filter_summary_for_panel or {})
    excel_sidebar_filter_summary["Total records"] = f"{len(excel_sidebar_df):,}"
    excel_sidebar_filter_summary["Fuente activa Excel"] = excel_sidebar_source
    if excel_using_manufacturing:
        excel_sidebar_filter_summary["Cruce de fabricación"] = "Incluido"
        st.caption("Excel preparado para usar antigüedad real por fecha de fabricación.")

    current_excel_signature = _excel_export_signature(
        excel_sidebar_df,
        excel_sidebar_filter_summary,
        active_dashboard_tab,
        excel_sidebar_source,
    )

    if st.button("📗 Preparar informe Excel", use_container_width=True, key="prepare_dashboard_excel_v50"):
        try:
            prepared_excel_bytes = build_dashboard_excel_export(
                excel_sidebar_df,
                excel_sidebar_filter_summary,
                active_dashboard_tab=active_dashboard_tab,
                source_label_value=excel_sidebar_source,
                stock_context=st.session_state.get("pdf_stock_context", {"available": False}),
            )
            st.session_state["prepared_dashboard_excel_bytes"] = prepared_excel_bytes
            st.session_state["prepared_dashboard_excel_signature"] = current_excel_signature
            st.session_state["prepared_dashboard_excel_name"] = (
                f"records_list_dashboard_professional_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            st.success("Informe Excel profesional preparado correctamente.")
        except Exception as e:
            st.session_state.pop("prepared_dashboard_excel_bytes", None)
            st.session_state.pop("prepared_dashboard_excel_signature", None)
            st.error(f"No fue posible generar el Excel: {e}")

    prepared_excel_is_current = (
        st.session_state.get("prepared_dashboard_excel_bytes") is not None
        and st.session_state.get("prepared_dashboard_excel_signature") == current_excel_signature
    )
    if prepared_excel_is_current:
        st.download_button(
            "⬇️ Descargar informe Excel",
            data=st.session_state["prepared_dashboard_excel_bytes"],
            file_name=st.session_state.get(
                "prepared_dashboard_excel_name",
                f"records_list_dashboard_professional_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            ),
            mime=EXCEL_MIME,
            use_container_width=True,
            key="download_prepared_dashboard_excel_v50",
        )
    elif st.session_state.get("prepared_dashboard_excel_bytes") is not None:
        st.info("Los filtros cambiaron. Prepara nuevamente el Excel para descargar la vista actual.")

st.markdown("---")
foot_l, foot_r = st.columns((0.75, 0.25))
with foot_l:
    st.markdown(
        '<div class="small-note">Filtros activos: región comercial, país, distribuidor, tipo de instrumento y estado operativo. Base instalada incluye análisis por ciudad. Sistema operativo prioriza la detección de equipos legacy que deben migrar a Windows 10. En stock, el dashboard intenta identificar automáticamente el distribuidor a partir del título del archivo cargado. La fuente activa de Records List conserva el último archivo que subas durante la sesión.</div>',
        unsafe_allow_html=True,
    )
with foot_r:
    # No construir Excel aquí. Streamlit evalúa ``data=...`` en cada rerun y una
    # llamada directa a openpyxl ralentiza todos los filtros, pestañas y gráficos.
    # El generador completo queda en el sidebar y se ejecuta solo bajo demanda.
    if prepared_excel_is_current:
        st.download_button(
            "⬇️ Descargar Excel preparado",
            data=st.session_state["prepared_dashboard_excel_bytes"],
            file_name=st.session_state.get(
                "prepared_dashboard_excel_name",
                f"records_list_dashboard_professional_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            ),
            mime=EXCEL_MIME,
            use_container_width=True,
            key="download_prepared_dashboard_excel_footer_v49",
        )
    else:
        st.caption("Prepara el informe Excel desde la barra lateral para habilitar la descarga.")
